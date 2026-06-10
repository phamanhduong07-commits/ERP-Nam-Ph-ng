"""Bổ sung HĐLĐ + Quan hệ gia đình từ file MSNV.xlsx.

Phải chạy SAU khi import nhân viên xong (`import_employees_msnv.py`).
Match employee theo ``ma_nv``. Skip row không tìm thấy NV trong DB.

Cấu trúc cột file MSNV (sheet 'DATA NHÂN SỰ 2026', header R2, data R3+):
- C1  MSNV
- C16 Mức lương đóng BHXH       → dùng làm luong_co_ban tham chiếu
- C17 Loại HĐ hiện tại           → 'HĐLĐ KXĐTH' / 'HĐLĐ XĐTH'
- C18 Thời hạn HDLD              → '12 tháng' / '36 tháng' / 0 / time
- C19 Từ ngày                    → ngay_hieu_luc + ngay_ky
- C20 Đến ngày kết thúc          → ngay_het_han
- C21 SỐ HĐ                      → so_hop_dong (UNIQUE — skip nếu 0)
- C43 Thông tin con nhỏ          → tên con (có thể multi-line)
- C44 Năm sinh con               → năm sinh (có thể '2019\\n2024' = 2 con)

Strategy:
- LaborContract: chỉ import row có ``so_hop_dong`` không rỗng/0. UPSERT theo
  (employee_id, so_hop_dong).
- FamilyRelation: import nếu có tên con HOẶC năm sinh con. moi_quan_he='Con'.
  Đối với row chỉ có năm (vd '2019\\n2024' = 2 năm) mà không có tên →
  tạo "Con NV {ma_nv} sinh {year}".

Usage:
    python scripts/import_contracts_family.py --src "C:/.../MSNV.xlsx" --dry-run
    python scripts/import_contracts_family.py --src "C:/.../MSNV.xlsx"
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime, time as dtime, timezone
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from app.database import SessionLocal
from app.models.hr import Employee, FamilyRelation, LaborContract


# ─── Helpers ───
def _parse_date(v) -> date | None:
    if v in (None, "", 0):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, dtime):
        return None
    s = str(v).strip()
    if not s or s == "0":
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_money(v) -> Decimal:
    if v in (None, "", 0):
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = re.sub(r"[^\d.,-]", "", str(v)).strip()
    if not s:
        return Decimal("0")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    try:
        return Decimal(s)
    except Exception:  # noqa: BLE001
        return Decimal("0")


def _norm_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, dtime):
        return None
    s = str(v).strip()
    if not s or s == "0":
        return None
    return s


def map_loai_hop_dong(s: str | None) -> str:
    if not s:
        return "xac_dinh_thoi_han"
    sl = s.lower()
    if "kxđth" in sl or "không xác định" in sl or "kxdth" in sl:
        return "khong_thoi_han"
    if "thử việc" in sl or "thu viec" in sl:
        return "thu_viec"
    if "khoán" in sl or "khoan" in sl:
        return "khoan_viec"
    return "xac_dinh_thoi_han"


def parse_years(v) -> list[int]:
    """Năm sinh con — có thể single int / multi-line string '2019\\n2024'."""
    if v in (None, "", 0):
        return []
    if isinstance(v, datetime):
        return [v.year]
    if isinstance(v, int):
        if 1900 < v < 2100:
            return [v]
        return []
    s = str(v).strip()
    years = []
    for part in re.split(r"[\n,;/]+", s):
        part = part.strip()
        if not part:
            continue
        m = re.search(r"(\d{4})", part)
        if m:
            y = int(m.group(1))
            if 1900 < y < 2100:
                years.append(y)
    return years


def parse_names(v) -> list[str]:
    """Tên con — có thể multi-line."""
    if not v:
        return []
    s = str(v).strip()
    return [p.strip() for p in re.split(r"[\n;]+", s) if p.strip()]


# ─── Main ───
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", required=True)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"❌ Không tìm thấy file: {src}")
        sys.exit(2)

    print(f"📂 Source: {src}")
    print(f"🔧 Mode  : {'DRY-RUN' if args.dry_run else 'COMMIT'}\n")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["DATA NHÂN SỰ 2026"]
    today = date.today()

    db = SessionLocal()
    try:
        # Cache employee lookup
        emp_map: dict[str, Employee] = {
            e.ma_nv: e for e in db.query(Employee).all()
        }

        # Cache existing contracts để skip duplicate
        existing_contract_keys: set[tuple[int, str]] = {
            (c.employee_id, c.so_hop_dong) for c in db.query(LaborContract).all()
        }

        # Cache existing family relations (tránh duplicate theo employee_id + ho_ten + nam_sinh)
        existing_family_keys: set[tuple[int, str, int | None]] = {
            (f.employee_id, (f.ho_ten or "").strip().lower(), f.nam_sinh)
            for f in db.query(FamilyRelation).all()
        }

        contract_added = 0
        contract_skipped = 0
        family_added = 0
        family_skipped = 0
        nv_not_found = 0
        errors: list[str] = []

        for r in range(3, ws.max_row + 1):
            ma_raw = ws.cell(r, 1).value
            if ma_raw in (None, ""):
                continue
            ma_nv = str(ma_raw).strip().zfill(5)
            emp = emp_map.get(ma_nv)
            if not emp:
                nv_not_found += 1
                continue

            # ─── LaborContract ───
            so_hd = _norm_str(ws.cell(r, 21).value)
            # Skip Excel formula errors / placeholder
            if so_hd and so_hd.upper() in ("#N/A", "#REF!", "#NAME?", "#VALUE!", "0"):
                so_hd = None
            if so_hd:
                # Check trùng so_hop_dong (DB + trong batch) — UNIQUE constraint
                if any(so_hd == existing_so for _, existing_so in existing_contract_keys):
                    contract_skipped += 1
                    continue
                key = (emp.id, so_hd)
                if key in existing_contract_keys:
                    contract_skipped += 1
                else:
                    try:
                        tu_ngay = _parse_date(ws.cell(r, 19).value)
                        den_ngay = _parse_date(ws.cell(r, 20).value)
                        luong = _parse_money(ws.cell(r, 16).value)
                        loai = map_loai_hop_dong(_norm_str(ws.cell(r, 17).value))

                        trang_thai = "hieu_luc"
                        if den_ngay and den_ngay < today:
                            trang_thai = "het_han"

                        c = LaborContract(
                            employee_id=emp.id,
                            so_hop_dong=so_hd,
                            loai_hop_dong=loai,
                            ngay_ky=tu_ngay or today,
                            ngay_hieu_luc=tu_ngay or today,
                            ngay_het_han=den_ngay,
                            luong_co_ban=luong,
                            phu_cap=Decimal("0"),
                            trang_thai=trang_thai,
                        )
                        db.add(c)
                        existing_contract_keys.add(key)
                        contract_added += 1
                    except Exception as exc:  # noqa: BLE001
                        errors.append(f"HĐ NV {ma_nv} ({so_hd}): {exc}")

            # ─── FamilyRelation (con) ───
            names = parse_names(ws.cell(r, 43).value)
            years = parse_years(ws.cell(r, 44).value)
            # Pair tên ↔ năm theo index. Nếu lệch số → bù None.
            if names or years:
                num = max(len(names), len(years))
                for i in range(num):
                    nm = names[i] if i < len(names) else f"Con {i + 1} - {emp.ho_ten}"
                    yr = years[i] if i < len(years) else None
                    key = (emp.id, nm.strip().lower(), yr)
                    if key in existing_family_keys:
                        family_skipped += 1
                        continue
                    try:
                        fr = FamilyRelation(
                            employee_id=emp.id,
                            ho_ten=nm,
                            nam_sinh=yr,
                            moi_quan_he="Con",
                            ghi_chu="Auto-import từ MSNV.xlsx",
                        )
                        db.add(fr)
                        existing_family_keys.add(key)
                        family_added += 1
                    except Exception as exc:  # noqa: BLE001
                        errors.append(f"Con NV {ma_nv} ({nm}): {exc}")

        if args.dry_run:
            db.rollback()
        else:
            db.commit()

        print("─" * 70)
        print(f"📄 HĐLĐ        : {contract_added} INSERT  ·  {contract_skipped} SKIP (đã có)")
        print(f"👨‍👩‍👧 Quan hệ gia đình: {family_added} INSERT  ·  {family_skipped} SKIP")
        if nv_not_found:
            print(f"⚠ NV không tìm thấy trong DB: {nv_not_found} row")
        if errors:
            print(f"\n⚠ Errors ({len(errors)}):")
            for e in errors[:10]:
                print(f"  - {e}")
            if len(errors) > 10:
                print(f"  ... ({len(errors) - 10} more)")
        if args.dry_run:
            print("\nDRY-RUN — Bỏ --dry-run để commit thật.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
