"""Import nhân viên từ file MSNV.xlsx (định dạng phẳng — 1 row 1 NV).

File source: ``C:\\Users\\khang\\Desktop\\MSNV.xlsx``
Sheet: ``DATA NHÂN SỰ 2026``
- R1: số thứ tự cột (1, 2, 3, ...)
- R2: header (MSNV, Họ và Tên, Giới tính, ...)
- R3+: data NV

Cấu trúc 80 cột, em chỉ import các trường có trong model Employee:
- C1   MSNV
- C2   Họ và Tên
- C4   Giới tính
- C5   Ngày sinh
- C7   Nơi sinh
- C8   Bộ phận
- C9   Tổ
- C10  Vị trí
- C11  Đơn vị công tác (cũng dùng làm pháp nhân nếu C14 trống)
- C12  Ngày vào làm việc
- C13  Tình trạng làm việc
- C14  Pháp nhân ký HĐ
- C16  Mức lương đóng BHXH
- C22  Địa chỉ thường trú
- C23  Địa chỉ tạm trú
- C27  CCCD gắn chíp
- C28  Ngày cấp CCCD gắn chíp
- C29  Nơi cấp CCCD gắn chíp
- C31  Số điện thoại mới (ưu tiên hơn C30)
- C30  Số điện thoại (fallback)
- C32  Email
- C33  Trình độ
- C34  Trường học
- C35  Chuyên ngành
- C38  Số sổ BHXH
- C39  STK VCB (ưu tiên)
- C40  STK ACB (fallback)
- C41  Ngân hàng
- C42  Chi nhánh

Dedup: nếu MSNV trùng (NV nghỉ rồi đi lại) → giữ row "Đang làm việc",
fallback row sau cùng.

Usage:
    python scripts/import_employees_msnv.py --src "C:/.../MSNV.xlsx" --dry-run
    python scripts/import_employees_msnv.py --src "C:/.../MSNV.xlsx" --wipe-old
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from datetime import date, datetime, time as dtime
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from app.database import SessionLocal
from app.models.hr import Department, Employee, Position
from app.models.master import PhapNhan


# ─── Helpers ───
def _slug(text: str, prefix: str = "", max_len: int = 18) -> str:
    if not text:
        return prefix
    s = unicodedata.normalize("NFD", text)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_").upper()
    return (prefix + s)[:max_len]


def _parse_date(v) -> date | None:
    if v in (None, "", 0):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, dtime):
        return None  # time(0,0) là placeholder "trống"
    s = str(v).strip()
    if not s or s == "0":
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_money(v) -> Decimal | None:
    if v in (None, "", 0):
        return None
    if isinstance(v, (int, float, Decimal)):
        d = Decimal(str(v))
        return d if d > 0 else None
    s = re.sub(r"[^\d.,-]", "", str(v)).strip()
    if not s:
        return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    elif "." in s and len(s.rsplit(".", 1)[-1]) > 2:
        s = s.replace(".", "")
    try:
        d = Decimal(s)
        return d if d > 0 else None
    except Exception:  # noqa: BLE001
        return None


def _norm_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, dtime):
        return None  # time placeholder
    s = str(v).strip()
    if not s or s == "0":
        return None
    # Bỏ dấu xuống dòng + nhiều khoảng trắng liên tiếp
    s = re.sub(r"\s+", " ", s)
    return s if s else None


def _norm_phone(v) -> str | None:
    """Normalize SĐT: bỏ dấu cách, dấu chấm, gạch ngang."""
    s = _norm_str(v)
    if not s:
        return None
    s = re.sub(r"[^\d+]", "", s)
    return s if s else None


def map_trang_thai(s: str | None) -> str:
    if not s:
        return "dang_lam"
    sl = s.lower()
    if "nghỉ việc" in sl or "đã nghỉ" in sl:
        return "da_nghi"
    if "tạm nghỉ" in sl or "tạm ngưng" in sl:
        return "tam_nghi"
    return "dang_lam"


def map_gender(s: str | None) -> str | None:
    if not s:
        return None
    sl = s.strip().lower()
    if sl in ("nam", "male", "m"):
        return "Nam"
    if sl in ("nữ", "nu", "female", "f"):
        return "Nữ"
    return s.strip()


# ─── Master data resolver ───
class Resolver:
    def __init__(self, db):
        self.db = db
        self.pn_cache: dict[str, PhapNhan] = {}
        self.dept_cache: dict[str, Department] = {}
        self.pos_cache: dict[str, Position] = {}
        self.created_pn = self.created_dept = self.created_pos = 0
        for p in db.query(PhapNhan).all():
            self.pn_cache[p.ten_phap_nhan.lower().strip()] = p
            if p.ten_viet_tat:
                self.pn_cache[p.ten_viet_tat.lower().strip()] = p
        for d in db.query(Department).all():
            self.dept_cache[d.ten_bo_phan.lower().strip()] = d
        for p in db.query(Position).all():
            self.pos_cache[p.ten_chuc_vu.lower().strip()] = p

    def _gen_unique_ma(self, model, col_name: str, base: str) -> str:
        ma = base
        suffix = 1
        while self.db.query(model).filter(getattr(model, col_name) == ma).first():
            ma = f"{base}{suffix}"
            suffix += 1
        return ma

    def get_phap_nhan(self, name: str | None) -> PhapNhan | None:
        if not name:
            return None
        key = name.strip().lower()
        if key in self.pn_cache:
            return self.pn_cache[key]
        ma = self._gen_unique_ma(PhapNhan, "ma_phap_nhan", _slug(name, max_len=20))
        pn = PhapNhan(ma_phap_nhan=ma, ten_phap_nhan=name.strip(), ten_viet_tat=name.strip()[:50])
        self.db.add(pn)
        self.db.flush()
        self.pn_cache[key] = pn
        self.created_pn += 1
        return pn

    def get_department(self, name: str | None) -> Department | None:
        if not name or name.strip() == "#REF!":
            return None
        key = name.strip().lower()
        if key in self.dept_cache:
            return self.dept_cache[key]
        ma = self._gen_unique_ma(Department, "ma_bo_phan", _slug(name, max_len=20))
        d = Department(ma_bo_phan=ma, ten_bo_phan=name.strip())
        self.db.add(d)
        self.db.flush()
        self.dept_cache[key] = d
        self.created_dept += 1
        return d

    def get_position(self, name: str | None) -> Position | None:
        if not name:
            return None
        key = name.strip().lower()
        if key in self.pos_cache:
            return self.pos_cache[key]
        ma = self._gen_unique_ma(Position, "ma_chuc_vu", _slug(name, max_len=20))
        p = Position(ma_chuc_vu=ma, ten_chuc_vu=name.strip())
        self.db.add(p)
        self.db.flush()
        self.pos_cache[key] = p
        self.created_pos += 1
        return p


# ─── Parse one row ───
def parse_row(ws, r: int) -> dict | None:
    def C(c):
        return ws.cell(r, c).value

    ho_ten = _norm_str(C(2))
    if not ho_ten:
        return None
    ma_nv_raw = C(1)
    ma_nv = str(ma_nv_raw).strip().zfill(5) if ma_nv_raw not in (None, "") else None

    # SĐT ưu tiên cột "Số điện thoại mới" (C31), fallback C30
    sdt = _norm_phone(C(31)) or _norm_phone(C(30))

    return {
        "ma_nv": ma_nv,
        "ho_ten": ho_ten,
        "gioi_tinh": map_gender(_norm_str(C(4))),
        "ngay_sinh": _parse_date(C(5)),
        "noi_sinh_tinh": _norm_str(C(7)),
        "bo_phan": _norm_str(C(8)),
        "to_nhom": _norm_str(C(9)),
        "chuc_vu": _norm_str(C(10)),
        "don_vi_ct": _norm_str(C(11)),
        "ngay_vao_lam": _parse_date(C(12)),
        "trang_thai_raw": _norm_str(C(13)),
        "phap_nhan_hd": _norm_str(C(14)) or _norm_str(C(11)),
        "muc_dong_bhxh": _parse_money(C(16)),
        "dia_chi_ho_khau": _norm_str(C(22)),
        "dia_chi_hien_tai": _norm_str(C(23)),
        "cccd": _norm_str(C(27)),
        "ngay_cap_cccd": _parse_date(C(28)),
        "noi_cap_cccd": _norm_str(C(29)),
        "so_dien_thoai": sdt,
        "email": _norm_str(C(32)),
        "trinh_do_hoc_van": _norm_str(C(33)),
        "truong_dao_tao": _norm_str(C(34)),
        "chuyen_nganh": _norm_str(C(35)),
        "so_so_bhxh": _norm_str(C(38)),
        "stk_vcb": _norm_str(C(39)),
        "stk_acb": _norm_str(C(40)),
        "ten_ngan_hang": _norm_str(C(41)),
        "chi_nhanh_ngan_hang": _norm_str(C(42)),
    }


# ─── Main ───
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--wipe-old", action="store_true",
        help="Xóa NV cũ không có MSNV trong file mới (giữ NV có user_id link)",
    )
    args = parser.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"❌ Không tìm thấy file: {src}")
        sys.exit(2)

    print(f"📂 Source : {src}")
    print(f"🔧 Mode   : {'DRY-RUN' if args.dry_run else 'COMMIT'}  wipe-old={args.wipe_old}\n")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["DATA NHÂN SỰ 2026"]

    # ─── Đọc + dedup ───
    rows_by_ma: dict[str, dict] = {}
    rows_no_ma: list[dict] = []
    skipped_no_name = 0

    for r in range(3, ws.max_row + 1):
        row = parse_row(ws, r)
        if not row:
            skipped_no_name += 1
            continue
        ma = row["ma_nv"]
        if not ma:
            rows_no_ma.append(row)
            continue
        # Dedup: ưu tiên row "Đang làm việc"
        existing = rows_by_ma.get(ma)
        if existing:
            new_active = (row["trang_thai_raw"] or "").lower().strip() == "đang làm việc"
            old_active = (existing["trang_thai_raw"] or "").lower().strip() == "đang làm việc"
            if new_active and not old_active:
                rows_by_ma[ma] = row
            elif new_active and old_active:
                # cả hai đều active → giữ row sau (latest)
                rows_by_ma[ma] = row
            # else: giữ cái cũ (đã active hoặc cả hai inactive)
        else:
            rows_by_ma[ma] = row

    print(f"📋 Tổng row có tên: {len(rows_by_ma) + len(rows_no_ma) + skipped_no_name - skipped_no_name}")
    print(f"  ↳ MSNV unique sau dedup : {len(rows_by_ma)}")
    print(f"  ↳ Thiếu MSNV            : {len(rows_no_ma)}")
    print()

    db = SessionLocal()
    try:
        resolver = Resolver(db)
        inserted = updated = wiped = 0
        cccd_skipped = 0
        errors: list[str] = []

        # CCCD owner map
        cccd_owner: dict[str, str] = {}
        for emp in db.query(Employee).filter(Employee.cccd.isnot(None)).all():
            if emp.cccd:
                cccd_owner[emp.cccd] = emp.ma_nv

        # ─── Wipe NV cũ không có trong file mới (nếu --wipe-old) ───
        if args.wipe_old:
            file_mas = set(rows_by_ma.keys())
            old_emps = db.query(Employee).filter(
                Employee.user_id.is_(None),
                ~Employee.ma_nv.in_(file_mas) if file_mas else True,
            ).all()
            for emp in old_emps:
                # Xoá thẳng các CCCD owner cache để khỏi xung đột
                if emp.cccd and cccd_owner.get(emp.cccd) == emp.ma_nv:
                    cccd_owner.pop(emp.cccd, None)
                db.delete(emp)
                wiped += 1
            db.flush()
            print(f"🗑  Wiped {wiped} NV cũ (không có MSNV trong file mới, chưa cấp tài khoản)\n")

        # ─── UPSERT ───
        for ma, data in rows_by_ma.items():
            try:
                # CCCD check
                cccd = data.get("cccd")
                if cccd:
                    owner = cccd_owner.get(cccd)
                    if owner and owner != ma:
                        errors.append(f"MSNV {ma} ({data['ho_ten']}): CCCD '{cccd}' trùng với NV {owner} — set NULL")
                        cccd = None
                        cccd_skipped += 1
                    else:
                        cccd_owner[cccd] = ma

                pn = resolver.get_phap_nhan(data.get("phap_nhan_hd"))
                dept = resolver.get_department(data.get("bo_phan"))
                pos = resolver.get_position(data.get("chuc_vu"))

                existing = db.query(Employee).filter(Employee.ma_nv == ma).first()
                action = "UPDATE" if existing else "INSERT"
                if existing:
                    emp = existing
                    updated += 1
                else:
                    emp = Employee(ma_nv=ma, ho_ten=data["ho_ten"])
                    db.add(emp)
                    inserted += 1

                emp.ho_ten = data["ho_ten"]
                if data.get("gioi_tinh"): emp.gioi_tinh = data["gioi_tinh"]
                if data.get("ngay_sinh"): emp.ngay_sinh = data["ngay_sinh"]
                if data.get("noi_sinh_tinh"): emp.noi_sinh_tinh = data["noi_sinh_tinh"]
                if data.get("ngay_vao_lam"): emp.ngay_vao_lam = data["ngay_vao_lam"]
                emp.trang_thai = map_trang_thai(data.get("trang_thai_raw"))
                if data.get("muc_dong_bhxh"): emp.muc_dong_bhxh = data["muc_dong_bhxh"]
                if data.get("dia_chi_ho_khau"): emp.dia_chi_ho_khau = data["dia_chi_ho_khau"]
                if data.get("dia_chi_hien_tai"): emp.dia_chi_hien_tai = data["dia_chi_hien_tai"]
                if cccd: emp.cccd = cccd
                if data.get("ngay_cap_cccd"): emp.ngay_cap = data["ngay_cap_cccd"]
                if data.get("noi_cap_cccd"): emp.noi_cap = data["noi_cap_cccd"]
                if data.get("so_dien_thoai"): emp.so_dien_thoai = data["so_dien_thoai"]
                if data.get("email"): emp.email = data["email"]
                if data.get("trinh_do_hoc_van"): emp.trinh_do_hoc_van = data["trinh_do_hoc_van"]
                if data.get("truong_dao_tao"): emp.truong_dao_tao = data["truong_dao_tao"]
                if data.get("chuyen_nganh"): emp.chuyen_nganh = data["chuyen_nganh"]
                if data.get("so_so_bhxh"): emp.so_so_bhxh = data["so_so_bhxh"]

                # STK: ưu tiên VCB, fallback ACB
                stk = data.get("stk_vcb") or data.get("stk_acb")
                if stk:
                    emp.so_tk_ngan_hang = stk
                    if not data.get("ten_ngan_hang"):
                        emp.ten_ngan_hang = "Vietcombank" if data.get("stk_vcb") else "ACB"
                if data.get("ten_ngan_hang"):
                    emp.ten_ngan_hang = data["ten_ngan_hang"]
                if data.get("chi_nhanh_ngan_hang"):
                    emp.chi_nhanh_ngan_hang = data["chi_nhanh_ngan_hang"]

                if pn: emp.phap_nhan_id = pn.id
                if dept: emp.bo_phan_id = dept.id
                if pos: emp.chuc_vu_id = pos.id

            except Exception as exc:  # noqa: BLE001
                errors.append(f"MSNV {ma}: {exc}")
                db.rollback()

        if args.dry_run:
            db.rollback()
        else:
            db.commit()

        print("─" * 70)
        print(f"📊 Master tạo mới : PhapNhan={resolver.created_pn}  Dept={resolver.created_dept}  Position={resolver.created_pos}")
        print(f"👥 Nhân viên      : {inserted} INSERT  ·  {updated} UPDATE  ·  {wiped} WIPED")
        if rows_no_ma:
            print(f"⚠ {len(rows_no_ma)} NV thiếu MSNV (chưa import): {[r['ho_ten'] for r in rows_no_ma[:5]]}{'...' if len(rows_no_ma) > 5 else ''}")
        if cccd_skipped:
            print(f"⚠ CCCD trùng (set NULL): {cccd_skipped}")
        if errors:
            print(f"\nWarnings ({len(errors)}):")
            for e in errors[:15]:
                print(f"  - {e}")
            if len(errors) > 15:
                print(f"  ... ({len(errors) - 15} more)")
        if args.dry_run:
            print("\nDRY-RUN — Bỏ --dry-run để commit thật.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
