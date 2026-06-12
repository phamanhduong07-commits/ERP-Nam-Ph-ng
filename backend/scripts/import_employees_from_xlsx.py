"""Import nhân viên từ file Excel Hồ sơ Nam Phương (file 194 NV).

Cấu trúc sheet chi tiết (20 rows × 10 cols) — file v2 (đã bỏ row Tên bí danh + SDT người thân):
- R3:  Mã NV (C2) | Họ đệm (C4) | Tên (C6) | Giới tính (C8)
- R4:  Ngày sinh (C2) | Quốc tịch (C4) | Dân tộc (C6) | Tôn giáo (C8)
- R5:  Nơi sinh (C2) | CCCD (C4) | Ngày cấp (C7) | Tình trạng HN (C9)
- R6:  Nơi cấp (C2) | Điện thoại (C5) | Email (C7)
- R7:  Địa chỉ TT (C2) | Địa chỉ tạm trú (C4)
- R9:  Trình độ HV (C2) | Chuyên ngành (C4) | Trường ĐT (C6)
- R11: Chức vụ (C2) | Bộ phận (C6) | Tổ/Nhóm (C9)
- R12: Đơn vị CT (C2) | Ngày vào làm (C5) | Thâm niên (C7) | Tình trạng (C9)
- R13: Pháp nhân HĐ (C2) | Loại HĐ (C5) | Lương BHXH (C9)
- R14: Từ ngày HĐ (C2) | Đến ngày HĐ (C4) | Số HĐ (C6)
- R16: Số sổ BHXH (C2) | MST TNCN (C4 or C5) | STK VCB (...)
- R17: STK ACB (C2) | Ngân hàng (C4) | Chi nhánh (C6)

Strategy:
- Auto-create master data (PhapNhan / Department / Position) nếu chưa có
- Mã NV "00000" → auto-gen TMP-XXX (NV chưa có mã chính thức)
- CCCD trùng → set NULL để tránh constraint, log warning
- UPSERT theo ma_nv (sau khi gen)

Usage:
    python scripts/import_employees_from_xlsx.py --src "C:/path/file.xlsx" --dry-run
    python scripts/import_employees_from_xlsx.py --src "C:/path/file.xlsx"
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from app.database import SessionLocal
from app.models.hr import Department, Employee, Position
from app.models.master import PhapNhan


# ─── Helpers ───
def _slug(text: str, prefix: str = "", max_len: int = 18) -> str:
    if not text:
        return prefix
    s = unicodedata.normalize("NFD", text)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_").upper()
    return (prefix + s)[:max_len]


def _parse_date(v) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_money(v) -> Decimal | None:
    """Parse số tiền VN format (12.000.000 = mười hai triệu)."""
    if v in (None, ""):
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = re.sub(r"[^\d.,-]", "", str(v)).strip()
    if not s:
        return None
    # VN format: '.' là thousand sep, ',' là decimal point
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        # Nhiều dấu chấm → đều là thousand sep
        s = s.replace(".", "")
    elif "." in s and len(s.rsplit(".", 1)[-1]) > 2:
        # 1 dấu . nhưng phần sau > 2 chữ số → thousand sep
        s = s.replace(".", "")
    try:
        return Decimal(s)
    except Exception:  # noqa: BLE001
        return None


def _norm_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# ─── Master data resolver ───
class Resolver:
    def __init__(self, db):
        self.db = db
        self.pn_cache: dict[str, PhapNhan] = {}
        self.dept_cache: dict[str, Department] = {}
        self.pos_cache: dict[str, Position] = {}
        self.created_pn = self.created_dept = self.created_pos = 0
        for p in db.query(PhapNhan).all():
            self.pn_cache[p.ten_phap_nhan.lower()] = p
            if p.ten_viet_tat:
                self.pn_cache[p.ten_viet_tat.lower()] = p
        for d in db.query(Department).all():
            self.dept_cache[d.ten_bo_phan.lower()] = d
        for p in db.query(Position).all():
            self.pos_cache[p.ten_chuc_vu.lower()] = p

    def _gen_unique_ma(self, model, col_name: str, base: str) -> str:
        ma = base
        suffix = 1
        while self.db.query(model).filter(getattr(model, col_name) == ma).first():
            ma = f"{base}{suffix}"
            suffix += 1
        return ma

    def get_phap_nhan(self, name: str | None) -> PhapNhan | None:
        if not name:
            return None
        key = name.strip().lower()
        if key in self.pn_cache:
            return self.pn_cache[key]
        ma = self._gen_unique_ma(PhapNhan, "ma_phap_nhan", _slug(name, max_len=20))
        pn = PhapNhan(ma_phap_nhan=ma, ten_phap_nhan=name.strip(), ten_viet_tat=name.strip()[:50])
        self.db.add(pn)
        self.db.flush()
        self.pn_cache[key] = pn
        self.created_pn += 1
        return pn

    def get_department(self, name: str | None) -> Department | None:
        if not name:
            return None
        key = name.strip().lower()
        if key in self.dept_cache:
            return self.dept_cache[key]
        ma = self._gen_unique_ma(Department, "ma_bo_phan", _slug(name, max_len=20))
        d = Department(ma_bo_phan=ma, ten_bo_phan=name.strip())
        self.db.add(d)
        self.db.flush()
        self.dept_cache[key] = d
        self.created_dept += 1
        return d

    def get_position(self, name: str | None) -> Position | None:
        if not name:
            return None
        key = name.strip().lower()
        if key in self.pos_cache:
            return self.pos_cache[key]
        ma = self._gen_unique_ma(Position, "ma_chuc_vu", _slug(name, max_len=20))
        p = Position(ma_chuc_vu=ma, ten_chuc_vu=name.strip())
        self.db.add(p)
        self.db.flush()
        self.pos_cache[key] = p
        self.created_pos += 1
        return p


# ─── Parse sheet (FILE V2 — 194 NV) ───
def parse_sheet(ws) -> dict:
    def C(r, c):
        return ws.cell(r, c).value

    return {
        # R3
        "ma_nv": _norm_str(C(3, 2)),
        "ho_dem": _norm_str(C(3, 4)),
        "ten": _norm_str(C(3, 6)),
        "gioi_tinh": _norm_str(C(3, 8)),
        # R4
        "ngay_sinh": _parse_date(C(4, 2)),
        "quoc_tich": _norm_str(C(4, 4)) or "Việt Nam",
        "dan_toc": _norm_str(C(4, 6)),
        "ton_giao": _norm_str(C(4, 8)),
        # R5 — nơi sinh C2 | cccd C4 | ngày cấp C7 | tình trạng HN C9
        "noi_sinh_tinh": _norm_str(C(5, 2)),
        "cccd": _norm_str(C(5, 4)),
        "ngay_cap": _parse_date(C(5, 7)),
        # R6 — nơi cấp C2 | điện thoại C6 | email C8
        "noi_cap": _norm_str(C(6, 2)),
        "so_dien_thoai": _norm_str(C(6, 6)),
        "email": _norm_str(C(6, 8)),
        # R7 — địa chỉ HK C2 | địa chỉ tạm trú C7
        "dia_chi_ho_khau": _norm_str(C(7, 2)),
        "dia_chi_hien_tai": _norm_str(C(7, 7)),
        # R9 — sơ yếu (trình độ C2 | chuyên ngành C4 | trường ĐT C8)
        "trinh_do_hoc_van": _norm_str(C(9, 2)),
        "chuyen_nganh": _norm_str(C(9, 4)),
        "truong_dao_tao": _norm_str(C(9, 8)),
        # R11 — chức vụ C2 | bộ phận C6 | tổ/nhóm C9
        "chuc_vu": _norm_str(C(11, 2)),
        "bo_phan": _norm_str(C(11, 6)),
        "to_nhom": _norm_str(C(11, 9)),
        # R12 — đơn vị CT C2 | ngày vào C5 | thâm niên C7 | tình trạng C9
        "don_vi_ct": _norm_str(C(12, 2)),
        "ngay_vao_lam": _parse_date(C(12, 5)),
        "trang_thai_raw": _norm_str(C(12, 9)),
        # R13 — pháp nhân C2 | loại HĐ C5 | thời hạn HĐ C7 | lương BHXH C9
        "phap_nhan_hd": _norm_str(C(13, 2)),
        "loai_hop_dong": _norm_str(C(13, 5)),
        "thoi_han_hop_dong": _norm_str(C(13, 7)),
        "muc_dong_bhxh": _parse_money(C(13, 9)),
        # R14 — từ ngày HĐ C2 | đến ngày HĐ C4 | số HĐ C6
        "ngay_bat_dau_hd": _parse_date(C(14, 2)),
        "ngay_ket_thuc_hd": _parse_date(C(14, 4)),
        "so_hop_dong": _norm_str(C(14, 6)),
        # R16 — số sổ BHXH C2 | MST TNCN C5 | STK VCB C8
        "so_so_bhxh": _norm_str(C(16, 2)),
        "mst_tncn": _norm_str(C(16, 5)),
        "stk_vcb": _norm_str(C(16, 8)),
        # R17 — STK ACB C2 | ngân hàng C4 | chi nhánh C6
        "stk_acb": _norm_str(C(17, 2)),
        "ten_ngan_hang": _norm_str(C(17, 4)),
        "chi_nhanh_ngan_hang": _norm_str(C(17, 6)),
    }


def map_trang_thai(s: str | None) -> str:
    if not s:
        return "dang_lam"
    sl = s.lower()
    if "nghỉ việc" in sl or "đã nghỉ" in sl:
        return "da_nghi"
    if "tạm nghỉ" in sl or "tạm ngưng" in sl:
        return "tam_nghi"
    return "dang_lam"


def map_gender(s: str | None) -> str | None:
    if not s:
        return None
    sl = s.strip().lower()
    if sl in ("nam", "male", "m"):
        return "Nam"
    if sl in ("nữ", "nu", "female", "f"):
        return "Nữ"
    return s.strip()


# ─── Main ───
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", required=True)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"❌ Không tìm thấy file: {src}")
        sys.exit(2)

    print(f"📂 Source : {src}")
    print(f"🔧 Mode   : {'DRY-RUN' if args.dry_run else 'COMMIT'}\n")

    wb = openpyxl.load_workbook(src, data_only=True)
    sheets = [s for s in wb.sheetnames if s != "DANH SÁCH"]
    print(f"📋 {len(sheets)} sheets to process\n")

    db = SessionLocal()
    try:
        resolver = Resolver(db)
        inserted = updated = 0
        cccd_skipped = 0
        ma_nv_generated = 0
        errors: list[str] = []

        # Track CCCD: map cccd → ma_nv (để biết CCCD đã thuộc NV nào)
        # Khi update chính NV đó thì KHÔNG coi là trùng.
        cccd_owner: dict[str, str] = {}
        for emp in db.query(Employee).filter(Employee.cccd.isnot(None)).all():
            if emp.cccd:
                cccd_owner[emp.cccd] = emp.ma_nv

        # Auto-gen mã cho NV chưa có (ma_nv == "00000")
        # KHÔNG skip existing TMP — vì thứ tự sheet trong file là cố định,
        # re-run sẽ map TMP-001 ↔ NV thứ 1 trong sheets → UPSERT đúng.
        tmp_counter = 1

        for sname in sheets:
            try:
                ws = wb[sname]
                data = parse_sheet(ws)

                ma_nv = data["ma_nv"]
                # NV chưa có mã chính thức → gen TMP-XXX (lấy tên từ sheet)
                if not ma_nv or ma_nv == "00000":
                    ma_nv = f"TMP-{tmp_counter:03d}"
                    tmp_counter += 1
                    ma_nv_generated += 1

                ho_ten = " ".join(filter(None, [data.get("ho_dem"), data.get("ten")])).strip()
                if not ho_ten:
                    # Lấy từ sheet name dạng "MaNV_HoTen"
                    parts = sname.split("_", 1)
                    if len(parts) > 1:
                        ho_ten = parts[1].strip()
                if not ho_ten:
                    errors.append(f"Sheet '{sname}': thiếu họ tên")
                    continue

                # Handle CCCD trùng — set NULL để tránh UNIQUE constraint
                # CCCD đã thuộc CHÍNH NV này (re-import) → không phải trùng.
                cccd = data.get("cccd")
                if cccd:
                    owner = cccd_owner.get(cccd)
                    if owner and owner != ma_nv:
                        errors.append(f"Sheet '{sname}' (ma={ma_nv}): CCCD '{cccd}' trùng với NV {owner} — set NULL")
                        cccd = None
                        cccd_skipped += 1
                    else:
                        cccd_owner[cccd] = ma_nv

                # Resolve master data
                pn = resolver.get_phap_nhan(data.get("phap_nhan_hd") or data.get("don_vi_ct"))
                dept = resolver.get_department(data.get("bo_phan"))
                pos = resolver.get_position(data.get("chuc_vu"))

                existing = db.query(Employee).filter(Employee.ma_nv == ma_nv).first()
                action = "UPDATE" if existing else "INSERT"

                if existing:
                    emp = existing
                    updated += 1
                else:
                    emp = Employee(ma_nv=ma_nv, ho_ten=ho_ten)
                    db.add(emp)
                    inserted += 1

                # Clean các giá trị NHIỄM LABEL từ lần import trước (mapping bị sai)
                LABEL_VALUES = {
                    "Điện thoại", "Email", "Địa chỉ tạm trú", "Trường ĐT",
                    "Chuyên ngành", "Bộ phận", "Tổ/Nhóm", "Đến ngày HĐ",
                    "Số HĐ", "STK VCB", "Ngân hàng", "Chi nhánh", "MST TNCN",
                }
                for fld in ("so_dien_thoai", "email", "dia_chi_hien_tai",
                            "truong_dao_tao", "chuyen_nganh",
                            "so_tk_ngan_hang", "ten_ngan_hang", "chi_nhanh_ngan_hang"):
                    if getattr(emp, fld, None) in LABEL_VALUES:
                        setattr(emp, fld, None)

                # Set fields (force overwrite — luôn ưu tiên data mới)
                emp.ho_ten = ho_ten
                if data.get("ho_dem"): emp.ho_dem = data["ho_dem"]
                if data.get("ten"): emp.ten = data["ten"]
                if map_gender(data.get("gioi_tinh")): emp.gioi_tinh = map_gender(data.get("gioi_tinh"))
                if data.get("ngay_sinh"): emp.ngay_sinh = data["ngay_sinh"]
                if data.get("quoc_tich"): emp.quoc_tich = data["quoc_tich"]
                if data.get("dan_toc"): emp.dan_toc = data["dan_toc"]
                if data.get("ton_giao"): emp.ton_giao = data["ton_giao"]
                if data.get("noi_sinh_tinh"): emp.noi_sinh_tinh = data["noi_sinh_tinh"]
                if cccd: emp.cccd = cccd
                if data.get("ngay_cap"): emp.ngay_cap = data["ngay_cap"]
                if data.get("noi_cap"): emp.noi_cap = data["noi_cap"]
                if data.get("so_dien_thoai"): emp.so_dien_thoai = data["so_dien_thoai"]
                if data.get("email"): emp.email = data["email"]
                if data.get("dia_chi_ho_khau"): emp.dia_chi_ho_khau = data["dia_chi_ho_khau"]
                if data.get("dia_chi_hien_tai"): emp.dia_chi_hien_tai = data["dia_chi_hien_tai"]
                if data.get("trinh_do_hoc_van"): emp.trinh_do_hoc_van = data["trinh_do_hoc_van"]
                if data.get("chuyen_nganh"): emp.chuyen_nganh = data["chuyen_nganh"]
                if data.get("truong_dao_tao"): emp.truong_dao_tao = data["truong_dao_tao"]
                if data.get("ngay_vao_lam"): emp.ngay_vao_lam = data["ngay_vao_lam"]
                emp.trang_thai = map_trang_thai(data.get("trang_thai_raw"))
                if data.get("so_so_bhxh"): emp.so_so_bhxh = data["so_so_bhxh"]
                if data.get("muc_dong_bhxh"): emp.muc_dong_bhxh = data["muc_dong_bhxh"]

                # STK: ưu tiên VCB, fallback ACB. Tự suy ten_ngan_hang nếu chưa có.
                stk = data.get("stk_vcb") or data.get("stk_acb")
                if stk:
                    emp.so_tk_ngan_hang = stk
                    if not data.get("ten_ngan_hang"):
                        emp.ten_ngan_hang = "Vietcombank" if data.get("stk_vcb") else "ACB"
                if data.get("ten_ngan_hang"):
                    emp.ten_ngan_hang = data["ten_ngan_hang"]
                if data.get("chi_nhanh_ngan_hang"):
                    emp.chi_nhanh_ngan_hang = data["chi_nhanh_ngan_hang"]

                if pn: emp.phap_nhan_id = pn.id
                if dept: emp.bo_phan_id = dept.id
                if pos: emp.chuc_vu_id = pos.id

                chuc_vu_str = (data.get("chuc_vu") or "")[:25]
                bo_phan_str = (data.get("bo_phan") or "")[:22]
                pn_str = (data.get("phap_nhan_hd") or "")[:15]
                print(f"  {action}  {ma_nv:<8} {ho_ten:<28} {chuc_vu_str:<25} {bo_phan_str:<22} {pn_str}")

            except Exception as exc:  # noqa: BLE001
                errors.append(f"Sheet '{sname}': {exc}")
                db.rollback()

        if args.dry_run:
            db.rollback()
        else:
            db.commit()

        print()
        print("─" * 70)
        print(f"📊 Master created: PhapNhan={resolver.created_pn}  Dept={resolver.created_dept}  Position={resolver.created_pos}")
        print(f"👥 Employees: {inserted} INSERT  ·  {updated} UPDATE")
        print(f"   - Mã NV auto-gen (TMP-XXX): {ma_nv_generated}")
        print(f"   - CCCD trùng (đã set NULL): {cccd_skipped}")
        if errors:
            print(f"\n⚠ Warnings ({len(errors)}):")
            for e in errors[:15]:
                print(f"  - {e}")
            if len(errors) > 15:
                print(f"  ... ({len(errors) - 15} more)")
        if args.dry_run:
            print("\nDRY-RUN — Bỏ --dry-run để commit thật.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
