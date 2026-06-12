"""Import Tổ (nhóm) từ cột C9 'Tổ' trong MSNV.xlsx.

Phải chạy SAU khi import nhân viên xong + migrate hr_teams.

Strategy:
- Mỗi NV có 1 ten_bp (C8) + ten_to (C9). Match Department theo tên (case-insensitive).
- Normalize ten_to: strip + lowercase first letter để dedup ('Thành phẩm' = 'thành phẩm').
- Tạo Team unique theo (bo_phan_id, normalized_ten_to). Lưu ten_to viết hoa đẹp nhất.
- Set Employee.to_id sau khi tạo.

Usage:
    python scripts/import_teams_msnv.py --src "C:/.../MSNV.xlsx" --dry-run
    python scripts/import_teams_msnv.py --src "C:/.../MSNV.xlsx"
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from app.database import SessionLocal
from app.models.hr import Department, Employee, Team


def _norm_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "0":
        return None
    return re.sub(r"\s+", " ", s)


def _norm_key(s: str) -> str:
    """Key dedup: lowercase + bỏ dấu để 'Thành phẩm' = 'thành phẩm' = 'tHÀNH phẩm'."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", required=True)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"❌ Không tìm thấy file: {src}")
        sys.exit(2)

    print(f"📂 Source: {src}")
    print(f"🔧 Mode  : {'DRY-RUN' if args.dry_run else 'COMMIT'}\n")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["DATA NHÂN SỰ 2026"]

    db = SessionLocal()
    try:
        # Map: tên BP (normalized) → Department object
        dept_by_key: dict[str, Department] = {}
        for d in db.query(Department).all():
            dept_by_key[_norm_key(d.ten_bo_phan)] = d

        # Existing teams: dict[(bo_phan_id, norm_ten_to)] = Team
        team_cache: dict[tuple[int | None, str], Team] = {}
        for t in db.query(Team).all():
            team_cache[(t.bo_phan_id, _norm_key(t.ten_to))] = t

        # Đếm để chọn variant đẹp nhất (capitalize) cho team name
        # variant_count[(bp_id, norm)] = {ten_variant: count}
        variant_count: dict[tuple[int | None, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))

        # First pass: gom tất cả tên tổ + đếm variant
        rows = []
        for r in range(3, ws.max_row + 1):
            ma_raw = ws.cell(r, 1).value
            ho_ten = _norm_str(ws.cell(r, 2).value)
            if not ho_ten or ma_raw in (None, ""):
                continue
            ma_nv = str(ma_raw).strip().zfill(5)
            ten_bp = _norm_str(ws.cell(r, 8).value)
            ten_to = _norm_str(ws.cell(r, 9).value)
            if not ten_to:
                continue
            dept = dept_by_key.get(_norm_key(ten_bp)) if ten_bp else None
            bp_id = dept.id if dept else None
            key = (bp_id, _norm_key(ten_to))
            variant_count[key][ten_to] += 1
            rows.append((ma_nv, bp_id, key, ten_to))

        # Tạo team mới + gắn employee
        created_teams = 0
        assigned_emps = 0
        skip_no_emp = 0
        errors: list[str] = []
        emp_map: dict[str, Employee] = {e.ma_nv: e for e in db.query(Employee).all()}

        for ma_nv, bp_id, key, ten_to_raw in rows:
            # Pick best variant (most common; if tie, prefer "Title Case" not "lowercase")
            variants = variant_count[key]
            best_variant = max(variants.items(), key=lambda x: (x[1], not x[0].islower()))[0]

            team = team_cache.get(key)
            if not team:
                team = Team(ten_to=best_variant, bo_phan_id=bp_id)
                db.add(team)
                try:
                    db.flush()
                    team_cache[key] = team
                    created_teams += 1
                except Exception as exc:  # noqa: BLE001
                    db.rollback()
                    errors.append(f"Tạo Team '{best_variant}' (bp={bp_id}): {exc}")
                    continue

            emp = emp_map.get(ma_nv)
            if not emp:
                skip_no_emp += 1
                continue
            if emp.to_id != team.id:
                emp.to_id = team.id
                assigned_emps += 1

        if args.dry_run:
            db.rollback()
        else:
            db.commit()

        print("─" * 70)
        print(f"📁 Tổ tạo mới        : {created_teams}")
        print(f"👤 NV gán vào tổ     : {assigned_emps}")
        if skip_no_emp:
            print(f"⚠ Skip NV không có trong DB: {skip_no_emp}")
        if errors:
            print(f"\n⚠ Errors ({len(errors)}):")
            for e in errors[:10]:
                print(f"  - {e}")

        # Summary: liệt kê các tổ + số NV
        if not args.dry_run:
            print("\n═══ Cơ cấu mới ═══")
            from sqlalchemy import func
            q = (
                db.query(Department.ten_bo_phan, Team.ten_to, func.count(Employee.id))
                .join(Team, Team.bo_phan_id == Department.id)
                .outerjoin(Employee, Employee.to_id == Team.id)
                .group_by(Department.ten_bo_phan, Team.ten_to)
                .order_by(Department.ten_bo_phan, Team.ten_to)
            )
            cur_bp = None
            for bp, to, cnt in q.all():
                if bp != cur_bp:
                    print(f"\n📁 {bp}")
                    cur_bp = bp
                print(f"  └─ {to} ({cnt} NV)")

        if args.dry_run:
            print("\nDRY-RUN — Bỏ --dry-run để commit thật.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
