"""
ExcelExportService — build styled .xlsx từ ExcelTemplate config + data.

Template config structure:
  column_config:  [{key, label, width}]
  header_config:  [{key, label}]  — info rows above table (e.g. so_phieu, ngay, ncc)
  footer_config:  {show_total, sum_columns, show_signatures, signatures}
  style_config:   {accent_color, alt_row_color, orientation, show_company_header, freeze_header}
"""
from __future__ import annotations
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers as xl_numbers
)
from openpyxl.utils import get_column_letter

_DEFAULT_ACCENT = "#1B5E20"
_DEFAULT_ALT = "#F1F8E9"


def _hex_to_rgb(hex_color: str) -> str:
    """Convert '#1B5E20' → 'FF1B5E20' (openpyxl ARGB format)."""
    h = hex_color.lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    return f"FF{h.upper()}"


def _thin_border() -> Border:
    thin = Side(style="thin", color="FFCCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build_xlsx(
    template: Any,       # ExcelTemplate ORM instance
    items: list[dict],   # rows of data
    meta: dict,          # doc-level info: so_phieu, ngay, ncc, kho, ...
    company_info: dict,  # ten, dia_chi, dien_thoai, ma_so_thue
) -> bytes:
    """Return xlsx bytes for the given template + data."""
    col_cfg: list[dict] = template.column_config or []
    hdr_cfg: list[dict] = template.header_config or []
    ftr_cfg: dict = template.footer_config or {}
    sty_cfg: dict = template.style_config or {}

    accent = _hex_to_rgb(sty_cfg.get("accent_color") or _DEFAULT_ACCENT)
    alt_fill_hex = sty_cfg.get("alt_row_color") or _DEFAULT_ALT
    alt_fill = PatternFill("solid", fgColor=_hex_to_rgb(alt_fill_hex)) if alt_fill_hex else None
    show_company = sty_cfg.get("show_company_header", True)
    freeze = sty_cfg.get("freeze_header", True)
    orientation = sty_cfg.get("orientation", "portrait")

    wb = Workbook()
    ws = wb.active
    ws.title = template.ten_mau[:31]

    # Page setup
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    total_cols = max(len(col_cfg), 1)
    row = 1

    # ── Company header ──────────────────────────────────────────────────────────
    if show_company and company_info.get("ten"):
        ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
        c = ws.cell(row=row, column=1, value=company_info["ten"])
        c.font = Font(name="Times New Roman", bold=True, size=13, color=accent)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        details = []
        if company_info.get("dia_chi"):
            details.append(f"Địa chỉ: {company_info['dia_chi']}")
        if company_info.get("dien_thoai"):
            details.append(f"ĐT: {company_info['dien_thoai']}")
        if company_info.get("ma_so_thue"):
            details.append(f"MST: {company_info['ma_so_thue']}")
        if details:
            ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
            c = ws.cell(row=row, column=1, value="  |  ".join(details))
            c.font = Font(name="Times New Roman", size=9, color="FF555555")
            c.alignment = Alignment(horizontal="left")
            ws.row_dimensions[row].height = 14
            row += 1

        row += 1  # blank

    # ── Document title ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    title_cell = ws.cell(row=row, column=1, value=template.ten_mau.upper())
    title_cell.font = Font(name="Times New Roman", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Info fields (header_config) ─────────────────────────────────────────────
    if hdr_cfg:
        mid = (total_cols + 1) // 2
        for i in range(0, len(hdr_cfg), 2):
            left = hdr_cfg[i]
            right = hdr_cfg[i + 1] if i + 1 < len(hdr_cfg) else None

            # Left field: cols 1..mid
            lbl_cell = ws.cell(row=row, column=1, value=f"{left['label']}:")
            lbl_cell.font = Font(name="Times New Roman", bold=True, size=10)
            val_cell = ws.cell(row=row, column=2, value=str(meta.get(left["key"], "") or ""))
            val_cell.font = Font(name="Times New Roman", size=10)

            if right and mid + 1 <= total_cols:
                rlbl = ws.cell(row=row, column=mid + 1, value=f"{right['label']}:")
                rlbl.font = Font(name="Times New Roman", bold=True, size=10)
                rval = ws.cell(row=row, column=mid + 2, value=str(meta.get(right["key"], "") or ""))
                rval.font = Font(name="Times New Roman", size=10)

            ws.row_dimensions[row].height = 16
            row += 1

        row += 1  # blank after info

    # ── Table header ─────────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor=accent)
    hdr_font = Font(name="Times New Roman", bold=True, size=10, color="FFFFFFFF")
    border = _thin_border()

    for col_idx, col in enumerate(col_cfg, 1):
        cell = ws.cell(row=row, column=col_idx, value=col.get("label", col["key"]))
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col.get("width", 15)
    ws.row_dimensions[row].height = 18
    table_header_row = row

    if freeze:
        ws.freeze_panes = f"A{row + 1}"

    row += 1

    # ── Data rows ────────────────────────────────────────────────────────────────
    num_fmt = '#,##0'
    sum_columns: set[str] = set(ftr_cfg.get("sum_columns") or [])
    col_totals: dict[str, float] = {k: 0.0 for k in sum_columns}

    data_font = Font(name="Times New Roman", size=10)
    for row_idx, item in enumerate(items):
        fill = alt_fill if (alt_fill and row_idx % 2 == 1) else None
        for col_idx, col in enumerate(col_cfg, 1):
            raw = item.get(col["key"])
            cell = ws.cell(row=row, column=col_idx)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill

            if raw is None:
                cell.value = ""
            elif isinstance(raw, (int, float)):
                cell.value = raw
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
                if col["key"] in sum_columns:
                    col_totals[col["key"]] += float(raw)
            else:
                cell.value = str(raw)
                cell.alignment = Alignment(horizontal="left", wrap_text=False)

        ws.row_dimensions[row].height = 16
        row += 1

    # ── Totals row ───────────────────────────────────────────────────────────────
    show_total = ftr_cfg.get("show_total", False)
    if show_total and col_cfg:
        total_fill = PatternFill("solid", fgColor="FFFFF3CD")
        total_font = Font(name="Times New Roman", bold=True, size=10)
        first_label_written = False
        for col_idx, col in enumerate(col_cfg, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            if col["key"] in sum_columns:
                cell.value = col_totals[col["key"]]
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
            elif not first_label_written:
                cell.value = "Tổng cộng:"
                cell.alignment = Alignment(horizontal="right")
                first_label_written = True
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Signatures ───────────────────────────────────────────────────────────────
    show_sigs = ftr_cfg.get("show_signatures", False)
    signatures: list[str] = ftr_cfg.get("signatures") or []
    if show_sigs and signatures:
        row += 2  # blank gap
        sig_count = len(signatures)
        cols_per_sig = max(1, total_cols // sig_count)
        sig_font = Font(name="Times New Roman", bold=True, size=10)
        sub_font = Font(name="Times New Roman", italic=True, size=9, color="FF777777")

        for i, sig in enumerate(signatures):
            start_col = i * cols_per_sig + 1
            end_col = start_col + cols_per_sig - 1 if i < sig_count - 1 else total_cols
            if start_col == end_col:
                c = ws.cell(row=row, column=start_col, value=sig)
            else:
                ws.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
                c = ws.cell(row=row, column=start_col, value=sig)
            c.font = sig_font
            c.alignment = Alignment(horizontal="center")

            if start_col == end_col:
                c2 = ws.cell(row=row + 1, column=start_col, value="(Ký, họ tên)")
            else:
                ws.merge_cells(f"{get_column_letter(start_col)}{row+1}:{get_column_letter(end_col)}{row+1}")
                c2 = ws.cell(row=row + 1, column=start_col, value="(Ký, họ tên)")
            c2.font = sub_font
            c2.alignment = Alignment(horizontal="center")

        # Name blank line 4 rows below
        ws.row_dimensions[row + 4].height = 20

    # ── Auto-filter on table header ──────────────────────────────────────────────
    if col_cfg:
        ws.auto_filter.ref = (
            f"A{table_header_row}:{get_column_letter(len(col_cfg))}{table_header_row}"
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
