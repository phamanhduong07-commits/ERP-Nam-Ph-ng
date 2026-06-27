"""Warehouse router — báo cáo tồn kho & lịch sử giao dịch.

Split out of app/routers/warehouse.py (pure structural extraction).
Shares the /api/warehouse prefix; mounted alongside warehouse.router.
"""
import io
from datetime import date
from decimal import Decimal
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, exists, func, or_
from sqlalchemy.orm import Session, aliased, joinedload
from app.database import get_db
from app.deps import get_current_user, get_sale_visible_nv_ids
from app.models.auth import User
from app.models.inventory import InventoryBalance, InventoryTransaction
from app.models.master import Warehouse, PaperMaterial, OtherMaterial, Product, PhanXuong, PhapNhan, Supplier, Customer, CustomerNhanVien
from app.models.purchase import PurchaseOrder, PurchaseOrderItem
from app.models.production import ProductionOrder, ProductionOrderItem
from app.models.bom import ProductionBOM, ProductionBOMItem
from app.models.sales import SalesOrder, SalesReturn, SalesReturnItem
from app.models.warehouse_doc import (
    GoodsReceipt, GoodsReceiptItem,
    GiayRoll,
    ProductionOutput, DeliveryOrder,
    DeliveryOrderItem,
)
from app.services.carton_metrics import production_item_metrics
from app.utils.log import get_logger

logger = get_logger(__name__)

router = APIRouter(prefix="/api/warehouse", tags=["warehouse"])

_SALE_STAFF_ROLES = {"SALE_ADMIN", "KINH_DOANH_NHAN_VIEN"}


@router.get("/giao-dich")
def get_giao_dich(
    warehouse_id: Optional[int] = Query(None),
    phan_xuong_id: Optional[int] = Query(None),
    phap_nhan_id: Optional[int] = Query(None),
    paper_material_id: Optional[int] = Query(None),
    other_material_id: Optional[int] = Query(None),
    product_id: Optional[int] = Query(None),
    loai_giao_dich: Optional[str] = Query(None),
    tu_ngay: Optional[date] = Query(None),
    den_ngay: Optional[date] = Query(None),
    limit: int = Query(200, le=1000),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(InventoryTransaction)
    if phan_xuong_id or phap_nhan_id:
        q = q.join(Warehouse, Warehouse.id == InventoryTransaction.warehouse_id)
    if warehouse_id:
        q = q.filter(InventoryTransaction.warehouse_id == warehouse_id)
    if phan_xuong_id:
        q = q.filter(Warehouse.phan_xuong_id == phan_xuong_id)
    if phap_nhan_id:
        q = q.join(PhanXuong, Warehouse.phan_xuong_id == PhanXuong.id).filter(PhanXuong.phap_nhan_id == phap_nhan_id)
    if paper_material_id:
        q = q.filter(InventoryTransaction.paper_material_id == paper_material_id)
    if other_material_id:
        q = q.filter(InventoryTransaction.other_material_id == other_material_id)
    if product_id:
        q = q.filter(InventoryTransaction.product_id == product_id)
    if loai_giao_dich:
        q = q.filter(InventoryTransaction.loai_giao_dich == loai_giao_dich)
    if tu_ngay:
        q = q.filter(InventoryTransaction.ngay_giao_dich >= tu_ngay)
    if den_ngay:
        q = q.filter(InventoryTransaction.ngay_giao_dich <= den_ngay)
    rows = q.order_by(InventoryTransaction.ngay_giao_dich.asc(), InventoryTransaction.id.asc()).limit(limit).all()
    # Batch-load names
    wh_ids = {r.warehouse_id for r in rows}
    pm_ids = {r.paper_material_id for r in rows if r.paper_material_id}
    om_ids = {r.other_material_id for r in rows if r.other_material_id}
    pd_ids = {r.product_id for r in rows if r.product_id}
    wh_rows = (
        db.query(Warehouse)
        .options(joinedload(Warehouse.phan_xuong_obj).joinedload(PhanXuong.phap_nhan))
        .filter(Warehouse.id.in_(wh_ids))
        .all()
    ) if wh_ids else []
    wh_map = {w.id: w for w in wh_rows}
    pm_map = {p.id: (p.ma_chinh, p.ten) for p in db.query(PaperMaterial).filter(PaperMaterial.id.in_(pm_ids)).all()} if pm_ids else {}
    om_map = {p.id: (p.ma_chinh, p.ten) for p in db.query(OtherMaterial).filter(OtherMaterial.id.in_(om_ids)).all()} if om_ids else {}
    pd_map = {p.id: (p.ma_hang or p.ma_amis, p.ten_hang) for p in db.query(Product).filter(Product.id.in_(pd_ids)).all()} if pd_ids else {}

    def _name(r) -> tuple[str, str]:
        if r.paper_material_id and r.paper_material_id in pm_map:
            return pm_map[r.paper_material_id]
        if r.other_material_id and r.other_material_id in om_map:
            return om_map[r.other_material_id]
        if r.product_id and r.product_id in pd_map:
            return pd_map[r.product_id]
        return ("", "")

    return [{
        "id": r.id,
        "ngay_giao_dich": r.ngay_giao_dich.isoformat() if r.ngay_giao_dich else None,
        "warehouse_id": r.warehouse_id,
        "ten_kho": wh_map.get(r.warehouse_id).ten_kho if wh_map.get(r.warehouse_id) else "",
        "loai_kho": wh_map.get(r.warehouse_id).loai_kho if wh_map.get(r.warehouse_id) else None,
        "phan_xuong_id": wh_map.get(r.warehouse_id).phan_xuong_id if wh_map.get(r.warehouse_id) else None,
        "ten_phan_xuong": wh_map.get(r.warehouse_id).phan_xuong_obj.ten_xuong if wh_map.get(r.warehouse_id) and wh_map.get(r.warehouse_id).phan_xuong_obj else None,
        "phap_nhan_id": wh_map.get(r.warehouse_id).phan_xuong_obj.phap_nhan_id if wh_map.get(r.warehouse_id) and wh_map.get(r.warehouse_id).phan_xuong_obj else None,
        "ten_phap_nhan": (
            wh_map.get(r.warehouse_id).phan_xuong_obj.phap_nhan.ten_viet_tat
            or wh_map.get(r.warehouse_id).phan_xuong_obj.phap_nhan.ten_phap_nhan
        ) if wh_map.get(r.warehouse_id) and wh_map.get(r.warehouse_id).phan_xuong_obj and wh_map.get(r.warehouse_id).phan_xuong_obj.phap_nhan else None,
        "paper_material_id": r.paper_material_id,
        "other_material_id": r.other_material_id,
        "product_id": r.product_id,
        "ma_hang": _name(r)[0],
        "ten_hang": _name(r)[1],
        "loai_giao_dich": r.loai_giao_dich,
        "so_luong": float(r.so_luong),
        "don_gia": float(r.don_gia),
        "gia_tri": float(r.gia_tri),
        "ton_sau_giao_dich": float(r.ton_sau_giao_dich),
        "chung_tu_loai": r.chung_tu_loai,
        "chung_tu_id": r.chung_tu_id,
        "ghi_chu": r.ghi_chu,
    } for r in rows]


@router.get("/ton-kho-tp-lsx")
def ton_kho_tp_lsx(
    ten_khach: Optional[str] = Query(default=None),
    so_lenh: Optional[str] = Query(default=None),
    ten_hang: Optional[str] = Query(default=None),
    nv_theo_doi_id: Optional[int] = Query(default=None),
    tu_ngay: Optional[str] = Query(default=None),
    den_ngay: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Tổng hợp tồn kho thành phẩm theo từng LSX."""
    role_code = current_user.role.ma_vai_tro if current_user.role else None
    # 1. Nhập TP: sum(so_luong_nhap) per production_order_id
    # Không dùng MAX(warehouse_id) vì 1 LSX có thể nằm ở 2+ kho sau khi chuyển kho
    nhap_rows = (
        db.query(
            ProductionOutput.production_order_id,
            func.coalesce(func.sum(ProductionOutput.so_luong_nhap), 0).label("tong_nhap"),
            func.max(ProductionOutput.dvt).label("dvt"),
            func.min(ProductionOutput.ngay_nhap).label("ngay_nhap_kho"),
        )
        .group_by(ProductionOutput.production_order_id)
        .all()
    )
    if not nhap_rows:
        return []

    order_ids = [r.production_order_id for r in nhap_rows]
    nhap_map = {
        r.production_order_id: {
            "tong_nhap": float(r.tong_nhap),
            "dvt": r.dvt or "Thùng",
            "ngay_nhap_kho": r.ngay_nhap_kho.isoformat() if r.ngay_nhap_kho else None,
        }
        for r in nhap_rows
    }

    # 2. Xuất TP: sum(so_luong) per production_order_id từ DeliveryOrderItem
    xuat_rows = (
        db.query(
            DeliveryOrderItem.production_order_id,
            func.coalesce(func.sum(DeliveryOrderItem.so_luong), 0).label("tong_xuat"),
        )
        .join(DeliveryOrder, DeliveryOrder.id == DeliveryOrderItem.delivery_id)
        .filter(DeliveryOrderItem.production_order_id.in_(order_ids))
        .filter(DeliveryOrder.trang_thai != "huy")
        .group_by(DeliveryOrderItem.production_order_id)
        .all()
    )
    xuat_map = {r.production_order_id: float(r.tong_xuat) for r in xuat_rows}

    tra_rows_all = (
        db.query(
            DeliveryOrderItem.production_order_id,
            DeliveryOrder.warehouse_id,
            func.coalesce(func.sum(SalesReturnItem.so_luong_tra), 0).label("tong_tra"),
        )
        .select_from(SalesReturnItem)
        .join(SalesReturn, SalesReturn.id == SalesReturnItem.sales_return_id)
        .join(DeliveryOrder, DeliveryOrder.id == SalesReturn.delivery_order_id)
        .join(
            DeliveryOrderItem,
            or_(
                SalesReturnItem.delivery_order_item_id == DeliveryOrderItem.id,
                and_(
                    SalesReturnItem.delivery_order_item_id.is_(None),
                    DeliveryOrderItem.delivery_id == SalesReturn.delivery_order_id,
                    DeliveryOrderItem.sales_order_item_id == SalesReturnItem.sales_order_item_id,
                ),
            ),
        )
        .filter(
            DeliveryOrderItem.production_order_id.in_(order_ids),
            SalesReturn.trang_thai != "huy",
        )
        .group_by(DeliveryOrderItem.production_order_id, DeliveryOrder.warehouse_id)
        .all()
    )
    tra_map: dict[int, dict[int | None, float]] = {}
    for r in tra_rows_all:
        by_wh = tra_map.setdefault(r.production_order_id, {})
        by_wh[r.warehouse_id] = by_wh.get(r.warehouse_id, 0.0) + float(r.tong_tra)

    tra_rows_approved = (
        db.query(
            DeliveryOrderItem.production_order_id,
            DeliveryOrder.warehouse_id,
            func.coalesce(func.sum(SalesReturnItem.so_luong_tra), 0).label("tong_tra"),
        )
        .select_from(SalesReturnItem)
        .join(SalesReturn, SalesReturn.id == SalesReturnItem.sales_return_id)
        .join(DeliveryOrder, DeliveryOrder.id == SalesReturn.delivery_order_id)
        .join(
            DeliveryOrderItem,
            or_(
                SalesReturnItem.delivery_order_item_id == DeliveryOrderItem.id,
                and_(
                    SalesReturnItem.delivery_order_item_id.is_(None),
                    DeliveryOrderItem.delivery_id == SalesReturn.delivery_order_id,
                    DeliveryOrderItem.sales_order_item_id == SalesReturnItem.sales_order_item_id,
                ),
            ),
        )
        .filter(
            DeliveryOrderItem.production_order_id.in_(order_ids),
            SalesReturn.trang_thai == "da_duyet",
        )
        .group_by(DeliveryOrderItem.production_order_id, DeliveryOrder.warehouse_id)
        .all()
    )
    tra_approved_map: dict[int, dict[int | None, float]] = {}
    for r in tra_rows_approved:
        by_wh = tra_approved_map.setdefault(r.production_order_id, {})
        by_wh[r.warehouse_id] = by_wh.get(r.warehouse_id, 0.0) + float(r.tong_tra)

    # 3. Phiếu xuất gần nhất per LSX
    phieu_xuat_rows = (
        db.query(
            DeliveryOrderItem.production_order_id,
            func.max(DeliveryOrder.id).label("delivery_id"),
        )
        .join(DeliveryOrder, DeliveryOrder.id == DeliveryOrderItem.delivery_id)
        .filter(DeliveryOrderItem.production_order_id.in_(order_ids))
        .filter(DeliveryOrder.trang_thai != "huy")
        .group_by(DeliveryOrderItem.production_order_id)
        .all()
    )
    delivery_ids = [r.delivery_id for r in phieu_xuat_rows]
    deliveries_map = {}
    if delivery_ids:
        for d in db.query(DeliveryOrder).filter(DeliveryOrder.id.in_(delivery_ids)).all():
            deliveries_map[d.id] = d
    phieu_xuat_map: dict[int, dict] = {}
    for r in phieu_xuat_rows:
        d = deliveries_map.get(r.delivery_id)
        if d:
            phieu_xuat_map[r.production_order_id] = {
                "so_phieu": d.so_phieu,
                "ngay_xuat": d.ngay_xuat.isoformat() if d.ngay_xuat else None,
            }

    # 4. Item info per LSX (ten_hang + sl_ke_hoach)
    item_rows = (
        db.query(ProductionOrderItem)
        .options(
            joinedload(ProductionOrderItem.sales_order_item),
            joinedload(ProductionOrderItem.product),
        )
        .filter(ProductionOrderItem.production_order_id.in_(order_ids))
        .order_by(ProductionOrderItem.production_order_id, ProductionOrderItem.id)
        .all()
    )
    item_map: dict[int, ProductionOrderItem] = {}
    sl_ke_hoach_map: dict[int, float] = {}
    for item in item_rows:
        oid = item.production_order_id
        if oid not in item_map:
            item_map[oid] = item
        sl_ke_hoach_map[oid] = sl_ke_hoach_map.get(oid, 0.0) + float(item.so_luong_ke_hoach)

    # 5. ProductionOrder với các join
    orders = (
        db.query(ProductionOrder)
        .options(
            joinedload(ProductionOrder.sales_order).joinedload(SalesOrder.customer),
            joinedload(ProductionOrder.phap_nhan),
            joinedload(ProductionOrder.phan_xuong),
            joinedload(ProductionOrder.nv_theo_doi),
        )
        .filter(ProductionOrder.id.in_(order_ids))
        .all()
    )

    # Scope theo vai trò: SALE_ADMIN/SALE_ADMIN_NHAN_VIEN/KINH_DOANH_NHAN_VIEN chỉ thấy KH được phân công
    _scope_nv_ids = get_sale_visible_nv_ids(current_user)
    if _scope_nv_ids is not None:
        scoped_cids = {row.id for row in db.query(Customer.id).filter(
            or_(
                Customer.nv_phu_trach_id.in_(_scope_nv_ids),
                exists().where(
                    (CustomerNhanVien.customer_id == Customer.id)
                    & (CustomerNhanVien.user_id.in_(_scope_nv_ids))
                ),
            )
        ).all()}
        orders = [
            o for o in orders
            if o.sales_order and o.sales_order.customer_id in scoped_cids
        ]

    # 6. Metadata bổ sung (Pháp nhân + Kho hiện tại)
    pn_map = {p.id: p.ten_viet_tat for p in db.query(PhapNhan).all()}

    # Kho hiện tại: lấy distinct warehouse_id từ ProductionOutput per LSX
    kho_output_rows = (
        db.query(
            ProductionOutput.production_order_id,
            ProductionOutput.warehouse_id,
        )
        .filter(
            ProductionOutput.production_order_id.in_(order_ids),
            ProductionOutput.warehouse_id.isnot(None),
        )
        .distinct()
        .all()
    )
    kho_wh_ids = {r.warehouse_id for r in kho_output_rows if r.warehouse_id}
    kho_wh_map = {}
    if kho_wh_ids:
        for wh in db.query(Warehouse).filter(Warehouse.id.in_(kho_wh_ids)).all():
            kho_wh_map[wh.id] = wh.ten_kho
    # Group: production_order_id → list of warehouse names + first warehouse_id
    kho_by_order: dict[int, list[str]] = {}
    kho_id_by_order: dict[int, int] = {}
    for r in kho_output_rows:
        if r.warehouse_id and r.warehouse_id in kho_wh_map:
            kho_by_order.setdefault(r.production_order_id, [])
            name = kho_wh_map[r.warehouse_id]
            if name not in kho_by_order[r.production_order_id]:
                kho_by_order[r.production_order_id].append(name)
            if r.production_order_id not in kho_id_by_order:
                kho_id_by_order[r.production_order_id] = r.warehouse_id

    result = []
    for o in orders:
        nh = nhap_map.get(o.id, {})
        tong_nhap = nh.get("tong_nhap", 0.0)
        tong_xuat = xuat_map.get(o.id, 0.0)
        # Tổng trả = cộng tất cả kho (vì 1 LSX có thể trả từ nhiều kho)
        tong_tra = sum(tra_map.get(o.id, {}).values())
        tong_tra_da_duyet = sum(tra_approved_map.get(o.id, {}).values())
        kh = o.sales_order.customer if o.sales_order else None
        first_item = item_map.get(o.id)
        ten_khach_hang = kh.ten_viet_tat if kh else None
        ngay_lenh_str = o.ngay_lenh.isoformat() if o.ngay_lenh else None
        sales_item = first_item.sales_order_item if first_item and first_item.sales_order_item else None

        if ten_khach and (not ten_khach_hang or ten_khach.lower() not in ten_khach_hang.lower()):
            continue
        if so_lenh and so_lenh.lower() not in o.so_lenh.lower():
            continue
        if ten_hang and (not (first_item and first_item.ten_hang) or ten_hang.lower() not in first_item.ten_hang.lower()):
            continue
        if nv_theo_doi_id and o.nv_theo_doi_id != nv_theo_doi_id:
            continue
        if tu_ngay and (not ngay_lenh_str or ngay_lenh_str < tu_ngay):
            continue
        if den_ngay and (not ngay_lenh_str or ngay_lenh_str > den_ngay):
            continue

        ton_kho = tong_nhap - tong_xuat + tong_tra_da_duyet
        metrics = production_item_metrics(first_item, Decimal(str(ton_kho)))

        result.append({
            "production_order_id": o.id,
            "so_lenh": o.so_lenh,
            "ngay_lenh": ngay_lenh_str,
            "sales_order_id": o.sales_order_id,
            "so_don": o.sales_order.so_don if o.sales_order else None,
            "customer_id": kh.id if kh else None,
            "ten_hang": first_item.ten_hang if first_item else None,
            "product_id": first_item.product_id if first_item else None,
            "sales_order_item_id": first_item.sales_order_item_id if first_item else None,
            "don_gia": float(sales_item.don_gia) if sales_item and sales_item.don_gia else (
                float(first_item.gia_ban_muc_tieu) if first_item and first_item.gia_ban_muc_tieu else 0.0
            ),
            "dia_chi_giao": o.sales_order.dia_chi_giao if o.sales_order else None,
            "ten_khach_hang": ten_khach_hang,
            "nv_theo_doi_id": o.nv_theo_doi_id,
            "ten_nv_theo_doi": o.nv_theo_doi.ho_ten if o.nv_theo_doi else None,
            "sl_ke_hoach": sl_ke_hoach_map.get(o.id, 0.0),
            "tong_nhap": tong_nhap,
            "tong_xuat": tong_xuat,
            "tong_tra": tong_tra,
            "tong_tra_da_duyet": tong_tra_da_duyet,
            "tinh_trang_hang": "hang_tra_ve" if tong_tra > 0 else "binh_thuong",
            "ton_kho": ton_kho,
            "dien_tich": float(metrics["dien_tich"]),
            "trong_luong": float(metrics["trong_luong"]),
            "the_tich": float(metrics["the_tich"]),
            "dvt": nh.get("dvt", "Thùng"),
            "warehouse_id": kho_id_by_order.get(o.id),
            "ten_phan_xuong": None,
            "phan_xuong_id": o.phan_xuong_id,
            "order_ten_phan_xuong": o.phan_xuong.ten_xuong if o.phan_xuong else None,
            "phap_nhan_id": o.phap_nhan_id,
            "ten_phap_nhan_sx": pn_map.get(o.phap_nhan_id) or (o.phap_nhan.ten_viet_tat if o.phap_nhan else None),
            "ten_kho_hien_tai": ", ".join(kho_by_order.get(o.id, [])) or None,
            "ngay_nhap_kho": nh.get("ngay_nhap_kho"),
            "phieu_xuat_gan_nhat": phieu_xuat_map.get(o.id),
            "loai_thung": first_item.loai_thung if first_item else None,
            "kho_tt": float(first_item.kho_tt) if first_item and first_item.kho_tt else None,
            "dai_tt": float(first_item.dai_tt) if first_item and first_item.dai_tt else None,
            "so_lop": first_item.so_lop if first_item else None,
        })

    result.sort(key=lambda x: x["so_lenh"])
    return result


@router.get("/kho-loi-tra-ve")
def kho_loi_tra_ve(
    phap_nhan_id: Optional[int] = Query(default=None),
    phan_xuong_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Kho ảo: tổng hợp hàng lỗi + hàng trả về xấu, đầy đủ thông tin như kho thực."""
    from app.models.production import ProductionOrderItem

    role_code = current_user.role.ma_vai_tro if current_user.role else None
    pn_map = {p.id: p.ten_viet_tat for p in db.query(PhapNhan).all()}

    def _item_specs(item) -> dict:
        """Trích specs kỹ thuật từ ProductionOrderItem."""
        if not item:
            return {}
        quy_cach = None
        if item.dai and item.rong and item.cao:
            quy_cach = f"{int(item.dai)}×{int(item.rong)}×{int(item.cao)}"
        return {
            "loai_thung": item.loai_thung,
            "dai": float(item.dai) if item.dai else None,
            "rong": float(item.rong) if item.rong else None,
            "cao": float(item.cao) if item.cao else None,
            "so_lop": item.so_lop,
            "to_hop_song": item.to_hop_song,
            "kho_tt": float(item.kho_tt) if item.kho_tt else None,
            "dai_tt": float(item.dai_tt) if item.dai_tt else None,
            "loai_in": item.loai_in,
            "so_mau": item.so_mau,
            "quy_cach": quy_cach,
            "ngay_giao_hang": item.ngay_giao_hang.isoformat() if item.ngay_giao_hang else None,
        }

    # ── Hàng lỗi từ sản xuất ─────────────────────────────────────────────────
    loi_q = (
        db.query(ProductionOutput)
        .options(
            joinedload(ProductionOutput.production_order)
                .joinedload(ProductionOrder.phan_xuong),
            joinedload(ProductionOutput.production_order)
                .joinedload(ProductionOrder.phap_nhan),
            joinedload(ProductionOutput.production_order)
                .joinedload(ProductionOrder.sales_order)
                .joinedload(SalesOrder.customer),
            joinedload(ProductionOutput.production_order)
                .joinedload(ProductionOrder.nv_theo_doi),
            joinedload(ProductionOutput.production_order)
                .joinedload(ProductionOrder.items),
        )
        .filter(ProductionOutput.so_luong_loi > 0)
    )
    if phan_xuong_id:
        loi_q = loi_q.join(
            ProductionOrder, ProductionOrder.id == ProductionOutput.production_order_id
        ).filter(ProductionOrder.phan_xuong_id == phan_xuong_id)
    elif phap_nhan_id:
        loi_q = loi_q.join(
            ProductionOrder, ProductionOrder.id == ProductionOutput.production_order_id
        ).filter(ProductionOrder.phap_nhan_id == phap_nhan_id)

    _scope_nv_ids = get_sale_visible_nv_ids(current_user)
    if _scope_nv_ids is not None:
        scoped_cids = db.query(Customer.id).filter(
            or_(
                Customer.nv_phu_trach_id.in_(_scope_nv_ids),
                exists().where(
                    (CustomerNhanVien.customer_id == Customer.id)
                    & (CustomerNhanVien.user_id.in_(_scope_nv_ids))
                ),
            )
        )
        scoped_po_ids = (
            db.query(ProductionOrder.id)
            .join(SalesOrder, SalesOrder.id == ProductionOrder.sales_order_id)
            .filter(SalesOrder.customer_id.in_(scoped_cids))
        )
        loi_q = loi_q.filter(ProductionOutput.production_order_id.in_(scoped_po_ids))

    hang_loi = []
    for po in loi_q.all():
        order = po.production_order
        kh = order.sales_order.customer if order and order.sales_order else None
        so_don = order.sales_order.so_don if order and order.sales_order else None
        first_item = order.items[0] if order and order.items else None
        specs = _item_specs(first_item)
        hang_loi.append({
            "id": po.id,
            "so_phieu": po.so_phieu,
            "ngay_nhap": po.ngay_nhap.isoformat() if po.ngay_nhap else None,
            "production_order_id": order.id if order else None,
            "so_lenh": order.so_lenh if order else None,
            "ngay_lenh": order.ngay_lenh.isoformat() if order and order.ngay_lenh else None,
            "so_don": so_don,
            "ten_hang": po.ten_hang or (first_item.ten_hang if first_item else None),
            "so_luong_ke_hoach": float(first_item.so_luong_ke_hoach) if first_item else None,
            "so_luong_nhap": float(po.so_luong_nhap),
            "so_luong_loi": float(po.so_luong_loi),
            "dvt": po.dvt or "Thùng",
            "ten_khach_hang": kh.ten_viet_tat if kh else None,
            "dia_chi_giao": order.sales_order.dia_chi_giao if order and order.sales_order else None,
            "ten_nv_theo_doi": order.nv_theo_doi.ho_ten if order and order.nv_theo_doi else None,
            "ten_phan_xuong": order.phan_xuong.ten_xuong if order and order.phan_xuong else None,
            "ten_phap_nhan": pn_map.get(order.phap_nhan_id) if order else None,
            "ghi_chu": po.ghi_chu,
            **specs,
        })

    # ── Hàng trả về xấu (hong/loi) đã duyệt ─────────────────────────────────
    tra_q = (
        db.query(SalesReturnItem)
        .options(
            joinedload(SalesReturnItem.sales_return).joinedload(SalesReturn.customer),
        )
        .join(SalesReturn, SalesReturn.id == SalesReturnItem.sales_return_id)
        .filter(
            SalesReturnItem.tinh_trang_hang.in_(["hong", "loi"]),
            SalesReturn.trang_thai == "da_duyet",
        )
    )
    if phan_xuong_id or phap_nhan_id:
        tra_q = tra_q.join(
            DeliveryOrderItem,
            or_(
                SalesReturnItem.delivery_order_item_id == DeliveryOrderItem.id,
                and_(
                    SalesReturnItem.delivery_order_item_id.is_(None),
                    DeliveryOrderItem.delivery_id == SalesReturn.delivery_order_id,
                    DeliveryOrderItem.sales_order_item_id == SalesReturnItem.sales_order_item_id,
                ),
            ),
            isouter=True,
        ).join(ProductionOrder, ProductionOrder.id == DeliveryOrderItem.production_order_id, isouter=True)
        if phan_xuong_id:
            tra_q = tra_q.filter(ProductionOrder.phan_xuong_id == phan_xuong_id)
        if phap_nhan_id:
            tra_q = tra_q.filter(ProductionOrder.phap_nhan_id == phap_nhan_id)

    if _scope_nv_ids is not None:
        tra_q = tra_q.filter(SalesReturn.customer_id.in_(scoped_cids))

    # Lấy ProductionOrderItem qua sales_order_item_id để có specs
    all_tra = tra_q.all()
    soi_ids = [i.sales_order_item_id for i in all_tra if i.sales_order_item_id]
    poi_by_soi: dict[int, ProductionOrderItem] = {}
    if soi_ids:
        for poi in db.query(ProductionOrderItem).filter(
            ProductionOrderItem.sales_order_item_id.in_(soi_ids)
        ).all():
            if poi.sales_order_item_id not in poi_by_soi:
                poi_by_soi[poi.sales_order_item_id] = poi

    # Lấy ProductionOrder để biết pháp nhân/xưởng cho hàng trả
    poi_order_ids = [v.production_order_id for v in poi_by_soi.values()]
    order_by_id: dict[int, ProductionOrder] = {}
    if poi_order_ids:
        orders_q = (
            db.query(ProductionOrder)
            .options(
                joinedload(ProductionOrder.phan_xuong),
                joinedload(ProductionOrder.nv_theo_doi),
            )
            .filter(ProductionOrder.id.in_(poi_order_ids))
        )
        order_by_id = {o.id: o for o in orders_q.all()}

    hang_tra_ve = []
    for item in all_tra:
        sr = item.sales_return
        kh = sr.customer if sr else None
        poi = poi_by_soi.get(item.sales_order_item_id) if item.sales_order_item_id else None
        order = order_by_id.get(poi.production_order_id) if poi else None
        specs = _item_specs(poi)
        hang_tra_ve.append({
            "id": item.id,
            "so_phieu_tra": sr.so_phieu_tra if sr else None,
            "ngay_tra": sr.ngay_tra.isoformat() if sr and sr.ngay_tra else None,
            "sales_return_id": sr.id if sr else None,
            "ten_hang": poi.ten_hang if poi else None,
            "so_luong_tra": float(item.so_luong_tra),
            "don_gia_tra": float(item.don_gia_tra) if item.don_gia_tra else 0.0,
            "tinh_trang_hang": item.tinh_trang_hang,
            "dvt": "Thùng",
            "ten_khach_hang": kh.ten_viet_tat if kh else None,
            "ly_do_tra": item.ly_do_tra or (sr.ly_do_tra if sr else None),
            "ten_phan_xuong": order.phan_xuong.ten_xuong if order and order.phan_xuong else None,
            "ten_phap_nhan": pn_map.get(order.phap_nhan_id) if order else None,
            "ten_nv_theo_doi": order.nv_theo_doi.ho_ten if order and order.nv_theo_doi else None,
            "so_lenh": order.so_lenh if order else None,
            "ghi_chu": item.ghi_chu,
            **specs,
        })

    return {"hang_loi": hang_loi, "hang_tra_ve": hang_tra_ve}


# ── Tồn kho giấy cuộn ────────────────────────────────────────────────────────

@router.get("/ton-kho-giay")
def ton_kho_giay(
    phan_xuong_id: Optional[int] = None,
    phap_nhan_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Tồn kho giấy cuộn — tính trực tiếp từ GiayRoll.trong_luong_con_lai để luôn chính xác."""
    from sqlalchemy import func as sql_func

    # Tính tổng KG + số cuộn còn lại, group by warehouse + paper_material
    agg = (
        db.query(
            GiayRoll.warehouse_id,
            GiayRoll.paper_material_id,
            sql_func.sum(GiayRoll.trong_luong_con_lai).label("ton_luong"),
            sql_func.count(GiayRoll.id).label("so_cuon"),
        )
        .filter(
            GiayRoll.warehouse_id.isnot(None),
            GiayRoll.paper_material_id.isnot(None),
            GiayRoll.trang_thai.in_(["trong_kho", "dang_dung"]),
        )
        .group_by(GiayRoll.warehouse_id, GiayRoll.paper_material_id)
        .all()
    )

    if not agg:
        return []

    wh_ids = {r.warehouse_id for r in agg}
    pm_ids = {r.paper_material_id for r in agg}

    wh_map: dict[int, Warehouse] = {
        wh.id: wh for wh in db.query(Warehouse).filter(Warehouse.id.in_(wh_ids)).all()
    }
    pm_map: dict[int, PaperMaterial] = {
        pm.id: pm for pm in db.query(PaperMaterial).filter(PaperMaterial.id.in_(pm_ids)).all()
    }

    px_ids = {wh.phan_xuong_id for wh in wh_map.values() if wh.phan_xuong_id}
    px_map: dict[int, PhanXuong] = {
        px.id: px for px in db.query(PhanXuong).filter(PhanXuong.id.in_(px_ids)).all()
    } if px_ids else {}

    # Lấy giá nhập gần nhất cho mỗi paper_material (1 query, không N+1)
    don_gia_rows = (
        db.query(
            GoodsReceiptItem.paper_material_id,
            func.max(GoodsReceiptItem.id).label("max_id"),
        )
        .filter(GoodsReceiptItem.paper_material_id.in_(pm_ids))
        .group_by(GoodsReceiptItem.paper_material_id)
        .all()
    )
    last_item_ids = [r.max_id for r in don_gia_rows]
    don_gia_map: dict[int, float] = {}
    if last_item_ids:
        for item in db.query(GoodsReceiptItem).filter(GoodsReceiptItem.id.in_(last_item_ids)).all():
            if item.paper_material_id and item.don_gia:
                don_gia_map[item.paper_material_id] = float(item.don_gia)

    # Batch-load NSX names
    nsx_ids = {pm.ma_nsx_id for pm in pm_map.values() if pm.ma_nsx_id}
    sup_map: dict[int, str] = {
        s.id: (s.ten_viet_tat or "")
        for s in db.query(Supplier).filter(Supplier.id.in_(nsx_ids)).all()
    } if nsx_ids else {}

    # Batch-load bien_dong từ InventoryBalance (ton_luong - ton_luong_truoc)
    bien_dong_map: dict[int, float | None] = {}
    if pm_ids:
        bal_rows = (
            db.query(
                InventoryBalance.paper_material_id,
                func.sum(InventoryBalance.ton_luong).label("ton_now"),
                func.sum(InventoryBalance.ton_luong_truoc).label("ton_prev"),
            )
            .filter(InventoryBalance.paper_material_id.in_(pm_ids))
            .group_by(InventoryBalance.paper_material_id)
            .all()
        )
        for b in bal_rows:
            if b.ton_prev is not None:
                bien_dong_map[b.paper_material_id] = float(b.ton_now or 0) - float(b.ton_prev)

    # Batch-load ngay_nhap_gan_nhat
    nhap_date_map: dict[int, str | None] = {}
    if pm_ids:
        nhap_rows = (
            db.query(
                GoodsReceiptItem.paper_material_id,
                func.max(GoodsReceipt.ngay_nhap).label("max_ngay"),
            )
            .join(GoodsReceipt, GoodsReceiptItem.receipt_id == GoodsReceipt.id)
            .filter(
                GoodsReceiptItem.paper_material_id.in_(pm_ids),
                GoodsReceipt.trang_thai == "da_duyet",
            )
            .group_by(GoodsReceiptItem.paper_material_id)
            .all()
        )
        nhap_date_map = {
            r.paper_material_id: r.max_ngay.isoformat() if r.max_ngay else None
            for r in nhap_rows
        }

    result = []
    for row in agg:
        wh = wh_map.get(row.warehouse_id)
        if not wh:
            continue
        px = px_map.get(wh.phan_xuong_id) if wh.phan_xuong_id else None
        if phan_xuong_id and wh.phan_xuong_id != phan_xuong_id:
            continue
        if phap_nhan_id and (not px or px.phap_nhan_id != phap_nhan_id):
            continue
        pm = pm_map.get(row.paper_material_id)
        ton_luong = float(row.ton_luong or 0)
        don_gia = don_gia_map.get(row.paper_material_id, 0)
        result.append({
            "paper_material_id": row.paper_material_id,
            "ma_chinh": pm.ma_chinh if pm else None,
            "ma_ky_hieu": pm.ma_ky_hieu if pm else None,
            "ten": pm.ten if pm else None,
            "kho": float(pm.kho) if pm and pm.kho else None,
            "dinh_luong": int(pm.dinh_luong) if pm and pm.dinh_luong else None,
            "loai_giay": pm.loai_giay if pm else None,
            "ton_toi_thieu": float(pm.ton_toi_thieu) if pm and pm.ton_toi_thieu else 0,
            "warehouse_id": row.warehouse_id,
            "ten_kho": wh.ten_kho,
            "phan_xuong_id": wh.phan_xuong_id,
            "ten_phan_xuong": px.ten_xuong if px else None,
            "phap_nhan_id": px.phap_nhan_id if px else None,
            "ton_luong": ton_luong,
            "so_cuon": int(row.so_cuon or 0),
            "gia_tri_ton": ton_luong * don_gia,
            "don_gia_binh_quan": don_gia,
            "ten_nsx": sup_map.get(pm.ma_nsx_id) if pm and pm.ma_nsx_id else None,
            "bien_dong": bien_dong_map.get(row.paper_material_id),
            "ngay_nhap_gan_nhat": nhap_date_map.get(row.paper_material_id),
        })

    result.sort(key=lambda x: (x["ma_chinh"] or "", x["ten_kho"] or ""))
    return result


# ── Dự trù nhu cầu giấy theo tuần ───────────────────────────────────────────

@router.get("/du-tru-giay")
def du_tru_giay(
    weeks: int = Query(4, ge=1, le=8),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """
    Dự trù nhu cầu giấy cuộn theo rolling n tuần.
    So sánh với cùng kỳ năm trước (LSX đã hoàn thành).
    """
    from datetime import timedelta

    today = date.today()
    week_start = today - timedelta(days=today.weekday())  # Monday của tuần hiện tại
    week_periods = []
    for i in range(weeks):
        ws = week_start + timedelta(weeks=i)
        we = ws + timedelta(days=6)
        week_periods.append((ws, we))

    last_year_periods = [
        (ws - timedelta(weeks=52), we - timedelta(weeks=52))
        for ws, we in week_periods
    ]

    def _fetch_demand(trang_thai_list: list[str], date_ranges: list[tuple]) -> dict:
        """Returns {paper_material_id: [kg_w0, kg_w1, ...]} for given statuses and week ranges."""
        all_ws = date_ranges[0][0]
        all_we = date_ranges[-1][1]

        rows = (
            db.query(
                ProductionBOMItem.paper_material_id,
                func.coalesce(
                    ProductionOrderItem.ngay_giao_hang,
                    ProductionOrder.ngay_hoan_thanh_ke_hoach,
                ).label("ref_date"),
                func.sum(ProductionBOMItem.trong_luong_can_tong).label("can_kg"),
            )
            .join(ProductionBOM, ProductionBOM.id == ProductionBOMItem.bom_id)
            .join(ProductionOrderItem,
                  ProductionOrderItem.id == ProductionBOM.production_order_item_id)
            .join(ProductionOrder,
                  ProductionOrder.id == ProductionOrderItem.production_order_id)
            .filter(
                ProductionOrder.trang_thai.in_(trang_thai_list),
                ProductionBOMItem.paper_material_id.isnot(None),
                ProductionBOMItem.trong_luong_can_tong.isnot(None),
                func.coalesce(
                    ProductionOrderItem.ngay_giao_hang,
                    ProductionOrder.ngay_hoan_thanh_ke_hoach,
                ).between(all_ws, all_we),
            )
            .group_by(
                ProductionBOMItem.paper_material_id,
                func.coalesce(
                    ProductionOrderItem.ngay_giao_hang,
                    ProductionOrder.ngay_hoan_thanh_ke_hoach,
                ),
            )
            .all()
        )

        result: dict[int, list[float]] = {}
        for r in rows:
            if r.ref_date is None:
                continue
            ref = r.ref_date if isinstance(r.ref_date, date) else r.ref_date.date()
            for i, (ws, we) in enumerate(date_ranges):
                if ws <= ref <= we:
                    if r.paper_material_id not in result:
                        result[r.paper_material_id] = [0.0] * len(date_ranges)
                    result[r.paper_material_id][i] += float(r.can_kg or 0)
                    break
        return result

    # Nhu cầu từ LSX đang chạy (tuần hiện tại trở đi)
    demand_map = _fetch_demand(["moi", "da_duyet", "dang_sx"], week_periods)

    # Cùng kỳ năm trước (LSX đã hoàn thành)
    last_year_map = _fetch_demand(["hoan_thanh"], last_year_periods)

    # Tồn kho hiện tại tổng hợp theo paper_material_id
    stock_rows = (
        db.query(
            InventoryBalance.paper_material_id,
            func.sum(InventoryBalance.ton_luong).label("ton_luong"),
            func.sum(InventoryBalance.don_gia_binh_quan * InventoryBalance.ton_luong).label("weighted"),
        )
        .filter(InventoryBalance.paper_material_id.isnot(None))
        .group_by(InventoryBalance.paper_material_id)
        .all()
    )
    stock_map = {r.paper_material_id: float(r.ton_luong) for r in stock_rows}
    dpbq_map = {
        r.paper_material_id: (float(r.weighted) / float(r.ton_luong) if float(r.ton_luong) > 0 else 0)
        for r in stock_rows
    }

    all_pm_ids = set(demand_map) | set(stock_map)
    if not all_pm_ids:
        return []

    pm_map2: dict[int, PaperMaterial] = {
        pm.id: pm
        for pm in db.query(PaperMaterial).filter(PaperMaterial.id.in_(all_pm_ids)).all()
    }

    result2 = []
    for pm_id in sorted(all_pm_ids):
        pm = pm_map2.get(pm_id)
        ton_hien_tai = stock_map.get(pm_id, 0.0)
        dpbq = dpbq_map.get(pm_id, 0.0)
        ton_toi_thieu = float(pm.ton_toi_thieu) if pm and pm.ton_toi_thieu else 0.0
        week_demand = demand_map.get(pm_id, [0.0] * weeks)
        week_ly = last_year_map.get(pm_id, [0.0] * weeks)

        periods = []
        running = ton_hien_tai
        for i, (ws, we) in enumerate(week_periods):
            can_kg = week_demand[i] if i < len(week_demand) else 0.0
            ly_kg = week_ly[i] if i < len(week_ly) else 0.0
            running -= can_kg
            tang_giam_pct = (
                round((can_kg - ly_kg) / ly_kg * 100, 1) if ly_kg > 0 else None
            )
            periods.append({
                "label": f"Tuần {i + 1} ({ws.strftime('%d/%m')}–{we.strftime('%d/%m/%Y')})",
                "date_from": ws.isoformat(),
                "date_to": we.isoformat(),
                "can_kg": round(can_kg, 2),
                "ton_sau_ky": round(running, 2),
                "am": running < ton_toi_thieu,
                "cung_ky_nam_truoc_kg": round(ly_kg, 2),
                "tang_giam_pct": tang_giam_pct,
            })

        tong_can = sum(p["can_kg"] for p in periods)
        # Lượng cần mua = deficit tích lũy lớn nhất (đủ cho toàn kỳ kể cả buffer)
        min_running = min((p["ton_sau_ky"] for p in periods), default=ton_hien_tai)
        can_mua_ngay = max(0.0, round(ton_toi_thieu - min_running, 2))

        result2.append({
            "paper_material_id": pm_id,
            "ma_chinh": pm.ma_chinh if pm else None,
            "ten": pm.ten if pm else None,
            "kho": float(pm.kho) if pm and pm.kho else None,
            "dinh_luong": int(pm.dinh_luong) if pm and pm.dinh_luong else None,
            "ton_toi_thieu": ton_toi_thieu,
            "ton_hien_tai": round(ton_hien_tai, 2),
            "don_gia_binh_quan": round(dpbq, 0),
            "periods": periods,
            "tong_can_kg": round(tong_can, 2),
            "can_mua_ngay": can_mua_ngay,
            "gia_tri_can_mua": round(can_mua_ngay * dpbq, 0),
        })

    result2.sort(key=lambda x: x["ma_chinh"] or "")
    return result2


# ─────────────────────────────────────────────────────────────────────────────
# Phôi sóng mua ngoài: KHSX line đánh dấu mua_phoi_ngoai
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/khsx-can-phoi-ngoai")
def khsx_can_phoi_ngoai(
    trang_thai: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """
    Trả về danh sách KHSX line có mua_phoi_ngoai=True (mua phôi sóng từ NCC ngoài).
    Join ProductionOrderItem để lấy đầy đủ cấu trúc giấy + kích thước thùng + QCCL.
    Default chỉ lấy line chưa hoàn thành (cho, dang_chay).
    """
    from app.models.production_plan import ProductionPlan, ProductionPlanLine

    trang_thai_list = (trang_thai or "cho,dang_chay").split(",")

    rows = (
        db.query(
            ProductionPlanLine.id.label("ppl_id"),
            ProductionPlan.so_ke_hoach,
            ProductionPlan.ngay_ke_hoach,
            ProductionPlanLine.ngay_chay,
            ProductionPlanLine.so_luong_ke_hoach,
            ProductionPlanLine.kho1,
            ProductionPlanLine.kho_giay,
            ProductionPlanLine.so_dao,
            ProductionPlanLine.kho_tt,
            ProductionOrder.so_lenh,
            ProductionOrderItem.id.label("poi_id"),
            ProductionOrderItem.ten_hang,
            # Cấu trúc giấy:
            ProductionOrderItem.so_lop, ProductionOrderItem.to_hop_song,
            ProductionOrderItem.mat, ProductionOrderItem.mat_dl,
            ProductionOrderItem.song_1, ProductionOrderItem.song_1_dl,
            ProductionOrderItem.mat_1, ProductionOrderItem.mat_1_dl,
            ProductionOrderItem.song_2, ProductionOrderItem.song_2_dl,
            ProductionOrderItem.mat_2, ProductionOrderItem.mat_2_dl,
            ProductionOrderItem.song_3, ProductionOrderItem.song_3_dl,
            ProductionOrderItem.mat_3, ProductionOrderItem.mat_3_dl,
            # Kích thước thùng:
            ProductionOrderItem.loai_thung,
            ProductionOrderItem.dai, ProductionOrderItem.rong, ProductionOrderItem.cao,
            ProductionOrderItem.dai_tt,
            # QCCL:
            ProductionOrderItem.c_tham, ProductionOrderItem.can_man,
            ProductionOrderItem.loai_lan, ProductionOrderItem.qccl,
        )
        .join(ProductionPlan, ProductionPlan.id == ProductionPlanLine.plan_id)
        .join(ProductionOrderItem,
              ProductionOrderItem.id == ProductionPlanLine.production_order_item_id)
        .join(ProductionOrder,
              ProductionOrder.id == ProductionOrderItem.production_order_id)
        .filter(
            ProductionPlanLine.mua_phoi_ngoai == True,  # noqa: E712
            ProductionPlanLine.trang_thai.in_(trang_thai_list),
        )
        .order_by(ProductionPlanLine.ngay_chay)
        .all()
    )

    # Số tấm đã đặt — từ PurchaseOrderItem có production_plan_line_id
    ordered_map: dict[int, float] = {}
    if rows:
        ppl_ids = [r.ppl_id for r in rows]
        ordered = (
            db.query(
                PurchaseOrderItem.production_plan_line_id,
                func.sum(PurchaseOrderItem.so_luong).label("da_dat"),
            )
            .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderItem.po_id)
            .filter(
                PurchaseOrderItem.production_plan_line_id.in_(ppl_ids),
                PurchaseOrder.trang_thai.in_([
                    "da_duyet", "da_gui_ncc", "dang_giao", "hoan_thanh"
                ]),
            )
            .group_by(PurchaseOrderItem.production_plan_line_id)
            .all()
        )
        ordered_map = {r.production_plan_line_id: float(r.da_dat or 0) for r in ordered}

    def _f(v):
        return float(v) if v is not None else None

    result = [
        {
            "ppl_id": r.ppl_id,
            "so_ke_hoach": r.so_ke_hoach,
            "ngay_ke_hoach": str(r.ngay_ke_hoach) if r.ngay_ke_hoach else None,
            "ngay_chay": str(r.ngay_chay) if r.ngay_chay else None,
            "so_lsx": r.so_lenh,
            "poi_id": r.poi_id,
            "ten_san_pham": r.ten_hang or "",
            "so_luong_thung": float(r.so_luong_ke_hoach or 0),
            # KHSX paper sizing:
            "kho1": _f(r.kho1),
            "kho_giay": _f(r.kho_giay),
            "so_dao": r.so_dao,
            "kho_tt": _f(r.kho_tt),
            "dai_tt": _f(r.dai_tt),
            # Cấu trúc giấy:
            "so_lop": r.so_lop, "to_hop_song": r.to_hop_song,
            "mat": r.mat, "mat_dl": _f(r.mat_dl),
            "song_1": r.song_1, "song_1_dl": _f(r.song_1_dl),
            "mat_1": r.mat_1, "mat_1_dl": _f(r.mat_1_dl),
            "song_2": r.song_2, "song_2_dl": _f(r.song_2_dl),
            "mat_2": r.mat_2, "mat_2_dl": _f(r.mat_2_dl),
            "song_3": r.song_3, "song_3_dl": _f(r.song_3_dl),
            "mat_3": r.mat_3, "mat_3_dl": _f(r.mat_3_dl),
            # Kích thước thùng:
            "loai_thung": r.loai_thung,
            "dai": _f(r.dai), "rong": _f(r.rong), "cao": _f(r.cao),
            # QCCL:
            "c_tham": r.c_tham, "can_man": r.can_man,
            "loai_lan": r.loai_lan, "qccl": r.qccl,
            # Số tấm đã đặt:
            "da_dat_so_tam": ordered_map.get(r.ppl_id, 0.0),
            "nguon": "khsx",
        }
        for r in rows
    ]

    # Thêm: lệnh SX có mua_phoi_ngoai=True trên ProductionOrderItem nhưng chưa có plan line
    from app.models.production import ProductionOrder as _PO, ProductionOrderItem as _POI
    from sqlalchemy import exists, select as _select
    poi_rows = (
        db.query(_POI, _PO.so_lenh)
        .join(_PO, _PO.id == _POI.production_order_id)
        .filter(
            _POI.mua_phoi_ngoai == True,  # noqa: E712
            _PO.trang_thai == "mua_ngoai",
            ~exists(_select(ProductionPlanLine.id).where(
                ProductionPlanLine.production_order_item_id == _POI.id
            )),
        )
        .all()
    )
    for poi, so_lenh in poi_rows:
        result.append({
            "ppl_id": None,
            "so_ke_hoach": None,
            "ngay_ke_hoach": None,
            "ngay_chay": None,
            "so_lsx": so_lenh,
            "poi_id": poi.id,
            "ten_san_pham": poi.ten_hang or "",
            "so_luong_thung": float(poi.so_luong_ke_hoach or 0),
            "kho1": None,
            "kho_giay": None,
            "so_dao": None,
            "kho_tt": _f(poi.kho_tt),
            "dai_tt": _f(poi.dai_tt),
            "so_lop": poi.so_lop, "to_hop_song": poi.to_hop_song,
            "mat": poi.mat, "mat_dl": _f(poi.mat_dl),
            "song_1": poi.song_1, "song_1_dl": _f(poi.song_1_dl),
            "mat_1": poi.mat_1, "mat_1_dl": _f(poi.mat_1_dl),
            "song_2": poi.song_2, "song_2_dl": _f(poi.song_2_dl),
            "mat_2": poi.mat_2, "mat_2_dl": _f(poi.mat_2_dl),
            "song_3": poi.song_3, "song_3_dl": _f(poi.song_3_dl),
            "mat_3": poi.mat_3, "mat_3_dl": _f(poi.mat_3_dl),
            "loai_thung": poi.loai_thung,
            "dai": _f(poi.dai), "rong": _f(poi.rong), "cao": _f(poi.cao),
            "c_tham": poi.c_tham, "can_man": poi.can_man,
            "loai_lan": poi.loai_lan, "qccl": poi.qccl,
            "da_dat_so_tam": 0.0,
            "nguon": "lenh_sx",
        })

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Đối soát tồn kho giấy cuộn: SQL Server vs ERP Receipts
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/doi-soat-giay")
def doi_soat_giay(
    ncc_id: Optional[int] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """
    So sánh tồn kho SQL Server (warehouse_id=9) với tổng nhập ERP (GoodsReceipt da_duyet, MUA_HANG).
    Sắp xếp theo |chênh lệch| giảm dần để dễ phát hiện sai số.
    """
    GIAY_WH_ID = 9

    # ── A: SQL Server inventory ───────────────────────────────────────────────
    inv_rows = (
        db.query(
            InventoryBalance.paper_material_id,
            func.sum(InventoryBalance.ton_luong).label("ton_sql"),
            func.sum(InventoryBalance.don_gia_binh_quan * InventoryBalance.ton_luong).label("weighted"),
        )
        .filter(
            InventoryBalance.warehouse_id == GIAY_WH_ID,
            InventoryBalance.paper_material_id.isnot(None),
        )
        .group_by(InventoryBalance.paper_material_id)
        .all()
    )
    inv_map: dict[int, dict] = {}
    for r in inv_rows:
        ton = float(r.ton_sql or 0)
        weighted = float(r.weighted or 0)
        inv_map[r.paper_material_id] = {
            "ton_sql": ton,
            "gia_sql": round(weighted / ton, 0) if ton > 0 else 0.0,
        }

    # ── B: ERP receipts (da_duyet, MUA_HANG) ─────────────────────────────────
    gr_q = (
        db.query(
            GoodsReceiptItem.paper_material_id,
            func.sum(GoodsReceiptItem.so_luong).label("tong_nhap_erp"),
            func.max(GoodsReceiptItem.don_gia).label("gia_erp"),
            func.max(GoodsReceipt.ngay_nhap).label("ngay_nhap_erp"),
        )
        .join(GoodsReceipt, GoodsReceiptItem.receipt_id == GoodsReceipt.id)
        .filter(
            GoodsReceipt.trang_thai == "da_duyet",
            GoodsReceipt.loai_nhap == "MUA_HANG",
            GoodsReceiptItem.paper_material_id.isnot(None),
        )
    )
    if ncc_id:
        gr_q = gr_q.filter(GoodsReceipt.supplier_id == ncc_id)
    if date_from:
        gr_q = gr_q.filter(GoodsReceipt.ngay_nhap >= date_from)
    if date_to:
        gr_q = gr_q.filter(GoodsReceipt.ngay_nhap <= date_to)
    gr_rows = gr_q.group_by(GoodsReceiptItem.paper_material_id).all()
    gr_map: dict[int, dict] = {
        r.paper_material_id: {
            "tong_nhap_erp": float(r.tong_nhap_erp or 0),
            "gia_erp": float(r.gia_erp or 0),
            "ngay_nhap_erp": r.ngay_nhap_erp.isoformat() if r.ngay_nhap_erp else None,
        }
        for r in gr_rows
    }

    all_pm_ids = set(inv_map) | set(gr_map)
    if not all_pm_ids:
        return []

    # ── PaperMaterial + NSX batch ─────────────────────────────────────────────
    sup_alias = aliased(Supplier)
    pm_rows = (
        db.query(PaperMaterial, sup_alias.ten_viet_tat.label("ten_nsx"))
        .outerjoin(sup_alias, PaperMaterial.ma_nsx_id == sup_alias.id)
        .filter(PaperMaterial.id.in_(all_pm_ids))
        .all()
    )
    pm_map: dict[int, tuple] = {pm.id: (pm, ten_nsx) for pm, ten_nsx in pm_rows}

    result = []
    for pm_id in all_pm_ids:
        inv = inv_map.get(pm_id, {"ton_sql": 0.0, "gia_sql": 0.0})
        gr = gr_map.get(pm_id, {"tong_nhap_erp": 0.0, "gia_erp": 0.0, "ngay_nhap_erp": None})
        pm_info = pm_map.get(pm_id)
        pm = pm_info[0] if pm_info else None
        ten_nsx = pm_info[1] if pm_info else None

        ton_sql = inv["ton_sql"]
        tong_nhap = gr["tong_nhap_erp"]
        chenh_lech = ton_sql - tong_nhap
        ty_le_khop = round(tong_nhap / ton_sql * 100, 1) if ton_sql > 0 else None
        chenh_gia = inv["gia_sql"] - gr["gia_erp"]

        result.append({
            "paper_material_id": pm_id,
            "ma_chinh": pm.ma_chinh if pm else None,
            "ma_ky_hieu": pm.ma_ky_hieu if pm else None,
            "ten": pm.ten if pm else None,
            "ten_nsx": ten_nsx,
            "loai_giay": pm.loai_giay if pm else None,
            "ton_sql": round(ton_sql, 2),
            "gia_sql": round(inv["gia_sql"], 0),
            "tong_nhap_erp": round(tong_nhap, 2),
            "gia_erp": round(gr["gia_erp"], 0),
            "chenh_lech": round(chenh_lech, 2),
            "ty_le_khop": ty_le_khop,
            "chenh_gia": round(chenh_gia, 0),
            "ngay_nhap_erp": gr["ngay_nhap_erp"],
        })

    result.sort(key=lambda x: abs(x["chenh_lech"]), reverse=True)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Sổ nhập xuất tồn theo kỳ (Period Movement Report)
# ─────────────────────────────────────────────────────────────────────────────

NHAP_TYPES = frozenset([
    "nhap_kho", "ton_dau_ky", "ton_dau_ky_bo_sung",
    "dieu_chinh_tang", "chuyen_kho_den",
    # upper-case variants stored by older code
    "NHAP_MUA", "NHAP_SX", "CHUYEN_KHO_NHAP", "DIEU_CHINH_TANG",
])

XUAT_TYPES = frozenset([
    "xuat_sx", "xuat_ban", "xuat_kho",
    "dieu_chinh_giam", "chuyen_kho_di", "xuat_huy",
    # upper-case variants
    "XUAT_SX", "XUAT_BAN", "CHUYEN_KHO_XUAT", "DIEU_CHINH_GIAM",
])


def _build_so_nhap_xuat_ton(
    db,
    tu_ngay: Optional[date],
    den_ngay: Optional[date],
    warehouse_id: Optional[int],
    loai_nvl: str,
) -> list[dict]:
    q = db.query(InventoryTransaction)
    if tu_ngay:
        q = q.filter(func.date(InventoryTransaction.ngay_giao_dich) >= tu_ngay)
    if den_ngay:
        q = q.filter(func.date(InventoryTransaction.ngay_giao_dich) <= den_ngay)
    if warehouse_id:
        q = q.filter(InventoryTransaction.warehouse_id == warehouse_id)
    if loai_nvl == "giay_cuon":
        q = q.filter(InventoryTransaction.paper_material_id.isnot(None))
    elif loai_nvl == "nvl_khac":
        q = q.filter(InventoryTransaction.other_material_id.isnot(None))

    txs = q.all()

    key_map: dict[tuple, dict] = {}
    for tx in txs:
        key = (tx.warehouse_id, tx.paper_material_id, tx.other_material_id)
        if key not in key_map:
            key_map[key] = {
                "warehouse_id": tx.warehouse_id,
                "paper_material_id": tx.paper_material_id,
                "other_material_id": tx.other_material_id,
                "sl_nhap": Decimal(0),
                "gt_nhap": Decimal(0),
                "sl_xuat": Decimal(0),
                "gt_xuat": Decimal(0),
            }
        entry = key_map[key]
        loai = tx.loai_giao_dich or ""
        sl = tx.so_luong or Decimal(0)
        gt = tx.gia_tri or Decimal(0)
        if loai in NHAP_TYPES:
            entry["sl_nhap"] += sl
            entry["gt_nhap"] += gt
        elif loai in XUAT_TYPES:
            entry["sl_xuat"] += sl
            entry["gt_xuat"] += gt

    if not key_map:
        return []

    wh_ids = {k[0] for k in key_map}
    pm_ids = {k[1] for k in key_map if k[1]}
    om_ids = {k[2] for k in key_map if k[2]}

    wh_map: dict[int, Warehouse] = {
        w.id: w for w in db.query(Warehouse).filter(Warehouse.id.in_(wh_ids)).all()
    }
    pm_map2: dict[int, PaperMaterial] = {
        p.id: p for p in db.query(PaperMaterial).filter(PaperMaterial.id.in_(pm_ids)).all()
    } if pm_ids else {}
    om_map2: dict[int, OtherMaterial] = {
        o.id: o for o in db.query(OtherMaterial).filter(OtherMaterial.id.in_(om_ids)).all()
    } if om_ids else {}

    bal_q = db.query(InventoryBalance).filter(
        InventoryBalance.warehouse_id.in_(wh_ids)
    )
    if pm_ids:
        bal_q = bal_q.filter(
            or_(
                InventoryBalance.paper_material_id.in_(pm_ids),
                InventoryBalance.other_material_id.in_(om_ids) if om_ids else False,
            )
        )
    bals = bal_q.all()
    bal_map: dict[tuple, InventoryBalance] = {
        (b.warehouse_id, b.paper_material_id, b.other_material_id): b
        for b in bals
    }

    result = []
    for key, entry in key_map.items():
        wh_id, pm_id, om_id = key
        wh = wh_map.get(wh_id)
        bal = bal_map.get(key)

        ton_cuoi = float(bal.ton_luong) if bal else 0.0
        gt_cuoi = float(bal.gia_tri_ton) if bal else 0.0
        don_gia = float(bal.don_gia_binh_quan) if bal else 0.0
        sl_nhap = float(entry["sl_nhap"])
        gt_nhap = float(entry["gt_nhap"])
        sl_xuat = float(entry["sl_xuat"])
        gt_xuat = float(entry["gt_xuat"])
        ton_dau = ton_cuoi - sl_nhap + sl_xuat
        gt_dau = ton_dau * don_gia if don_gia else 0.0

        if pm_id and pm_id in pm_map2:
            pm = pm_map2[pm_id]
            ma_hang = pm.ma_chinh or ""
            ten_hang = pm.ten or ""
            don_vi = "Kg"
        elif om_id and om_id in om_map2:
            om = om_map2[om_id]
            ma_hang = getattr(om, "ma_chinh", "") or ""
            ten_hang = om.ten or ""
            don_vi = getattr(om, "don_vi", "") or ""
        else:
            ma_hang = ""
            ten_hang = ""
            don_vi = ""

        result.append({
            "warehouse_id": wh_id,
            "warehouse_name": wh.ten_kho if wh else "",
            "warehouse_code": wh.ma_kho if wh else "",
            "paper_material_id": pm_id,
            "other_material_id": om_id,
            "ma_hang": ma_hang,
            "ten_hang": ten_hang,
            "don_vi": don_vi,
            "don_gia": round(don_gia, 0),
            "ton_dau": round(ton_dau, 3),
            "gia_tri_dau": round(gt_dau, 0),
            "so_luong_nhap": round(sl_nhap, 3),
            "gia_tri_nhap": round(gt_nhap, 0),
            "so_luong_xuat": round(sl_xuat, 3),
            "gia_tri_xuat": round(gt_xuat, 0),
            "ton_cuoi": round(ton_cuoi, 3),
            "gia_tri_cuoi": round(gt_cuoi, 0),
        })

    result.sort(key=lambda x: (x["ma_hang"], x["warehouse_name"]))
    return result


@router.get("/so-nhap-xuat-ton")
def so_nhap_xuat_ton(
    tu_ngay: Optional[date] = Query(None),
    den_ngay: Optional[date] = Query(None),
    warehouse_id: Optional[int] = Query(None),
    loai_nvl: str = Query("all"),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return _build_so_nhap_xuat_ton(db, tu_ngay, den_ngay, warehouse_id, loai_nvl)


@router.get("/so-nhap-xuat-ton/export")
def export_so_nhap_xuat_ton(
    tu_ngay: Optional[date] = Query(None),
    den_ngay: Optional[date] = Query(None),
    warehouse_id: Optional[int] = Query(None),
    loai_nvl: str = Query("all"),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    rows = _build_so_nhap_xuat_ton(db, tu_ngay, den_ngay, warehouse_id, loai_nvl)

    wb = Workbook()
    ws = wb.active
    ws.title = "Nhập xuất tồn"

    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    subheader_fill = PatternFill("solid", fgColor="4472C4")
    subheader_font = Font(bold=True, color="FFFFFF", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    period_str = ""
    if tu_ngay and den_ngay:
        period_str = f"Từ {tu_ngay.strftime('%d/%m/%Y')} đến {den_ngay.strftime('%d/%m/%Y')}"
    elif tu_ngay:
        period_str = f"Từ {tu_ngay.strftime('%d/%m/%Y')}"
    elif den_ngay:
        period_str = f"Đến {den_ngay.strftime('%d/%m/%Y')}"

    ws.merge_cells("A1:N1")
    ws["A1"] = "SỔ NHẬP XUẤT TỒN NGUYÊN VẬT LIỆU"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center

    ws.merge_cells("A2:N2")
    ws["A2"] = period_str
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = center

    ws.merge_cells("A4:A5"); ws["A4"] = "STT"
    ws.merge_cells("B4:B5"); ws["B4"] = "Mã hàng"
    ws.merge_cells("C4:C5"); ws["C4"] = "Tên hàng"
    ws.merge_cells("D4:D5"); ws["D4"] = "ĐVT"
    ws.merge_cells("E4:E5"); ws["E4"] = "Đơn giá"
    ws.merge_cells("F4:G4"); ws["F4"] = "Tồn đầu kỳ"
    ws["F5"] = "Số lượng"; ws["G5"] = "Giá trị"
    ws.merge_cells("H4:I4"); ws["H4"] = "Nhập trong kỳ"
    ws["H5"] = "Số lượng"; ws["I5"] = "Giá trị"
    ws.merge_cells("J4:K4"); ws["J4"] = "Xuất trong kỳ"
    ws["J5"] = "Số lượng"; ws["K5"] = "Giá trị"
    ws.merge_cells("L4:M4"); ws["L4"] = "Tồn cuối kỳ"
    ws["L5"] = "Số lượng"; ws["M5"] = "Giá trị"
    ws.merge_cells("N4:N5"); ws["N4"] = "Kho"

    header_cells_row4 = ["A4","B4","C4","D4","E4","F4","H4","J4","L4","N4"]
    header_cells_row5 = ["F5","G5","H5","I5","J5","K5","L5","M5"]

    for c in header_cells_row4:
        ws[c].font = header_font
        ws[c].fill = header_fill
        ws[c].alignment = center
        ws[c].border = border

    for c in header_cells_row5:
        ws[c].font = subheader_font
        ws[c].fill = subheader_fill
        ws[c].alignment = center
        ws[c].border = border

    col_widths = [5, 18, 30, 6, 12, 13, 14, 13, 14, 13, 14, 13, 14, 18]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    totals = {k: 0.0 for k in ["ton_dau","gia_tri_dau","so_luong_nhap","gia_tri_nhap","so_luong_xuat","gia_tri_xuat","ton_cuoi","gia_tri_cuoi"]}

    for idx, row in enumerate(rows, 1):
        r = idx + 5
        ws.cell(r, 1, idx)
        ws.cell(r, 2, row["ma_hang"])
        ws.cell(r, 3, row["ten_hang"])
        ws.cell(r, 4, row["don_vi"])
        ws.cell(r, 5, row["don_gia"])
        ws.cell(r, 6, row["ton_dau"])
        ws.cell(r, 7, row["gia_tri_dau"])
        ws.cell(r, 8, row["so_luong_nhap"])
        ws.cell(r, 9, row["gia_tri_nhap"])
        ws.cell(r, 10, row["so_luong_xuat"])
        ws.cell(r, 11, row["gia_tri_xuat"])
        ws.cell(r, 12, row["ton_cuoi"])
        ws.cell(r, 13, row["gia_tri_cuoi"])
        ws.cell(r, 14, row["warehouse_name"])

        for col in [5, 6, 7, 8, 9, 10, 11, 12, 13]:
            ws.cell(r, col).alignment = right
        ws.cell(r, 1).alignment = center

        for col in range(1, 15):
            ws.cell(r, col).border = border

        for k in totals:
            totals[k] += row.get(k, 0.0)

    tr = len(rows) + 6
    ws.cell(tr, 1, "Tổng cộng")
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
    ws.cell(tr, 1).font = Font(bold=True)
    ws.cell(tr, 1).alignment = center

    for col, key in [(6,"ton_dau"),(7,"gia_tri_dau"),(8,"so_luong_nhap"),(9,"gia_tri_nhap"),
                     (10,"so_luong_xuat"),(11,"gia_tri_xuat"),(12,"ton_cuoi"),(13,"gia_tri_cuoi")]:
        c = ws.cell(tr, col, round(totals[key], 0))
        c.font = Font(bold=True)
        c.alignment = right
        c.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tu_str = tu_ngay.strftime("%Y%m%d") if tu_ngay else "all"
    den_str = den_ngay.strftime("%Y%m%d") if den_ngay else "all"
    filename = f"so_nhap_xuat_ton_{tu_str}_{den_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
