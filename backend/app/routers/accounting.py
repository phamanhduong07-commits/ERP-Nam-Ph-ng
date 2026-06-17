from datetime import date, datetime, timezone
from decimal import Decimal
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from app.utils.template import apply_template, standard_vars
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user, require_roles
from app.models.auth import User
from app.models.auth import AuditLog
from app.models.accounting import BankTransaction, CashReceipt, CashPayment, OpeningBalance, PurchaseInvoice, KheUocVay, KheUocChoVay, LichTraNo, InternalTransfer
from app.models.master import BankAccount, Customer, PhapNhan, Supplier
from app.models.warehouse_doc import GoodsReceipt
from sqlalchemy import desc, func
from app.services.accounting_service import AccountingService
from app.schemas.accounting import (
    PurchaseInvoiceCreate,
    PurchaseInvoiceResponse, CashReceiptCreate,
    CashReceiptResponse, CashPaymentCreate,
    CashPaymentResponse, OpeningBalanceCreate,
    ARCustomerSummaryRow,
    WorkshopPayrollCreate,
    WorkshopPayrollResponse, OverheadAllocationRequest,
    OverheadAllocationResponse, ClosingResult,
    ProductionCostPeriodCreate,
    FixedAssetCreate,
    FixedAssetResponse, ManualJournalEntryCreate,
    KheUocVayCreate, KheUocVayUpdate, KheUocVayResponse,
    KheUocChoVayCreate, KheUocChoVayUpdate, KheUocChoVayResponse,
    LichTraNoResponse, TraNoRequest, CapNhatLaiSuatRequest, TraTruocHanRequest, TatToanRequest,
    CashFlowForecastResponse, ForecastDayItem,
    InternalTransferCreate, BatchReceiptCreate,
    CashReceiptUpdate, CashPaymentUpdate,
    TaxPaymentCreate, TaxPaymentBatchResponse, TaxObligationItem,
    InsuranceObligationItem, InsurancePaymentCreate, InsuranceBatchResponse,
    SalaryObligationItem, SalaryPaymentCreate, SalaryBatchResponse,
)
from app.services.excel_import_service import (
    ImportField, build_template_response, import_excel, parse_date, parse_decimal, parse_text,
)
from app.utils.format_utils import so_thanh_chu, ngay_str
import html as _html_mod
import logging
import io


def _logo_img(pn, settings: dict) -> str:
    src = (
        f"/api/phap-nhan/logo/{pn.ma_phap_nhan}" if pn and pn.ma_phap_nhan
        else settings.get("logo_url") or ""
    )
    return f'<img src="{_html_mod.escape(src)}" style="max-height:50px;max-width:100%;object-fit:contain"/>' if src else ""
import pandas as pd
from enum import Enum as _Enum


class ReceiptStatus(str, _Enum):
    CHO_DUYET = "cho_duyet"
    DA_DUYET = "da_duyet"
    HUY = "huy"


class PaymentStatus(str, _Enum):
    CHO_CHOT = "cho_chot"
    DA_CHOT = "da_chot"
    DA_DUYET = "da_duyet"
    HUY = "huy"


class PurchaseInvoiceStatus(str, _Enum):
    NHAP = "nhap"
    DA_DUYET = "da_duyet"
    HUY = "huy"


router = APIRouter(prefix="/api/accounting", tags=["accounting"])

logger = logging.getLogger(__name__)

KE_TOAN_ROLES = ("KE_TOAN_TRUONG", "KE_TOAN_CONG_NO", "KE_TOAN_MUA_HANG", "KETOAN_NHAN_VIEN", "BGD_GIAM_DOC")
ACCOUNTING_AUDIT_TABLES = {
    "bank_transactions",
    "cash_receipts",
    "cash_payments",
    "purchase_invoices",
    "production_cost_periods",
    "journal_entries",
}

BANK_TRANSACTION_IMPORT_FIELDS = [
    ImportField("ngay_giao_dich", "Ngay giao dich", required=True, parser=parse_date, help_text="YYYY-MM-DD hoac DD/MM/YYYY"),
    ImportField("so_tai_khoan", "So tai khoan", required=True, parser=parse_text),
    ImportField("so_tham_chieu", "So tham chieu", parser=parse_text),
    ImportField("mo_ta", "Mo ta", parser=parse_text),
    ImportField("thu", "Thu", parser=parse_decimal, default=Decimal("0")),
    ImportField("chi", "Chi", parser=parse_decimal, default=Decimal("0")),
    ImportField("so_du", "So du", parser=parse_decimal),
]


def _bank_transaction_resolver(db: Session, values: dict) -> tuple[dict, list[str]]:
    errors: list[str] = []
    so_tai_khoan = values.get("so_tai_khoan")
    account = None
    if so_tai_khoan:
        account = db.query(BankAccount).filter(BankAccount.so_tai_khoan == so_tai_khoan).first()
        if account:
            values["bank_account_id"] = account.id
            values["phap_nhan_id"] = account.phap_nhan_id

    thu = Decimal(str(values.get("thu") or 0))
    chi = Decimal(str(values.get("chi") or 0))
    if thu < 0 or chi < 0:
        errors.append("Thu/Chi khong duoc am")
    if (thu > 0 and chi > 0) or (thu == 0 and chi == 0):
        errors.append("Moi giao dich phai co dung mot cot Thu hoac Chi")

    ngay = values.get("ngay_giao_dich")
    parts = [
        str(account.id if account else so_tai_khoan or ""),
        ngay.isoformat() if hasattr(ngay, "isoformat") else str(ngay or ""),
        str(values.get("so_tham_chieu") or ""),
        str(thu),
        str(chi),
        str(values.get("mo_ta") or "")[:80],
    ]
    values["import_key"] = "|".join(parts)[:255]
    values["trang_thai"] = "chua_doi_soat"
    return values, errors


@router.post("/chart-of-accounts/seed")
def seed_chart_of_accounts(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Seed idempotent he thong tai khoan loi, gom tai khoan noi bo va thue."""
    result = AccountingService(db).ensure_core_chart_of_accounts()
    db.commit()
    return result


@router.get("/audit-logs")
def list_accounting_audit_logs(
    bang: str | None = Query(None),
    ban_ghi_id: str | None = Query(None),
    user_id: int | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    if bang:
        if bang not in ACCOUNTING_AUDIT_TABLES:
            raise HTTPException(400, "Bang audit khong thuoc pham vi ke toan")
    q = db.query(AuditLog).filter(AuditLog.bang.in_(ACCOUNTING_AUDIT_TABLES))
    if bang:
        q = q.filter(AuditLog.bang == bang)
    if ban_ghi_id:
        q = q.filter(AuditLog.ban_ghi_id == str(ban_ghi_id))
    if user_id:
        q = q.filter(AuditLog.user_id == user_id)
    if tu_ngay:
        q = q.filter(func.date(AuditLog.created_at) >= tu_ngay)
    if den_ngay:
        q = q.filter(func.date(AuditLog.created_at) <= den_ngay)

    total = q.count()
    items = q.order_by(desc(AuditLog.created_at), desc(AuditLog.id))\
             .offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.get("/audit/dimensions")
def get_accounting_dimension_audit(
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).get_dimension_audit(
        tu_ngay=tu_ngay,
        den_ngay=den_ngay,
        phap_nhan_id=phap_nhan_id,
        phan_xuong_id=phan_xuong_id,
        limit=limit,
    )


@router.get("/documents/{bang}/{ban_ghi_id}/audit")
def get_document_audit(
    bang: str,
    ban_ghi_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    if bang not in ACCOUNTING_AUDIT_TABLES:
        raise HTTPException(400, "Bang audit khong thuoc pham vi ke toan")
    return db.query(AuditLog).filter(
        AuditLog.bang == bang,
        AuditLog.ban_ghi_id == str(ban_ghi_id),
    ).order_by(AuditLog.created_at, AuditLog.id).all()


@router.get("/documents/{chung_tu_loai}/{chung_tu_id}/journal-entries")
def get_document_journal_entries(
    chung_tu_loai: str,
    chung_tu_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from app.models.accounting import JournalEntry
    return db.query(JournalEntry).filter(
        JournalEntry.chung_tu_loai == chung_tu_loai,
        JournalEntry.chung_tu_id == chung_tu_id,
    ).order_by(JournalEntry.ngay_but_toan, JournalEntry.id).all()


def _bank_transaction_payload(tx: BankTransaction) -> dict:
    return {
        "id": tx.id,
        "bank_account_id": tx.bank_account_id,
        "phap_nhan_id": tx.phap_nhan_id,
        "ngay_giao_dich": tx.ngay_giao_dich,
        "so_tai_khoan": tx.so_tai_khoan or (tx.bank_account.so_tai_khoan if tx.bank_account else None),
        "so_tham_chieu": tx.so_tham_chieu,
        "mo_ta": tx.mo_ta,
        "thu": tx.thu,
        "chi": tx.chi,
        "so_du": tx.so_du,
        "trang_thai": tx.trang_thai,
        "matched_chung_tu_loai": tx.matched_chung_tu_loai,
        "matched_chung_tu_id": tx.matched_chung_tu_id,
        "matched_at": tx.matched_at,
        "matched_by": tx.matched_by,
        "created_at": tx.created_at,
    }


def _candidate_payload(chung_tu_loai: str, obj, doi_tuong: str | None) -> dict:
    return {
        "chung_tu_loai": chung_tu_loai,
        "chung_tu_id": obj.id,
        "so_chung_tu": obj.so_phieu,
        "ngay": obj.ngay_phieu,
        "doi_tuong": doi_tuong,
        "dien_giai": obj.dien_giai,
        "so_tien": obj.so_tien,
    }


def _get_bank_transaction_or_404(db: Session, tx_id: int) -> BankTransaction:
    tx = db.get(BankTransaction, tx_id)
    if not tx:
        raise HTTPException(404, "Khong tim thay giao dich ngan hang")
    return tx


def _validate_reconcile_target(db: Session, tx: BankTransaction, payload: dict) -> tuple[str, int]:
    chung_tu_loai = payload.get("chung_tu_loai")
    chung_tu_id = payload.get("chung_tu_id")
    if chung_tu_loai not in {"phieu_thu", "phieu_chi"} or not chung_tu_id:
        raise HTTPException(400, "Doi tuong doi soat khong hop le")

    model = CashReceipt if chung_tu_loai == "phieu_thu" else CashPayment
    obj = db.get(model, int(chung_tu_id))
    if not obj:
        raise HTTPException(404, "Khong tim thay chung tu doi soat")
    approved_status = ReceiptStatus.DA_DUYET if chung_tu_loai == "phieu_thu" else PaymentStatus.DA_DUYET
    if obj.trang_thai != approved_status:
        raise HTTPException(400, "Chi doi soat chung tu da duyet")

    tx_amount = Decimal(str(tx.thu or 0)) if tx.thu and tx.thu > 0 else Decimal(str(tx.chi or 0))
    if Decimal(str(obj.so_tien)) != tx_amount:
        raise HTTPException(400, "So tien giao dich khong khop chung tu")
    if chung_tu_loai == "phieu_thu" and not (tx.thu and tx.thu > 0):
        raise HTTPException(400, "Giao dich thu chi duoc doi soat voi phieu thu")
    if chung_tu_loai == "phieu_chi" and not (tx.chi and tx.chi > 0):
        raise HTTPException(400, "Giao dich chi chi duoc doi soat voi phieu chi")
    if tx.phap_nhan_id and obj.phap_nhan_id and tx.phap_nhan_id != obj.phap_nhan_id:
        raise HTTPException(400, "Phap nhan giao dich khong khop chung tu")
    return chung_tu_loai, int(chung_tu_id)


@router.get("/bank-transactions/import-template")
def download_bank_transaction_import_template(_: User = Depends(get_current_user)):
    return build_template_response("mau_import_sao_ke_ngan_hang.xlsx", BANK_TRANSACTION_IMPORT_FIELDS)


@router.post("/bank-transactions/import")
async def import_bank_transactions(
    commit: bool = Query(default=False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return await import_excel(
        db=db,
        file=file,
        model=BankTransaction,
        fields=BANK_TRANSACTION_IMPORT_FIELDS,
        key_field="import_key",
        commit=commit,
        resolver=_bank_transaction_resolver,
        user=current_user,
        loai_du_lieu="sao_ke_ngan_hang",
    )


@router.get("/bank-transactions")
def list_bank_transactions(
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    bank_account_id: int | None = Query(None),
    trang_thai: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(BankTransaction)
    if tu_ngay:
        q = q.filter(BankTransaction.ngay_giao_dich >= tu_ngay)
    if den_ngay:
        q = q.filter(BankTransaction.ngay_giao_dich <= den_ngay)
    if phap_nhan_id:
        q = q.filter(BankTransaction.phap_nhan_id == phap_nhan_id)
    if bank_account_id:
        q = q.filter(BankTransaction.bank_account_id == bank_account_id)
    if trang_thai:
        q = q.filter(BankTransaction.trang_thai == trang_thai)

    total = q.count()
    items = (
        q.order_by(desc(BankTransaction.ngay_giao_dich), desc(BankTransaction.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {"total": total, "page": page, "page_size": page_size, "items": [_bank_transaction_payload(tx) for tx in items]}


@router.get("/bank-transactions/{tx_id}/candidates")
def get_bank_transaction_candidates(
    tx_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    tx = _get_bank_transaction_or_404(db, tx_id)
    amount = Decimal(str(tx.thu or 0)) if tx.thu and tx.thu > 0 else Decimal(str(tx.chi or 0))
    candidates = []
    if tx.thu and tx.thu > 0:
        q = db.query(CashReceipt).join(Customer, CashReceipt.customer_id == Customer.id).filter(
            CashReceipt.trang_thai == ReceiptStatus.DA_DUYET,
            CashReceipt.so_tien == amount,
        )
        if tx.phap_nhan_id:
            q = q.filter(CashReceipt.phap_nhan_id == tx.phap_nhan_id)
        for obj in q.order_by(desc(CashReceipt.ngay_phieu), desc(CashReceipt.id)).limit(20).all():
            candidates.append(_candidate_payload("phieu_thu", obj, obj.customer.ten_viet_tat if obj.customer else None))
    elif tx.chi and tx.chi > 0:
        q = db.query(CashPayment).join(Supplier, CashPayment.supplier_id == Supplier.id).filter(
            CashPayment.trang_thai == PaymentStatus.DA_DUYET,
            CashPayment.so_tien == amount,
        )
        if tx.phap_nhan_id:
            q = q.filter(CashPayment.phap_nhan_id == tx.phap_nhan_id)
        for obj in q.order_by(desc(CashPayment.ngay_phieu), desc(CashPayment.id)).limit(20).all():
            candidates.append(_candidate_payload("phieu_chi", obj, obj.supplier.ten_viet_tat if obj.supplier else None))
    return candidates


@router.post("/bank-transactions/{tx_id}/reconcile")
def reconcile_bank_transaction(
    tx_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    tx = _get_bank_transaction_or_404(db, tx_id)
    if tx.trang_thai == "da_doi_soat":
        raise HTTPException(400, "Giao dich da doi soat")
    chung_tu_loai, chung_tu_id = _validate_reconcile_target(db, tx, payload)
    tx.trang_thai = "da_doi_soat"
    tx.matched_chung_tu_loai = chung_tu_loai
    tx.matched_chung_tu_id = chung_tu_id
    tx.matched_at = datetime.now(timezone.utc)
    tx.matched_by = current_user.id
    AccountingService(db)._audit(
        "reconcile",
        "bank_transactions",
        tx.id,
        user_id=current_user.id,
        du_lieu_moi={"chung_tu_loai": chung_tu_loai, "chung_tu_id": chung_tu_id},
    )
    db.commit()
    db.refresh(tx)
    return _bank_transaction_payload(tx)


@router.post("/bank-transactions/{tx_id}/unreconcile")
def unreconcile_bank_transaction(
    tx_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    tx = _get_bank_transaction_or_404(db, tx_id)
    old = {"chung_tu_loai": tx.matched_chung_tu_loai, "chung_tu_id": tx.matched_chung_tu_id}
    tx.trang_thai = "chua_doi_soat"
    tx.matched_chung_tu_loai = None
    tx.matched_chung_tu_id = None
    tx.matched_at = None
    tx.matched_by = None
    AccountingService(db)._audit("unreconcile", "bank_transactions", tx.id, user_id=current_user.id, du_lieu_cu=old)
    db.commit()
    db.refresh(tx)
    return _bank_transaction_payload(tx)


@router.post("/bank-transactions/{tx_id}/ignore")
def ignore_bank_transaction(
    tx_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    tx = _get_bank_transaction_or_404(db, tx_id)
    old_status = tx.trang_thai
    tx.trang_thai = "bo_qua"
    AccountingService(db)._audit(
        "ignore",
        "bank_transactions",
        tx.id,
        user_id=current_user.id,
        du_lieu_cu={"trang_thai": old_status},
        du_lieu_moi={"trang_thai": tx.trang_thai},
    )
    db.commit()
    db.refresh(tx)
    return _bank_transaction_payload(tx)


# ─────────────────────────────────────────────
# PHIẾU THU
# ─────────────────────────────────────────────

@router.get("/receipts")
def list_receipts(
    customer_id: int | None = Query(None),
    trang_thai: str | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).list_receipts(
        customer_id=customer_id, trang_thai=trang_thai,
        tu_ngay=tu_ngay, den_ngay=den_ngay,
        phap_nhan_id=phap_nhan_id, phan_xuong_id=phan_xuong_id,
        page=page, page_size=page_size,
    )


@router.post("/receipts", response_model=CashReceiptResponse, status_code=201)
def create_receipt(
    data: CashReceiptCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).create_cash_receipt(data, current_user.id)


@router.get("/receipts/{receipt_id}", response_model=CashReceiptResponse)
def get_receipt(
    receipt_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_receipt(receipt_id)


@router.patch("/receipts/{receipt_id}/approve", response_model=CashReceiptResponse)
def approve_receipt(
    receipt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    logger.info("approve_receipt id=%s user=%s", receipt_id, current_user.id)
    return AccountingService(db).approve_receipt(receipt_id, current_user.id)


@router.patch("/receipts/{receipt_id}/cancel", response_model=CashReceiptResponse)
def cancel_receipt(
    receipt_id: int,
    ly_do: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    logger.info("cancel_receipt id=%s user=%s", receipt_id, current_user.id)
    return AccountingService(db).cancel_receipt(receipt_id, current_user.id, ly_do)


@router.put("/receipts/{receipt_id}", response_model=CashReceiptResponse)
def update_receipt(
    receipt_id: int,
    data: CashReceiptUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).update_receipt(receipt_id, data, current_user.id)


# ─────────────────────────────────────────────
# THU TIỀN NHIỀU KHÁCH HÀNG (BATCH)
# ─────────────────────────────────────────────

@router.post("/receipts/batch")
def batch_create_receipts(
    data: BatchReceiptCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    from app.schemas.accounting import (
        BatchReceiptResponse,
        BatchReceiptResultItem, CashReceiptCreate,
    )
    results = []
    service = AccountingService(db)
    thanh_cong = 0
    that_bai = 0
    for idx, item in enumerate(data.items):
        try:
            receipt_data = CashReceiptCreate(
                customer_id=item.customer_id,
                sales_invoice_id=item.sales_invoice_id,
                ngay_phieu=data.ngay_phieu,
                hinh_thuc_tt=item.hinh_thuc_tt,
                so_tai_khoan=data.so_tai_khoan,
                so_tham_chieu=item.so_tham_chieu,
                dien_giai=item.dien_giai,
                so_tien=item.so_tien,
                phap_nhan_id=data.phap_nhan_id,
                phan_xuong_id=data.phan_xuong_id,
            )
            receipt = service.create_cash_receipt(receipt_data, current_user.id)
            results.append(BatchReceiptResultItem(
                index=idx,
                customer_id=item.customer_id,
                so_phieu=receipt.so_phieu,
                so_tien=item.so_tien,
                success=True,
            ))
            thanh_cong += 1
        except Exception as e:
            db.rollback()
            service = AccountingService(db)
            results.append(BatchReceiptResultItem(
                index=idx,
                customer_id=item.customer_id,
                so_phieu=None,
                so_tien=item.so_tien,
                success=False,
                error=str(e),
            ))
            that_bai += 1
    return BatchReceiptResponse(
        tong_so=len(data.items),
        thanh_cong=thanh_cong,
        that_bai=that_bai,
        items=results,
    )


@router.get("/receipts/import-template")
def download_receipts_import_template(_: User = Depends(get_current_user)):
    """Tải mẫu Excel để import phiếu thu (tương thích MISA)."""
    return build_template_response("mau_import_phieu_thu.xlsx", [
        ImportField("ngay_chung_tu",  "Ngày chứng từ",          required=True,  parser=parse_date,    help_text="DD/MM/YYYY"),
        ImportField("ngay_hach_toan", "Ngày hạch toán",                          parser=parse_date,    help_text="Để trống = dùng Ngày chứng từ"),
        ImportField("so_chung_tu",    "Số chứng từ",                             parser=parse_text,    help_text="Để trống - hệ thống tự sinh"),
        ImportField("ma_doi_tuong",   "Mã đối tượng",           required=True,  parser=parse_text,    help_text="Mã khách hàng (ma_kh)"),
        ImportField("ten_doi_tuong",  "Tên đối tượng",                           parser=parse_text),
        ImportField("dien_giai_ht",   "Diễn giải (hạch toán)",                   parser=parse_text,    help_text="Diễn giải ghi vào phiếu"),
        ImportField("tk_no",          "TK Nợ",                                   parser=parse_text,    help_text="Mặc định: 112 (CK) hoặc 111 (TM)"),
        ImportField("tk_co",          "TK Có",                                   parser=parse_text,    help_text="Mặc định: 131"),
        ImportField("so_tien",        "Số tiền",                required=True,  parser=parse_decimal, help_text="Số tiền thu (VND)"),
        ImportField("so_tk_nh",       "Số TK ngân hàng",                         parser=parse_text,    help_text="Số tài khoản ngân hàng"),
        ImportField("ten_nh",         "Tên ngân hàng",                            parser=parse_text),
        ImportField("httt",           "Hình thức TT",                             parser=parse_text,    help_text="TM=Tiền mặt, CK=Chuyển khoản"),
        ImportField("ly_do",          "Lý do thu",                               parser=parse_text,    help_text="Lý do thu tiền"),
        ImportField("co_hoa_don",     "Có hóa đơn",                               parser=parse_text,    help_text="Có hoặc Không"),
        ImportField("so_hoa_don",     "Số hóa đơn",                               parser=parse_text),
        ImportField("ngay_hoa_don",   "Ngày hóa đơn",                            parser=parse_date),
    ])


@router.post("/receipts/import-excel")
async def import_receipts_excel(
    file: UploadFile = File(...),
    ngay_phieu: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    dry_run: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Import phiếu thu từ Excel. Cột bắt buộc: ma_kh/customer_id, so_tien.
    Cột tùy chọn: so_hoa_don/sales_invoice_id, hinh_thuc_tt, dien_giai, so_tham_chieu."""
    from openpyxl import load_workbook
    from app.models.master import Customer
    from app.models.billing import SalesInvoice
    from app.schemas.accounting import (
        BatchReceiptCreate, BatchReceiptResponse,
        BatchReceiptResultItem, BatchReceiptItem, CashReceiptCreate,
    )
    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "File Excel không hợp lệ")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "File Excel trống")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    def find_col(*names: str) -> int | None:
        for n in names:
            if n.lower() in headers:
                return headers.index(n.lower())
        return None

    col_cust     = find_col("customer_id", "ma_kh", "mã kh", "doi_tuong", "mã đối tượng", "ma doi tuong", "đối tượng", "doi tuong", "khách hàng")
    col_tien     = find_col("so_tien", "số tiền", "so tien", "tien")
    col_inv      = find_col("sales_invoice_id", "so_hoa_don", "số hóa đơn")
    col_httt     = find_col("hình thức tt", "httt", "hinh_thuc_tt", "loai_chung_tu", "loại chứng từ", "loai chung tu")
    col_dg       = find_col("diễn giải (hạch toán)", "dien_giai_ht", "dien_giai", "diễn giải", "dien giai")
    col_ref      = find_col("so_tham_chieu", "số tham chiếu")
    col_tk       = find_col("số tk ngân hàng", "so_tk_nh", "so tai khoan", "so_tai_khoan", "số tài khoản nh", "so tk nh")
    col_ly_do    = find_col("lý do thu", "ly do thu", "ly_do_thu", "ly_do")
    col_ngay_row = find_col("ngày chứng từ", "ngay chung tu", "ngay_chung_tu", "ngay_hach_toan", "ngày hạch toán", "ngay hach toan")
    col_so_phieu = find_col("số chứng từ", "so chung tu", "so_chung_tu", "so_phieu")
    col_tk_no    = find_col("tk nợ", "tk_no", "tk no")
    col_tk_co    = find_col("tk có", "tk_co", "tk co")

    if col_cust is None or col_tien is None:
        raise HTTPException(400, "File thiếu cột bắt buộc: Mã đối tượng (ma_kh) và Số tiền")

    today = ngay_phieu or date.today()
    service = AccountingService(db)
    results = []
    thanh_cong = 0
    that_bai = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(c is not None for c in row):
            continue
        raw_cust = row[col_cust]
        raw_tien = row[col_tien]
        if raw_cust is None or raw_tien is None:
            continue

        try:
            # Resolve customer
            cust_id = None
            if isinstance(raw_cust, (int, float)):
                cust_id = int(raw_cust)
            else:
                cust = db.query(Customer).filter(Customer.ma_kh == str(raw_cust).strip()).first()
                if cust:
                    cust_id = cust.id
            if not cust_id:
                raise ValueError(f"Không tìm thấy khách hàng '{raw_cust}'")

            so_tien = Decimal(str(raw_tien).replace(",", "").strip())
            row_ngay = today
            if col_ngay_row is not None and row[col_ngay_row]:
                from datetime import datetime as _dt
                raw_ng = row[col_ngay_row]
                if isinstance(raw_ng, date) and not isinstance(raw_ng, datetime):
                    row_ngay = raw_ng
                elif isinstance(raw_ng, datetime):
                    row_ngay = raw_ng.date()
                else:
                    for _fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                        try:
                            row_ngay = _dt.strptime(str(raw_ng).strip(), _fmt).date()
                            break
                        except ValueError:
                            pass

            httt = "CK"
            if col_httt is not None and row[col_httt]:
                raw_ht = str(row[col_httt]).strip().upper()
                if raw_ht in ("TM", "TIỀN MẶT", "TIEN MAT", "TIEN_MAT"):
                    httt = "TM"

            inv_id = None
            if col_inv is not None and row[col_inv]:
                raw_inv = row[col_inv]
                if isinstance(raw_inv, (int, float)):
                    inv_id = int(raw_inv)
                else:
                    inv = db.query(SalesInvoice).filter(SalesInvoice.so_hoa_don == str(raw_inv).strip()).first()
                    if inv:
                        inv_id = inv.id

            dg_val = None
            if col_dg is not None and row[col_dg]:
                dg_val = str(row[col_dg]).strip()
            elif col_ly_do is not None and row[col_ly_do]:
                dg_val = str(row[col_ly_do]).strip()

            so_phieu_val = None
            if col_so_phieu is not None and row[col_so_phieu]:
                so_phieu_val = str(row[col_so_phieu]).strip() or None

            tk_no_val = "112"
            if col_tk_no is not None and row[col_tk_no]:
                tk_no_val = str(row[col_tk_no]).strip()

            tk_co_val = "131"
            if col_tk_co is not None and row[col_tk_co]:
                tk_co_val = str(row[col_tk_co]).strip()

            receipt_data = CashReceiptCreate(
                customer_id=cust_id,
                sales_invoice_id=inv_id,
                so_phieu=so_phieu_val,
                ngay_phieu=row_ngay,
                hinh_thuc_tt=httt,
                dien_giai=dg_val,
                so_tham_chieu=str(row[col_ref]).strip() if col_ref is not None and row[col_ref] else None,
                so_tai_khoan=str(row[col_tk]).strip() if col_tk is not None and row[col_tk] else None,
                so_tien=so_tien,
                tk_no=tk_no_val,
                tk_co=tk_co_val,
                phap_nhan_id=phap_nhan_id,
                phan_xuong_id=phan_xuong_id,
            )
            if dry_run:
                results.append(BatchReceiptResultItem(
                    index=row_idx, customer_id=cust_id,
                    so_phieu=None, so_tien=so_tien, success=True,
                ))
            else:
                receipt = service.create_cash_receipt(receipt_data, current_user.id)
                results.append(BatchReceiptResultItem(
                    index=row_idx, customer_id=cust_id,
                    so_phieu=receipt.so_phieu, so_tien=so_tien, success=True,
                ))
            thanh_cong += 1
        except Exception as e:
            db.rollback()
            service = AccountingService(db)
            results.append(BatchReceiptResultItem(
                index=row_idx,
                customer_id=0,
                so_phieu=None,
                so_tien=Decimal("0"),
                success=False,
                error=f"Dòng {row_idx}: {e}",
            ))
            that_bai += 1

    return BatchReceiptResponse(
        tong_so=len(results),
        thanh_cong=thanh_cong,
        that_bai=that_bai,
        items=results,
    )


# ─────────────────────────────────────────────
# CHUYỂN TIỀN NỘI BỘ (INTERNAL TRANSFER)
# ─────────────────────────────────────────────

@router.get("/internal-transfers")
def list_internal_transfers(
    trang_thai: str | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from app.models.accounting import InternalTransfer
    q = db.query(InternalTransfer)
    if trang_thai:
        q = q.filter(InternalTransfer.trang_thai == trang_thai)
    if tu_ngay:
        q = q.filter(InternalTransfer.ngay_phieu >= tu_ngay)
    if den_ngay:
        q = q.filter(InternalTransfer.ngay_phieu <= den_ngay)
    if phap_nhan_id:
        q = q.filter(
            (InternalTransfer.tu_phap_nhan_id == phap_nhan_id) |
            (InternalTransfer.den_phap_nhan_id == phap_nhan_id)
        )
    total = q.count()
    items = q.order_by(desc(InternalTransfer.ngay_phieu), desc(InternalTransfer.id)) \
              .offset((page - 1) * page_size).limit(page_size).all()

    from app.schemas.accounting import InternalTransferResponse
    result = []
    for it in items:
        r = InternalTransferResponse.model_validate(it)
        r.tu_phap_nhan_ten = it.tu_phap_nhan.ten_phap_nhan if it.tu_phap_nhan else None
        r.den_phap_nhan_ten = it.den_phap_nhan.ten_phap_nhan if it.den_phap_nhan else None
        result.append(r)
    return {"total": total, "items": result}


@router.post("/internal-transfers", status_code=201)
def create_internal_transfer(
    data: InternalTransferCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    from app.schemas.accounting import InternalTransferResponse
    transfer = InternalTransfer(
        so_phieu=AccountingService(db)._gen_so_phieu("CTN", InternalTransfer),
        ngay_phieu=data.ngay_phieu,
        tu_phap_nhan_id=data.tu_phap_nhan_id,
        den_phap_nhan_id=data.den_phap_nhan_id,
        tu_tai_khoan=data.tu_tai_khoan,
        den_tai_khoan=data.den_tai_khoan,
        so_tien=data.so_tien,
        hinh_thuc_tt=data.hinh_thuc_tt,
        so_tham_chieu=data.so_tham_chieu,
        dien_giai=data.dien_giai,
        tk_no=data.tk_no,
        tk_co=data.tk_co,
        created_by=current_user.id,
    )
    db.add(transfer)
    db.commit()
    db.refresh(transfer)
    r = InternalTransferResponse.model_validate(transfer)
    r.tu_phap_nhan_ten = transfer.tu_phap_nhan.ten_phap_nhan if transfer.tu_phap_nhan else None
    r.den_phap_nhan_ten = transfer.den_phap_nhan.ten_phap_nhan if transfer.den_phap_nhan else None
    return r


@router.get("/internal-transfers/{transfer_id}")
def get_internal_transfer(
    transfer_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from app.models.accounting import InternalTransfer
    from app.schemas.accounting import InternalTransferResponse
    transfer = db.get(InternalTransfer, transfer_id)
    if not transfer:
        raise HTTPException(404, "Không tìm thấy phiếu chuyển tiền")
    r = InternalTransferResponse.model_validate(transfer)
    r.tu_phap_nhan_ten = transfer.tu_phap_nhan.ten_phap_nhan if transfer.tu_phap_nhan else None
    r.den_phap_nhan_ten = transfer.den_phap_nhan.ten_phap_nhan if transfer.den_phap_nhan else None
    return r


@router.patch("/internal-transfers/{transfer_id}/approve")
def approve_internal_transfer(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    from app.models.accounting import InternalTransfer
    from app.schemas.accounting import InternalTransferResponse
    transfer = db.get(InternalTransfer, transfer_id)
    if not transfer:
        raise HTTPException(404, "Không tìm thấy phiếu chuyển tiền")
    if transfer.trang_thai != "cho_duyet":
        raise HTTPException(400, "Chỉ duyệt phiếu đang ở trạng thái Chờ duyệt")
    transfer.trang_thai = "da_duyet"
    transfer.nguoi_duyet_id = current_user.id
    transfer.ngay_duyet = datetime.now(timezone.utc)
    db.commit()
    db.refresh(transfer)
    r = InternalTransferResponse.model_validate(transfer)
    r.tu_phap_nhan_ten = transfer.tu_phap_nhan.ten_phap_nhan if transfer.tu_phap_nhan else None
    r.den_phap_nhan_ten = transfer.den_phap_nhan.ten_phap_nhan if transfer.den_phap_nhan else None
    return r


@router.patch("/internal-transfers/{transfer_id}/cancel")
def cancel_internal_transfer(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    from app.models.accounting import InternalTransfer
    from app.schemas.accounting import InternalTransferResponse
    transfer = db.get(InternalTransfer, transfer_id)
    if not transfer:
        raise HTTPException(404, "Không tìm thấy phiếu chuyển tiền")
    if transfer.trang_thai == "huy":
        raise HTTPException(400, "Phiếu đã bị hủy")
    transfer.trang_thai = "huy"
    db.commit()
    db.refresh(transfer)
    r = InternalTransferResponse.model_validate(transfer)
    r.tu_phap_nhan_ten = transfer.tu_phap_nhan.ten_phap_nhan if transfer.tu_phap_nhan else None
    r.den_phap_nhan_ten = transfer.den_phap_nhan.ten_phap_nhan if transfer.den_phap_nhan else None
    return r


# ─────────────────────────────────────────────
# HÓA ĐƠN MUA HÀNG
# ─────────────────────────────────────────────

@router.get("/purchase-invoices")
def list_purchase_invoices(
    supplier_id: int | None = Query(None),
    trang_thai: str | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    qua_han_only: bool = Query(False),
    phap_nhan_id: int | None = Query(None),
    so_hoa_don: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).list_purchase_invoices(
        supplier_id=supplier_id, trang_thai=trang_thai,
        tu_ngay=tu_ngay, den_ngay=den_ngay,
        qua_han_only=qua_han_only, phap_nhan_id=phap_nhan_id,
        so_hoa_don=so_hoa_don, page=page, page_size=page_size,
    )


@router.post("/purchase-invoices", response_model=PurchaseInvoiceResponse, status_code=201)
def create_purchase_invoice(
    data: PurchaseInvoiceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).create_purchase_invoice(data, current_user.id)


@router.get("/purchase-invoices/{inv_id}", response_model=PurchaseInvoiceResponse)
def get_purchase_invoice(
    inv_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_purchase_invoice(inv_id)


@router.post("/purchase-invoices/from-po/{po_id}", response_model=PurchaseInvoiceResponse, status_code=201)
def create_purchase_invoice_from_po(
    po_id: int,
    thue_suat: Decimal = Query(Decimal("8"), ge=Decimal("0"), le=Decimal("100"), description="VAT: 0, 5, 8, 10"),
    co_vat: bool = Query(True, description="Co hoa don VAT hay khong"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).create_purchase_invoice_from_po(
        po_id, current_user.id, thue_suat=thue_suat, co_vat=co_vat
    )


@router.post("/purchase-invoices/from-gr/{gr_id}", response_model=PurchaseInvoiceResponse, status_code=201)
def create_purchase_invoice_from_gr(
    gr_id: int,
    thue_suat: Decimal = Query(Decimal("8"), ge=Decimal("0"), le=Decimal("100"), description="VAT: 0, 5, 8, 10"),
    co_vat: bool = Query(True, description="Co hoa don VAT hay khong"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).create_purchase_invoice_from_gr(
        gr_id, current_user.id, thue_suat=thue_suat, co_vat=co_vat
    )


@router.patch("/purchase-invoices/{inv_id}/cancel", response_model=PurchaseInvoiceResponse)
def cancel_purchase_invoice(
    inv_id: int,
    ly_do: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    logger.info("cancel_purchase_invoice id=%s user=%s", inv_id, current_user.id)
    return AccountingService(db).cancel_purchase_invoice(inv_id, current_user.id, ly_do)


@router.get("/purchase-invoices/{inv_id}/print", response_class=HTMLResponse)
def print_purchase_invoice(
    inv_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    import html as _html_mod
    from app.models.system import PrintTemplate, SystemSetting

    inv = db.query(PurchaseInvoice).options(
        __import__("sqlalchemy.orm", fromlist=["joinedload"]).joinedload(PurchaseInvoice.payments)
    ).filter(PurchaseInvoice.id == inv_id).first()
    if not inv:
        raise HTTPException(404, "Không tìm thấy hóa đơn mua")

    pn: PhapNhan | None = (
        db.get(PhapNhan, inv.phap_nhan_id) if inv.phap_nhan_id
        else db.query(PhapNhan).filter(PhapNhan.trang_thai.is_(True)).first()
    )

    tpl_q = db.query(PrintTemplate).filter(PrintTemplate.ma_mau == "PURCHASE_INVOICE")
    tpl = tpl_q.filter(PrintTemplate.phap_nhan_id == pn.id).first() if pn else None
    if not tpl:
        tpl = tpl_q.filter(PrintTemplate.phap_nhan_id.is_(None)).first() or tpl_q.first()
    if not tpl:
        raise HTTPException(404, "Chưa có mẫu in PURCHASE_INVOICE — vui lòng cấu hình trong Hệ thống > Mẫu in")

    settings = {s.key: s.value for s in db.query(SystemSetting).all()}
    accent = "#E65100"
    ten_cty = pn.ten_phap_nhan if pn else "CÔNG TY TNHH NAM PHƯƠNG BAO BÌ"
    parts = []
    if pn and pn.dia_chi: parts.append(f"Địa chỉ: {pn.dia_chi}")
    if pn and pn.ma_so_thue: parts.append(f"MST: {pn.ma_so_thue}")
    if pn and pn.so_dien_thoai: parts.append(f"ĐT: {pn.so_dien_thoai}")

    payments_rows = ""
    for p in (inv.payments or []):
        hinh_thuc = HINH_THUC_LABEL.get(p.hinh_thuc_tt, p.hinh_thuc_tt or "")
        payments_rows += (
            f"<tr><td>{_html_mod.escape(p.so_phieu or '')}</td>"
            f"<td>{p.ngay_phieu.strftime('%d/%m/%Y') if p.ngay_phieu else ''}</td>"
            f"<td>{_html_mod.escape(hinh_thuc)}</td>"
            f"<td style='text-align:right'>{float(p.so_tien):,.0f}</td></tr>"
        )
    payments_table = (
        f"<h4 style='margin:16px 0 6px'>Phiếu chi đã tạo</h4>"
        f"<table style='width:100%;border-collapse:collapse;font-size:11px'>"
        f"<thead><tr style='background:{accent};color:#fff'>"
        f"<th style='padding:4px 6px;text-align:left'>Số phiếu</th>"
        f"<th style='padding:4px 6px;text-align:left'>Ngày</th>"
        f"<th style='padding:4px 6px;text-align:left'>Hình thức</th>"
        f"<th style='padding:4px 6px;text-align:right'>Số tiền</th></tr></thead>"
        f"<tbody>{payments_rows}</tbody></table>"
    ) if payments_rows else ""

    replacements = {
        **standard_vars(subtitle="HÓA ĐƠN MUA VÀO", customer_name=_html_mod.escape(inv.ten_don_vi or "")),
        "{{document_number}}": _html_mod.escape(inv.so_hoa_don or f"#{inv.id}"),
        "{{mau_so}}": _html_mod.escape(inv.mau_so or ""),
        "{{ky_hieu}}": _html_mod.escape(inv.ky_hieu or ""),
        "{{document_date}}": ngay_str(inv.ngay_lap),
        "{{han_tt}}": ngay_str(inv.han_tt) if inv.han_tt else "-",
        "{{company_name}}": _html_mod.escape(ten_cty),
        "{{company_details}}": _html_mod.escape(" | ".join(parts)),
        "{{logo_img}}": _logo_img(pn, settings),
        "{{accent}}": accent,
        "{{nha_cung_cap}}": _html_mod.escape(inv.ten_don_vi or ""),
        "{{ma_so_thue}}": _html_mod.escape(inv.ma_so_thue or ""),
        "{{thue_suat}}": str(int(inv.thue_suat)),
        "{{tong_tien_hang}}": f"{float(inv.tong_tien_hang):,.0f}",
        "{{tien_thue}}": f"{float(inv.tien_thue):,.0f}",
        "{{tong_thanh_toan}}": f"{float(inv.tong_thanh_toan):,.0f}",
        "{{da_thanh_toan}}": f"{float(inv.da_thanh_toan):,.0f}",
        "{{con_lai}}": f"{float(inv.con_lai):,.0f}",
        "{{ghi_chu}}": _html_mod.escape(inv.ghi_chu or ""),
        "{{payments_table}}": payments_table,
    }
    content = apply_template(tpl.html_content, replacements)
    page = (
        "<!DOCTYPE html><html lang='vi'><head><meta charset='UTF-8'>"
        f"<title>Hóa đơn mua {_html_mod.escape(inv.so_hoa_don or str(inv.id))}</title>"
        "<style>body{margin:0;padding:0}@media print{.no-print{display:none!important}}</style>"
        "</head><body>"
        "<div class='no-print' style='padding:10px;background:#f0f0f0;display:flex;gap:10px'>"
        f"<button onclick='window.print()' style='padding:7px 18px;background:{accent};color:#fff;border:none;border-radius:4px;cursor:pointer'>🖨️ In phiếu</button>"
        "<button onclick='window.close()' style='padding:7px 14px;border:1px solid #ccc;border-radius:4px;cursor:pointer'>Đóng</button>"
        "</div>"
        f"{content}</body></html>"
    )
    return HTMLResponse(content=page)


# ─────────────────────────────────────────────
# PHIẾU CHI
# ─────────────────────────────────────────────

@router.get("/payments")
def list_payments(
    supplier_id: int | None = Query(None),
    trang_thai: str | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).list_payments(
        supplier_id=supplier_id, trang_thai=trang_thai,
        tu_ngay=tu_ngay, den_ngay=den_ngay,
        phap_nhan_id=phap_nhan_id, phan_xuong_id=phan_xuong_id,
        page=page, page_size=page_size,
    )


@router.post("/payments", response_model=CashPaymentResponse, status_code=201)
def create_payment(
    data: CashPaymentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).create_cash_payment(data, current_user.id)


@router.get("/payments/{payment_id}", response_model=CashPaymentResponse)
def get_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_payment(payment_id)


@router.patch("/payments/{payment_id}/approve", response_model=CashPaymentResponse)
def approve_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).approve_payment(payment_id, current_user.id)


@router.patch("/payments/{payment_id}/cancel", response_model=CashPaymentResponse)
def cancel_payment(
    payment_id: int,
    ly_do: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    logger.info("cancel_payment id=%s user=%s", payment_id, current_user.id)
    return AccountingService(db).cancel_payment(payment_id, current_user.id, ly_do)


@router.put("/payments/{payment_id}", response_model=CashPaymentResponse)
def update_payment(
    payment_id: int,
    data: CashPaymentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).update_payment(payment_id, data, current_user.id)


@router.get("/tax-obligations", response_model=list[TaxObligationItem])
def get_tax_obligations(
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Tính toán các khoản thuế phải nộp từ dữ liệu hệ thống."""
    from app.models.billing import SalesInvoice

    q = db.query(func.coalesce(func.sum(SalesInvoice.tien_vat), 0)).filter(
        SalesInvoice.trang_thai.in_(["da_phat_hanh", "da_tt_mot_phan", "da_tt_du", "qua_han"])
    )
    if tu_ngay:
        q = q.filter(SalesInvoice.ngay_hoa_don >= tu_ngay)
    if den_ngay:
        q = q.filter(SalesInvoice.ngay_hoa_don <= den_ngay)
    if phap_nhan_id:
        q = q.filter(SalesInvoice.phap_nhan_id == phap_nhan_id)

    tong_vat = Decimal(str(q.scalar() or 0))

    return [
        TaxObligationItem(
            loai_thue="gtgt_dau_ra",
            ten_khoan="Thuế GTGT đầu ra",
            tk_no="3331",
            so_phai_nop=tong_vat,
        )
    ]


@router.post("/tax-payments", response_model=TaxPaymentBatchResponse, status_code=201)
def create_tax_payments(
    data: TaxPaymentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Tạo phiếu chi nộp thuế (không cần nhà cung cấp)."""
    svc = AccountingService(db)
    created_ids: list[int] = []
    thanh_cong = 0
    that_bai = 0
    tong_tien = Decimal("0")

    tk_co_default = "111" if data.hinh_thuc_tt in ("tien_mat", "TM") else "112"

    for item in data.items:
        if item.so_tien <= 0:
            continue
        payment_data = CashPaymentCreate(
            supplier_id=None,
            ngay_phieu=data.ngay_phieu,
            hinh_thuc_tt=data.hinh_thuc_tt,
            so_tai_khoan=data.so_tai_khoan,
            dien_giai=item.dien_giai or f"Nop {item.ten_khoan}",
            so_tien=item.so_tien,
            tk_no=item.tk_no,
            tk_co=item.tk_co if item.tk_co != "112" else tk_co_default,
            loai_chi="nop_thue",
            phap_nhan_id=data.phap_nhan_id,
            phan_xuong_id=data.phan_xuong_id,
        )
        try:
            p = svc.create_cash_payment(payment_data, current_user.id)
            created_ids.append(p.id)
            tong_tien += item.so_tien
            thanh_cong += 1
        except Exception:
            that_bai += 1

    return TaxPaymentBatchResponse(
        tong_so=len(data.items),
        thanh_cong=thanh_cong,
        that_bai=that_bai,
        tong_tien=tong_tien,
        phieu_chi_ids=created_ids,
    )


@router.get("/insurance-obligations", response_model=list[InsuranceObligationItem])
def get_insurance_obligations(
    thang: int | None = Query(None),
    nam: int | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Lấy khoản bảo hiểm phải nộp từ bảng lương (PayrollAdjustment)."""
    from app.models.hr import PayrollAdjustment

    BH_MAP = {
        "bhxh":          ("Bảo hiểm xã hội",     "3383"),
        "bhyt":          ("Bảo hiểm y tế",        "3384"),
        "bhtn":          ("Bảo hiểm thất nghiệp", "3385"),
        "cong_doan_phi": ("Kinh phí công đoàn",   "3382"),
    }

    q = db.query(
        PayrollAdjustment.sub_loai,
        func.coalesce(func.sum(PayrollAdjustment.so_tien), 0),
    ).filter(
        PayrollAdjustment.loai == "khau_tru",
        PayrollAdjustment.sub_loai.in_(list(BH_MAP.keys())),
    )
    if thang:
        q = q.filter(PayrollAdjustment.thang == thang)
    if nam:
        q = q.filter(PayrollAdjustment.nam == nam)
    q = q.group_by(PayrollAdjustment.sub_loai)

    rows = q.all()
    result = []
    for sub_loai, tong in rows:
        ten, tk_no = BH_MAP.get(sub_loai, (sub_loai, "338"))
        if Decimal(str(tong)) > 0:
            result.append(InsuranceObligationItem(
                loai_bh=sub_loai,
                ten_khoan=ten,
                tk_no=tk_no,
                so_phai_nop=Decimal(str(tong)),
            ))
    return result


@router.post("/insurance-payments", response_model=InsuranceBatchResponse, status_code=201)
def create_insurance_payments(
    data: InsurancePaymentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Tạo phiếu chi nộp bảo hiểm (không cần nhà cung cấp)."""
    svc = AccountingService(db)
    created_ids: list[int] = []
    thanh_cong = 0
    that_bai = 0
    tong_tien = Decimal("0")

    tk_co_default = "111" if data.hinh_thuc_tt in ("tien_mat", "TM") else "112"

    for item in data.items:
        if item.so_tien <= 0:
            continue
        payment_data = CashPaymentCreate(
            supplier_id=None,
            ngay_phieu=data.ngay_phieu,
            hinh_thuc_tt=data.hinh_thuc_tt,
            so_tai_khoan=data.so_tai_khoan,
            dien_giai=item.dien_giai or f"Nop {item.ten_khoan} thang {data.thang}/{data.nam}",
            so_tien=item.so_tien,
            tk_no=item.tk_no,
            tk_co=item.tk_co if item.tk_co != "112" else tk_co_default,
            loai_chi="nop_bh",
            phap_nhan_id=data.phap_nhan_id,
            phan_xuong_id=data.phan_xuong_id,
        )
        try:
            p = svc.create_cash_payment(payment_data, current_user.id)
            created_ids.append(p.id)
            tong_tien += item.so_tien
            thanh_cong += 1
        except Exception:
            that_bai += 1

    return InsuranceBatchResponse(
        tong_so=len(data.items),
        thanh_cong=thanh_cong,
        that_bai=that_bai,
        tong_tien=tong_tien,
        phieu_chi_ids=created_ids,
    )


@router.get("/salary-obligations", response_model=list[SalaryObligationItem])
def get_salary_obligations(
    thang: int | None = Query(None),
    nam: int | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Lấy danh sách lương chưa trả (PayrollRun trang_thai=da_chot)."""
    from app.models.hr import PayrollRun, Employee, Department

    q = (
        db.query(PayrollRun, Employee, Department)
        .join(Employee, PayrollRun.employee_id == Employee.id)
        .outerjoin(Department, Employee.bo_phan_id == Department.id)
        .filter(PayrollRun.trang_thai == "da_chot")
    )
    if thang:
        q = q.filter(PayrollRun.thang == thang)
    if nam:
        q = q.filter(PayrollRun.nam == nam)
    if phap_nhan_id:
        q = q.filter(Employee.phap_nhan_id == phap_nhan_id)

    result = []
    for run, emp, dept in q.all():
        if run.thuc_linh > 0:
            result.append(SalaryObligationItem(
                payroll_run_id=run.id,
                employee_id=emp.id,
                ma_nv=emp.ma_nv,
                ho_ten=emp.ho_ten,
                don_vi=dept.ten_bo_phan if dept else None,
                so_tai_khoan=emp.so_tk_ngan_hang,
                ten_ngan_hang=emp.ten_ngan_hang,
                so_phai_tra=run.thuc_linh,
                so_tra=run.thuc_linh,
            ))
    return result


@router.post("/salary-payments", response_model=SalaryBatchResponse, status_code=201)
def create_salary_payments(
    data: SalaryPaymentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Tạo phiếu chi trả lương từng nhân viên (không cần nhà cung cấp)."""
    from app.models.hr import PayrollRun

    svc = AccountingService(db)
    created_ids: list[int] = []
    thanh_cong = 0
    that_bai = 0
    tong_tien = Decimal("0")

    tk_co_default = "111" if data.hinh_thuc_tt in ("tien_mat", "TM") else "112"

    for item in data.items:
        if item.so_tien <= 0:
            continue
        payment_data = CashPaymentCreate(
            supplier_id=None,
            ngay_phieu=data.ngay_phieu,
            hinh_thuc_tt=data.hinh_thuc_tt,
            so_tai_khoan=data.so_tai_khoan,
            dien_giai=item.dien_giai or f"Tra luong {item.ho_ten} thang {data.thang}/{data.nam}",
            so_tien=item.so_tien,
            tk_no=item.tk_no,
            tk_co=item.tk_co if item.tk_co != "112" else tk_co_default,
            loai_chi="tra_luong",
            phap_nhan_id=data.phap_nhan_id,
            phan_xuong_id=data.phan_xuong_id,
        )
        try:
            p = svc.create_cash_payment(payment_data, current_user.id)
            created_ids.append(p.id)
            tong_tien += item.so_tien
            thanh_cong += 1
            # Đánh dấu PayrollRun đã thanh toán
            run = db.query(PayrollRun).filter(PayrollRun.id == item.payroll_run_id).first()
            if run:
                run.trang_thai = "da_thanh_toan"
                db.add(run)
        except Exception:
            that_bai += 1

    db.commit()
    return SalaryBatchResponse(
        tong_so=len(data.items),
        thanh_cong=thanh_cong,
        that_bai=that_bai,
        tong_tien=tong_tien,
        phieu_chi_ids=created_ids,
    )


@router.get("/payments/import-template")
def download_payments_import_template(_: User = Depends(get_current_user)):
    """Tải mẫu Excel để import phiếu chi (tương thích MISA)."""
    return build_template_response("mau_import_phieu_chi.xlsx", [
        ImportField("ngay_chung_tu",   "Ngày chứng từ",              required=True,  parser=parse_date,    help_text="DD/MM/YYYY"),
        ImportField("ngay_hach_toan",  "Ngày hạch toán",                              parser=parse_date,    help_text="Để trống = dùng Ngày chứng từ"),
        ImportField("so_chung_tu",     "Số chứng từ",                                 parser=parse_text,    help_text="Để trống - hệ thống tự sinh"),
        ImportField("ma_doi_tuong",    "Mã đối tượng",               required=True,  parser=parse_text,    help_text="Mã NCC (ma_ncc)"),
        ImportField("ten_doi_tuong",   "Tên đối tượng",                               parser=parse_text),
        ImportField("dia_chi",         "Địa chỉ",                                     parser=parse_text),
        ImportField("ly_do_chi",       "Lý do chi",                                   parser=parse_text,    help_text="Lý do chi tiền"),
        ImportField("dien_giai_ld",    "Diễn giải lý do chi",                         parser=parse_text),
        ImportField("loai_tien",       "Loại tiền",                                   parser=parse_text,    help_text="VND"),
        ImportField("ty_gia",          "Tỷ giá",                                      parser=parse_decimal, help_text="1"),
        ImportField("dien_giai_ht",    "Diễn giải (hạch toán)",                       parser=parse_text,    help_text="Diễn giải ghi vào phiếu"),
        ImportField("tk_no",           "TK Nợ",                                       parser=parse_text,    help_text="Mặc định: 331"),
        ImportField("tk_co",           "TK Có",                                       parser=parse_text,    help_text="Mặc định: 112 (CK) / 111 (TM)"),
        ImportField("so_tien",         "Số tiền",                    required=True,  parser=parse_decimal, help_text="Số tiền chi (VND)"),
        ImportField("so_tk_nh",        "Số TK ngân hàng",                             parser=parse_text),
        ImportField("ten_nh",          "Tên ngân hàng",                               parser=parse_text),
        ImportField("httt",            "Hình thức TT",                                parser=parse_text,    help_text="TM=Tiền mặt, CK=Chuyển khoản"),
        ImportField("co_hoa_don",      "Có hóa đơn",                                  parser=parse_text,    help_text="Có hoặc Không"),
        ImportField("tong_tien_hang",  "Giá trị HHDV chưa thuế",                     parser=parse_decimal, help_text="Giá chưa VAT"),
        ImportField("pct_thue_gtgt",   "% thuế GTGT",                                 parser=parse_decimal, help_text="0, 5, 8 hoặc 10"),
        ImportField("tien_thue",       "Tiền thuế GTGT",                              parser=parse_decimal),
        ImportField("ngay_hoa_don",    "Ngày hóa đơn",                               parser=parse_date),
        ImportField("so_hoa_don",      "Số hóa đơn",                                  parser=parse_text),
        ImportField("ma_ncc_hd",       "Mã NCC",                                      parser=parse_text,    help_text="Mã NCC trên hóa đơn"),
        ImportField("ten_ncc_hd",      "Tên NCC",                                     parser=parse_text),
        ImportField("mst_ncc",         "Mã số thuế NCC",                              parser=parse_text),
    ])


@router.post("/payments/import-excel")
async def import_payments_excel(
    file: UploadFile = File(...),
    ngay_phieu: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    dry_run: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Import phiếu chi từ Excel. Cột bắt buộc: ma_ncc (hoặc supplier_id), so_tien.
    Cột tùy chọn: hinh_thuc_tt, dien_giai, so_tham_chieu."""
    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "File Excel không hợp lệ")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "File Excel trống")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    def find_col(*names: str) -> int | None:
        for n in names:
            if n.lower() in headers:
                return headers.index(n.lower())
        return None

    col_supp      = find_col("supplier_id", "ma_ncc", "mã ncc", "mã đối tượng", "ma doi tuong", "doi_tuong", "đối tượng", "doi tuong", "nhà cung cấp")
    col_tien      = find_col("so_tien", "số tiền", "so tien", "tien")
    col_httt      = find_col("hình thức tt", "httt", "hinh_thuc_tt", "loai_chung_tu", "loại chứng từ", "loai chung tu")
    col_dg        = find_col("diễn giải (hạch toán)", "dien_giai_ht", "dien_giai", "diễn giải", "dien giai")
    col_ref       = find_col("so_tham_chieu", "số tham chiếu")
    col_tk        = find_col("số tk ngân hàng", "so_tk_nh", "so tai khoan", "so_tai_khoan", "số tài khoản nh", "so tk nh")
    col_ly_do     = find_col("lý do chi", "ly do chi", "ly_do_chi", "ly_do")
    col_ngay_row  = find_col("ngày chứng từ", "ngay chung tu", "ngay_chung_tu", "ngay_hach_toan", "ngày hạch toán", "ngay hach toan")
    col_so_phieu  = find_col("số chứng từ", "so chung tu", "so_chung_tu", "so_phieu")
    col_tk_no     = find_col("tk nợ", "tk_no", "tk no")
    col_tk_co     = find_col("tk có", "tk_co", "tk co")
    col_co_hd     = find_col("có hóa đơn", "co hoa don", "co_hoa_don")
    col_tien_hang = find_col("giá trị hhdv chưa thuế", "gia tri hhdv chua thue", "tong_tien_hang", "tien_hang")
    col_pct_thue  = find_col("% thuế gtgt", "pct_thue_gtgt", "thue_suat_gtgt", "% thue gtgt")
    col_tien_thue = find_col("tiền thuế gtgt", "tien_thue", "tien thue gtgt")
    col_ngay_hd   = find_col("ngày hóa đơn", "ngay_hoa_don", "ngay hoa don")
    col_so_hd     = find_col("số hóa đơn", "so_hoa_don", "so hoa don")
    col_mst_ncc   = find_col("mã số thuế ncc", "mst_ncc", "ma_so_thue_ncc", "mã số thuế")

    if col_supp is None or col_tien is None:
        raise HTTPException(400, "File thiếu cột bắt buộc: Mã đối tượng (ma_ncc) và Số tiền")

    today = ngay_phieu or date.today()
    service = AccountingService(db)
    results = []
    thanh_cong = 0
    that_bai = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(c is not None for c in row):
            continue
        raw_supp = row[col_supp]
        raw_tien = row[col_tien]
        if raw_supp is None or raw_tien is None:
            continue

        try:
            supp_id = None
            if isinstance(raw_supp, (int, float)):
                supp_id = int(raw_supp)
            else:
                supp = db.query(Supplier).filter(Supplier.ma_ncc == str(raw_supp).strip()).first()
                if supp:
                    supp_id = supp.id
            if not supp_id:
                raise ValueError(f"Không tìm thấy nhà cung cấp '{raw_supp}'")

            so_tien = Decimal(str(raw_tien).replace(",", "").strip())
            row_ngay = today
            if col_ngay_row is not None and row[col_ngay_row]:
                from datetime import datetime as _dt
                raw_ng = row[col_ngay_row]
                if isinstance(raw_ng, date) and not isinstance(raw_ng, datetime):
                    row_ngay = raw_ng
                elif isinstance(raw_ng, datetime):
                    row_ngay = raw_ng.date()
                else:
                    for _fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                        try:
                            row_ngay = _dt.strptime(str(raw_ng).strip(), _fmt).date()
                            break
                        except ValueError:
                            pass

            httt = "CK"
            if col_httt is not None and row[col_httt]:
                raw_ht = str(row[col_httt]).strip().upper()
                if raw_ht in ("TM", "TIỀN MẶT", "TIEN MAT", "TIEN_MAT"):
                    httt = "TM"

            dg_val = None
            if col_dg is not None and row[col_dg]:
                dg_val = str(row[col_dg]).strip()
            elif col_ly_do is not None and row[col_ly_do]:
                dg_val = str(row[col_ly_do]).strip()

            so_phieu_val = None
            if col_so_phieu is not None and row[col_so_phieu]:
                so_phieu_val = str(row[col_so_phieu]).strip() or None

            tk_no_val = "331"
            if col_tk_no is not None and row[col_tk_no]:
                tk_no_val = str(row[col_tk_no]).strip()

            tk_co_val = "112"
            if col_tk_co is not None and row[col_tk_co]:
                tk_co_val = str(row[col_tk_co]).strip()
            elif httt in ("TM", "tien_mat"):
                tk_co_val = "111"

            # Tạo hóa đơn mua hàng nếu có
            purchase_invoice_id = None
            co_hoa_don = False
            if col_co_hd is not None and row[col_co_hd]:
                raw_co = str(row[col_co_hd]).strip().lower()
                co_hoa_don = raw_co in ("có", "co", "1", "true", "x", "yes", "y")

            if co_hoa_don and supp_id:
                tien_hang = None
                if col_tien_hang is not None and row[col_tien_hang]:
                    try:
                        tien_hang = Decimal(str(row[col_tien_hang]).replace(",", "").strip())
                    except Exception:
                        pass

                if tien_hang and tien_hang > 0:
                    thue_suat_hd = Decimal("8")
                    if col_pct_thue is not None and row[col_pct_thue]:
                        try:
                            ts_raw = str(row[col_pct_thue]).replace("%", "").strip()
                            ts_val = Decimal(ts_raw)
                            if ts_val in (Decimal("0"), Decimal("5"), Decimal("8"), Decimal("10")):
                                thue_suat_hd = ts_val
                        except Exception:
                            pass

                    tien_thue_hd = None
                    if col_tien_thue is not None and row[col_tien_thue]:
                        try:
                            tien_thue_hd = Decimal(str(row[col_tien_thue]).replace(",", "").strip())
                        except Exception:
                            pass

                    ngay_hd_val = row_ngay
                    if col_ngay_hd is not None and row[col_ngay_hd]:
                        from datetime import datetime as _dt2
                        raw_nhd = row[col_ngay_hd]
                        if isinstance(raw_nhd, date) and not isinstance(raw_nhd, datetime):
                            ngay_hd_val = raw_nhd
                        elif isinstance(raw_nhd, datetime):
                            ngay_hd_val = raw_nhd.date()
                        else:
                            for _fmt2 in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                                try:
                                    ngay_hd_val = _dt2.strptime(str(raw_nhd).strip(), _fmt2).date()
                                    break
                                except ValueError:
                                    pass

                    so_hd_val = None
                    if col_so_hd is not None and row[col_so_hd]:
                        so_hd_val = str(row[col_so_hd]).strip()

                    mst_val = None
                    if col_mst_ncc is not None and row[col_mst_ncc]:
                        mst_val = str(row[col_mst_ncc]).strip()

                    from app.schemas.accounting import PurchaseInvoiceCreate
                    inv_data = PurchaseInvoiceCreate(
                        supplier_id=supp_id,
                        ngay_lap=row_ngay,
                        ngay_hoa_don=ngay_hd_val,
                        so_hoa_don=so_hd_val,
                        tong_tien_hang=tien_hang,
                        thue_suat=thue_suat_hd,
                        tien_thue=tien_thue_hd,
                        ma_so_thue=mst_val,
                        phap_nhan_id=phap_nhan_id,
                        phan_xuong_id=phan_xuong_id,
                    )
                    try:
                        if not dry_run:
                            inv_obj = service.create_purchase_invoice(inv_data, current_user.id)
                            purchase_invoice_id = inv_obj.id
                    except Exception:
                        pass  # hóa đơn là best-effort, không block payment

            payment_data = CashPaymentCreate(
                supplier_id=supp_id,
                purchase_invoice_id=purchase_invoice_id,
                so_phieu=so_phieu_val,
                ngay_phieu=row_ngay,
                hinh_thuc_tt=httt,
                dien_giai=dg_val,
                so_tham_chieu=str(row[col_ref]).strip() if col_ref is not None and row[col_ref] else None,
                so_tai_khoan=str(row[col_tk]).strip() if col_tk is not None and row[col_tk] else None,
                so_tien=so_tien,
                tk_no=tk_no_val,
                tk_co=tk_co_val,
                phap_nhan_id=phap_nhan_id,
                phan_xuong_id=phan_xuong_id,
            )
            if dry_run:
                results.append({
                    "index": row_idx, "supplier_id": supp_id,
                    "so_phieu": None, "so_tien": float(so_tien),
                    "success": True, "error": None,
                })
            else:
                payment = service.create_cash_payment(payment_data, current_user.id)
                results.append({
                    "index": row_idx, "supplier_id": supp_id,
                    "so_phieu": payment.so_phieu, "so_tien": float(so_tien),
                    "success": True, "error": None,
                })
            thanh_cong += 1
        except Exception as e:
            db.rollback()
            service = AccountingService(db)
            results.append({
                "index": row_idx,
                "supplier_id": 0,
                "so_phieu": None,
                "so_tien": 0.0,
                "success": False,
                "error": f"Dòng {row_idx}: {e}",
            })
            that_bai += 1

    return {
        "tong_so": len(results),
        "thanh_cong": thanh_cong,
        "that_bai": that_bai,
        "items": results,
    }


# ─────────────────────────────────────────────
# SỔ CÔNG NỢ — AR (phải thu)
# ─────────────────────────────────────────────

@router.get("/ar/ledger")
def ar_ledger(
    customer_id: int | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    trang_thai: str | None = Query(None),
    qua_han_only: bool = Query(False),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ar_ledger(
        customer_id=customer_id, tu_ngay=tu_ngay, den_ngay=den_ngay,
        trang_thai=trang_thai, qua_han_only=qua_han_only, phap_nhan_id=phap_nhan_id,
    )


@router.get("/ar/aging")
def ar_aging(
    as_of_date: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ar_aging(as_of_date, phap_nhan_id=phap_nhan_id)


@router.get("/ar/balance")
def ar_balance(
    customer_id: int | None = Query(None),
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ar_balance(customer_id, tu_ngay, den_ngay, phap_nhan_id=phap_nhan_id)


@router.get("/ar/customer-summary", response_model=list[ARCustomerSummaryRow])
def ar_customer_summary(
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ar_customer_summary(phap_nhan_id=phap_nhan_id)


# ─────────────────────────────────────────────
# SỔ CÔNG NỢ — AP (phải trả)
# ─────────────────────────────────────────────

@router.get("/ar/ledger-entries")
def ar_ledger_entries(
    customer_id: int | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ar_ledger_entries(
        customer_id, tu_ngay, den_ngay, phap_nhan_id=phap_nhan_id,
        page=page, page_size=page_size,
    )


@router.get("/ap/ledger")
def ap_ledger(
    supplier_id: int | None = Query(None),
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    trang_thai: str | None = Query(None),
    qua_han_only: bool = Query(False),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ap_ledger(
        supplier_id=supplier_id, tu_ngay=tu_ngay, den_ngay=den_ngay,
        trang_thai=trang_thai, qua_han_only=qua_han_only, phap_nhan_id=phap_nhan_id,
    )


@router.get("/ap/aging")
def ap_aging(
    as_of_date: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ap_aging(as_of_date, phap_nhan_id=phap_nhan_id)


@router.get("/ap/balance")
def ap_balance(
    supplier_id: int | None = Query(None),
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_ap_balance(supplier_id, tu_ngay, den_ngay, phap_nhan_id=phap_nhan_id)


# ─────────────────────────────────────────────
# SỔ CHI TIẾT MUA HÀNG
# ─────────────────────────────────────────────

@router.get("/debt/overdue-alerts")
def debt_overdue_alerts(
    as_of_date: date | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_debt_overdue_alerts(
        as_of_date=as_of_date,
        phap_nhan_id=phap_nhan_id,
        limit=limit,
    )


@router.get("/purchase/so-chi-tiet")
def so_chi_tiet_mua_hang(
    supplier_id: int | None = Query(None),
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_so_chi_tiet_mua_hang(
        supplier_id, tu_ngay, den_ngay,
        phap_nhan_id=phap_nhan_id,
        page=page, page_size=page_size,
    )


# ─────────────────────────────────────────────
# BIÊN BẢN ĐỐI CHIẾU CÔNG NỢ
# ─────────────────────────────────────────────

@router.get("/ap/doi-chieu-phai-tra")
def doi_chieu_phai_tra(
    supplier_id: int | None = Query(default=None),
    tu_ngay: date | None = Query(default=None),
    den_ngay: date | None = Query(default=None),
    phap_nhan_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Đối chiếu phải trả: so sánh GR (hàng nhận) với HĐ mua hàng theo nhà cung cấp.

    Trả về danh sách theo nhà cung cấp với tổng GR, tổng HĐ, chênh lệch.
    """
    # Aggregate GRs
    gr_q = db.query(
        GoodsReceipt.supplier_id,
        func.sum(GoodsReceipt.tong_gia_tri).label("tong_gr"),
        func.count(GoodsReceipt.id).label("so_phieu_gr"),
    ).filter(GoodsReceipt.trang_thai == "da_duyet")
    if supplier_id:
        gr_q = gr_q.filter(GoodsReceipt.supplier_id == supplier_id)
    if phap_nhan_id:
        gr_q = gr_q.filter(GoodsReceipt.phap_nhan_id == phap_nhan_id)
    if tu_ngay:
        gr_q = gr_q.filter(GoodsReceipt.ngay_nhap >= tu_ngay)
    if den_ngay:
        gr_q = gr_q.filter(GoodsReceipt.ngay_nhap <= den_ngay)
    gr_by_sup = {
        row.supplier_id: {"tong_gr": float(row.tong_gr or 0), "so_phieu_gr": row.so_phieu_gr}
        for row in gr_q.group_by(GoodsReceipt.supplier_id).all()
    }

    # Aggregate PurchaseInvoices
    inv_q = db.query(
        PurchaseInvoice.supplier_id,
        func.sum(PurchaseInvoice.tong_thanh_toan).label("tong_hd"),
        func.count(PurchaseInvoice.id).label("so_hoa_don"),
    ).filter(PurchaseInvoice.trang_thai != PurchaseInvoiceStatus.HUY)
    if supplier_id:
        inv_q = inv_q.filter(PurchaseInvoice.supplier_id == supplier_id)
    if phap_nhan_id:
        inv_q = inv_q.filter(PurchaseInvoice.phap_nhan_id == phap_nhan_id)
    if tu_ngay:
        inv_q = inv_q.filter(PurchaseInvoice.ngay_lap >= tu_ngay)
    if den_ngay:
        inv_q = inv_q.filter(PurchaseInvoice.ngay_lap <= den_ngay)
    inv_by_sup = {
        row.supplier_id: {"tong_hd": float(row.tong_hd or 0), "so_hoa_don": row.so_hoa_don}
        for row in inv_q.group_by(PurchaseInvoice.supplier_id).all()
    }

    # Merge by supplier
    all_sup_ids = set(gr_by_sup) | set(inv_by_sup)
    # Batch-load all suppliers up front to avoid an N+1 query inside the loop.
    suppliers_map: dict[int, Supplier] = {}
    if all_sup_ids:
        suppliers_map = {
            s.id: s
            for s in db.query(Supplier).filter(Supplier.id.in_(all_sup_ids)).all()
        }
    rows = []
    for sid in all_sup_ids:
        sup = suppliers_map.get(sid)
        gr = gr_by_sup.get(sid, {"tong_gr": 0.0, "so_phieu_gr": 0})
        inv = inv_by_sup.get(sid, {"tong_hd": 0.0, "so_hoa_don": 0})
        tong_gr = gr["tong_gr"]
        tong_hd = inv["tong_hd"]
        chenh_lech = tong_gr - tong_hd  # >0: GR chưa có HĐ; <0: HĐ vượt GR
        rows.append({
            "supplier_id": sid,
            "ten_ncc": sup.ten_viet_tat if sup else "",
            "ma_ncc": sup.ma_ncc if sup else "",
            "so_phieu_gr": gr["so_phieu_gr"],
            "tong_gia_tri_gr": tong_gr,
            "so_hoa_don": inv["so_hoa_don"],
            "tong_gia_tri_hd": tong_hd,
            "chenh_lech": chenh_lech,
        })
    rows.sort(key=lambda x: abs(x["chenh_lech"]), reverse=True)
    return rows


@router.get("/ap/doi-chieu/{supplier_id}")
def doi_chieu_cong_no(
    supplier_id: int,
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_doi_chieu_cong_no(supplier_id, tu_ngay, den_ngay)


# ─────────────────────────────────────────────
# SỐ DƯ ĐẦU KỲ (nhập từ AMIS khi chuyển đổi)
# ─────────────────────────────────────────────

@router.post("/opening-balances", status_code=201)
def create_opening_balance(
    data: OpeningBalanceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    logger.info("create_opening_balance doi_tuong=%s user=%s", data.doi_tuong, current_user.id)
    return AccountingService(db).create_opening_balance(data, current_user.id)


_OB_AR_FIELDS = [
    ImportField("ma_kh", "Ma KH", required=True, parser=parse_text,
                help_text="Ma khach hang phai ton tai trong he thong"),
    ImportField("ky_mo_so", "Ngay mo so", required=True, parser=parse_text, help_text="YYYY-MM-DD"),
    ImportField("so_du_dau_ky", "So du (VND)", required=True, parser=parse_decimal, help_text="So no phai thu dau ky"),
    ImportField("ghi_chu", "Ghi chu", parser=parse_text),
]

_OB_AP_FIELDS = [
    ImportField("ma_ncc", "Ma NCC", required=True, parser=parse_text,
                help_text="Ma nha cung cap phai ton tai trong he thong"),
    ImportField("ky_mo_so", "Ngay mo so", required=True, parser=parse_text, help_text="YYYY-MM-DD"),
    ImportField("so_du_dau_ky", "So du (VND)", required=True, parser=parse_decimal, help_text="So no phai tra dau ky"),
    ImportField("ghi_chu", "Ghi chu", parser=parse_text),
]


@router.get("/opening-balances/template-ar")
def download_ob_ar_template(_: User = Depends(get_current_user)):
    return build_template_response("mau_import_cong_no_phai_thu_dau_ky.xlsx", _OB_AR_FIELDS)


@router.get("/opening-balances/template-ap")
def download_ob_ap_template(_: User = Depends(get_current_user)):
    return build_template_response("mau_import_cong_no_phai_tra_dau_ky.xlsx", _OB_AP_FIELDS)


@router.post("/opening-balances/import-ar")
async def import_opening_balances_ar(
    commit: bool = Query(default=False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chi chap nhan file Excel .xlsx/.xls")
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "File rong")
    df = pd.read_excel(io.BytesIO(raw), dtype=object)

    # Batch load all customers referenced in the file to avoid N+1 queries
    all_ma_kh = {
        str(src.get("Ma KH", "") or "").strip()
        for _, src in df.iterrows()
        if str(src.get("Ma KH", "") or "").strip()
    }
    kh_map: dict[str, Customer] = {}
    if all_ma_kh:
        kh_map = {
            c.ma_kh: c
            for c in db.query(Customer).filter(Customer.ma_kh.in_(all_ma_kh)).all()
        }

    rows, created, errors_count = [], 0, 0
    objects_to_save = []
    for idx, src in df.iterrows():
        row_no = int(idx) + 2
        errs = []
        ma_kh = str(src.get("Ma KH", "") or "").strip()
        ky_mo_so_str = str(src.get("Ngay mo so", "") or "").strip()
        so_du_raw = src.get("So du (VND)")
        ghi_chu = str(src.get("Ghi chu", "") or "").strip() or None
        if not ma_kh:
            errs.append("Ma KH: bat buoc")
        if not ky_mo_so_str:
            errs.append("Ngay mo so: bat buoc")
        try:
            ky_mo_so = date.fromisoformat(ky_mo_so_str[:10])
        except Exception:
            errs.append("Ngay mo so: sai dinh dang (phai la YYYY-MM-DD)")
            ky_mo_so = None
        try:
            so_du = Decimal(str(so_du_raw).replace(",", "")) if so_du_raw is not None else None
        except Exception:
            errs.append("So du: phai la so")
            so_du = None
        if so_du is None:
            errs.append("So du: bat buoc")
        kh = kh_map.get(ma_kh) if ma_kh else None
        if ma_kh and not kh:
            errs.append(f"Ma KH: khong ton tai '{ma_kh}'")
        if errs:
            errors_count += 1
            rows.append({"row": row_no, "status": "error", "errors": errs, "data": {}})
            continue
        ob = db.query(OpeningBalance).filter(
            OpeningBalance.doi_tuong == "khach_hang",
            OpeningBalance.customer_id == kh.id,
            OpeningBalance.ky_mo_so == ky_mo_so,
        ).first()
        status = "update" if ob else "create"
        objects_to_save.append((ob,
                                {"doi_tuong": "khach_hang",
                                 "customer_id": kh.id,
                                 "ky_mo_so": ky_mo_so,
                                 "so_du_dau_ky": so_du,
                                 "ghi_chu": ghi_chu,
                                 "created_by": current_user.id}))
        created += 1
        rows.append({"row": row_no, "status": status, "errors": [], "data": {"ma_kh": ma_kh, "so_du": str(so_du)}})
    if commit and errors_count == 0:
        for ob, vals in objects_to_save:
            if ob:
                for k, v in vals.items():
                    setattr(ob, k, v)
            else:
                db.add(OpeningBalance(**vals))
        db.commit()
    logger.info("import_ob_ar commit=%s total=%s errors=%s user=%s", commit, len(rows), errors_count, current_user.id)
    return {"commit": commit, "total": len(rows), "created": created, "updated": 0,
            "skipped": 0, "errors": errors_count, "rows": rows[:200]}


@router.post("/opening-balances/import-ap")
async def import_opening_balances_ap(
    commit: bool = Query(default=False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chi chap nhan file Excel .xlsx/.xls")
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "File rong")
    df = pd.read_excel(io.BytesIO(raw), dtype=object)

    # Batch load all suppliers referenced in the file to avoid N+1 queries
    all_ma_ncc = {
        str(src.get("Ma NCC", "") or "").strip()
        for _, src in df.iterrows()
        if str(src.get("Ma NCC", "") or "").strip()
    }
    ncc_map: dict[str, Supplier] = {}
    if all_ma_ncc:
        ncc_map = {
            s.ma_ncc: s
            for s in db.query(Supplier).filter(Supplier.ma_ncc.in_(all_ma_ncc)).all()
        }

    rows, created, errors_count = [], 0, 0
    objects_to_save = []
    for idx, src in df.iterrows():
        row_no = int(idx) + 2
        errs = []
        ma_ncc = str(src.get("Ma NCC", "") or "").strip()
        ky_mo_so_str = str(src.get("Ngay mo so", "") or "").strip()
        so_du_raw = src.get("So du (VND)")
        ghi_chu = str(src.get("Ghi chu", "") or "").strip() or None
        if not ma_ncc:
            errs.append("Ma NCC: bat buoc")
        if not ky_mo_so_str:
            errs.append("Ngay mo so: bat buoc")
        try:
            ky_mo_so = date.fromisoformat(ky_mo_so_str[:10])
        except Exception:
            errs.append("Ngay mo so: sai dinh dang (phai la YYYY-MM-DD)")
            ky_mo_so = None
        try:
            so_du = Decimal(str(so_du_raw).replace(",", "")) if so_du_raw is not None else None
        except Exception:
            errs.append("So du: phai la so")
            so_du = None
        if so_du is None:
            errs.append("So du: bat buoc")
        ncc = ncc_map.get(ma_ncc) if ma_ncc else None
        if ma_ncc and not ncc:
            errs.append(f"Ma NCC: khong ton tai '{ma_ncc}'")
        if errs:
            errors_count += 1
            rows.append({"row": row_no, "status": "error", "errors": errs, "data": {}})
            continue
        ob = db.query(OpeningBalance).filter(
            OpeningBalance.doi_tuong == "nha_cung_cap",
            OpeningBalance.supplier_id == ncc.id,
            OpeningBalance.ky_mo_so == ky_mo_so,
        ).first()
        status = "update" if ob else "create"
        objects_to_save.append((ob,
                                {"doi_tuong": "nha_cung_cap",
                                 "supplier_id": ncc.id,
                                 "ky_mo_so": ky_mo_so,
                                 "so_du_dau_ky": so_du,
                                 "ghi_chu": ghi_chu,
                                 "created_by": current_user.id}))
        created += 1
        rows.append({"row": row_no, "status": status, "errors": [], "data": {"ma_ncc": ma_ncc, "so_du": str(so_du)}})
    if commit and errors_count == 0:
        for ob, vals in objects_to_save:
            if ob:
                for k, v in vals.items():
                    setattr(ob, k, v)
            else:
                db.add(OpeningBalance(**vals))
        db.commit()
    logger.info("import_ob_ap commit=%s total=%s errors=%s user=%s", commit, len(rows), errors_count, current_user.id)
    return {"commit": commit, "total": len(rows), "created": created, "updated": 0,
            "skipped": 0, "errors": errors_count, "rows": rows[:200]}


# ─────────────────────────────────────────────
# SỐ DƯ ĐẦU KỲ — QUỸ TIỀN MẶT
# ─────────────────────────────────────────────

@router.get("/opening-balances/cash/import-template")
def download_ob_cash_template(_: User = Depends(get_current_user)):
    return build_template_response(
        "mau_import_so_du_quy_tien_mat.xlsx",
        [
            ImportField("ky_mo_so", "Ngay mo so", required=True, parser=parse_text,
                        help_text="YYYY-MM-DD — ngay go-live (bat dau mo so quy)"),
            ImportField(
                "so_du_dau_ky",
                "So du (VND)",
                required=True,
                parser=parse_decimal,
                help_text="So du quy tien mat dau ky"),
            ImportField("ghi_chu", "Ghi chu", parser=parse_text),
        ],
    )


@router.post("/opening-balances/cash/import")
async def import_ob_cash(
    commit: bool = Query(default=False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chi chap nhan file Excel .xlsx/.xls")
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "File rong")
    df = pd.read_excel(io.BytesIO(raw), dtype=object)
    rows, created, errors_count = [], 0, 0
    objects_to_save = []
    for idx, src in df.iterrows():
        row_no = int(idx) + 2
        errs = []
        ky_mo_so_str = str(src.get("Ngay mo so", "") or "").strip()
        so_du_raw = src.get("So du (VND)")
        ghi_chu = str(src.get("Ghi chu", "") or "").strip() or None
        if not ky_mo_so_str:
            errs.append("Ngay mo so: bat buoc")
        try:
            ky_mo_so = date.fromisoformat(ky_mo_so_str[:10])
        except Exception:
            errs.append("Ngay mo so: sai dinh dang (phai la YYYY-MM-DD)")
            ky_mo_so = None
        try:
            so_du = Decimal(str(so_du_raw).replace(",", "")) if so_du_raw is not None else None
        except Exception:
            errs.append("So du: phai la so")
            so_du = None
        if so_du is None:
            errs.append("So du: bat buoc")
        if errs:
            errors_count += 1
            rows.append({"row": row_no, "status": "error", "errors": errs, "data": {}})
            continue
        ob = db.query(OpeningBalance).filter(
            OpeningBalance.doi_tuong == "quy_tien_mat",
            OpeningBalance.ky_mo_so == ky_mo_so,
        ).first()
        status = "update" if ob else "create"
        objects_to_save.append((ob, {
            "doi_tuong": "quy_tien_mat",
            "ky_mo_so": ky_mo_so,
            "so_du_dau_ky": so_du,
            "ghi_chu": ghi_chu,
            "created_by": current_user.id,
        }))
        created += 1
        rows.append({"row": row_no, "status": status, "errors": [], "data": {
                    "ky_mo_so": str(ky_mo_so), "so_du": str(so_du)}})
    if commit and errors_count == 0:
        for ob, vals in objects_to_save:
            if ob:
                for k, v in vals.items():
                    setattr(ob, k, v)
            else:
                db.add(OpeningBalance(**vals))
        db.commit()
    return {"commit": commit, "total": len(rows), "created": created, "updated": 0,
            "skipped": 0, "errors": errors_count, "rows": rows[:200]}


# ─────────────────────────────────────────────
# SỔ QUỸ TIỀN MẶT
# ─────────────────────────────────────────────

@router.get("/cash-book")
def cash_book(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_cash_book(tu_ngay, den_ngay, phap_nhan_id, phan_xuong_id)


# ─────────────────────────────────────────────
# SỔ NGÂN HÀNG
# ─────────────────────────────────────────────

@router.get("/bank-ledger")
def bank_ledger(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    so_tai_khoan: str | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_bank_ledger(tu_ngay, den_ngay, so_tai_khoan, phap_nhan_id, phan_xuong_id)


# ─────────────────────────────────────────────
# BÁO CÁO TÀI CHÍNH
# ─────────────────────────────────────────────

@router.get("/reports/pnl")
def get_pnl_report(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Báo cáo Kết quả kinh doanh (P&L)"""
    return AccountingService(db).get_pnl(tu_ngay, den_ngay, phap_nhan_id, phan_xuong_id)


@router.get("/reports/balance-sheet")
def get_balance_sheet(
    ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Bảng cân đối kế toán"""
    return AccountingService(db).get_balance_sheet(ngay, phap_nhan_id)


@router.post("/reports/perform-closing", response_model=ClosingResult)
def perform_closing(
    thang: int = Query(...),
    nam: int = Query(...),
    phap_nhan_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Thực hiện kết chuyển lãi lỗ cuối kỳ"""
    return AccountingService(db).perform_closing(thang, nam, phap_nhan_id, current_user.id)


@router.get("/reports/closing-readiness")
def get_closing_readiness(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2000),
    phap_nhan_id: int = Query(...),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Checklist truoc khi ket chuyen/khoa so."""
    return AccountingService(db).get_closing_readiness(thang, nam, phap_nhan_id, limit)


@router.get("/period-locks")
def list_period_locks(
    phap_nhan_id: int | None = Query(None),
    nam: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Danh sach ky ke toan da khoa/mo theo phap nhan."""
    return AccountingService(db).list_period_locks(phap_nhan_id=phap_nhan_id, nam=nam)


@router.post("/period-locks/unlock")
def unlock_period(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2000),
    phap_nhan_id: int = Query(...),
    ly_do_mo_khoa: str = Query(..., min_length=3),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Mo khoa ky ke toan de sua/chay lai ket chuyen co audit log."""
    return AccountingService(db).unlock_period(
        thang=thang,
        nam=nam,
        phap_nhan_id=phap_nhan_id,
        user_id=current_user.id,
        ly_do_mo_khoa=ly_do_mo_khoa,
    )


# ─────────────────────────────────────────────
# IN PHIẾU
# ─────────────────────────────────────────────

HINH_THUC_LABEL = {
    "tien_mat": "Tiền mặt",
    "chuyen_khoan": "Chuyển khoản",
    "TM": "Tiền mặt",
    "CK": "Chuyển khoản",
    "bu_tru_cong_no": "Bù trừ công nợ",
    "khac": "Khác",
}

@router.get("/receipts/{receipt_id}/print", response_class=HTMLResponse)
def print_receipt(
    receipt_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    import html as _html_mod
    from app.models.system import PrintTemplate, SystemSetting

    r = db.get(CashReceipt, receipt_id)
    if not r:
        raise HTTPException(404, "Không tìm thấy phiếu thu")

    pn: PhapNhan | None = (
        db.get(PhapNhan, r.phap_nhan_id) if r.phap_nhan_id
        else db.query(PhapNhan).filter(PhapNhan.trang_thai.is_(True)).first()
    )

    tpl_q = db.query(PrintTemplate).filter(PrintTemplate.ma_mau == "CASH_RECEIPT")
    tpl = tpl_q.filter(PrintTemplate.phap_nhan_id == pn.id).first() if pn else None
    if not tpl:
        tpl = tpl_q.filter(PrintTemplate.phap_nhan_id.is_(None)).first() or tpl_q.first()
    if not tpl:
        raise HTTPException(404, "Chưa có mẫu in CASH_RECEIPT — vui lòng cấu hình trong Hệ thống > Mẫu in")

    settings = {s.key: s.value for s in db.query(SystemSetting).all()}
    accent = pn.mau_sac_chinh if pn and pn.mau_sac_chinh else "#1565C0"
    ten_cty = pn.ten_phap_nhan if pn else "CÔNG TY TNHH NAM PHƯƠNG BAO BÌ"
    parts = []
    if pn and pn.dia_chi: parts.append(f"Địa chỉ: {pn.dia_chi}")
    if pn and pn.ma_so_thue: parts.append(f"MST: {pn.ma_so_thue}")
    if pn and pn.so_dien_thoai: parts.append(f"ĐT: {pn.so_dien_thoai}")
    ten_kh = r.customer.ten_viet_tat if r.customer else ""
    hinh_thuc = HINH_THUC_LABEL.get(r.hinh_thuc_tt, r.hinh_thuc_tt or "")

    replacements = {
        **standard_vars(subtitle="PHIẾU THU", customer_name=_html_mod.escape(ten_kh), delivery_address=_html_mod.escape((r.customer.dia_chi or "") if r.customer else "")),
        "{{document_number}}": _html_mod.escape(r.so_phieu or ""),
        "{{document_date}}": ngay_str(r.ngay_phieu),
        "{{company_name}}": _html_mod.escape(ten_cty),
        "{{company_details}}": _html_mod.escape(" | ".join(parts)),
        "{{logo_img}}": _logo_img(pn, settings),
        "{{accent}}": accent,
        "{{nguoi_nop}}": _html_mod.escape(ten_kh),
        "{{khach_hang}}": _html_mod.escape(ten_kh),
        "{{dia_chi_kh}}": _html_mod.escape((r.customer.dia_chi or "") if r.customer else ""),
        "{{ly_do_thu}}": _html_mod.escape(r.dien_giai or ""),
        "{{hinh_thuc}}": _html_mod.escape(hinh_thuc),
        "{{so_tien}}": f"{float(r.so_tien):,.0f} đồng",
        "{{so_tien_bang_chu}}": so_thanh_chu(float(r.so_tien)),
        "{{tk_no}}": _html_mod.escape(r.tk_no or ""),
        "{{tk_co}}": _html_mod.escape(r.tk_co or ""),
    }
    content = apply_template(tpl.html_content, replacements)
    page = (
        "<!DOCTYPE html><html lang='vi'><head><meta charset='UTF-8'>"
        f"<title>Phiếu thu {_html_mod.escape(r.so_phieu or '')}</title>"
        "<style>body{margin:0;padding:0}@media print{.no-print{display:none!important}}</style>"
        "</head><body>"
        "<div class='no-print' style='padding:10px;background:#f0f0f0;display:flex;gap:10px'>"
        "<button onclick='window.print()' style='padding:7px 18px;background:#1565C0;color:#fff;border:none;border-radius:4px;cursor:pointer'>🖨️ In phiếu</button>"
        "<button onclick='window.close()' style='padding:7px 14px;border:1px solid #ccc;border-radius:4px;cursor:pointer'>Đóng</button>"
        "</div>"
        f"{content}</body></html>"
    )
    return HTMLResponse(content=page)


@router.get("/payments/{payment_id}/print", response_class=HTMLResponse)
def print_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    import html as _html_mod
    from app.models.system import PrintTemplate, SystemSetting

    p = db.get(CashPayment, payment_id)
    if not p:
        raise HTTPException(404, "Không tìm thấy phiếu chi")

    pn: PhapNhan | None = (
        db.get(PhapNhan, p.phap_nhan_id) if p.phap_nhan_id
        else db.query(PhapNhan).filter(PhapNhan.trang_thai.is_(True)).first()
    )

    tpl_q = db.query(PrintTemplate).filter(PrintTemplate.ma_mau == "CASH_PAYMENT")
    tpl = tpl_q.filter(PrintTemplate.phap_nhan_id == pn.id).first() if pn else None
    if not tpl:
        tpl = tpl_q.filter(PrintTemplate.phap_nhan_id.is_(None)).first() or tpl_q.first()
    if not tpl:
        raise HTTPException(404, "Chưa có mẫu in CASH_PAYMENT — vui lòng cấu hình trong Hệ thống > Mẫu in")

    settings = {s.key: s.value for s in db.query(SystemSetting).all()}
    accent = pn.mau_sac_chinh if pn and pn.mau_sac_chinh else "#B71C1C"
    ten_cty = pn.ten_phap_nhan if pn else "CÔNG TY TNHH NAM PHƯƠNG BAO BÌ"
    parts = []
    if pn and pn.dia_chi: parts.append(f"Địa chỉ: {pn.dia_chi}")
    if pn and pn.ma_so_thue: parts.append(f"MST: {pn.ma_so_thue}")
    if pn and pn.so_dien_thoai: parts.append(f"ĐT: {pn.so_dien_thoai}")
    ten_ncc = p.supplier.ten_viet_tat if p.supplier else ""
    hinh_thuc = HINH_THUC_LABEL.get(p.hinh_thuc_tt, p.hinh_thuc_tt or "")

    replacements = {
        **standard_vars(subtitle="PHIẾU CHI", customer_name=_html_mod.escape(ten_ncc), delivery_address=_html_mod.escape((p.supplier.dia_chi or "") if p.supplier else "")),
        "{{document_number}}": _html_mod.escape(p.so_phieu or ""),
        "{{document_date}}": ngay_str(p.ngay_phieu),
        "{{company_name}}": _html_mod.escape(ten_cty),
        "{{company_details}}": _html_mod.escape(" | ".join(parts)),
        "{{logo_img}}": _logo_img(pn, settings),
        "{{accent}}": accent,
        "{{nguoi_nhan}}": _html_mod.escape(ten_ncc),
        "{{nha_cung_cap}}": _html_mod.escape(ten_ncc),
        "{{dia_chi_ncc}}": _html_mod.escape((p.supplier.dia_chi or "") if p.supplier else ""),
        "{{ly_do_chi}}": _html_mod.escape(p.dien_giai or ""),
        "{{hinh_thuc}}": _html_mod.escape(hinh_thuc),
        "{{so_tien}}": f"{float(p.so_tien):,.0f} đồng",
        "{{so_tien_bang_chu}}": so_thanh_chu(float(p.so_tien)),
        "{{tk_no}}": _html_mod.escape(p.tk_no or ""),
        "{{tk_co}}": _html_mod.escape(p.tk_co or ""),
    }
    content = apply_template(tpl.html_content, replacements)
    page = (
        "<!DOCTYPE html><html lang='vi'><head><meta charset='UTF-8'>"
        f"<title>Phiếu chi {_html_mod.escape(p.so_phieu or '')}</title>"
        "<style>body{margin:0;padding:0}@media print{.no-print{display:none!important}}</style>"
        "</head><body>"
        "<div class='no-print' style='padding:10px;background:#f0f0f0;display:flex;gap:10px'>"
        "<button onclick='window.print()' style='padding:7px 18px;background:#B71C1C;color:#fff;border:none;border-radius:4px;cursor:pointer'>🖨️ In phiếu</button>"
        "<button onclick='window.close()' style='padding:7px 14px;border:1px solid #ccc;border-radius:4px;cursor:pointer'>Đóng</button>"
        "</div>"
        f"{content}</body></html>"
    )
    return HTMLResponse(content=page)


@router.get("/ar/reconciliation/{customer_id}")
def get_customer_reconciliation(
    customer_id: int,
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    """Lấy dữ liệu đối chiếu công nợ khách hàng dựa trên giao hàng và thanh toán."""
    return AccountingService(db).get_customer_reconciliation(customer_id, tu_ngay, den_ngay, phap_nhan_id)


@router.get("/ap/reconciliation/{supplier_id}")
def get_supplier_reconciliation(
    supplier_id: int,
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    """Lấy dữ liệu đối chiếu công nợ nhà cung cấp dựa trên nhập kho và phiếu chi."""
    return AccountingService(db).get_supplier_reconciliation(supplier_id, tu_ngay, den_ngay, phap_nhan_id)


@router.get("/general-ledger")
def get_general_ledger(
    so_tk: str = Query(...),
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    """Lấy sổ cái chi tiết tài khoản."""
    return AccountingService(db).get_general_ledger(so_tk, tu_ngay, den_ngay, phap_nhan_id, phan_xuong_id)


@router.get("/trial-balance")
def get_trial_balance(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    """Lấy bảng cân đối số phát sinh."""
    return AccountingService(db).get_trial_balance(tu_ngay, den_ngay, phap_nhan_id, phan_xuong_id)


def _wb_stream(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/trial-balance/export")
def export_trial_balance(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Xuất bảng CĐPS ra Excel."""
    from app.services.excel_export_service import build_xlsx
    from app.models.system import ExcelTemplate

    tpl = db.query(ExcelTemplate).filter(ExcelTemplate.ma_mau == "TRIAL_BALANCE").first()
    if not tpl:
        raise HTTPException(404, "Chưa cấu hình mẫu Excel TRIAL_BALANCE")

    rows = AccountingService(db).get_trial_balance(tu_ngay, den_ngay, phap_nhan_id, phan_xuong_id)

    items_data = []
    for row_data in rows:
        so_du_dau = float(row_data.get("so_du_dau", 0) or 0)
        so_du_cuoi = float(row_data.get("so_du_cuoi", 0) or 0)
        items_data.append({
            "so_tk": row_data.get("so_tk", ""),
            "ten_tk": row_data.get("ten_tk", ""),
            "du_dau_no": so_du_dau if so_du_dau > 0 else 0,
            "du_dau_co": abs(so_du_dau) if so_du_dau < 0 else 0,
            "phat_sinh_no": float(row_data.get("phat_sinh_no", 0) or 0),
            "phat_sinh_co": float(row_data.get("phat_sinh_co", 0) or 0),
            "du_cuoi_no": so_du_cuoi if so_du_cuoi > 0 else 0,
            "du_cuoi_co": abs(so_du_cuoi) if so_du_cuoi < 0 else 0,
        })

    meta = {"document_number": f"CĐPS {tu_ngay} – {den_ngay}"}
    xlsx_bytes = build_xlsx(tpl, items_data, meta, {})
    filename = f"cdps_{tu_ngay}_{den_ngay}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────
# BÁO CÁO QUẢN TRỊ
# ─────────────────────────────────────────────

@router.get("/reports/workshop-pnl")
def get_workshop_pnl(
    phan_xuong_id: int | None = Query(None),
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Báo cáo Lãi/Lỗ theo Phân xưởng"""
    if not phan_xuong_id:
        raise HTTPException(400, "Vui lòng chọn phân xưởng")
    return AccountingService(db).get_workshop_pnl(phan_xuong_id, tu_ngay, den_ngay)


@router.get("/reports/legal-entity-cashflow")
def get_legal_entity_cashflow(
    phap_nhan_id: int,
    tu_ngay: date,
    den_ngay: date,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Báo cáo Dòng tiền theo Pháp nhân"""
    return AccountingService(db).get_legal_entity_cashflow(phap_nhan_id, tu_ngay, den_ngay)


@router.get("/reports/cashflow-daily")
def get_cashflow_daily(
    ngay: date = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Dòng tiền Group ngày — 3 pháp nhân × per-bank-account + tổng group"""
    return AccountingService(db).get_cashflow_daily(ngay)


@router.get("/reports/group-pnl")
def get_group_pnl(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """P&L Group — 3 pháp nhân side by side + cột Tổng Group"""
    return AccountingService(db).get_group_pnl(tu_ngay, den_ngay)


@router.get("/reports/group-debt")
def get_group_debt(
    as_of_date: date = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Công nợ Group — AR + AP per pháp nhân + tổng Group"""
    return AccountingService(db).get_group_debt_summary(as_of_date)


@router.get("/ar/aging-7bucket")
def get_ar_aging_7bucket(
    as_of_date: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """AR Aging 7 bucket (chua_den_han/trong_han/sap_den_han_10/qua_han_14/qua_han_30/qua_han_60/kho_doi)"""
    return AccountingService(db).get_ar_aging_7bucket(as_of_date, phap_nhan_id)


@router.get("/ap/payment-plan")
def get_ap_payment_plan(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """AP kế hoạch thanh toán theo loại hàng"""
    return AccountingService(db).get_ap_payment_plan(tu_ngay, den_ngay, phap_nhan_id)


@router.get("/reports/production-costing")
def get_production_costing(
    tu_ngay: date,
    den_ngay: date,
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Báo cáo Giá thành Sản xuất thực tế"""
    return AccountingService(db).get_production_costing(tu_ngay, den_ngay, phan_xuong_id)


@router.get("/reports/production-costing/export")
def export_production_costing(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phan_xuong_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Xuất báo cáo Giá thành Sản xuất ra Excel."""
    from app.services.excel_export_service import build_xlsx
    from app.models.system import ExcelTemplate

    tpl = db.query(ExcelTemplate).filter(ExcelTemplate.ma_mau == "PRODUCTION_COSTING").first()
    if not tpl:
        raise HTTPException(404, "Chưa cấu hình mẫu Excel PRODUCTION_COSTING")

    rows = AccountingService(db).get_production_costing(tu_ngay, den_ngay, phan_xuong_id)

    items_data = [
        {
            "so_lenh": row_data.get("so_lenh", ""),
            "ten_hang": row_data.get("ten_hang", ""),
            "dvt": row_data.get("dvt", ""),
            "so_luong": row_data.get("so_luong", 0),
            "cp_nvl": row_data.get("cp_nvl", 0),
            "cp_nhan_cong": row_data.get("cp_nhan_cong", 0),
            "cp_chung": row_data.get("cp_chung", 0),
            "tong_chi_phi": row_data.get("tong_chi_phi", 0),
            "gia_thanh_don_vi": row_data.get("gia_thanh_don_vi", 0),
            "standard_cost": row_data.get("standard_cost", 0),
        }
        for row_data in rows
    ]

    meta = {"document_number": f"Giá thành SX {tu_ngay} – {den_ngay}"}
    xlsx_bytes = build_xlsx(tpl, items_data, meta, {})
    filename = f"gia_thanh_{tu_ngay}_{den_ngay}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reports/workshop-pnl-export")
def export_workshop_pnl(
    phan_xuong_id: int | None = Query(None),
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Xuất báo cáo Lãi/Lỗ Phân xưởng ra Excel."""
    if not phan_xuong_id:
        raise HTTPException(400, "Vui lòng chọn phân xưởng")

    from app.services.excel_export_service import build_xlsx
    from app.models.system import ExcelTemplate

    data = AccountingService(db).get_workshop_pnl(phan_xuong_id, tu_ngay, den_ngay)

    tpl = db.query(ExcelTemplate).filter(ExcelTemplate.ma_mau == "WORKSHOP_PNL").first()
    if not tpl:
        raise HTTPException(404, "Chưa cấu hình mẫu Excel WORKSHOP_PNL — vui lòng cấu hình trong Hệ thống > Mẫu Excel")

    pnl_keys = [
        ("doanh_thu_ngoai", "Doanh thu ngoại"),
        ("doanh_thu_noi_bo", "Doanh thu nội bộ"),
        ("tong_doanh_thu", "Tổng doanh thu"),
        ("gia_von_ngoai", "Giá vốn ngoại"),
        ("gia_von_noi_bo", "Giá vốn nội bộ"),
        ("tong_gia_von", "Tổng giá vốn"),
        ("loi_nhuan_gop", "Lợi nhuận gộp"),
        ("bien_dong_dinh_muc", "Biến động định mức"),
        ("cp_nhan_cong", "CP nhân công"),
        ("cp_khau_hao", "CP khấu hao"),
        ("cp_phan_bo", "CP phân bổ"),
        ("cp_ban_hang", "CP bán hàng"),
        ("cp_quan_ly", "CP quản lý"),
        ("loi_nhuan_thuan", "Lợi nhuận thuần"),
    ]
    items_data = [
        {"chi_tieu": label, "gia_tri": float(data.get(key, 0) or 0)}
        for key, label in pnl_keys
    ]
    meta = {"document_number": f"PNL Phân xưởng {tu_ngay} – {den_ngay}"}
    xlsx_bytes = build_xlsx(tpl, items_data, meta, {})
    filename = f"workshop_pnl_{tu_ngay}_{den_ngay}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────
# BÁO CÁO THUẾ
# ─────────────────────────────────────────────

@router.get("/reports/trial-balance-tax")
def get_trial_balance_tax(
    tu_ngay: date = Query(...),
    den_ngay: date = Query(...),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Bảng CĐPS dùng cho kê khai thuế/BCTC — loại TK nội bộ 5112/6322/1368/3368."""
    return AccountingService(db).get_trial_balance_tax(tu_ngay, den_ngay, phap_nhan_id)


@router.get("/reports/vat-summary")
def get_vat_summary(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2020),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Tổng hợp thuế GTGT theo tháng — đầu ra/đầu vào/số phải nộp (mẫu 01/GTGT)."""
    return AccountingService(db).get_vat_summary(thang, nam, phap_nhan_id)


@router.get("/reports/vat-audit")
def get_vat_audit(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2020),
    phap_nhan_id: int | None = Query(None),
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Kiem soat du lieu VAT truoc khi ke khai."""
    return AccountingService(db).get_vat_audit(thang, nam, phap_nhan_id, limit)


@router.get("/reports/vat-export-output")
def export_vat_output(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2020),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Xuất bảng kê 01-1/GTGT (hóa đơn bán ra) dạng Excel."""
    data = AccountingService(db).export_vat_output_excel(thang, nam, phap_nhan_id)
    filename = f"bang_ke_dau_ra_{thang:02d}_{nam}.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reports/vat-export-input")
def export_vat_input(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2020),
    phap_nhan_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Xuất bảng kê 01-2/GTGT (hóa đơn mua vào) dạng Excel."""
    data = AccountingService(db).export_vat_input_excel(thang, nam, phap_nhan_id)
    filename = f"bang_ke_dau_vao_{thang:02d}_{nam}.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tax/cit-preview")
def preview_cit(
    quy: int = Query(..., ge=1, le=4),
    nam: int = Query(..., ge=2020),
    phap_nhan_id: int = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Preview thuế TNDN theo quý — không tạo bút toán."""
    return AccountingService(db).preview_cit(quy, nam, phap_nhan_id)


@router.post("/tax/provision-cit", status_code=201)
def provision_cit(
    quy: int = Query(..., ge=1, le=4),
    nam: int = Query(..., ge=2020),
    phap_nhan_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Trích lập thuế TNDN theo quý. Tạo bút toán Nợ 821 / Có 3334."""
    try:
        return AccountingService(db).provision_cit(quy, nam, phap_nhan_id, current_user.id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ─────────────────────────────────────────────
# BẢNG LƯƠNG XƯỞNG
# ─────────────────────────────────────────────

@router.get("/workshop-payroll")
def list_workshop_payroll(
    phan_xuong_id: int | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Danh sách bảng lương xưởng"""
    return AccountingService(db).list_workshop_payroll(
        phan_xuong_id, phap_nhan_id, page=page, page_size=page_size,
    )


@router.post("/workshop-payroll", response_model=WorkshopPayrollResponse, status_code=201)
def create_workshop_payroll(
    data: WorkshopPayrollCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Tạo bảng lương xưởng"""
    return AccountingService(db).create_workshop_payroll(data, current_user.id)


@router.get("/journal-entries")
def list_journal_entries(
    tu_ngay: date | None = Query(None),
    den_ngay: date | None = Query(None),
    loai_but_toan: str | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    chung_tu_loai: str | None = Query(None),
    chung_tu_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from app.models.accounting import JournalEntry
    from sqlalchemy import desc
    q = db.query(JournalEntry)
    if tu_ngay:
        q = q.filter(JournalEntry.ngay_but_toan >= tu_ngay)
    if den_ngay:
        q = q.filter(JournalEntry.ngay_but_toan <= den_ngay)
    if loai_but_toan:
        q = q.filter(JournalEntry.loai_but_toan == loai_but_toan)
    if phap_nhan_id:
        q = q.filter(JournalEntry.phap_nhan_id == phap_nhan_id)
    if phan_xuong_id:
        q = q.filter(JournalEntry.phan_xuong_id == phan_xuong_id)
    if chung_tu_loai:
        q = q.filter(JournalEntry.chung_tu_loai == chung_tu_loai)
    if chung_tu_id:
        q = q.filter(JournalEntry.chung_tu_id == chung_tu_id)

    total = q.count()
    items = q.order_by(desc(JournalEntry.ngay_but_toan), desc(JournalEntry.id))\
             .offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.post("/journal-entries", status_code=201)
def create_manual_journal_entry(
    data: ManualJournalEntryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Tạo bút toán tổng hợp thủ công"""
    return AccountingService(db)._create_journal_entry(
        ngay=data.ngay_but_toan,
        dien_giai=data.dien_giai,
        loai_but_toan='tong_hop',
        chung_tu_loai='tong_hop',
        chung_tu_id=None,
        lines=[line.model_dump() for line in data.lines],
        phap_nhan_id=data.phap_nhan_id,
        phan_xuong_id=data.phan_xuong_id,
        user_id=current_user.id,
    )


@router.patch("/workshop-payroll/{wp_id}/approve", response_model=WorkshopPayrollResponse)
def approve_workshop_payroll(
    wp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Duyệt bảng lương và hạch toán vào chi phí xưởng (154)"""
    return AccountingService(db).approve_workshop_payroll(wp_id, current_user.id)


@router.post("/allocate-overhead", response_model=OverheadAllocationResponse)
def allocate_overhead(
    data: OverheadAllocationRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Thực hiện phân bổ chi phí chung cho các xưởng"""
    return AccountingService(db).allocate_overhead(
        tu_ngay=data.tu_ngay,
        den_ngay=data.den_ngay,
        so_tk=data.so_tk,
        allocations=[a.dict() for a in data.allocations],
        phap_nhan_id=data.phap_nhan_id,
        user_id=current_user.id
    )


@router.get("/production-cost-periods")
def list_production_cost_periods(
    phap_nhan_id: int | None = Query(None),
    phan_xuong_id: int | None = Query(None),
    trang_thai: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).list_production_cost_periods(
        phap_nhan_id=phap_nhan_id,
        phan_xuong_id=phan_xuong_id,
        trang_thai=trang_thai,
        page=page,
        page_size=page_size,
    )


@router.post("/production-cost-periods", status_code=201)
def create_production_cost_period(
    data: ProductionCostPeriodCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).create_production_cost_period(data, current_user.id)


@router.get("/production-cost-periods/{period_id}")
def get_production_cost_period(
    period_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).get_production_cost_period(period_id)


@router.post("/production-cost-periods/{period_id}/collect-inputs")
def collect_production_cost_inputs(
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).collect_production_cost_inputs(period_id, current_user.id)


@router.get("/production-cost-periods/{period_id}/allocation-preview")
def preview_production_cost_allocations(
    period_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return AccountingService(db).preview_production_cost_allocations(period_id)


@router.post("/production-cost-periods/{period_id}/calculate")
def calculate_production_cost_period(
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).calculate_production_cost_period(period_id, current_user.id)


@router.post("/production-cost-periods/{period_id}/close")
def close_production_cost_period(
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    return AccountingService(db).close_production_cost_period(period_id, current_user.id)


# ─────────────────────────────────────────────
# TÀI SẢN CỐ ĐỊNH & KHẤU HAO
# ─────────────────────────────────────────────

@router.get("/fixed-assets")
def list_fixed_assets(
    phan_xuong_id: int | None = Query(None),
    phap_nhan_id: int | None = Query(None),
    trang_thai: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from app.models.accounting import FixedAsset
    q = db.query(FixedAsset)
    if phan_xuong_id:
        q = q.filter(FixedAsset.phan_xuong_id == phan_xuong_id)
    if phap_nhan_id:
        q = q.filter(FixedAsset.phap_nhan_id == phap_nhan_id)
    if trang_thai:
        q = q.filter(FixedAsset.trang_thai == trang_thai)
    total = q.count()
    items = (
        q.order_by(FixedAsset.ngay_mua.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [FixedAssetResponse.model_validate(item) for item in items],
    }


@router.get("/fixed-assets/{asset_id}", response_model=FixedAssetResponse)
def get_fixed_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from app.models.accounting import FixedAsset
    asset = db.get(FixedAsset, asset_id)
    if not asset:
        raise HTTPException(404, "Không tìm thấy tài sản")
    return asset


@router.post("/fixed-assets", response_model=FixedAssetResponse, status_code=201)
def create_fixed_asset(
    data: FixedAssetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Đăng ký tài sản cố định mới"""
    return AccountingService(db).create_fixed_asset(data)


@router.post("/fixed-assets/run-depreciation")
def run_depreciation(
    thang: int = Query(...),
    nam: int = Query(...),
    phap_nhan_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    """Chạy khấu hao hàng tháng cho toàn bộ tài sản"""
    return AccountingService(db).run_monthly_depreciation(thang, nam, phap_nhan_id, current_user.id)


# ─────────────────────────────────────────────
# IMPORT EXCEL
# ─────────────────────────────────────────────

@router.get("/fixed-assets/import-template")
def get_fixed_asset_template(_: User = Depends(get_current_user)):
    from app.services.excel_import_service import build_template_response
    from app.services.accounting_import_service import FIXED_ASSET_FIELDS
    return build_template_response("Mau_Import_Tai_San.xlsx", FIXED_ASSET_FIELDS)


@router.post("/fixed-assets/import")
async def import_fixed_assets(
    file: UploadFile = File(...),
    commit: bool = Query(False),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    from app.services.excel_import_service import import_excel
    from app.services.accounting_import_service import FIXED_ASSET_FIELDS, fixed_asset_resolver
    from app.models.accounting import FixedAsset
    return await import_excel(
        db=db, file=file, model=FixedAsset, fields=FIXED_ASSET_FIELDS,
        key_field="ma_ts", commit=commit, resolver=fixed_asset_resolver,
        user=user, loai_du_lieu="tai_san_co_dinh"
    )


@router.get("/workshop-payroll/import-template")
def get_workshop_payroll_template(_: User = Depends(get_current_user)):
    from app.services.excel_import_service import build_template_response
    from app.services.accounting_import_service import WORKSHOP_PAYROLL_FIELDS
    return build_template_response("Mau_Import_Luong_Xuong.xlsx", WORKSHOP_PAYROLL_FIELDS)


@router.post("/workshop-payroll/import")
async def import_workshop_payroll(
    file: UploadFile = File(...),
    commit: bool = Query(False),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    from app.services.excel_import_service import import_excel
    from app.services.accounting_import_service import WORKSHOP_PAYROLL_FIELDS, workshop_payroll_resolver
    from app.models.accounting import WorkshopPayroll
    return await import_excel(
        db=db, file=file, model=WorkshopPayroll, fields=WORKSHOP_PAYROLL_FIELDS,
        key_field="id",  # Payroll usually creates new records
        commit=commit, resolver=workshop_payroll_resolver,
        user=user, loai_du_lieu="luong_xuong"
    )


# ══════════════════════════════════════════════════════
# KHẾ ƯỚC — helpers
# ══════════════════════════════════════════════════════

def _next_so_khe_uoc_vay(db: Session) -> str:
    from datetime import date as _date
    today = _date.today().strftime("%Y%m%d")
    prefix = f"KUV-{today}-"
    last = (db.query(KheUocVay)
              .filter(KheUocVay.so_khe_uoc.like(f"{prefix}%"))
              .order_by(KheUocVay.id.desc()).first())
    seq = int(last.so_khe_uoc.split("-")[-1]) + 1 if last else 1
    return f"{prefix}{seq:03d}"


def _next_so_khe_uoc_cho_vay(db: Session) -> str:
    from datetime import date as _date
    today = _date.today().strftime("%Y%m%d")
    prefix = f"KUC-{today}-"
    last = (db.query(KheUocChoVay)
              .filter(KheUocChoVay.so_khe_uoc.like(f"{prefix}%"))
              .order_by(KheUocChoVay.id.desc()).first())
    seq = int(last.so_khe_uoc.split("-")[-1]) + 1 if last else 1
    return f"{prefix}{seq:03d}"


def _generate_lich_tra(
    so_tien: Decimal,
    lai_suat_nam: Decimal,
    ky_tinh_lai: str,
    phuong_thuc_tra: str,
    ngay_hieu_luc: date,
    ngay_ket_thuc: date,
    loai_khe_uoc: str,
    khe_uoc_id: int,
) -> list[LichTraNo]:
    from dateutil.relativedelta import relativedelta
    import math

    # Số kỳ và bước nhảy
    if ky_tinh_lai == "thang":
        step = relativedelta(months=1)
    elif ky_tinh_lai == "quy":
        step = relativedelta(months=3)
    else:  # nam
        step = relativedelta(years=1)

    # Đếm số kỳ
    n = 0
    d = ngay_hieu_luc
    while True:
        d = d + step
        if d > ngay_ket_thuc:
            break
        n += 1
    if n == 0:
        n = 1

    r = float(lai_suat_nam) / 100.0 / (12 if ky_tinh_lai == "thang" else 4 if ky_tinh_lai == "quy" else 1)
    P = float(so_tien)

    records = []
    remaining = P
    ngay_ky = ngay_hieu_luc

    for k in range(1, n + 1):
        ngay_den_han = ngay_hieu_luc + step * k
        if k == n:
            ngay_den_han = ngay_ket_thuc

        if phuong_thuc_tra == "gop_deu":
            if r == 0:
                goc = P / n
                lai = 0.0
            else:
                M = P * r * (1 + r) ** n / ((1 + r) ** n - 1)
                lai = remaining * r
                goc = M - lai
            remaining -= goc
            if k == n:
                goc += remaining  # rounding correction
                remaining = 0

        elif phuong_thuc_tra == "goc_deu":
            goc = P / n
            lai = remaining * r
            remaining -= goc
            if k == n:
                goc += remaining
                remaining = 0

        else:  # cuoi_ky
            if k < n:
                goc = 0.0
                lai = remaining * r
            else:
                goc = P
                lai = remaining * r

        tong = goc + lai
        records.append(LichTraNo(
            loai_khe_uoc=loai_khe_uoc,
            khe_uoc_id=khe_uoc_id,
            ky_so=k,
            ngay_den_han=ngay_den_han,
            so_tien_goc=Decimal(str(round(goc, 0))),
            so_tien_lai=Decimal(str(round(lai, 0))),
            tong_cong=Decimal(str(round(tong, 0))),
            trang_thai="chua_tra",
        ))
    return records


# ══════════════════════════════════════════════════════
# KHẾ ƯỚC ĐI VAY — CRUD
# ══════════════════════════════════════════════════════

@router.get("/khe-uoc-vay", response_model=list[KheUocVayResponse])
def list_khe_uoc_vay(
    trang_thai: str | None = None,
    phap_nhan_id: int | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(KheUocVay)
    if trang_thai:
        q = q.filter(KheUocVay.trang_thai == trang_thai)
    if phap_nhan_id:
        q = q.filter(KheUocVay.phap_nhan_id == phap_nhan_id)
    items = q.order_by(KheUocVay.id.desc()).all()
    result = []
    for item in items:
        lich = db.query(LichTraNo).filter(
            LichTraNo.loai_khe_uoc == "di_vay",
            LichTraNo.khe_uoc_id == item.id
        ).order_by(LichTraNo.ky_so).all()
        d = KheUocVayResponse.model_validate(item)
        d.lich_tra = [LichTraNoResponse.model_validate(l) for l in lich]
        result.append(d)
    return result


@router.post("/khe-uoc-vay", response_model=KheUocVayResponse, status_code=201)
def create_khe_uoc_vay(
    body: KheUocVayCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = KheUocVay(
        **body.model_dump(),
        so_khe_uoc=_next_so_khe_uoc_vay(db),
        created_by=user.id,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return KheUocVayResponse.model_validate(obj)


# ── Khế ước đi vay: Import / Export ──────────────────────────────────────────

_KUV_IMPORT_FIELDS = [
    ImportField("so_khe_uoc", "Số hợp đồng/Số khế ước", required=False, parser=parse_text,
                aliases=("Số khế ước",),
                help_text="Để trống = tạo mới. Điền vào = cập nhật nếu trùng số."),
    ImportField("ngay_ky", "Ngày ký", required=True, parser=parse_date,
                help_text="dd/mm/yyyy"),
    ImportField("ngay_hieu_luc", "Ngày giải ngân", required=True, parser=parse_date,
                aliases=("Ngày hiệu lực",),
                help_text="dd/mm/yyyy — ngày tiền được giải ngân vào tài khoản"),
    ImportField("ngay_ket_thuc", "Ngày kết thúc", required=True, parser=parse_date,
                help_text="dd/mm/yyyy — ngày đáo hạn cuối cùng"),
    ImportField("to_chuc_cho_vay", "Đối tượng cho vay", required=True, parser=parse_text,
                aliases=("Tổ chức cho vay",),
                help_text="Tên ngân hàng / tổ chức tín dụng"),
    ImportField("so_tien_vay", "Giá trị khoản vay", required=True, parser=parse_decimal,
                aliases=("Hạn mức tín dụng", "Số tiền vay"),
                help_text="Đơn vị: đồng"),
    ImportField("lai_suat", "Lãi suất hiện tại", required=True, parser=parse_decimal,
                aliases=("Lãi suất (%/năm)",),
                help_text="VD: 8.5 = 8.5%/năm"),
    ImportField("ky_tinh_lai", "Kỳ tính lãi", required=False, parser=parse_text,
                default="thang", help_text="thang | quy | nam"),
    ImportField("phuong_thuc_tra", "Phương thức trả", required=False, parser=parse_text,
                default="gop_deu", help_text="goc_deu | gop_deu | cuoi_ky"),
    ImportField("tai_khoan_nhan", "TK nhận tiền", required=False, parser=parse_text,
                help_text="VD: 112.01"),
    ImportField("tai_san_the_chap", "Tài sản thế chấp", required=False, parser=parse_text),
    ImportField("ghi_chu", "Mục đích vay", required=False, parser=parse_text,
                aliases=("Mục đích vay / Ghi chú",),
                help_text="Mục đích sử dụng vốn vay"),
]


@router.get("/khe-uoc-vay/import-template")
def download_kuv_import_template(_: User = Depends(get_current_user)):
    return build_template_response("mau_import_khe_uoc_di_vay.xlsx", _KUV_IMPORT_FIELDS)


@router.post("/khe-uoc-vay/import")
async def import_khe_uoc_vay(
    file: UploadFile = File(...),
    commit: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    _VALID_KYL = {"thang", "quy", "nam"}
    _VALID_PT = {"goc_deu", "gop_deu", "cuoi_ky"}

    def _resolver(db: Session, values: dict):
        if values.get("ky_tinh_lai") not in _VALID_KYL:
            values["ky_tinh_lai"] = "thang"
        if values.get("phuong_thuc_tra") not in _VALID_PT:
            values["phuong_thuc_tra"] = "gop_deu"
        values.setdefault("trang_thai", "hieu_luc")
        values["created_by"] = current_user.id
        return values, []

    return await import_excel(
        db=db,
        file=file,
        model=KheUocVay,
        fields=_KUV_IMPORT_FIELDS,
        key_field="so_khe_uoc",
        commit=commit,
        resolver=_resolver,
        user=current_user,
        loai_du_lieu="khe_uoc_vay",
    )


@router.get("/khe-uoc-vay/export")
def export_khe_uoc_vay(
    phap_nhan_id: int | None = Query(None),
    trang_thai: str | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from datetime import date as _date
    q = db.query(KheUocVay)
    if phap_nhan_id:
        q = q.filter(KheUocVay.phap_nhan_id == phap_nhan_id)
    if trang_thai:
        q = q.filter(KheUocVay.trang_thai == trang_thai)
    items = q.order_by(KheUocVay.ngay_ky.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Khe uoc di vay"

    hdr_fill = PatternFill("solid", fgColor="1565C0")
    hdr_font = Font(bold=True, color="FFFFFF")

    HEADERS = [
        "STT",
        "Số hợp đồng/Số khế ước",
        "Ngày giải ngân",
        "Ngày ký",
        "Thời hạn vay",
        "Đối tượng cho vay",
        "Loại tiền",
        "Hạn mức tín dụng",
        "Giá trị khoản vay",
        "Giá trị đã giải ngân",
        "Nợ gốc đã trả",
        "Dư nợ hiện tại",
        "Lãi phải trả",
        "Lãi đã trả",
        "Lãi suất hiện tại",
        "Mục đích vay",
        "Trạng thái",
        "Ngày trả gốc tiếp theo",
        "Ngày trả lãi tiếp theo",
        "Thuộc hợp đồng tín dụng",
    ]

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    _TRANG_THAI_LABEL = {"hieu_luc": "Hiệu lực", "da_tra": "Đã trả", "huy": "Hủy"}

    for stt, obj in enumerate(items, 1):
        _pending = ("chua_tra", "qua_han")
        lich = db.query(LichTraNo).filter(
            LichTraNo.khe_uoc_id == obj.id,
            LichTraNo.loai_khe_uoc == "di_vay",
        ).order_by(LichTraNo.ngay_den_han).all()

        goc_da_tra = sum(Decimal(str(r.so_tien_goc)) for r in lich if r.trang_thai == "da_tra")
        lai_da_tra = sum(Decimal(str(r.so_tien_lai)) for r in lich if r.trang_thai == "da_tra")
        lai_con    = sum(Decimal(str(r.so_tien_lai)) for r in lich if r.trang_thai in _pending)
        du_no      = Decimal(str(obj.so_tien_vay)) - goc_da_tra

        # Ngày trả gốc và lãi tiếp theo (có thể khác nhau với cuoi_ky)
        next_goc = next((r for r in lich if r.trang_thai in _pending and Decimal(str(r.so_tien_goc)) > 0), None)
        next_lai = next((r for r in lich if r.trang_thai in _pending and Decimal(str(r.so_tien_lai)) > 0), None)

        thoi_han = (
            (obj.ngay_ket_thuc.year - obj.ngay_hieu_luc.year) * 12
            + (obj.ngay_ket_thuc.month - obj.ngay_hieu_luc.month)
        )

        pn = db.get(PhapNhan, obj.phap_nhan_id) if obj.phap_nhan_id else None

        ws.append([
            stt,
            obj.so_khe_uoc,
            obj.ngay_hieu_luc.strftime("%d/%m/%Y") if obj.ngay_hieu_luc else "",
            obj.ngay_ky.strftime("%d/%m/%Y") if obj.ngay_ky else "",
            f"{thoi_han} tháng",
            obj.to_chuc_cho_vay,
            "VNĐ",
            float(obj.so_tien_vay),
            float(obj.so_tien_vay),
            float(obj.so_tien_vay),
            float(goc_da_tra),
            float(du_no),
            float(lai_con),
            float(lai_da_tra),
            float(obj.lai_suat),
            obj.ghi_chu or "",
            _TRANG_THAI_LABEL.get(obj.trang_thai, obj.trang_thai),
            next_goc.ngay_den_han.strftime("%d/%m/%Y") if next_goc else "",
            next_lai.ngay_den_han.strftime("%d/%m/%Y") if next_lai else "",
            pn.ten_phap_nhan if pn else "",
        ])

    for col in ws.columns:
        max_w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_w + 2, 12), 38)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="khe_uoc_di_vay.xlsx"'},
    )


@router.get("/khe-uoc-vay/{id}", response_model=KheUocVayResponse)
def get_khe_uoc_vay(
    id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    lich = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay", LichTraNo.khe_uoc_id == id
    ).order_by(LichTraNo.ky_so).all()
    d = KheUocVayResponse.model_validate(obj)
    d.lich_tra = [LichTraNoResponse.model_validate(l) for l in lich]
    return d


@router.put("/khe-uoc-vay/{id}", response_model=KheUocVayResponse)
def update_khe_uoc_vay(
    id: int,
    body: KheUocVayUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    if obj.trang_thai != "hieu_luc":
        raise HTTPException(400, "Chỉ cập nhật khế ước đang hiệu lực")
    changes = body.model_dump(exclude_none=True)
    lai_suat_changed = "lai_suat" in changes and changes["lai_suat"] != obj.lai_suat
    for k, v in changes.items():
        setattr(obj, k, v)
    if lai_suat_changed:
        db.query(LichTraNo).filter(
            LichTraNo.loai_khe_uoc == "di_vay",
            LichTraNo.khe_uoc_id == id,
            LichTraNo.trang_thai != "da_tra",
        ).delete()
    db.commit()
    db.refresh(obj)
    return KheUocVayResponse.model_validate(obj)


@router.post("/khe-uoc-vay/{id}/generate-schedule", response_model=list[LichTraNoResponse])
def generate_schedule_vay(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    # Xóa lịch cũ nếu có
    db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay", LichTraNo.khe_uoc_id == id
    ).delete()
    records = _generate_lich_tra(
        so_tien=obj.so_tien_vay,
        lai_suat_nam=obj.lai_suat,
        ky_tinh_lai=obj.ky_tinh_lai,
        phuong_thuc_tra=obj.phuong_thuc_tra,
        ngay_hieu_luc=obj.ngay_hieu_luc,
        ngay_ket_thuc=obj.ngay_ket_thuc,
        loai_khe_uoc="di_vay",
        khe_uoc_id=id,
    )
    db.add_all(records)
    db.commit()
    return [LichTraNoResponse.model_validate(r) for r in records]


@router.delete("/khe-uoc-vay/{id}/schedule", status_code=204)
def delete_schedule_vay(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay", LichTraNo.khe_uoc_id == id
    ).delete()
    db.commit()


@router.patch("/khe-uoc-vay/{id}/tra-no")
def tra_no_vay(
    id: int,
    body: TraNoRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    ky = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay",
        LichTraNo.khe_uoc_id == id,
        LichTraNo.ky_so == body.ky_so,
    ).first()
    if not ky:
        raise HTTPException(404, f"Không tìm thấy kỳ {body.ky_so}")
    ky.trang_thai = "da_tra"
    ky.ngay_tra_thuc = body.ngay_tra_thuc
    ky.so_tien_tra_thuc = body.so_tien_tra_thuc
    db.commit()
    return {"ok": True}


@router.patch("/khe-uoc-vay/{id}/ket-thuc")
def ket_thuc_khe_uoc_vay(
    id: int,
    body: TatToanRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    if obj.trang_thai != "hieu_luc":
        raise HTTPException(400, "Khế ước không còn hiệu lực")
    remaining = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay",
        LichTraNo.khe_uoc_id == id,
        LichTraNo.trang_thai.in_(["chua_tra", "qua_han"]),
    ).all()
    for ky in remaining:
        ky.trang_thai = "da_tra"
        ky.ngay_tra_thuc = body.ngay_tat_toan
        ky.so_tien_tra_thuc = ky.tong_cong
    obj.trang_thai = "da_tra"
    db.commit()
    return {"ok": True, "ky_count": len(remaining)}


@router.delete("/khe-uoc-vay/{id}", status_code=204)
def delete_khe_uoc_vay(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    paid = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay",
        LichTraNo.khe_uoc_id == id,
        LichTraNo.trang_thai == "da_tra",
    ).count()
    if paid > 0:
        raise HTTPException(400, "Không thể xóa khế ước đã có kỳ thanh toán")
    db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay",
        LichTraNo.khe_uoc_id == id,
    ).delete()
    db.delete(obj)
    db.commit()


@router.patch("/khe-uoc-vay/{id}/lai-suat")
def cap_nhat_lai_suat_vay(
    id: int,
    body: CapNhatLaiSuatRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    if obj.trang_thai != "hieu_luc":
        raise HTTPException(400, "Chỉ thay đổi lãi suất khế ước đang hiệu lực")
    obj.lai_suat = body.lai_suat
    # Xóa các kỳ chưa trả để tái tạo lịch sau
    db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay",
        LichTraNo.khe_uoc_id == id,
        LichTraNo.trang_thai.in_(["chua_tra", "qua_han"]),
    ).delete()
    db.commit()
    return {"ok": True}


@router.post("/khe-uoc-vay/{id}/nhan-ban", response_model=KheUocVayResponse, status_code=201)
def nhan_ban_khe_uoc_vay(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    new_obj = KheUocVay(
        so_khe_uoc=_next_so_khe_uoc_vay(db),
        ngay_ky=obj.ngay_ky,
        ngay_hieu_luc=obj.ngay_hieu_luc,
        ngay_ket_thuc=obj.ngay_ket_thuc,
        to_chuc_cho_vay=obj.to_chuc_cho_vay,
        so_tien_vay=obj.so_tien_vay,
        lai_suat=obj.lai_suat,
        ky_tinh_lai=obj.ky_tinh_lai,
        phuong_thuc_tra=obj.phuong_thuc_tra,
        tai_khoan_nhan=obj.tai_khoan_nhan,
        tai_san_the_chap=obj.tai_san_the_chap,
        ghi_chu=obj.ghi_chu,
        trang_thai="hieu_luc",
        phap_nhan_id=obj.phap_nhan_id,
        created_by=user.id,
    )
    db.add(new_obj)
    db.commit()
    db.refresh(new_obj)
    return KheUocVayResponse.model_validate(new_obj)


@router.patch("/khe-uoc-vay/{id}/tra-truoc-han")
def tra_truoc_han_vay(
    id: int,
    body: TraTruocHanRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocVay).filter(KheUocVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước đi vay")
    if obj.trang_thai != "hieu_luc":
        raise HTTPException(400, "Khế ước không còn hiệu lực")
    remaining = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay",
        LichTraNo.khe_uoc_id == id,
        LichTraNo.trang_thai.in_(["chua_tra", "qua_han"]),
    ).all()
    for ky in remaining:
        ky.trang_thai = "da_tra"
        ky.ngay_tra_thuc = body.ngay_tra_thuc
        ky.so_tien_tra_thuc = ky.tong_cong
    obj.trang_thai = "da_tra"
    db.commit()
    return {"ok": True, "ky_count": len(remaining)}


# ══════════════════════════════════════════════════════
# KHẾ ƯỢC CHO VAY — CRUD
# ══════════════════════════════════════════════════════

@router.get("/khe-uoc-cho-vay", response_model=list[KheUocChoVayResponse])
def list_khe_uoc_cho_vay(
    trang_thai: str | None = None,
    phap_nhan_id: int | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(KheUocChoVay)
    if trang_thai:
        q = q.filter(KheUocChoVay.trang_thai == trang_thai)
    if phap_nhan_id:
        q = q.filter(KheUocChoVay.phap_nhan_id == phap_nhan_id)
    items = q.order_by(KheUocChoVay.id.desc()).all()
    result = []
    for item in items:
        lich = db.query(LichTraNo).filter(
            LichTraNo.loai_khe_uoc == "cho_vay",
            LichTraNo.khe_uoc_id == item.id
        ).order_by(LichTraNo.ky_so).all()
        d = KheUocChoVayResponse.model_validate(item)
        d.lich_tra = [LichTraNoResponse.model_validate(l) for l in lich]
        result.append(d)
    return result


@router.post("/khe-uoc-cho-vay", response_model=KheUocChoVayResponse, status_code=201)
def create_khe_uoc_cho_vay(
    body: KheUocChoVayCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = KheUocChoVay(
        **body.model_dump(),
        so_khe_uoc=_next_so_khe_uoc_cho_vay(db),
        created_by=user.id,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return KheUocChoVayResponse.model_validate(obj)


@router.get("/khe-uoc-cho-vay/{id}", response_model=KheUocChoVayResponse)
def get_khe_uoc_cho_vay(
    id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    obj = db.query(KheUocChoVay).filter(KheUocChoVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước cho vay")
    lich = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "cho_vay", LichTraNo.khe_uoc_id == id
    ).order_by(LichTraNo.ky_so).all()
    d = KheUocChoVayResponse.model_validate(obj)
    d.lich_tra = [LichTraNoResponse.model_validate(l) for l in lich]
    return d


@router.put("/khe-uoc-cho-vay/{id}", response_model=KheUocChoVayResponse)
def update_khe_uoc_cho_vay(
    id: int,
    body: KheUocChoVayUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocChoVay).filter(KheUocChoVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước cho vay")
    if obj.trang_thai != "hieu_luc":
        raise HTTPException(400, "Chỉ cập nhật khế ước đang hiệu lực")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return KheUocChoVayResponse.model_validate(obj)


@router.post("/khe-uoc-cho-vay/{id}/generate-schedule", response_model=list[LichTraNoResponse])
def generate_schedule_cho_vay(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocChoVay).filter(KheUocChoVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước cho vay")
    db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "cho_vay", LichTraNo.khe_uoc_id == id
    ).delete()
    records = _generate_lich_tra(
        so_tien=obj.so_tien_cho_vay,
        lai_suat_nam=obj.lai_suat,
        ky_tinh_lai=obj.ky_tinh_lai,
        phuong_thuc_tra=obj.phuong_thuc_tra,
        ngay_hieu_luc=obj.ngay_hieu_luc,
        ngay_ket_thuc=obj.ngay_ket_thuc,
        loai_khe_uoc="cho_vay",
        khe_uoc_id=id,
    )
    db.add_all(records)
    db.commit()
    return [LichTraNoResponse.model_validate(r) for r in records]


@router.delete("/khe-uoc-cho-vay/{id}/schedule", status_code=204)
def delete_schedule_cho_vay(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "cho_vay", LichTraNo.khe_uoc_id == id
    ).delete()
    db.commit()


@router.patch("/khe-uoc-cho-vay/{id}/tra-no")
def tra_no_cho_vay(
    id: int,
    body: TraNoRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    ky = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "cho_vay",
        LichTraNo.khe_uoc_id == id,
        LichTraNo.ky_so == body.ky_so,
    ).first()
    if not ky:
        raise HTTPException(404, f"Không tìm thấy kỳ {body.ky_so}")
    ky.trang_thai = "da_tra"
    ky.ngay_tra_thuc = body.ngay_tra_thuc
    ky.so_tien_tra_thuc = body.so_tien_tra_thuc
    db.commit()
    return {"ok": True}


@router.patch("/khe-uoc-cho-vay/{id}/ket-thuc")
def ket_thuc_khe_uoc_cho_vay(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*KE_TOAN_ROLES)),
):
    obj = db.query(KheUocChoVay).filter(KheUocChoVay.id == id).first()
    if not obj:
        raise HTTPException(404, "Không tìm thấy khế ước cho vay")
    obj.trang_thai = "da_tra"
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════
# DỰ BÁO DÒNG TIỀN
# ══════════════════════════════════════════════════════

@router.get("/cash-flow/forecast", response_model=CashFlowForecastResponse)
def cash_flow_forecast(
    days: int = Query(30, ge=7, le=365),
    phap_nhan_id: int | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from datetime import date as _date, timedelta
    from collections import defaultdict

    today = _date.today()
    end_date = today + timedelta(days=days)

    # Khởi tạo map ngày → numbers
    day_map: dict[_date, dict] = defaultdict(lambda: {
        "thu": Decimal("0"),
        "chi": Decimal("0"),
        "tra_no": Decimal("0"),
        "thu_no": Decimal("0"),
    })

    # Phiếu thu chưa duyệt (thu dự kiến)
    q_thu = db.query(CashReceipt).filter(
        CashReceipt.trang_thai == "cho_duyet",
        CashReceipt.ngay_phieu >= today,
        CashReceipt.ngay_phieu <= end_date,
    )
    if phap_nhan_id:
        q_thu = q_thu.filter(CashReceipt.phap_nhan_id == phap_nhan_id)
    for r in q_thu.all():
        day_map[r.ngay_phieu]["thu"] += r.so_tien or Decimal("0")

    # Phiếu chi chưa chốt (chi dự kiến)
    q_chi = db.query(CashPayment).filter(
        CashPayment.trang_thai == "cho_chot",
        CashPayment.ngay_phieu >= today,
        CashPayment.ngay_phieu <= end_date,
    )
    if phap_nhan_id:
        q_chi = q_chi.filter(CashPayment.phap_nhan_id == phap_nhan_id)
    for p in q_chi.all():
        day_map[p.ngay_phieu]["chi"] += p.so_tien or Decimal("0")

    # Lịch trả nợ đi vay sắp đến hạn
    q_tra = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "di_vay",
        LichTraNo.trang_thai == "chua_tra",
        LichTraNo.ngay_den_han >= today,
        LichTraNo.ngay_den_han <= end_date,
    )
    for l in q_tra.all():
        day_map[l.ngay_den_han]["tra_no"] += l.tong_cong or Decimal("0")

    # Lịch thu nợ cho vay sắp đến hạn
    q_thu_no = db.query(LichTraNo).filter(
        LichTraNo.loai_khe_uoc == "cho_vay",
        LichTraNo.trang_thai == "chua_tra",
        LichTraNo.ngay_den_han >= today,
        LichTraNo.ngay_den_han <= end_date,
    )
    for l in q_thu_no.all():
        day_map[l.ngay_den_han]["thu_no"] += l.tong_cong or Decimal("0")

    # Tạo array từ today đến end_date
    items = []
    luy_ke = Decimal("0")
    cur = today
    while cur <= end_date:
        d = day_map[cur]
        thu = d["thu"]
        chi = d["chi"]
        tra_no = d["tra_no"]
        thu_no = d["thu_no"]
        net = thu + thu_no - chi - tra_no
        luy_ke += net
        items.append(ForecastDayItem(
            ngay=cur,
            thu=thu,
            chi=chi,
            tra_no=tra_no,
            thu_no=thu_no,
            net=net,
            luy_ke=luy_ke,
        ))
        cur += timedelta(days=1)

    return CashFlowForecastResponse(
        days=days,
        phap_nhan_id=phap_nhan_id,
        items=items,
        tong_thu=sum(i.thu for i in items),
        tong_chi=sum(i.chi for i in items),
        tong_tra_no=sum(i.tra_no for i in items),
        tong_thu_no=sum(i.thu_no for i in items),
    )
