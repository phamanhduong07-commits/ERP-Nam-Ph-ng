"""Unified defect records — polymorphic kho ảo hàng lỗi.

Gộp hai bảng cũ (hang_loi_kho_ao = TP lỗi, hang_loi_phoi_kho_ao = phôi lỗi CD1)
thành một bảng defect_records, phân biệt nguồn bằng (ref_type, ref_id) và khâu (khau).

Mỗi bản ghi trỏ về 1 nguồn lỗi:
    ref_type='production_output'          → ProductionOutput  (khâu 'tp')
    ref_type='phieu_nhap_phoi_song_item' → PhieuNhapPhoiSongItem (khâu 'cd1')
    ref_type='sales_return_item'         → SalesReturnItem   (khâu 'tra_ve', auto + thủ công)

Response là superset của cả hai phiếu cũ — field nào không áp dụng cho nguồn
hiện tại thì trả None.
"""
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user, get_sale_visible_nv_ids
from app.models.auth import User
from app.models.defect_records import DefectRecord
from app.models.warehouse_doc import ProductionOutput
from app.models.phieu_nhap_phoi_song import PhieuNhapPhoiSong, PhieuNhapPhoiSongItem
from app.models.production import ProductionOrder, ProductionOrderItem
from app.models.master import PhanXuong, PhapNhan, Warehouse

router = APIRouter(prefix="/api/defect-records", tags=["defect-records"])

# Nguồn lỗi hợp lệ khi nhập kho ảo qua endpoint thống nhất
REF_TYPE_PRODUCTION_OUTPUT = "production_output"
REF_TYPE_PHOI_ITEM = "phieu_nhap_phoi_song_item"
REF_TYPE_SALES_RETURN_ITEM = "sales_return_item"
KHAU_BY_REF_TYPE = {
    REF_TYPE_PRODUCTION_OUTPUT: "tp",
    REF_TYPE_PHOI_ITEM: "cd1",
    REF_TYPE_SALES_RETURN_ITEM: "tra_ve",
}

_TRANG_THAI_LABEL = {
    "cho_xu_ly": "Chờ xử lý",
    "ban_phe": "Bán phế phẩm",
    "tan_dung": "Tận dụng",
    "da_xu_ly": "Đã xử lý",
    "huy": "Huỷ",
}

# Trạng thái cho phép chuyển sang qua PATCH /{id}/trang-thai
ALLOWED_TRANG_THAI = {"ban_phe", "tan_dung", "da_xu_ly", "huy"}


class NhapDefectIn(BaseModel):
    ref_type: str
    ref_id: int


class UpdateTrangThaiIn(BaseModel):
    trang_thai: str          # ban_phe | tan_dung | da_xu_ly | huy
    ghi_chu: Optional[str] = None
    production_order_id_tan_dung: Optional[int] = None


def _empty_context() -> dict:
    """Các field ngữ cảnh khi không resolve được nguồn — giữ shape ổn định."""
    return {
        "customer_id": None,
        "so_lenh": None,
        "ten_hang": None,
        "ngay": None,
        "ca": None,
        "so_phieu": None,
        "dvt": None,
        "quy_cach": None,
        "loai_thung": None,
        "so_lop": None,
        "ten_phan_xuong": None,
        "ten_phap_nhan": None,
        "phan_xuong_id": None,
        "phap_nhan_id": None,
        "ten_khach_hang": None,
        "ly_do_tra": None,
    }


def _context_production_output(ref_id: int, db: Session) -> dict:
    """Ngữ cảnh cho TP lỗi: ProductionOutput → Order → Item → PhanXuong → PhapNhan."""
    from app.models.sales import SalesOrder as _SO
    po = db.get(ProductionOutput, ref_id)
    if not po:
        return _empty_context()

    order = db.get(ProductionOrder, po.production_order_id)
    item = (
        db.query(ProductionOrderItem)
        .filter(ProductionOrderItem.production_order_id == order.id)
        .first()
    ) if order else None
    px = db.get(PhanXuong, order.phan_xuong_id) if order and order.phan_xuong_id else None
    pn = db.get(PhapNhan, order.phap_nhan_id) if order and order.phap_nhan_id else None
    so = db.get(_SO, order.sales_order_id) if order and order.sales_order_id else None

    quy_cach = None
    if item and item.dai and item.rong and item.cao:
        quy_cach = f"{int(item.dai)}×{int(item.rong)}×{int(item.cao)}"

    return {
        "customer_id": so.customer_id if so else None,
        "so_lenh": order.so_lenh if order else None,
        "ten_hang": po.ten_hang,
        "ngay": str(po.ngay_nhap) if po.ngay_nhap else None,
        "ca": None,
        "so_phieu": po.so_phieu,
        "dvt": po.dvt,
        "quy_cach": quy_cach,
        "loai_thung": item.loai_thung if item else None,
        "so_lop": item.so_lop if item else None,
        "ten_phan_xuong": px.ten_xuong if px else None,
        "ten_phap_nhan": pn.ten_viet_tat if pn else None,
        "phan_xuong_id": order.phan_xuong_id if order else None,
        "phap_nhan_id": order.phap_nhan_id if order else None,
        "ten_khach_hang": None,
        "ly_do_tra": None,
    }


def _context_phoi_item(ref_id: int, db: Session) -> dict:
    """Ngữ cảnh cho phôi lỗi CD1: Item → Phieu → Order → POI → PhanXuong → Warehouse → PhapNhan."""
    from app.models.sales import SalesOrder as _SO
    item = db.get(PhieuNhapPhoiSongItem, ref_id)
    if not item:
        return _empty_context()

    phieu = db.get(PhieuNhapPhoiSong, item.phieu_id) if item.phieu_id else None
    # poi = ProductionOrderItem: nguồn ten_hang cho phôi (PhieuNhapPhoiSongItem không có ten_hang)
    poi = db.get(ProductionOrderItem, item.production_order_item_id) if item.production_order_item_id else None
    order = db.get(ProductionOrder, phieu.production_order_id) if phieu else None
    px = db.get(PhanXuong, order.phan_xuong_id) if order and order.phan_xuong_id else None
    so = db.get(_SO, order.sales_order_id) if order and order.sales_order_id else None

    wh = db.get(Warehouse, phieu.warehouse_id) if phieu and phieu.warehouse_id else None
    pn_id = None
    if wh and wh.phan_xuong_obj:
        pn_id = wh.phan_xuong_obj.phap_nhan_id
    elif order and order.phap_nhan_id:
        pn_id = order.phap_nhan_id
    pn = db.get(PhapNhan, pn_id) if pn_id else None

    return {
        "customer_id": so.customer_id if so else None,
        "so_lenh": order.so_lenh if order else None,
        "ten_hang": poi.ten_hang if poi else None,
        "ngay": str(phieu.ngay) if phieu and phieu.ngay else None,
        "ca": phieu.ca if phieu else None,
        "so_phieu": phieu.so_phieu if phieu else None,
        "dvt": "Tấm",
        "quy_cach": None,
        "loai_thung": None,
        "so_lop": None,
        "ten_phan_xuong": px.ten_xuong if px else None,
        "ten_phap_nhan": pn.ten_viet_tat if pn else None,
        "phan_xuong_id": order.phan_xuong_id if order else None,
        "phap_nhan_id": pn_id,
        "ten_khach_hang": None,
        "ly_do_tra": None,
    }


def _context_sales_return_item(ref_id: int, db: Session) -> dict:
    """Ngữ cảnh cho hàng trả về lỗi: SalesReturnItem → SalesReturn / SalesOrderItem / Customer.

    Resolve phan_xuong/phap_nhan qua: SalesOrderItem → ProductionOrderItem → ProductionOrder.
    """
    from app.models.sales import SalesReturnItem, SalesReturn, SalesOrderItem
    from app.models.master import Customer
    from app.models.production import ProductionOrderItem

    sri = db.get(SalesReturnItem, ref_id)
    if not sri:
        return _empty_context()

    sr = db.get(SalesReturn, sri.sales_return_id) if sri.sales_return_id else None
    soi = db.get(SalesOrderItem, sri.sales_order_item_id) if sri.sales_order_item_id else None

    ten_khach_hang = None
    if sr and sr.customer_id:
        cust = db.get(Customer, sr.customer_id)
        if cust:
            ten_khach_hang = cust.ten_viet_tat or cust.ten_don_vi

    # Resolve phan_xuong / phap_nhan từ POI → PO
    phan_xuong_id = None
    phap_nhan_id = None
    ten_phan_xuong = None
    ten_phap_nhan = None
    if soi:
        poi = (
            db.query(ProductionOrderItem)
            .filter(ProductionOrderItem.sales_order_item_id == soi.id)
            .first()
        )
        if poi:
            order = db.get(ProductionOrder, poi.production_order_id)
            if order:
                phan_xuong_id = order.phan_xuong_id
                phap_nhan_id = order.phap_nhan_id
                px = db.get(PhanXuong, order.phan_xuong_id) if order.phan_xuong_id else None
                pn = db.get(PhapNhan, order.phap_nhan_id) if order.phap_nhan_id else None
                ten_phan_xuong = px.ten_xuong if px else None
                ten_phap_nhan = pn.ten_viet_tat if pn else None

    ctx = _empty_context()
    ctx.update({
        "customer_id": sr.customer_id if sr else None,
        "so_lenh": sr.so_phieu_tra if sr else None,
        "ten_hang": soi.ten_hang if soi else None,
        "ngay": str(sr.ngay_tra) if sr and sr.ngay_tra else None,
        "so_phieu": sr.so_phieu_tra if sr else None,
        "dvt": (soi.dvt if soi else None) or "Thùng",
        "ten_khach_hang": ten_khach_hang,
        "ly_do_tra": sri.ly_do_tra,
        "phan_xuong_id": phan_xuong_id,
        "phap_nhan_id": phap_nhan_id,
        "ten_phan_xuong": ten_phan_xuong,
        "ten_phap_nhan": ten_phap_nhan,
    })
    return ctx


def _resolve_context(entry: DefectRecord, db: Session) -> dict:
    """Lấy ngữ cảnh nguồn theo ref_type. ref_type lạ → context rỗng (không vỡ shape)."""
    if entry.ref_type == REF_TYPE_PRODUCTION_OUTPUT:
        return _context_production_output(entry.ref_id, db)
    if entry.ref_type == REF_TYPE_PHOI_ITEM:
        return _context_phoi_item(entry.ref_id, db)
    if entry.ref_type == REF_TYPE_SALES_RETURN_ITEM:
        return _context_sales_return_item(entry.ref_id, db)
    return _empty_context()


def _to_response(entry: DefectRecord, db: Session) -> dict:
    ctx = _resolve_context(entry, db)
    lsx_td = (
        db.get(ProductionOrder, entry.production_order_id_tan_dung)
        if entry.production_order_id_tan_dung else None
    )

    return {
        "id": entry.id,
        "ref_type": entry.ref_type,
        "ref_id": entry.ref_id,
        "khau": entry.khau,
        "so_luong": float(entry.so_luong),
        "trang_thai": entry.trang_thai,
        "ghi_chu": entry.ghi_chu,
        "so_lenh": ctx["so_lenh"],
        "ten_hang": ctx["ten_hang"],
        "ngay": ctx["ngay"],
        "ca": ctx["ca"],
        "dvt": ctx["dvt"],
        "quy_cach": ctx["quy_cach"],
        "loai_thung": ctx["loai_thung"],
        "so_lop": ctx["so_lop"],
        "ten_phan_xuong": ctx["ten_phan_xuong"],
        "ten_phap_nhan": ctx["ten_phap_nhan"],
        "phan_xuong_id": ctx["phan_xuong_id"],
        "phap_nhan_id": ctx["phap_nhan_id"],
        "ten_khach_hang": ctx["ten_khach_hang"],
        "ly_do_tra": ctx["ly_do_tra"],
        "production_order_id_tan_dung": entry.production_order_id_tan_dung,
        "so_lenh_tan_dung": lsx_td.so_lenh if lsx_td else None,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
    }


@router.get("")
def list_defect_records(
    khau: Optional[str] = Query(None),
    trang_thai: Optional[str] = Query(None),
    phap_nhan_id: Optional[int] = Query(None),
    phan_xuong_id: Optional[int] = Query(None),
    tu_ngay: Optional[date] = Query(None),
    den_ngay: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Liệt kê bản ghi lỗi.

    khau/trang_thai lọc thẳng trên defect_records (có index).
    phan_xuong_id/phap_nhan_id/tu_ngay/den_ngay/customer scope lọc ở Python sau khi resolve
    ngữ cảnh — vì nguồn polymorphic nên join SQL không thống nhất được.
    """
    q = db.query(DefectRecord)
    if khau:
        q = q.filter(DefectRecord.khau == khau)
    if trang_thai:
        q = q.filter(DefectRecord.trang_thai == trang_thai)
    # phan_xuong_id / phap_nhan_id giờ được lưu trực tiếp → dùng SQL filter
    if phan_xuong_id is not None:
        q = q.filter(DefectRecord.phan_xuong_id == phan_xuong_id)
    if phap_nhan_id is not None:
        q = q.filter(DefectRecord.phap_nhan_id == phap_nhan_id)
    rows = q.order_by(DefectRecord.created_at.desc()).limit(300).all()

    results = [_to_response(r, db) for r in rows]

    # SA scope: chỉ thấy hàng lỗi/trả về của KH được phân công
    scope_nv_ids = get_sale_visible_nv_ids(current_user)
    if scope_nv_ids is not None:
        from sqlalchemy import exists, or_
        from app.models.master import Customer, CustomerNhanVien
        visible_cids = {r.id for r in db.query(Customer.id).filter(
            or_(
                Customer.nv_phu_trach_id.in_(scope_nv_ids),
                exists().where(
                    (CustomerNhanVien.customer_id == Customer.id)
                    & (CustomerNhanVien.user_id.in_(scope_nv_ids))
                ),
            )
        ).all()}
        results = [r for r in results if r.get("customer_id") in visible_cids]

    if tu_ngay is not None:
        tu_str = tu_ngay.isoformat()
        results = [r for r in results if r["ngay"] is not None and r["ngay"] >= tu_str]
    if den_ngay is not None:
        den_str = den_ngay.isoformat()
        results = [r for r in results if r["ngay"] is not None and r["ngay"] <= den_str]

    return results


def _resolve_context_for_ref(ref_type: str, ref_id: int, db: Session) -> dict:
    """Resolve context chỉ để lấy phan_xuong_id/phap_nhan_id — gọi trước khi tạo DefectRecord."""
    if ref_type == REF_TYPE_PRODUCTION_OUTPUT:
        return _context_production_output(ref_id, db)
    if ref_type == REF_TYPE_PHOI_ITEM:
        return _context_phoi_item(ref_id, db)
    if ref_type == REF_TYPE_SALES_RETURN_ITEM:
        return _context_sales_return_item(ref_id, db)
    return _empty_context()


@router.post("/nhap", status_code=201)
def nhap_defect_record(
    body: NhapDefectIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Nhập một nguồn lỗi vào kho ảo thống nhất.

    Validate nguồn tồn tại, có so_luong_loi > 0 và trang_thai_loi == 'cho_xu_ly',
    chưa có DefectRecord cho cùng (ref_type, ref_id). Sau đó tạo bản ghi và đặt
    trang_thai_loi của nguồn = 'da_nhap_kho_ao'.
    """
    if body.ref_type not in KHAU_BY_REF_TYPE:
        raise HTTPException(
            400,
            f"ref_type không hợp lệ. Cho phép: {', '.join(sorted(KHAU_BY_REF_TYPE))}",
        )

    existing = db.query(DefectRecord).filter(
        DefectRecord.ref_type == body.ref_type,
        DefectRecord.ref_id == body.ref_id,
    ).first()
    if existing:
        raise HTTPException(400, "Nguồn này đã có trong kho ảo")

    khau = KHAU_BY_REF_TYPE[body.ref_type]

    if body.ref_type == REF_TYPE_PRODUCTION_OUTPUT:
        po = db.get(ProductionOutput, body.ref_id)
        if not po:
            raise HTTPException(404, "Không tìm thấy phiếu nhập thành phẩm")
        if not po.so_luong_loi or po.so_luong_loi <= 0:
            raise HTTPException(400, "Phiếu này không có hàng lỗi")
        if po.trang_thai_loi != "cho_xu_ly":
            raise HTTPException(400, "Hàng lỗi đã được nhập kho ảo hoặc không hợp lệ")
        so_luong_loi = po.so_luong_loi
        source = po
        source_attr = "trang_thai_loi"
    elif body.ref_type == REF_TYPE_PHOI_ITEM:
        item = db.get(PhieuNhapPhoiSongItem, body.ref_id)
        if not item:
            raise HTTPException(404, "Không tìm thấy dòng phôi")
        if not item.so_luong_loi or item.so_luong_loi <= 0:
            raise HTTPException(400, "Dòng này không có phôi lỗi")
        if item.trang_thai_loi != "cho_xu_ly":
            raise HTTPException(400, "Phôi lỗi đã được nhập kho ảo hoặc không hợp lệ")
        so_luong_loi = item.so_luong_loi
        source = item
        source_attr = "trang_thai_loi"
    else:  # REF_TYPE_SALES_RETURN_ITEM — nhập thủ công
        from app.models.sales import SalesReturnItem, SalesReturn
        sri = db.get(SalesReturnItem, body.ref_id)
        if not sri:
            raise HTTPException(404, "Không tìm thấy dòng hàng trả về")
        if sri.tinh_trang_hang not in ("hong", "loi"):
            raise HTTPException(400, "Dòng này không phải hàng hỏng/lỗi")
        sr = db.get(SalesReturn, sri.sales_return_id) if sri.sales_return_id else None
        if not sr or sr.trang_thai != "da_duyet":
            raise HTTPException(400, "Phiếu trả chưa được duyệt")
        so_luong_loi = sri.so_luong_tra
        source = None
        source_attr = None

    # Resolve phan_xuong/phap_nhan tại thời điểm nhập để lưu trực tiếp vào bảng
    ctx = _resolve_context_for_ref(body.ref_type, body.ref_id, db)

    entry = DefectRecord(
        ref_type=body.ref_type,
        ref_id=body.ref_id,
        khau=khau,
        so_luong=so_luong_loi,
        trang_thai="cho_xu_ly",
        phan_xuong_id=ctx.get("phan_xuong_id"),
        phap_nhan_id=ctx.get("phap_nhan_id"),
        created_by=current_user.id,
    )
    db.add(entry)
    if source and source_attr:
        setattr(source, source_attr, "da_nhap_kho_ao")
    db.commit()
    db.refresh(entry)
    return _to_response(entry, db)


@router.get("/export")
def export_defect_records(
    khau: Optional[str] = Query(None),
    trang_thai: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Xuất danh sách kho ảo hàng lỗi ra Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(DefectRecord)
    if khau:
        q = q.filter(DefectRecord.khau == khau)
    if trang_thai:
        q = q.filter(DefectRecord.trang_thai == trang_thai)
    rows = q.order_by(DefectRecord.created_at.desc()).limit(1000).all()

    records = [_to_response(r, db) for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    khau_label = {"tp": "Thành phẩm lỗi", "cd1": "Phôi lỗi", "tra_ve": "Hàng trả về"}.get(khau or "", "Tất cả")
    ws.title = "Kho lỗi"

    # Company header
    ws.merge_cells("A1:K1")
    ws["A1"] = "CÔNG TY TNHH NAM PHƯƠNG BAO BÌ"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:K2")
    ws["A2"] = f"DANH SÁCH KHO ẢO HÀNG LỖI — {khau_label.upper()}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["STT", "Mã phiếu nguồn", "Khâu", "Tên hàng", "Quy cách", "SL", "ĐVT",
               "Ngày", "Khách hàng / Xưởng", "Trạng thái", "Ghi chú"]
    col_widths = [6, 20, 14, 30, 16, 8, 8, 12, 26, 16, 30]

    hdr_fill = PatternFill(fill_type="solid", fgColor="B71C1C")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, rec in enumerate(records, start=1):
        row_idx = ri + 4
        khau_str = {"tp": "Thành phẩm", "cd1": "Phôi (CD1)", "tra_ve": "Trả về"}.get(rec["khau"] or "", rec["khau"] or "")
        nguon_label = rec.get("ten_phan_xuong") or rec.get("ten_khach_hang") or ""
        tt_label = _TRANG_THAI_LABEL.get(rec["trang_thai"] or "", rec["trang_thai"] or "")
        vals = [
            ri,
            rec.get("so_phieu") or "",
            khau_str,
            rec.get("ten_hang") or "",
            rec.get("quy_cach") or "",
            rec.get("so_luong"),
            rec.get("dvt") or "",
            rec.get("ngay") or "",
            nguon_label,
            tt_label,
            rec.get("ghi_chu") or "",
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.border = border
            if ci == 6:
                cell.alignment = Alignment(horizontal="right")

    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"kho-loi-{khau or 'tat-ca'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.patch("/{entry_id}/trang-thai")
def update_trang_thai(
    entry_id: int,
    body: UpdateTrangThaiIn,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Chuyển trạng thái xử lý: ban_phe | tan_dung | da_xu_ly | huy."""
    if body.trang_thai not in ALLOWED_TRANG_THAI:
        raise HTTPException(
            400,
            f"Trạng thái không hợp lệ. Cho phép: {', '.join(sorted(ALLOWED_TRANG_THAI))}",
        )
    entry = db.get(DefectRecord, entry_id)
    if not entry:
        raise HTTPException(404, "Không tìm thấy bản ghi lỗi")

    entry.trang_thai = body.trang_thai
    if body.ghi_chu is not None:
        entry.ghi_chu = body.ghi_chu
    if body.trang_thai == "tan_dung" and body.production_order_id_tan_dung:
        entry.production_order_id_tan_dung = body.production_order_id_tan_dung
    entry.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(entry)
    return _to_response(entry, db)


@router.get("/{entry_id}/print", response_class=HTMLResponse)
def print_defect_record(
    entry_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Trả HTML phiếu xử lý hàng lỗi/trả về (A5, đỏ #B71C1C)."""
    entry = db.get(DefectRecord, entry_id)
    if not entry:
        raise HTTPException(404, "Không tìm thấy bản ghi lỗi")

    rec = _to_response(entry, db)
    so_phieu_xl = f"PXHL-{entry.id:06d}"
    tt_label = _TRANG_THAI_LABEL.get(rec["trang_thai"] or "", rec["trang_thai"] or "")
    khau_label = {"tp": "Thành phẩm lỗi", "cd1": "Phôi lỗi (CD1)", "tra_ve": "Hàng trả về"}.get(rec["khau"] or "", rec["khau"] or "")

    ngay_str = ""
    if rec.get("ngay"):
        parts = str(rec["ngay"]).split("-")
        if len(parts) == 3:
            ngay_str = f"Ngày {parts[2]} tháng {parts[1]} năm {parts[0]}"

    nguon_phieu = rec.get("so_phieu") or rec.get("so_lenh") or "—"
    ten_hang = rec.get("ten_hang") or "—"
    so_luong = rec.get("so_luong") or 0
    dvt = rec.get("dvt") or ""
    quy_cach = rec.get("quy_cach") or ""
    ten_xuong = rec.get("ten_phan_xuong") or ""
    ten_kh = rec.get("ten_khach_hang") or ""
    ly_do = rec.get("ly_do_tra") or ""
    ghi_chu = rec.get("ghi_chu") or ""
    so_lenh_td = rec.get("so_lenh_tan_dung") or ""

    rendered = f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<title>Phiếu xử lý hàng lỗi {so_phieu_xl}</title>
<style>
  @page {{ size: A5 portrait; margin: 10mm 10mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Times New Roman', serif; font-size: 10pt; color: #222; }}
  .noprint {{ margin-bottom: 6mm; }}
  @media print {{ .noprint {{ display: none; }} }}
  button {{ padding: 6px 18px; background: #B71C1C; color: #fff; border: none; border-radius: 3px; cursor: pointer; font-size: 10pt; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 3mm; }}
  .company-name {{ font-size: 11pt; font-weight: bold; color: #B71C1C; }}
  .company-info {{ font-size: 8.5pt; line-height: 1.5; margin-top: 1mm; }}
  .mau {{ font-size: 8pt; color: #555; text-align: right; }}
  .divider {{ border-top: 2px solid #B71C1C; margin: 2mm 0; }}
  .title {{ text-align: center; margin: 2mm 0 3mm; }}
  .title h2 {{ font-size: 14pt; font-weight: bold; letter-spacing: 1px; text-transform: uppercase; }}
  .title .so {{ font-size: 9pt; margin-top: 1mm; }}
  .title .date {{ font-size: 9pt; font-style: italic; }}
  .info-block {{ font-size: 10pt; line-height: 1.9; margin-bottom: 3mm; }}
  .row {{ display: flex; margin: 1px 0; }}
  .row .label {{ min-width: 120px; font-weight: bold; flex-shrink: 0; }}
  .row .dots {{ flex: 1; border-bottom: 1px dotted #888; padding-left: 4px; }}
  .amount-box {{ border: 1.5px solid #B71C1C; border-radius: 3px; padding: 4px 8px; font-weight: bold; text-align: center; font-size: 11pt; margin: 3mm 0; }}
  .chu {{ font-size: 9.5pt; margin: 2mm 0; }}
  .sig-table {{ width: 100%; border-collapse: collapse; margin-top: 5mm; }}
  .sig-table td {{ border: none; text-align: center; vertical-align: top; width: 33%; padding: 2px; }}
  .sig-label {{ font-weight: bold; font-size: 9pt; }}
  .sig-sub {{ font-style: italic; font-size: 8pt; color: #555; }}
  .sig-name {{ margin-top: 28px; font-weight: bold; }}
  .badge {{ display: inline-block; padding: 1px 8px; border-radius: 3px; font-size: 9pt; color: #fff; background: #B71C1C; }}
</style>
</head>
<body>
<div class="noprint"><button onclick="window.print()">🖨️ In phiếu</button></div>

<div class="header">
  <div>
    <div class="company-name">CÔNG TY TNHH NAM PHƯƠNG BAO BÌ</div>
    <div class="company-info">
      ĐC: TP. Hồ Chí Minh<br>
      MST: 0315xxxxxx
    </div>
  </div>
  <div class="mau">Nội bộ</div>
</div>

<div class="divider"></div>

<div class="title">
  <h2>Phiếu xử lý hàng lỗi/trả về</h2>
  <div class="so">Số: <strong>{so_phieu_xl}</strong> &nbsp;|&nbsp; Khâu: {khau_label}</div>
  <div class="date">{ngay_str}</div>
</div>

<div class="info-block">
  <div class="row"><span class="label">Phiếu nguồn:</span><span class="dots">{nguon_phieu}</span></div>
  <div class="row"><span class="label">Tên hàng:</span><span class="dots">{ten_hang}</span></div>
  {"<div class='row'><span class='label'>Quy cách:</span><span class='dots'>" + quy_cach + "</span></div>" if quy_cach else ""}
  {"<div class='row'><span class='label'>Xưởng:</span><span class='dots'>" + ten_xuong + "</span></div>" if ten_xuong else ""}
  {"<div class='row'><span class='label'>Khách hàng:</span><span class='dots'>" + ten_kh + "</span></div>" if ten_kh else ""}
  {"<div class='row'><span class='label'>Lý do trả:</span><span class='dots'>" + ly_do + "</span></div>" if ly_do else ""}
</div>

<div class="amount-box">Số lượng: {so_luong:,.0f} {dvt}</div>

<div class="info-block">
  <div class="row"><span class="label">Trạng thái:</span><span class="dots"><span class="badge">{tt_label}</span></span></div>
  {"<div class='row'><span class='label'>LSX tận dụng:</span><span class='dots'>" + so_lenh_td + "</span></div>" if so_lenh_td else ""}
</div>

{"<div class='chu'>Ghi chú: " + ghi_chu + "</div>" if ghi_chu else ""}

<table class="sig-table">
  <tr>
    <td><div class="sig-label">Người lập phiếu</div><div class="sig-sub">(Ký, họ tên)</div><div class="sig-name">&nbsp;</div></td>
    <td><div class="sig-label">Thủ kho</div><div class="sig-sub">(Ký, họ tên)</div><div class="sig-name">&nbsp;</div></td>
    <td><div class="sig-label">Giám đốc</div><div class="sig-sub">(Ký, họ tên)</div><div class="sig-name">&nbsp;</div></td>
  </tr>
</table>
</body>
</html>"""
    return HTMLResponse(content=rendered)
