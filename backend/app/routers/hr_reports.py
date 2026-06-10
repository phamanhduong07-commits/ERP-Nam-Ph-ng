"""Router: Báo cáo HR & Compliance (Phase 1.5).

5 báo cáo Excel cho HR + cơ quan nhà nước:
1. /employees-list     — Danh sách NV (Sổ quản lý lao động) — mẫu BLĐ
2. /labor-report       — Báo cáo lao động Sở LĐ-TBXH (quý/năm)
3. /gender-equality    — Bình đẳng giới (NĐ 145/2020/NĐ-CP Chương III)
4. /hr-costs           — Chi phí nhân sự (lương BHXH + phúc lợi + đào tạo + TNLĐ)
5. /summary-report     — Báo cáo HR tổng hợp tháng/quý

Tất cả trả về file Excel via StreamingResponse.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import require_roles
from app.models.auth import User
from app.models.hr import (
    Employee, Department, LaborContract, FamilyRelation,
    HealthCheck, SafetyEquipmentIssue, SafetyTraining, SafetyTrainingParticipant,
    WorkAccident, KPIEvaluation, KPICycle,
)
from app.models.master import PhapNhan

logger = logging.getLogger("erp")
router = APIRouter(prefix="/api/hr/reports", tags=["hr-reports"])


# ─── Excel helpers ───
HEADER_FILL = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1677FF")
SUBTITLE_FONT = Font(name="Arial", size=11, italic=True, color="595959")
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _style_header_row(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _style_data_cell(ws, row: int, col: int, align: Alignment = LEFT):
    cell = ws.cell(row=row, column=col)
    cell.border = BORDER
    cell.alignment = align


def _auto_fit_columns(ws, max_cols: int = 30):
    for col in range(1, max_cols + 1):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            if cell.value is not None:
                ln = len(str(cell.value))
                if ln > max_len:
                    max_len = ln
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 50)


def _build_response(wb: Workbook, filename: str) -> StreamingResponse:
    """Trả file Excel. Header Content-Disposition phải ASCII-safe,
    tên file Unicode (có dấu tiếng Việt) dùng RFC 5987 `filename*=UTF-8''...`.
    """
    import unicodedata
    from urllib.parse import quote
    # ASCII fallback (bỏ dấu) cho client cũ
    ascii_name = (
        unicodedata.normalize("NFD", filename)
        .encode("ascii", "ignore").decode("ascii")
    ) or "report.xlsx"
    encoded = quote(filename, safe="")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            # filename=ASCII (fallback) + filename*=UTF-8 (chuẩn RFC 5987)
            "Content-Disposition": f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{encoded}',
        },
    )


# ═══════════════════════════════════════════════════════════════
# 1) Sổ quản lý lao động — Danh sách NV
# ═══════════════════════════════════════════════════════════════
@router.get("/employees-list")
def report_employees_list(
    phap_nhan_id: Optional[int] = None,
    trang_thai: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")),
):
    """Sổ quản lý lao động — danh sách NV đầy đủ theo mẫu của Bộ LĐ-TBXH."""
    q = db.query(Employee)
    if phap_nhan_id:
        q = q.filter(Employee.phap_nhan_id == phap_nhan_id)
    if trang_thai:
        q = q.filter(Employee.trang_thai == trang_thai)
    employees = q.order_by(Employee.ma_nv).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sổ quản lý LĐ"

    # Title
    ws["A1"] = "SỔ QUẢN LÝ LAO ĐỘNG"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")
    ws["A1"].alignment = CENTER
    ws["A2"] = f"Xuất ngày {date.today().strftime('%d/%m/%Y')} — Tổng số: {len(employees)} NV"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:N2")
    ws["A2"].alignment = CENTER

    # Header
    headers = ["STT", "Mã NV", "Họ và Tên", "Ngày sinh", "Giới tính", "CCCD",
               "Pháp nhân", "Bộ phận", "Chức vụ", "Ngày vào làm",
               "SĐT", "Địa chỉ hộ khẩu", "Số sổ BHXH", "Trạng thái"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, len(headers))

    trang_thai_map = {"dang_lam": "Đang làm", "tam_nghi": "Tạm nghỉ", "da_nghi": "Đã nghỉ"}
    for i, e in enumerate(employees, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=e.ma_nv)
        ws.cell(row=row, column=3, value=e.ho_ten)
        ws.cell(row=row, column=4, value=e.ngay_sinh.strftime("%d/%m/%Y") if e.ngay_sinh else "")
        ws.cell(row=row, column=5, value=e.gioi_tinh or "")
        ws.cell(row=row, column=6, value=e.cccd or "")
        ws.cell(row=row, column=7, value=e.phap_nhan.ten_phap_nhan if e.phap_nhan else "")
        ws.cell(row=row, column=8, value=e.bo_phan.ten_bo_phan if e.bo_phan else "")
        ws.cell(row=row, column=9, value=e.chuc_vu.ten_chuc_vu if e.chuc_vu else "")
        ws.cell(row=row, column=10, value=e.ngay_vao_lam.strftime("%d/%m/%Y") if e.ngay_vao_lam else "")
        ws.cell(row=row, column=11, value=e.so_dien_thoai or "")
        ws.cell(row=row, column=12, value=e.dia_chi_ho_khau or "")
        ws.cell(row=row, column=13, value=e.so_so_bhxh or "")
        ws.cell(row=row, column=14, value=trang_thai_map.get(e.trang_thai, e.trang_thai))
        for c in range(1, len(headers) + 1):
            _style_data_cell(ws, row, c, CENTER if c in (1, 2, 4, 5, 10, 14) else LEFT)

    _auto_fit_columns(ws, len(headers))
    logger.info("HR report employees-list exported %s rows by user=%s", len(employees), current_user.id)
    return _build_response(wb, f"so-quan-ly-lao-dong-{date.today().strftime('%Y%m%d')}.xlsx")


# ═══════════════════════════════════════════════════════════════
# 2) Báo cáo lao động Sở LĐ-TBXH
# ═══════════════════════════════════════════════════════════════
@router.get("/labor-report")
def report_labor(
    year: int = Query(default=date.today().year),
    quarter: Optional[int] = Query(default=None, ge=1, le=4),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")),
):
    """Báo cáo định kỳ Sở LĐ-TBXH (quý/năm).

    Mẫu báo cáo: Tổng + theo giới tính + theo HĐLĐ + biến động trong kỳ.
    """
    # Date range
    if quarter:
        q_start_month = (quarter - 1) * 3 + 1
        start = date(year, q_start_month, 1)
        if quarter == 4:
            end = date(year, 12, 31)
        else:
            end = date(year, q_start_month + 3, 1) - timedelta(days=1)
        period_label = f"Quý {quarter}/{year}"
    else:
        start, end = date(year, 1, 1), date(year, 12, 31)
        period_label = f"Năm {year}"

    # Stats
    total = db.query(func.count(Employee.id)).filter(Employee.trang_thai == "dang_lam").scalar() or 0
    by_gender = dict(
        db.query(Employee.gioi_tinh, func.count(Employee.id))
        .filter(Employee.trang_thai == "dang_lam")
        .group_by(Employee.gioi_tinh).all()
    )
    by_phap_nhan = db.query(PhapNhan.ten_phap_nhan, func.count(Employee.id)) \
        .join(Employee, Employee.phap_nhan_id == PhapNhan.id) \
        .filter(Employee.trang_thai == "dang_lam") \
        .group_by(PhapNhan.id, PhapNhan.ten_phap_nhan).all()

    # HĐLĐ thống kê
    contracts_by_loai = dict(
        db.query(LaborContract.loai_hop_dong, func.count(LaborContract.id))
        .filter(LaborContract.trang_thai == "hieu_luc")
        .group_by(LaborContract.loai_hop_dong).all()
    )

    # Biến động trong kỳ
    new_hires = db.query(func.count(Employee.id)).filter(
        Employee.ngay_vao_lam.between(start, end),
    ).scalar() or 0
    resigned = db.query(func.count(Employee.id)).filter(
        Employee.ngay_nghi_viec.between(start, end),
    ).scalar() or 0

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo LĐ"

    ws["A1"] = f"BÁO CÁO TÌNH HÌNH SỬ DỤNG LAO ĐỘNG — {period_label.upper()}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = CENTER
    ws["A2"] = f"Kỳ báo cáo: {start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')} · Xuất ngày {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = CENTER

    r = 4
    # Section 1
    ws.cell(row=r, column=1, value="I. TỔNG QUAN").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Tổng số lao động đang làm việc:")
    ws.cell(row=r, column=2, value=total).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="  - Nam:")
    ws.cell(row=r, column=2, value=by_gender.get("Nam", 0))
    r += 1
    ws.cell(row=r, column=1, value="  - Nữ:")
    ws.cell(row=r, column=2, value=by_gender.get("Nữ", 0))
    r += 2

    # Section 2: by pháp nhân
    ws.cell(row=r, column=1, value="II. CƠ CẤU THEO PHÁP NHÂN").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Pháp nhân").font = HEADER_FONT
    ws.cell(row=r, column=2, value="Số lao động").font = HEADER_FONT
    for c in (1, 2):
        ws.cell(row=r, column=c).fill = HEADER_FILL
        ws.cell(row=r, column=c).alignment = CENTER
    r += 1
    for ten, cnt in by_phap_nhan:
        ws.cell(row=r, column=1, value=ten)
        ws.cell(row=r, column=2, value=cnt).alignment = CENTER
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    r += 1

    # Section 3: HĐLĐ
    ws.cell(row=r, column=1, value="III. CƠ CẤU HỢP ĐỒNG LAO ĐỘNG").font = Font(bold=True, size=12)
    r += 1
    loai_label = {
        "khong_thoi_han": "Không xác định thời hạn",
        "xac_dinh_thoi_han": "Xác định thời hạn",
        "thu_viec": "Thử việc",
        "khoan_viec": "Khoán việc",
    }
    for loai, cnt in contracts_by_loai.items():
        ws.cell(row=r, column=1, value=f"  - {loai_label.get(loai, loai)}:")
        ws.cell(row=r, column=2, value=cnt)
        r += 1
    r += 1

    # Section 4: biến động
    ws.cell(row=r, column=1, value="IV. BIẾN ĐỘNG TRONG KỲ").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Tuyển mới:")
    ws.cell(row=r, column=2, value=new_hires).font = Font(bold=True, color="52C41A")
    r += 1
    ws.cell(row=r, column=1, value="Nghỉ việc:")
    ws.cell(row=r, column=2, value=resigned).font = Font(bold=True, color="CF1322")
    r += 1
    net = new_hires - resigned
    ws.cell(row=r, column=1, value="Biến động ròng:")
    ws.cell(row=r, column=2, value=net).font = Font(bold=True)

    _auto_fit_columns(ws, 4)
    logger.info("HR report labor exported period=%s by user=%s", period_label, current_user.id)
    return _build_response(wb, f"bao-cao-lao-dong-{period_label.replace('/', '-')}.xlsx")


# ═══════════════════════════════════════════════════════════════
# 3) Bình đẳng giới
# ═══════════════════════════════════════════════════════════════
@router.get("/gender-equality")
def report_gender_equality(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")),
):
    """Báo cáo bình đẳng giới (NĐ 145/2020/NĐ-CP Chương III)."""
    # Theo bộ phận
    by_dept_gender = db.query(
        Department.ten_bo_phan, Employee.gioi_tinh, func.count(Employee.id),
    ).join(Employee, Employee.bo_phan_id == Department.id) \
     .filter(Employee.trang_thai == "dang_lam") \
     .group_by(Department.id, Department.ten_bo_phan, Employee.gioi_tinh) \
     .order_by(Department.ten_bo_phan).all()

    # Pivot
    by_dept: dict[str, dict[str, int]] = {}
    for ten_bp, gt, cnt in by_dept_gender:
        if ten_bp not in by_dept:
            by_dept[ten_bp] = {"Nam": 0, "Nữ": 0, "Khác": 0, "Total": 0}
        key = gt if gt in ("Nam", "Nữ") else "Khác"
        by_dept[ten_bp][key] = cnt
        by_dept[ten_bp]["Total"] += cnt

    wb = Workbook()
    ws = wb.active
    ws.title = "Bình đẳng giới"

    ws["A1"] = "BÁO CÁO BÌNH ĐẲNG GIỚI THEO BỘ PHẬN"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = CENTER
    ws["A2"] = f"Theo Nghị định 145/2020/NĐ-CP Chương III · Xuất ngày {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")
    ws["A2"].alignment = CENTER

    headers = ["STT", "Bộ phận", "Nam", "Nữ", "Tổng", "% Nữ"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, len(headers))

    total_nam, total_nu, total_all = 0, 0, 0
    for i, (ten_bp, counts) in enumerate(sorted(by_dept.items()), 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=ten_bp)
        ws.cell(row=row, column=3, value=counts["Nam"])
        ws.cell(row=row, column=4, value=counts["Nữ"])
        ws.cell(row=row, column=5, value=counts["Total"])
        pct = (counts["Nữ"] * 100.0 / counts["Total"]) if counts["Total"] else 0
        ws.cell(row=row, column=6, value=f"{pct:.1f}%")
        for c in range(1, 7):
            _style_data_cell(ws, row, c, CENTER if c != 2 else LEFT)
        total_nam += counts["Nam"]; total_nu += counts["Nữ"]; total_all += counts["Total"]

    # Total row
    last_row = 4 + len(by_dept) + 1
    ws.cell(row=last_row, column=2, value="TỔNG CỘNG").font = Font(bold=True)
    ws.cell(row=last_row, column=3, value=total_nam).font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total_nu).font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=total_all).font = Font(bold=True)
    pct_total = (total_nu * 100.0 / total_all) if total_all else 0
    ws.cell(row=last_row, column=6, value=f"{pct_total:.1f}%").font = Font(bold=True)
    for c in range(1, 7):
        ws.cell(row=last_row, column=c).fill = PatternFill(start_color="E6F4FF", end_color="E6F4FF", fill_type="solid")
        ws.cell(row=last_row, column=c).border = BORDER
        ws.cell(row=last_row, column=c).alignment = CENTER if c != 2 else LEFT

    _auto_fit_columns(ws, 6)
    logger.info("HR report gender-equality exported by user=%s", current_user.id)
    return _build_response(wb, f"binh-dang-gioi-{date.today().strftime('%Y%m%d')}.xlsx")


# ═══════════════════════════════════════════════════════════════
# 4) Chi phí nhân sự
# ═══════════════════════════════════════════════════════════════
@router.get("/hr-costs")
def report_hr_costs(
    year: int = Query(default=date.today().year),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")),
):
    """Chi phí nhân sự năm — lương BHXH + huấn luyện ATVSLĐ + TNLĐ + khám sức khỏe + BHLĐ."""
    start = date(year, 1, 1)
    end = date(year, 12, 31)

    # Lương đóng BHXH (tổng mức lương BHXH × 12 tháng — ước tính)
    total_bhxh_monthly = db.query(func.coalesce(func.sum(Employee.muc_dong_bhxh), 0)).filter(
        Employee.trang_thai == "dang_lam",
    ).scalar() or 0
    total_bhxh_yearly = float(total_bhxh_monthly) * 12

    # Chi phí huấn luyện ATVSLĐ
    training_cost = db.query(func.coalesce(func.sum(SafetyTraining.chi_phi), 0)).filter(
        SafetyTraining.ngay_bat_dau.between(start, end),
    ).scalar() or 0

    # Chi phí TNLĐ (y tế công ty trả, sau khi trừ BH chi)
    accident_cost_q = db.query(
        func.coalesce(func.sum(WorkAccident.chi_phi_y_te), 0),
        func.coalesce(func.sum(WorkAccident.bao_hiem_chi_tra), 0),
    ).filter(WorkAccident.ngay_xay_ra.between(start, end)).first()
    accident_med = float(accident_cost_q[0] or 0)
    accident_bh = float(accident_cost_q[1] or 0)
    accident_net = max(0, accident_med - accident_bh)

    # Chi phí khám sức khỏe
    health_cost = db.query(func.coalesce(func.sum(HealthCheck.chi_phi), 0)).filter(
        HealthCheck.ngay_kham.between(start, end),
    ).scalar() or 0

    # Chi phí BHLĐ: tổng (số lượng × đơn giá) của các lần cấp trong năm
    from app.models.hr import SafetyEquipment as SE
    bhld_cost = db.query(
        func.coalesce(func.sum(SafetyEquipmentIssue.so_luong * SE.don_gia), 0),
    ).join(SE, SE.id == SafetyEquipmentIssue.equipment_id).filter(
        SafetyEquipmentIssue.ngay_cap.between(start, end),
    ).scalar() or 0
    bhld_cost = float(bhld_cost)

    total_cost = total_bhxh_yearly + float(training_cost) + accident_net + float(health_cost) + bhld_cost

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Chi phí NS"

    ws["A1"] = f"BÁO CÁO CHI PHÍ NHÂN SỰ — NĂM {year}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = CENTER
    ws["A2"] = f"Xuất ngày {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = CENTER

    headers = ["STT", "Khoản chi phí", "Số tiền (VNĐ)", "Tỷ trọng %"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, len(headers))

    items = [
        ("Lương đóng BHXH (12 tháng)", total_bhxh_yearly),
        ("Huấn luyện ATVSLĐ", float(training_cost)),
        ("Tai nạn lao động (sau BH)", accident_net),
        ("Khám sức khỏe định kỳ", float(health_cost)),
        ("Bảo hộ lao động (BHLĐ)", bhld_cost),
    ]
    for i, (label, val) in enumerate(items, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=val).number_format = '#,##0'
        pct = (val * 100 / total_cost) if total_cost else 0
        ws.cell(row=row, column=4, value=f"{pct:.1f}%")
        for c in range(1, 5):
            _style_data_cell(ws, row, c, RIGHT if c in (3, 4) else (CENTER if c == 1 else LEFT))

    # Total row
    total_row = 4 + len(items) + 1
    ws.cell(row=total_row, column=2, value="TỔNG CỘNG").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=total_cost).font = Font(bold=True, color="1677FF")
    ws.cell(row=total_row, column=3).number_format = '#,##0'
    ws.cell(row=total_row, column=4, value="100%").font = Font(bold=True)
    for c in range(1, 5):
        ws.cell(row=total_row, column=c).fill = PatternFill(start_color="E6F4FF", end_color="E6F4FF", fill_type="solid")
        ws.cell(row=total_row, column=c).border = BORDER

    _auto_fit_columns(ws, 4)
    logger.info("HR report hr-costs year=%s total=%s by user=%s", year, total_cost, current_user.id)
    return _build_response(wb, f"chi-phi-nhan-su-{year}.xlsx")


# ═══════════════════════════════════════════════════════════════
# 5) Báo cáo tổng hợp HR theo tháng
# ═══════════════════════════════════════════════════════════════
@router.get("/summary-report")
def report_summary(
    year: int = Query(default=date.today().year),
    month: int = Query(default=date.today().month, ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")),
):
    """Báo cáo HR tổng hợp tháng: tuyển mới, nghỉ việc, sinh nhật, sự kiện."""
    start = date(year, month, 1)
    end = (date(year, month + 1, 1) - timedelta(days=1)) if month < 12 else date(year, 12, 31)

    # New hires
    new_emps = db.query(Employee).filter(
        Employee.ngay_vao_lam.between(start, end),
    ).order_by(Employee.ngay_vao_lam).all()

    # Resigned
    resigned_emps = db.query(Employee).filter(
        Employee.ngay_nghi_viec.between(start, end),
    ).order_by(Employee.ngay_nghi_viec).all()

    # Birthdays
    birthdays = db.query(Employee).filter(
        Employee.trang_thai == "dang_lam",
        func.extract("month", Employee.ngay_sinh) == month,
    ).order_by(func.extract("day", Employee.ngay_sinh)).all()

    # Contracts expiring this month
    expiring = db.query(LaborContract).filter(
        LaborContract.trang_thai == "hieu_luc",
        LaborContract.ngay_het_han.between(start, end),
    ).all()

    # Excel với 4 sheet
    wb = Workbook()
    wb.remove(wb.active)

    def _make_sheet(title: str, headers: list[str], rows: list[list]):
        ws = wb.create_sheet(title=title)
        ws.cell(row=1, column=1, value=f"{title.upper()} — Tháng {month}/{year}").font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1).alignment = CENTER
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        _style_header_row(ws, 3, len(headers))
        for i, r in enumerate(rows, 1):
            for c, v in enumerate(r, 1):
                ws.cell(row=3 + i, column=c, value=v)
                _style_data_cell(ws, 3 + i, c, CENTER if c == 1 else LEFT)
        _auto_fit_columns(ws, len(headers))

    _make_sheet(
        "Tuyển mới",
        ["STT", "Mã NV", "Họ tên", "Ngày vào", "Bộ phận", "Pháp nhân"],
        [[i, e.ma_nv, e.ho_ten,
          e.ngay_vao_lam.strftime("%d/%m/%Y") if e.ngay_vao_lam else "",
          e.bo_phan.ten_bo_phan if e.bo_phan else "",
          e.phap_nhan.ten_phap_nhan if e.phap_nhan else ""]
         for i, e in enumerate(new_emps, 1)],
    )
    _make_sheet(
        "Nghỉ việc",
        ["STT", "Mã NV", "Họ tên", "Ngày nghỉ", "Bộ phận"],
        [[i, e.ma_nv, e.ho_ten,
          e.ngay_nghi_viec.strftime("%d/%m/%Y") if e.ngay_nghi_viec else "",
          e.bo_phan.ten_bo_phan if e.bo_phan else ""]
         for i, e in enumerate(resigned_emps, 1)],
    )
    _make_sheet(
        "Sinh nhật",
        ["STT", "Mã NV", "Họ tên", "Ngày sinh", "Bộ phận"],
        [[i, e.ma_nv, e.ho_ten,
          e.ngay_sinh.strftime("%d/%m/%Y") if e.ngay_sinh else "",
          e.bo_phan.ten_bo_phan if e.bo_phan else ""]
         for i, e in enumerate(birthdays, 1)],
    )
    _make_sheet(
        "HĐLĐ hết hạn",
        ["STT", "Số HĐ", "NV", "Mã NV", "Ngày hết hạn"],
        [[i, c.so_hop_dong, c.employee.ho_ten if c.employee else "",
          c.employee.ma_nv if c.employee else "",
          c.ngay_het_han.strftime("%d/%m/%Y") if c.ngay_het_han else ""]
         for i, c in enumerate(expiring, 1)],
    )

    logger.info("HR report summary month=%s/%s by user=%s", month, year, current_user.id)
    return _build_response(wb, f"bao-cao-hr-{month:02d}-{year}.xlsx")
