from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from typing import List, Optional
import re
import unicodedata
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from app.database import get_db
from app.deps import get_current_user, require_roles, require_permissions, require_any_permission, assert_has_permission
from app.models.auth import User
from app.models.hr import Department, Position, Employee, AttendanceLog, LeaveRequest, PayrollConfig, PayrollHoliday, EmployeeHistory, EmployeeDocument, LaborContract, PayrollRun, FamilyRelation, CheckInLocation, Team, HealthCheck
from app.models.master import PhapNhan
from app.services.hr_service import PayrollService
from app.schemas import hr as schemas
from app.utils.log import get_logger
from app.services.excel_import_service import ImportField, build_template_response

logger = get_logger(__name__)

router = APIRouter(prefix="/api/hr", tags=["hr"])

HR_EMPLOYEE_IMPORT_FIELDS = [
    ImportField("ma_nv", "Ma NV", required=True, help_text="Ma nhan vien (duy nhat)"),
    ImportField("ho_ten", "Ho ten", required=True),
    ImportField("ngay_sinh", "Ngay sinh", help_text="DD/MM/YYYY"),
    ImportField("gioi_tinh", "Gioi tinh", help_text="Nam / Nu"),
    ImportField("so_dien_thoai", "So dien thoai"),
    ImportField("email", "Email"),
    ImportField("ngay_vao_lam", "Ngay vao lam", help_text="DD/MM/YYYY"),
    ImportField("ngay_nghi_viec", "Ngay nghi viec", help_text="DD/MM/YYYY, de trong neu con lam"),
    ImportField("trang_thai", "Trang thai", help_text="dang_lam / tam_nghi / da_nghi"),
    ImportField("ten_chuc_vu", "Chuc vu", help_text="Ten chuc vu (lookup theo ten)"),
    ImportField("ten_bo_phan", "Bo phan", help_text="Ten bo phan (lookup theo ten)"),
    ImportField("he_so_ca_nhan", "He so ca nhan", help_text="So thap phan, VD: 1.5"),
]


def _role_code(user: User) -> str | None:
    return user.role.ma_vai_tro if getattr(user, "role", None) else None

def _sync_leave_to_attendance(req: LeaveRequest, db: Session):
    if req.loai_don != "nghi_phep" or req.trang_thai != "bgd_duyet":
        return
    current = req.ngay_bat_dau.date()
    end = req.ngay_ket_thuc.date()
    while current <= end:
        log = db.query(AttendanceLog).filter(
            AttendanceLog.employee_id == req.employee_id,
            AttendanceLog.ngay == current,
        ).first()
        if not log:
            log = AttendanceLog(employee_id=req.employee_id, ngay=current)
            db.add(log)
        log.trang_thai = "nghi_phep"
        log.so_cong = Decimal("1")
        log.ghi_chu = req.ly_do or "Tu dong cap nhat tu don nghi phep da duyet"
        current += timedelta(days=1)

def _normalize_attendance_key(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")

def _attendance_value(row: dict, *keys: str):
    normalized = {_normalize_attendance_key(k): v for k, v in row.items()}
    for key in keys:
        value = normalized.get(_normalize_attendance_key(key))
        if value not in (None, ""):
            return value
    return None

def _parse_attendance_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        raise ValueError("empty date")
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass

    match = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if match:
        first = int(match.group(1))
        second = int(match.group(2))
        year = int(match.group(3))
        if year < 100:
            year += 2000
        day = first if first > 12 or second <= 12 else second
        month = second if first > 12 or second <= 12 else first
        return date(year, month, day)
    raise ValueError("invalid date")

def _parse_attendance_datetime(value: object, ngay: date) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        pass
    match = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if match:
        return datetime(
            ngay.year,
            ngay.month,
            ngay.day,
            int(match.group(1)),
            int(match.group(2)),
            int(match.group(3) or 0),
        )
    return None

def _decimal_value(row: dict, *keys: str) -> Decimal:
    value = _attendance_value(row, *keys)
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value).replace(",", "."))

def _contract_allowance_total(contract: LaborContract) -> Decimal:
    detailed = sum((
        contract.phu_cap_chuyen_can or Decimal("0"),
        contract.phu_cap_trach_nhiem or Decimal("0"),
        contract.phu_cap_nha_o_com or Decimal("0"),
        contract.phu_cap_dien_thoai or Decimal("0"),
        contract.phu_cap_khac or Decimal("0"),
    ), Decimal("0"))
    legacy = contract.phu_cap or Decimal("0")
    return detailed if detailed > 0 else legacy

# --- Departments ---
# NOTE: Sprint A security review P1 — write endpoints require HR/Admin (master data
# feed vào matrix-role assignment, không cho user thường tạo phòng/chức vụ tùy ý).
@router.get("/departments", response_model=List[schemas.Department])
def list_departments(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Department).all()

@router.post("/departments", response_model=schemas.Department)
def create_department(body: schemas.DepartmentCreate, db: Session = Depends(get_db), current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    if db.query(Department).filter(Department.ma_bo_phan == body.ma_bo_phan).first():
        raise HTTPException(400, "Mã bộ phận đã tồn tại")
    db_dept = Department(**body.model_dump())
    db.add(db_dept)
    db.commit()
    db.refresh(db_dept)
    logger.info("HR department created id=%s ma_bo_phan=%s by user=%s", db_dept.id, db_dept.ma_bo_phan, current_user.id)
    return db_dept

@router.put("/departments/{id}", response_model=schemas.Department)
def update_department(id: int, body: schemas.DepartmentUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    db_dept = db.get(Department, id)
    if not db_dept:
        logger.warning("department id=%s not found", id)
        raise HTTPException(404, "Không tìm thấy bộ phận")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(db_dept, k, v)
    db.commit()
    db.refresh(db_dept)
    logger.info("HR department updated id=%s by user=%s", id, current_user.id)
    return db_dept

@router.delete("/departments/{id}")
def delete_department(id: int, db: Session = Depends(get_db), current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    db_dept = db.get(Department, id)
    if not db_dept:
        raise HTTPException(404, "Không tìm thấy bộ phận")
    # Check có nhân viên/phòng con thuộc bộ phận này không
    has_employees = db.query(Employee).filter(Employee.bo_phan_id == id).first() is not None
    if has_employees:
        raise HTTPException(400, "Không thể xóa: còn nhân viên thuộc bộ phận này. Vui lòng chuyển nhân viên sang bộ phận khác trước.")
    has_children = db.query(Department).filter(Department.parent_id == id).first() is not None
    if has_children:
        raise HTTPException(400, "Không thể xóa: bộ phận này có bộ phận con. Vui lòng xóa hoặc chuyển bộ phận con trước.")
    # Lưu thông tin TRƯỚC khi delete để audit forensic
    ma = db_dept.ma_bo_phan
    ten = db_dept.ten_bo_phan
    db.delete(db_dept)
    db.commit()
    logger.info("HR department deleted id=%s ma=%s ten=%s by user=%s", id, ma, ten, current_user.id)
    return {"ok": True}

# --- Positions ---
@router.get("/positions", response_model=List[schemas.Position])
def list_positions(
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """List chức vụ. Mặc định chỉ trả active — soft-deleted ẩn khỏi dropdown.

    `include_inactive=true` cho HR/Admin xem history.
    """
    q = db.query(Position)
    if not include_inactive:
        q = q.filter(Position.trang_thai == True)
    return q.all()

@router.post("/positions", response_model=schemas.Position)
def create_position(body: schemas.PositionCreate, db: Session = Depends(get_db), current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    if db.query(Position).filter(Position.ma_chuc_vu == body.ma_chuc_vu).first():
        raise HTTPException(400, "Mã chức vụ đã tồn tại")
    db_pos = Position(**body.model_dump())
    db.add(db_pos)
    db.commit()
    db.refresh(db_pos)
    logger.info("HR position created id=%s ma_chuc_vu=%s by user=%s", db_pos.id, db_pos.ma_chuc_vu, current_user.id)
    return db_pos


@router.put("/positions/{id}", response_model=schemas.Position)
def update_position(id: int, body: schemas.PositionUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    db_pos = db.get(Position, id)
    if not db_pos:
        raise HTTPException(404, "Không tìm thấy chức vụ")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(db_pos, k, v)
    db.commit()
    db.refresh(db_pos)
    logger.info("HR position updated id=%s by user=%s", id, current_user.id)
    return db_pos


@router.delete("/positions/{id}")
def delete_position(id: int, db: Session = Depends(get_db), current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    """Soft-delete chức vụ — set trang_thai=False thay vì xóa cứng.

    Thống nhất pattern với Department / Team / CheckInLocation / SafetyEquipment.
    Lý do: giữ history NV từng có chức vụ này → các báo cáo cũ vẫn truy ngược được.
    """
    db_pos = db.get(Position, id)
    if not db_pos:
        raise HTTPException(404, "Không tìm thấy chức vụ")
    # Integrity check — không soft-delete nếu còn NV đang dùng chức vụ này
    has_employees = db.query(Employee).filter(Employee.chuc_vu_id == id).first() is not None
    if has_employees:
        raise HTTPException(400, "Không thể xóa: còn nhân viên thuộc chức vụ này. Hãy chuyển NV sang chức vụ khác trước.")
    ma = db_pos.ma_chuc_vu
    ten = db_pos.ten_chuc_vu
    db_pos.trang_thai = False  # soft-delete
    db.commit()
    logger.info("HR position soft-deleted id=%s ma=%s ten=%s by user=%s", id, ma, ten, current_user.id)
    return {"ok": True, "soft_deleted": True}

# --- Teams (Tổ / Nhóm — cấp dưới Bộ phận) ---
@router.get("/teams", response_model=List[schemas.Team])
def list_teams(
    bo_phan_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """List tổ kèm số NV, tên BP, tên tổ trưởng (cho dropdown + cây tổ chức)."""
    q = db.query(Team)
    if bo_phan_id is not None:
        q = q.filter(Team.bo_phan_id == bo_phan_id)
    teams = q.order_by(Team.ten_to).all()
    # Count NV theo to_id
    counts = dict(
        db.query(Employee.to_id, func.count(Employee.id))
        .filter(Employee.to_id.isnot(None))
        .group_by(Employee.to_id).all()
    )
    result = []
    for t in teams:
        result.append({
            "id": t.id,
            "ten_to": t.ten_to,
            "bo_phan_id": t.bo_phan_id,
            "to_truong_id": t.to_truong_id,
            "mo_ta": t.mo_ta,
            "trang_thai": t.trang_thai,
            "created_at": t.created_at,
            "ten_bo_phan": t.bo_phan.ten_bo_phan if t.bo_phan else None,
            "ho_ten_to_truong": t.to_truong.ho_ten if t.to_truong else None,
            "so_nv": counts.get(t.id, 0),
        })
    return result


@router.post("/teams", response_model=schemas.Team)
def create_team(
    body: schemas.TeamCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    # Validate FK trước (rõ ràng hơn là để bùng IntegrityError 500)
    if body.bo_phan_id is not None and not db.get(Department, body.bo_phan_id):
        raise HTTPException(400, "bo_phan_id không tồn tại")
    if body.to_truong_id is not None and not db.get(Employee, body.to_truong_id):
        raise HTTPException(400, "to_truong_id không tồn tại")
    # Check duplicate trong cùng bộ phận (case-insensitive)
    norm = body.ten_to.strip().lower()
    existing = (
        db.query(Team)
        .filter(Team.bo_phan_id == body.bo_phan_id, func.lower(Team.ten_to) == norm)
        .first()
    )
    if existing:
        raise HTTPException(400, f"Tổ '{body.ten_to}' đã tồn tại trong bộ phận này")
    t = Team(**body.model_dump())
    db.add(t)
    db.commit()
    db.refresh(t)
    logger.info("HR team created id=%s ten=%s by user=%s", t.id, t.ten_to, current_user.id)
    return {
        "id": t.id, "ten_to": t.ten_to, "bo_phan_id": t.bo_phan_id,
        "to_truong_id": t.to_truong_id, "mo_ta": t.mo_ta, "trang_thai": t.trang_thai,
        "created_at": t.created_at, "ten_bo_phan": None, "ho_ten_to_truong": None, "so_nv": 0,
    }


@router.put("/teams/{id}", response_model=schemas.Team)
def update_team(
    id: int,
    body: schemas.TeamUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    t = db.get(Team, id)
    if not t:
        raise HTTPException(404, "Không tìm thấy tổ")
    data = body.model_dump(exclude_unset=True)
    # Check duplicate nếu đổi tên hoặc đổi bộ phận
    if "ten_to" in data or "bo_phan_id" in data:
        new_name = data.get("ten_to", t.ten_to).strip().lower()
        new_bp = data.get("bo_phan_id", t.bo_phan_id)
        dup = (
            db.query(Team)
            .filter(Team.bo_phan_id == new_bp, func.lower(Team.ten_to) == new_name, Team.id != id)
            .first()
        )
        if dup:
            raise HTTPException(400, "Tên tổ đã tồn tại trong bộ phận này")
    for k, v in data.items():
        setattr(t, k, v)
    db.commit()
    db.refresh(t)
    so_nv = db.query(func.count(Employee.id)).filter(Employee.to_id == t.id).scalar() or 0
    logger.info("HR team updated id=%s by user=%s", id, current_user.id)
    return {
        "id": t.id, "ten_to": t.ten_to, "bo_phan_id": t.bo_phan_id,
        "to_truong_id": t.to_truong_id, "mo_ta": t.mo_ta, "trang_thai": t.trang_thai,
        "created_at": t.created_at,
        "ten_bo_phan": t.bo_phan.ten_bo_phan if t.bo_phan else None,
        "ho_ten_to_truong": t.to_truong.ho_ten if t.to_truong else None,
        "so_nv": so_nv,
    }


@router.delete("/teams/{id}")
def delete_team(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    t = db.get(Team, id)
    if not t:
        raise HTTPException(404, "Không tìm thấy tổ")
    # Integrity check: nếu còn NV thuộc tổ này → từ chối (HR phải chuyển NV trước)
    nv_count = db.query(func.count(Employee.id)).filter(Employee.to_id == id).scalar() or 0
    if nv_count > 0:
        raise HTTPException(
            400,
            f"Không thể xóa: còn {nv_count} NV thuộc tổ này. Hãy chuyển NV sang tổ khác trước.",
        )
    db.delete(t)
    db.commit()
    logger.info("HR team deleted id=%s by user=%s", id, current_user.id)
    return {"ok": True}


# --- Employees ---
@router.get("/employees")
def list_employees(
    search: str = Query(default=""),
    phan_xuong_id: Optional[int] = None,
    phap_nhan_id: Optional[int] = None,
    bo_phan_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_has_permission("hr.view", current_user, db)
    q = db.query(Employee)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(Employee.ma_nv.ilike(like), Employee.ho_ten.ilike(like)))
    if phan_xuong_id:
        q = q.filter(Employee.phan_xuong_id == phan_xuong_id)
    if phap_nhan_id:
        q = q.filter(Employee.phap_nhan_id == phap_nhan_id)
    is_hr_admin = _role_code(current_user) in ("ADMIN", "NHAN_SU")

    # Scoped manager: auto-restrict to own department, ignore caller-supplied filter
    if not is_hr_admin and current_user.bo_phan_id:
        q = q.filter(Employee.bo_phan_id == current_user.bo_phan_id)
    elif bo_phan_id:
        q = q.filter(Employee.bo_phan_id == bo_phan_id)

    employees = q.all()
    result = []
    for e in employees:
        row = {
            "id": e.id,
            "ma_nv": e.ma_nv,
            "ho_ten": e.ho_ten,
            "ngay_sinh": e.ngay_sinh.isoformat() if e.ngay_sinh else None,
            "gioi_tinh": e.gioi_tinh,
            "so_dien_thoai": e.so_dien_thoai,
            "email": e.email,
            "trang_thai": e.trang_thai,
            "ngay_vao_lam": e.ngay_vao_lam.isoformat() if e.ngay_vao_lam else None,
            "ngay_nghi_viec": e.ngay_nghi_viec.isoformat() if e.ngay_nghi_viec else None,
            "phap_nhan_id": e.phap_nhan_id,
            "phan_xuong_id": e.phan_xuong_id,
            "bo_phan_id": e.bo_phan_id,
            "chuc_vu_id": e.chuc_vu_id,
            "ten_bo_phan": e.bo_phan.ten_bo_phan if e.bo_phan else None,
            "ten_chuc_vu": e.chuc_vu.ten_chuc_vu if e.chuc_vu else None,
            "ten_phan_xuong": e.phan_xuong.ten_xuong if e.phan_xuong else None,
            "ten_phap_nhan": e.phap_nhan.ten_phap_nhan if e.phap_nhan else None,
            "to_id": e.to_id,
            "ten_to": e.to_nhom.ten_to if e.to_nhom else None,
            "has_account": e.user_id is not None,
        }
        # Wage + PII + account: HR/Admin only — salary coeff, CCCD, bank, fingerprint
        if is_hr_admin:
            row["he_so_ca_nhan"] = float(e.he_so_ca_nhan or 0)
            row["cccd"] = e.cccd
            row["ngay_cap"] = e.ngay_cap.isoformat() if e.ngay_cap else None
            row["noi_cap"] = e.noi_cap
            row["dia_chi"] = e.dia_chi
            row["que_quan"] = e.que_quan
            row["so_tk_ngan_hang"] = e.so_tk_ngan_hang
            row["ten_ngan_hang"] = e.ten_ngan_hang
            row["chi_nhanh_ngan_hang"] = e.chi_nhanh_ngan_hang
            row["ma_van_tay"] = e.ma_van_tay
            row["username"] = e.user.username if e.user else None
            row["user_status"] = e.user.trang_thai if e.user else None
        result.append(row)
    return result

@router.get("/employees/import-template")
def employee_import_template(_: User = Depends(get_current_user)):
    """Tải file Excel mẫu để import nhân viên."""
    return build_template_response("mau_import_nhan_vien.xlsx", HR_EMPLOYEE_IMPORT_FIELDS)


@router.get("/employees/{id}")
def get_employee(id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Detail endpoint trả ĐẦY ĐỦ extended fields (sơ yếu, BHXH, hộ khẩu...).

    UI EmployeeProfileModal gọi endpoint này để fill toàn bộ form. List endpoint
    chỉ trả basic fields đủ render table để tiết kiệm payload.
    """
    assert_has_permission("hr.view", current_user, db)
    e = db.query(Employee).filter(Employee.id == id).first()
    if not e:
        raise HTTPException(404, "Không tìm thấy nhân viên")
    is_hr_admin = _role_code(current_user) in ("ADMIN", "NHAN_SU")
    row: dict = {
        "id": e.id,
        "ma_nv": e.ma_nv,
        "ho_ten": e.ho_ten,
        "ngay_sinh": e.ngay_sinh.isoformat() if e.ngay_sinh else None,
        "gioi_tinh": e.gioi_tinh,
        "so_dien_thoai": e.so_dien_thoai,
        "email": e.email,
        "trang_thai": e.trang_thai,
        "ngay_vao_lam": e.ngay_vao_lam.isoformat() if e.ngay_vao_lam else None,
        "ngay_nghi_viec": e.ngay_nghi_viec.isoformat() if e.ngay_nghi_viec else None,
        "phap_nhan_id": e.phap_nhan_id,
        "phan_xuong_id": e.phan_xuong_id,
        "bo_phan_id": e.bo_phan_id,
        "chuc_vu_id": e.chuc_vu_id,
        "ten_bo_phan": e.bo_phan.ten_bo_phan if e.bo_phan else None,
        "ten_chuc_vu": e.chuc_vu.ten_chuc_vu if e.chuc_vu else None,
        "ten_phan_xuong": e.phan_xuong.ten_xuong if e.phan_xuong else None,
        "ten_phap_nhan": e.phap_nhan.ten_phap_nhan if e.phap_nhan else None,
        "to_id": e.to_id,
        "ten_to": e.to_nhom.ten_to if e.to_nhom else None,
        "has_account": e.user_id is not None,
    }
    if is_hr_admin:
        # PII + bank + salary coeff
        row["he_so_ca_nhan"] = float(e.he_so_ca_nhan or 0)
        row["cccd"] = e.cccd
        row["ngay_cap"] = e.ngay_cap.isoformat() if e.ngay_cap else None
        row["noi_cap"] = e.noi_cap
        row["dia_chi"] = e.dia_chi
        row["que_quan"] = e.que_quan
        row["so_tk_ngan_hang"] = e.so_tk_ngan_hang
        row["ten_ngan_hang"] = e.ten_ngan_hang
        row["chi_nhanh_ngan_hang"] = e.chi_nhanh_ngan_hang
        row["ma_van_tay"] = e.ma_van_tay
        row["username"] = e.user.username if e.user else None
        row["user_status"] = e.user.trang_thai if e.user else None
        # Extended fields cho form chi tiết (HR-only — nhiều thông tin nhạy cảm)
        for fld in (
            "ho_dem", "ten", "ten_bi_danh",
            "quoc_tich", "dan_toc", "ton_giao",
            "noi_sinh_tinh", "noi_sinh_dia_chi",
            "tinh_que_quan", "huyen_que_quan", "phuong_que_quan", "dia_chi_que_quan",
            "tinh_ho_khau", "huyen_ho_khau", "phuong_ho_khau", "dia_chi_ho_khau",
            "dia_chi_hien_tai", "dien_thoai_ban", "avatar_url",
            "trinh_do_hoc_van", "chuyen_nganh", "truong_dao_tao",
            "ngoai_ngu", "tin_hoc", "ky_nang_khac", "so_yeu_tom_tat",
            "so_so_bhxh", "ma_bhyt", "noi_kham_chua_benh",
        ):
            row[fld] = getattr(e, fld, None)
        # Numeric / date fields
        row["nam_tot_nghiep"] = e.nam_tot_nghiep
        row["ngay_tham_gia_bhxh"] = e.ngay_tham_gia_bhxh.isoformat() if e.ngay_tham_gia_bhxh else None
        row["muc_dong_bhxh"] = float(e.muc_dong_bhxh) if e.muc_dong_bhxh is not None else None
    return row

@router.post("/employees", response_model=schemas.Employee)
def create_employee(body: schemas.EmployeeCreate, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    if db.query(Employee).filter(Employee.ma_nv == body.ma_nv).first():
        raise HTTPException(400, "Mã nhân viên đã tồn tại")
    db_emp = Employee(**body.model_dump())
    db.add(db_emp)
    db.commit()
    db.refresh(db_emp)
    logger.info("created employee id=%s ma_nv=%s", db_emp.id, db_emp.ma_nv)
    return db_emp

@router.post("/employees/bulk")
def bulk_create_employees(body: schemas.EmployeeBulkCreate, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    created = 0
    updated = 0
    errors = []
    for idx, item in enumerate(body.items, start=1):
        data = item.model_dump()
        try:
            db_emp = db.query(Employee).filter(Employee.ma_nv == item.ma_nv).first()
            if db_emp:
                for k, v in data.items():
                    setattr(db_emp, k, v)
                updated += 1
            else:
                db.add(Employee(**data))
                created += 1
        except Exception as exc:
            errors.append({"row": idx, "ma_nv": item.ma_nv, "error": str(exc)})
    if errors:
        db.rollback()
        raise HTTPException(400, {"message": "Import nhan vien co loi", "errors": errors})
    db.commit()
    return {"ok": True, "created": created, "updated": updated}

@router.put("/employees/{id}", response_model=schemas.Employee)
def update_employee(
    id: int,
    body: schemas.EmployeeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))
):
    from app.models.auth import Role, User as AuthUser
    db_emp = db.query(Employee).filter(Employee.id == id).first()
    if not db_emp:
        logger.warning("employee id=%s not found", id)
        raise HTTPException(404, "Không tìm thấy nhân viên")
    
    # Ghi log nếu thay đổi hệ số
    if body.he_so_ca_nhan is not None and body.he_so_ca_nhan != db_emp.he_so_ca_nhan:
        history = EmployeeHistory(
            employee_id=id,
            loai="he_so",
            gia_tri_cu=str(db_emp.he_so_ca_nhan),
            gia_tri_moi=str(body.he_so_ca_nhan),
            ly_do="Cập nhật hệ số lương",
            created_by=current_user.id
        )
        db.add(history)

    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(db_emp, k, v)
    
    # Matrix Role Assignment
    if db_emp.user_id and db_emp.bo_phan_id and db_emp.chuc_vu_id:
        user = db.get(AuthUser, db_emp.user_id)
        if user and user.role and user.role.ma_vai_tro != 'ADMIN':
            dept = db.get(Department, db_emp.bo_phan_id)
            pos = db.get(Position, db_emp.chuc_vu_id)
            if dept and pos:
                role_code = f"{dept.ma_bo_phan}_{pos.ma_chuc_vu}"
                new_role = db.query(Role).filter(Role.ma_vai_tro == role_code).first()
                if new_role:
                    user.role_id = new_role.id

    db.commit()
    db.refresh(db_emp)
    logger.info("updated employee id=%s", id)
    return db_emp

@router.delete("/employees/{id}")
def delete_employee(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN")),
):
    """Xóa cứng nhân viên. Chỉ ADMIN. Bị chặn nếu NV đã có dữ liệu lương/hợp đồng/chấm công."""
    emp = db.query(Employee).filter(Employee.id == id).first()
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")

    # Guard: còn hợp đồng
    if db.query(LaborContract).filter(LaborContract.employee_id == id).first():
        raise HTTPException(400, "Không thể xóa: nhân viên đã có hợp đồng lao động")

    # Guard: còn dữ liệu lương
    if db.query(PayrollRun).filter(PayrollRun.employee_id == id).first():
        raise HTTPException(400, "Không thể xóa: nhân viên đã có dữ liệu tính lương")

    ma_nv = emp.ma_nv
    ho_ten = emp.ho_ten
    db.delete(emp)
    db.commit()
    logger.info("employee hard-deleted id=%s ma_nv=%s ho_ten=%s by user=%s", id, ma_nv, ho_ten, current_user.id)
    return {"ok": True, "deleted": {"id": id, "ma_nv": ma_nv, "ho_ten": ho_ten}}


@router.get("/employees/{id}/history")
def get_employee_history(id: int, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    return db.query(EmployeeHistory).filter(EmployeeHistory.employee_id == id).order_by(EmployeeHistory.created_at.desc()).all()


# --- Check-in Locations (Sprint B — geo-fence chấm công) ---
@router.get("/checkin-locations", response_model=List[schemas.CheckInLocation])
def list_checkin_locations(
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """List địa điểm chấm công. Mọi user authenticated đọc được (cần cho mobile)."""
    q = db.query(CheckInLocation)
    if not include_inactive:
        q = q.filter(CheckInLocation.is_active.is_(True))
    return q.order_by(CheckInLocation.ten).all()


@router.post("/checkin-locations", response_model=schemas.CheckInLocation)
def create_checkin_location(
    payload: schemas.CheckInLocationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    loc = CheckInLocation(**payload.model_dump())
    db.add(loc)
    db.commit()
    db.refresh(loc)
    logger.info("HR checkin-location created id=%s ten=%s by user=%s", loc.id, loc.ten, current_user.id)
    return loc


@router.put("/checkin-locations/{loc_id}", response_model=schemas.CheckInLocation)
def update_checkin_location(
    loc_id: int,
    payload: schemas.CheckInLocationUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    loc = db.query(CheckInLocation).filter(CheckInLocation.id == loc_id).first()
    if not loc:
        raise HTTPException(404, "Không tìm thấy địa điểm")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(loc, k, v)
    db.commit()
    db.refresh(loc)
    logger.info("HR checkin-location updated id=%s by user=%s", loc_id, current_user.id)
    return loc


@router.delete("/checkin-locations/{loc_id}")
def delete_checkin_location(
    loc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    """Soft-delete: set is_active=False thay vì xóa thật để giữ history
    AttendanceLog.checkin_location_id (FK SET NULL nhưng location record vẫn cần để
    truy ngược ai chấm ở đâu trong quá khứ)."""
    loc = db.query(CheckInLocation).filter(CheckInLocation.id == loc_id).first()
    if not loc:
        raise HTTPException(404, "Không tìm thấy địa điểm")
    loc.is_active = False
    db.commit()
    logger.info("HR checkin-location soft-deleted id=%s ten=%s by user=%s", loc_id, loc.ten, current_user.id)
    return {"ok": True, "message": "Đã chuyển sang tạm dừng (soft-delete để giữ lịch sử chấm công)"}


@router.get("/attendance/today")
def list_attendance_today(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    """HR view: danh sách chấm công hôm nay (realtime). Join với Employee + Location."""
    today = date.today()
    rows = (
        db.query(AttendanceLog, Employee, CheckInLocation)
        .join(Employee, Employee.id == AttendanceLog.employee_id)
        .outerjoin(CheckInLocation, CheckInLocation.id == AttendanceLog.checkin_location_id)
        .filter(AttendanceLog.ngay == today)
        .order_by(AttendanceLog.gio_vao.desc().nullslast())
        .all()
    )
    return [
        {
            "log_id": log.id,
            "employee_id": emp.id,
            "ma_nv": emp.ma_nv,
            "ho_ten": emp.ho_ten,
            "gio_vao": log.gio_vao.isoformat() if log.gio_vao else None,
            "gio_ra": log.gio_ra.isoformat() if log.gio_ra else None,
            "loai": log.loai,
            "checkin_lat": log.checkin_lat,
            "checkin_lng": log.checkin_lng,
            "checkin_address": log.checkin_address,
            "checkin_selfie_url": log.checkin_selfie_url,
            "checkin_distance_m": log.checkin_distance_m,
            "checkout_address": log.checkout_address,
            "location_id": loc.id if loc else None,
            "location_name": loc.ten if loc else None,
            "trang_thai": log.trang_thai,
        }
        for (log, emp, loc) in rows
    ]


# --- Family Relations ---
@router.get("/employees/{id}/family-relations", response_model=List[schemas.FamilyRelation])
def list_family_relations(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List quan hệ gia đình.

    PII của người thân (không phải nhân viên consent) → giới hạn:
    - ADMIN / NHAN_SU: xem được tất cả
    - User thường: chỉ xem được CỦA CHÍNH MÌNH (qua Employee.user_id)
    """
    # Check role — current_user.role là Role ORM object, dùng ma_vai_tro
    role = (current_user.role.ma_vai_tro or "").upper() if getattr(current_user, "role", None) else ""
    is_hr_admin = role in ("ADMIN", "NHAN_SU")
    if not is_hr_admin:
        emp = db.query(Employee).filter(Employee.id == id).first()
        if not emp or emp.user_id != current_user.id:
            raise HTTPException(403, "Bạn chỉ xem được quan hệ gia đình của chính mình")
    return db.query(FamilyRelation).filter(FamilyRelation.employee_id == id).order_by(FamilyRelation.id).all()


@router.post("/employees/{id}/family-relations", response_model=schemas.FamilyRelation)
def create_family_relation(
    id: int,
    payload: schemas.FamilyRelationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    emp = db.query(Employee).filter(Employee.id == id).first()
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")
    fr = FamilyRelation(employee_id=id, **payload.model_dump())
    db.add(fr)
    db.commit()
    db.refresh(fr)
    logger.info("HR family-relation created id=%s employee=%s by user=%s", fr.id, id, current_user.id)
    return fr


@router.put("/family-relations/{fr_id}", response_model=schemas.FamilyRelation)
def update_family_relation(
    fr_id: int,
    payload: schemas.FamilyRelationUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    fr = db.query(FamilyRelation).filter(FamilyRelation.id == fr_id).first()
    if not fr:
        raise HTTPException(404, "Không tìm thấy quan hệ gia đình")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(fr, k, v)
    db.commit()
    db.refresh(fr)
    logger.info("HR family-relation updated id=%s by user=%s", fr.id, current_user.id)
    return fr


@router.delete("/family-relations/{fr_id}")
def delete_family_relation(
    fr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    fr = db.query(FamilyRelation).filter(FamilyRelation.id == fr_id).first()
    if not fr:
        raise HTTPException(404, "Không tìm thấy quan hệ gia đình")
    employee_id = fr.employee_id
    db.delete(fr)
    db.commit()
    logger.info("HR family-relation deleted id=%s employee=%s by user=%s", fr_id, employee_id, current_user.id)
    return {"ok": True}


# --- Employee Documents (File hồ sơ) ---
@router.get("/employees/{id}/documents", response_model=List[schemas.EmployeeDocument])
def list_employee_documents(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List file hồ sơ. User thường chỉ xem được của mình; HR/Admin xem tất cả."""
    role = (current_user.role.ma_vai_tro or "").upper() if getattr(current_user, "role", None) else ""
    is_hr_admin = role in ("ADMIN", "NHAN_SU")
    if not is_hr_admin:
        emp = db.query(Employee).filter(Employee.id == id).first()
        if not emp or emp.user_id != current_user.id:
            raise HTTPException(403, "Bạn chỉ xem được hồ sơ của chính mình")
    return db.query(EmployeeDocument).filter(EmployeeDocument.employee_id == id).order_by(EmployeeDocument.id.desc()).all()


@router.post("/employees/{id}/documents", response_model=schemas.EmployeeDocument)
def create_employee_document(
    id: int,
    payload: schemas.EmployeeDocumentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    """Tạo bản ghi metadata cho tài liệu. File upload thật phải qua /api/media/upload trước,
    rồi truyền file_path vào đây."""
    emp = db.query(Employee).filter(Employee.id == id).first()
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")
    doc = EmployeeDocument(employee_id=id, **payload.model_dump())
    db.add(doc)
    db.commit()
    db.refresh(doc)
    logger.info("HR document created id=%s employee=%s by user=%s", doc.id, id, current_user.id)
    return doc


@router.delete("/employee-documents/{doc_id}")
def delete_employee_document(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    doc = db.query(EmployeeDocument).filter(EmployeeDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(404, "Không tìm thấy tài liệu")
    employee_id = doc.employee_id
    db.delete(doc)
    db.commit()
    logger.info("HR document deleted id=%s employee=%s by user=%s", doc_id, employee_id, current_user.id)
    return {"ok": True}


# --- Employee History (Thuyên chuyển, Chức vụ, Lương, Phụ cấp) ---
@router.get("/employees/{id}/history-typed", response_model=List[schemas.EmployeeHistory])
def list_employee_history_typed(
    id: int,
    loai: Optional[str] = Query(None, description="Filter: he_so | chuc_vu | bo_phan | luong_cb | phu_cap"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List lịch sử thay đổi. User thường chỉ xem được của mình."""
    role = (current_user.role.ma_vai_tro or "").upper() if getattr(current_user, "role", None) else ""
    is_hr_admin = role in ("ADMIN", "NHAN_SU")
    if not is_hr_admin:
        emp = db.query(Employee).filter(Employee.id == id).first()
        if not emp or emp.user_id != current_user.id:
            raise HTTPException(403, "Bạn chỉ xem được lịch sử của chính mình")
    q = db.query(EmployeeHistory).filter(EmployeeHistory.employee_id == id)
    if loai:
        q = q.filter(EmployeeHistory.loai == loai)
    return q.order_by(EmployeeHistory.ngay_hieu_luc.desc(), EmployeeHistory.id.desc()).all()


@router.post("/employees/{id}/history-typed", response_model=schemas.EmployeeHistory)
def create_employee_history(
    id: int,
    payload: schemas.EmployeeHistoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    emp = db.query(Employee).filter(Employee.id == id).first()
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")
    hist = EmployeeHistory(employee_id=id, created_by=current_user.id, **payload.model_dump())
    db.add(hist)
    db.commit()
    db.refresh(hist)
    logger.info("HR history created id=%s employee=%s loai=%s by user=%s", hist.id, id, hist.loai, current_user.id)
    return hist


@router.delete("/employee-history/{hist_id}")
def delete_employee_history(
    hist_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    hist = db.query(EmployeeHistory).filter(EmployeeHistory.id == hist_id).first()
    if not hist:
        raise HTTPException(404, "Không tìm thấy bản ghi lịch sử")
    employee_id = hist.employee_id
    db.delete(hist)
    db.commit()
    logger.info("HR history deleted id=%s employee=%s by user=%s", hist_id, employee_id, current_user.id)
    return {"ok": True}


# --- Labor Contracts (Quá trình hợp đồng) ---
@router.get("/employees/{id}/contracts", response_model=List[schemas.LaborContract])
def list_employee_contracts(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List hợp đồng của nhân viên. User thường chỉ xem được của mình."""
    role = (current_user.role.ma_vai_tro or "").upper() if getattr(current_user, "role", None) else ""
    is_hr_admin = role in ("ADMIN", "NHAN_SU")
    if not is_hr_admin:
        emp = db.query(Employee).filter(Employee.id == id).first()
        if not emp or emp.user_id != current_user.id:
            raise HTTPException(403, "Bạn chỉ xem được hợp đồng của chính mình")
    return db.query(LaborContract).filter(LaborContract.employee_id == id).order_by(LaborContract.ngay_hieu_luc.desc()).all()


# --- Contracts & Warnings ---
@router.get("/contracts/expiring")
def list_expiring_contracts(days: int = 30, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    from datetime import timedelta
    limit_date = date.today() + timedelta(days=days)
    contracts = db.query(LaborContract).filter(
        LaborContract.trang_thai == "hieu_luc",
        LaborContract.ngay_het_han <= limit_date,
        LaborContract.ngay_het_han >= date.today()
    ).all()
    
    result = []
    for c in contracts:
        result.append({
            "id": c.id,
            "employee_id": c.employee_id,
            "ho_ten": c.employee.ho_ten,
            "so_hop_dong": c.so_hop_dong,
            "ngay_het_han": c.ngay_het_han.isoformat(),
            "con_lai": (c.ngay_het_han - date.today()).days
        })
    return result

# --- Health Checks (Khám sức khỏe định kỳ) ---
@router.get("/health-checks")
def list_health_checks(
    employee_id: Optional[int] = None,
    phan_loai: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    due_soon_days: Optional[int] = None,
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List khám sức khỏe.

    - `employee_id`: lọc theo 1 NV
    - `phan_loai`: I/II/III/IV/V
    - `from_date`, `to_date`: lọc theo khoảng ngày khám
    - `due_soon_days`: trả về NV có ngay_kham_tiep_theo trong N ngày tới
    """
    is_hr_admin = _role_code(current_user) in ("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")
    if not is_hr_admin:
        # NV thường chỉ xem được của chính mình
        my_emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not my_emp:
            return []
        employee_id = my_emp.id

    q = db.query(HealthCheck)
    if employee_id:
        q = q.filter(HealthCheck.employee_id == employee_id)
    if phan_loai:
        q = q.filter(HealthCheck.phan_loai_suc_khoe == phan_loai)
    if from_date:
        q = q.filter(HealthCheck.ngay_kham >= from_date)
    if to_date:
        q = q.filter(HealthCheck.ngay_kham <= to_date)
    if due_soon_days is not None:
        from datetime import timedelta
        limit = date.today() + timedelta(days=due_soon_days)
        q = q.filter(
            HealthCheck.ngay_kham_tiep_theo.isnot(None),
            HealthCheck.ngay_kham_tiep_theo <= limit,
        )

    records = q.order_by(HealthCheck.ngay_kham.desc()).offset(offset).limit(limit).all()
    result = []
    for hc in records:
        emp = hc.employee
        result.append({
            "id": hc.id,
            "employee_id": hc.employee_id,
            "ngay_kham": hc.ngay_kham.isoformat() if hc.ngay_kham else None,
            "loai_kham": hc.loai_kham,
            "phan_loai_suc_khoe": hc.phan_loai_suc_khoe,
            "noi_kham": hc.noi_kham,
            "bac_si": hc.bac_si,
            "ket_luan": hc.ket_luan,
            "benh_man_tinh": hc.benh_man_tinh,
            "file_url": hc.file_url,
            "chi_phi": float(hc.chi_phi) if hc.chi_phi is not None else 0,
            "ngay_kham_tiep_theo": hc.ngay_kham_tiep_theo.isoformat() if hc.ngay_kham_tiep_theo else None,
            "ghi_chu": hc.ghi_chu,
            "created_at": hc.created_at.isoformat() if hc.created_at else None,
            "ho_ten": emp.ho_ten if emp else None,
            "ma_nv": emp.ma_nv if emp else None,
            "ten_bo_phan": emp.bo_phan.ten_bo_phan if emp and emp.bo_phan else None,
            "ten_phap_nhan": emp.phap_nhan.ten_phap_nhan if emp and emp.phap_nhan else None,
        })
    return result


@router.get("/health-checks/summary")
def health_check_summary(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")),
):
    """Stats tổng quan cho dashboard sức khỏe."""
    from datetime import timedelta
    today = date.today()
    days_30 = today + timedelta(days=30)
    days_60 = today + timedelta(days=60)

    total_records = db.query(func.count(HealthCheck.id)).scalar() or 0

    # NV có ít nhất 1 lần khám (DISTINCT)
    nv_da_kham = db.query(func.count(func.distinct(HealthCheck.employee_id))).scalar() or 0
    total_nv = db.query(func.count(Employee.id)).filter(Employee.trang_thai == "dang_lam").scalar() or 0
    nv_chua_kham = max(0, total_nv - nv_da_kham)

    # Sắp đến hạn / quá hạn (dựa trên ngay_kham_tiep_theo của LẦN KHÁM GẦN NHẤT mỗi NV)
    # Subquery: ngày khám tiếp theo gần nhất của mỗi NV
    latest_next = (
        db.query(
            HealthCheck.employee_id,
            func.max(HealthCheck.ngay_kham_tiep_theo).label("next_due"),
        )
        .filter(HealthCheck.ngay_kham_tiep_theo.isnot(None))
        .group_by(HealthCheck.employee_id)
        .subquery()
    )
    due_30 = db.query(func.count(latest_next.c.employee_id)).filter(
        latest_next.c.next_due >= today, latest_next.c.next_due <= days_30,
    ).scalar() or 0
    due_60 = db.query(func.count(latest_next.c.employee_id)).filter(
        latest_next.c.next_due >= today, latest_next.c.next_due <= days_60,
    ).scalar() or 0
    overdue = db.query(func.count(latest_next.c.employee_id)).filter(
        latest_next.c.next_due < today,
    ).scalar() or 0

    # Phân loại sức khỏe (theo lần khám gần nhất mỗi NV)
    by_phan_loai_raw = db.query(
        HealthCheck.phan_loai_suc_khoe, func.count(HealthCheck.id)
    ).filter(HealthCheck.phan_loai_suc_khoe.isnot(None)).group_by(HealthCheck.phan_loai_suc_khoe).all()
    by_phan_loai = [{"name": p, "value": c} for p, c in by_phan_loai_raw]

    return {
        "total_records": total_records,
        "total_nv": total_nv,
        "nv_da_kham": nv_da_kham,
        "nv_chua_kham": nv_chua_kham,
        "due_30": due_30,
        "due_60": due_60,
        "overdue": overdue,
        "by_phan_loai": by_phan_loai,
    }


@router.post("/health-checks", response_model=schemas.HealthCheck)
def create_health_check(
    body: schemas.HealthCheckCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    # Validate FK
    emp = db.get(Employee, body.employee_id)
    if not emp:
        raise HTTPException(400, "employee_id không tồn tại")
    # Auto-tính ngày khám tiếp theo nếu chưa set (default 12 tháng)
    data = body.model_dump()
    if not data.get("ngay_kham_tiep_theo"):
        from dateutil.relativedelta import relativedelta
        try:
            data["ngay_kham_tiep_theo"] = data["ngay_kham"] + relativedelta(months=12)
        except Exception:  # noqa: BLE001
            from datetime import timedelta
            data["ngay_kham_tiep_theo"] = data["ngay_kham"] + timedelta(days=365)

    hc = HealthCheck(**data, created_by_id=current_user.id)
    db.add(hc)
    db.commit()
    db.refresh(hc)
    logger.info("HR health_check created id=%s emp=%s by user=%s", hc.id, body.employee_id, current_user.id)
    return _serialize_health_check(hc)


@router.put("/health-checks/{id}", response_model=schemas.HealthCheck)
def update_health_check(
    id: int,
    body: schemas.HealthCheckUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    hc = db.get(HealthCheck, id)
    if not hc:
        raise HTTPException(404, "Không tìm thấy lần khám")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(hc, k, v)
    db.commit()
    db.refresh(hc)
    logger.info("HR health_check updated id=%s by user=%s", id, current_user.id)
    return _serialize_health_check(hc)


@router.delete("/health-checks/{id}")
def delete_health_check(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    hc = db.get(HealthCheck, id)
    if not hc:
        raise HTTPException(404, "Không tìm thấy lần khám")
    db.delete(hc)
    db.commit()
    logger.info("HR health_check deleted id=%s by user=%s", id, current_user.id)
    return {"ok": True}


def _serialize_health_check(hc: HealthCheck) -> dict:
    """Trả về dict đầy đủ kèm enriched fields (ho_ten, ten_bo_phan…)."""
    emp = hc.employee
    return {
        "id": hc.id,
        "employee_id": hc.employee_id,
        "ngay_kham": hc.ngay_kham,
        "loai_kham": hc.loai_kham,
        "phan_loai_suc_khoe": hc.phan_loai_suc_khoe,
        "noi_kham": hc.noi_kham,
        "bac_si": hc.bac_si,
        "ket_luan": hc.ket_luan,
        "benh_man_tinh": hc.benh_man_tinh,
        "file_url": hc.file_url,
        "chi_phi": hc.chi_phi,
        "ngay_kham_tiep_theo": hc.ngay_kham_tiep_theo,
        "ghi_chu": hc.ghi_chu,
        "created_at": hc.created_at,
        "created_by_id": hc.created_by_id,
        "ho_ten": emp.ho_ten if emp else None,
        "ma_nv": emp.ma_nv if emp else None,
        "ten_bo_phan": emp.bo_phan.ten_bo_phan if emp and emp.bo_phan else None,
        "ten_phap_nhan": emp.phap_nhan.ten_phap_nhan if emp and emp.phap_nhan else None,
    }


# --- HR Dashboard Overview ---
@router.get("/dashboard/overview")
def hr_dashboard_overview(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU", "GIAM_DOC", "BGD")),
):
    """Dashboard tổng quan HR cho BGĐ — headcount + cơ cấu + sự kiện sắp tới.

    Visible: ADMIN/NHAN_SU/GIAM_DOC/BGD. PII không trả về (chỉ số tổng hợp).
    """
    from datetime import timedelta
    today = date.today()
    days_30 = today + timedelta(days=30)
    days_60 = today + timedelta(days=60)
    today_year_start = date(today.year, 1, 1)

    # ─── 1) Tổng quan headcount ───
    total = db.query(func.count(Employee.id)).scalar() or 0
    by_status = dict(
        db.query(Employee.trang_thai, func.count(Employee.id))
        .group_by(Employee.trang_thai).all()
    )
    by_gender = dict(
        db.query(Employee.gioi_tinh, func.count(Employee.id))
        .filter(Employee.trang_thai == "dang_lam")
        .group_by(Employee.gioi_tinh).all()
    )

    # ─── 2) Phân bổ theo pháp nhân + bộ phận ───
    by_phap_nhan = [
        {"name": pn or "Chưa gán", "value": cnt}
        for pn, cnt in db.query(
            func.coalesce(PhapNhan.ten_viet_tat, PhapNhan.ten_phap_nhan), func.count(Employee.id)
        ).join(Employee, Employee.phap_nhan_id == PhapNhan.id)
         .filter(Employee.trang_thai == "dang_lam")
         .group_by(PhapNhan.id, PhapNhan.ten_viet_tat, PhapNhan.ten_phap_nhan)
         .order_by(func.count(Employee.id).desc()).all()
    ]
    by_bo_phan = [
        {"name": bp, "value": cnt}
        for bp, cnt in db.query(Department.ten_bo_phan, func.count(Employee.id))
        .join(Employee, Employee.bo_phan_id == Department.id)
        .filter(Employee.trang_thai == "dang_lam")
        .group_by(Department.id, Department.ten_bo_phan)
        .order_by(func.count(Employee.id).desc()).limit(10).all()
    ]

    # ─── 3) Cơ cấu độ tuổi (5 nhóm) ───
    age_buckets = {"≤25": 0, "26-35": 0, "36-45": 0, "46-55": 0, ">55": 0}
    for ngay_sinh, in db.query(Employee.ngay_sinh).filter(
        Employee.trang_thai == "dang_lam",
        Employee.ngay_sinh.isnot(None),
    ).all():
        if not ngay_sinh:
            continue
        age = today.year - ngay_sinh.year - (
            (today.month, today.day) < (ngay_sinh.month, ngay_sinh.day)
        )
        if age <= 25: age_buckets["≤25"] += 1
        elif age <= 35: age_buckets["26-35"] += 1
        elif age <= 45: age_buckets["36-45"] += 1
        elif age <= 55: age_buckets["46-55"] += 1
        else: age_buckets[">55"] += 1
    age_distribution = [{"name": k, "value": v} for k, v in age_buckets.items()]

    # ─── 4) Cơ cấu thâm niên (3 nhóm) ───
    tenure_buckets = {"<1 năm": 0, "1-5 năm": 0, "5-10 năm": 0, ">10 năm": 0}
    for ngay_vao, in db.query(Employee.ngay_vao_lam).filter(
        Employee.trang_thai == "dang_lam",
        Employee.ngay_vao_lam.isnot(None),
    ).all():
        if not ngay_vao:
            continue
        years = (today - ngay_vao).days / 365.25
        if years < 1: tenure_buckets["<1 năm"] += 1
        elif years < 5: tenure_buckets["1-5 năm"] += 1
        elif years < 10: tenure_buckets["5-10 năm"] += 1
        else: tenure_buckets[">10 năm"] += 1
    tenure_distribution = [{"name": k, "value": v} for k, v in tenure_buckets.items()]

    # ─── 5) Biến động YTD (tuyển mới + nghỉ việc từ đầu năm) ───
    new_hires_ytd = db.query(func.count(Employee.id)).filter(
        Employee.ngay_vao_lam >= today_year_start,
    ).scalar() or 0
    resigned_ytd = db.query(func.count(Employee.id)).filter(
        Employee.ngay_nghi_viec >= today_year_start,
    ).scalar() or 0
    avg_headcount = total + (resigned_ytd / 2)  # ước lượng
    turnover_pct = round(resigned_ytd * 100 / avg_headcount, 1) if avg_headcount > 0 else 0

    # ─── 6) Cảnh báo sắp tới ───
    # Sinh nhật trong 30 ngày
    bday_30 = 0
    for ngay_sinh, in db.query(Employee.ngay_sinh).filter(
        Employee.trang_thai == "dang_lam",
        Employee.ngay_sinh.isnot(None),
    ).all():
        if not ngay_sinh:
            continue
        try:
            bday_this_year = date(today.year, ngay_sinh.month, ngay_sinh.day)
            if bday_this_year < today:
                bday_this_year = date(today.year + 1, ngay_sinh.month, ngay_sinh.day)
            if (bday_this_year - today).days <= 30:
                bday_30 += 1
        except ValueError:  # 29/2
            continue

    # HĐLĐ sắp hết hạn trong 60 ngày
    contracts_expiring_60 = db.query(func.count(LaborContract.id)).filter(
        LaborContract.trang_thai == "hieu_luc",
        LaborContract.ngay_het_han <= days_60,
        LaborContract.ngay_het_han >= today,
    ).scalar() or 0

    # NV chưa cấp tài khoản
    no_account = db.query(func.count(Employee.id)).filter(
        Employee.trang_thai == "dang_lam",
        Employee.user_id.is_(None),
    ).scalar() or 0

    # NV thiếu thông tin (no CCCD or no SĐT or no BHXH)
    missing_info = db.query(func.count(Employee.id)).filter(
        Employee.trang_thai == "dang_lam",
        or_(
            Employee.cccd.is_(None),
            Employee.so_dien_thoai.is_(None),
            Employee.so_so_bhxh.is_(None),
        ),
    ).scalar() or 0

    # ─── Phase 1 alerts: Khám SK + ATLĐ + TNLĐ ───
    from app.models.hr import (
        HealthCheck, SafetyTrainingParticipant, WorkAccident,
    )
    # Khám SK quá hạn (NV có ngày khám tiếp theo < today)
    latest_health = (
        db.query(
            HealthCheck.employee_id,
            func.max(HealthCheck.ngay_kham_tiep_theo).label("next_due"),
        ).filter(HealthCheck.ngay_kham_tiep_theo.isnot(None))
        .group_by(HealthCheck.employee_id).subquery()
    )
    health_overdue = db.query(func.count(latest_health.c.employee_id)).filter(
        latest_health.c.next_due < today,
    ).scalar() or 0

    # Chứng chỉ ATVSLĐ sắp hết hạn (60 ngày)
    cert_expiring_60d = db.query(func.count(SafetyTrainingParticipant.id)).filter(
        SafetyTrainingParticipant.han_chung_chi.isnot(None),
        SafetyTrainingParticipant.han_chung_chi >= today,
        SafetyTrainingParticipant.han_chung_chi <= days_60,
    ).scalar() or 0

    # TNLĐ nặng/tử vong chưa báo Sở LĐ
    tnld_unreported = db.query(func.count(WorkAccident.id)).filter(
        WorkAccident.muc_do.in_(["nang", "tu_vong"]),
        WorkAccident.da_bao_cao_so_lao_dong == False,
    ).scalar() or 0

    return {
        "summary": {
            "total": total,
            "dang_lam": by_status.get("dang_lam", 0),
            "tam_nghi": by_status.get("tam_nghi", 0),
            "da_nghi": by_status.get("da_nghi", 0),
            "new_hires_ytd": new_hires_ytd,
            "resigned_ytd": resigned_ytd,
            "turnover_pct": turnover_pct,
        },
        "by_gender": [
            {"name": k or "Không rõ", "value": v} for k, v in by_gender.items()
        ],
        "by_phap_nhan": by_phap_nhan,
        "by_bo_phan": by_bo_phan,
        "age_distribution": age_distribution,
        "tenure_distribution": tenure_distribution,
        "alerts": {
            "birthdays_30d": bday_30,
            "contracts_expiring_60d": contracts_expiring_60,
            "no_account": no_account,
            "missing_info": missing_info,
            # Phase 1 alerts:
            "health_overdue": health_overdue,
            "cert_expiring_60d": cert_expiring_60d,
            "tnld_unreported": tnld_unreported,
        },
    }


# --- Attendance Logs ---
@router.get("/attendance", response_model=List[schemas.AttendanceLog])
def list_attendance(
    employee_id: Optional[int] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(AttendanceLog)
    if employee_id:
        q = q.filter(AttendanceLog.employee_id == employee_id)
    if from_date:
        q = q.filter(AttendanceLog.ngay >= from_date)
    if to_date:
        q = q.filter(AttendanceLog.ngay <= to_date)
    return q.order_by(AttendanceLog.ngay.desc()).all()

@router.post("/attendance/bulk")
def bulk_create_attendance(logs: List[schemas.AttendanceLogCreate], db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    """Dùng để import dữ liệu từ máy chấm công — chỉ HR/Admin."""
    for log in logs:
        # Check if exists
        existing = db.query(AttendanceLog).filter(
            AttendanceLog.employee_id == log.employee_id,
            AttendanceLog.ngay == log.ngay
        ).first()
        if existing:
            for k, v in log.model_dump().items():
                setattr(existing, k, v)
        else:
            db.add(AttendanceLog(**log.model_dump()))
    db.commit()
    return {"ok": True, "count": len(logs)}

@router.post("/attendance/import")
def import_attendance(rows: List[dict], db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    saved = 0
    errors = []
    for idx, row in enumerate(rows, start=1):
        ma_nv = _attendance_value(row, "ma_nv", "MA_NV", "Mã Nhân Viên", "Ma Nhan Vien", "ma_van_tay")
        ma_nv = str(ma_nv).strip() if ma_nv is not None else None
        emp = db.query(Employee).filter(
            or_(Employee.ma_nv == ma_nv, Employee.ma_van_tay == ma_nv)
        ).first() if ma_nv else None
        if not emp:
            errors.append({"row": idx, "ma_nv": ma_nv, "error": "Khong tim thay ma NV"})
            continue

        ngay_raw = _attendance_value(row, "ngay", "Ngày")
        try:
            ngay = _parse_attendance_date(ngay_raw)
        except Exception:
            errors.append({"row": idx, "ngay": ngay_raw, "error": "Ngay khong hop le"})
            continue

        log = db.query(AttendanceLog).filter(
            AttendanceLog.employee_id == emp.id,
            AttendanceLog.ngay == ngay,
        ).first()
        if not log:
            log = AttendanceLog(employee_id=emp.id, ngay=ngay)
            db.add(log)

        gio_vao = _parse_attendance_datetime(_attendance_value(row, "gio_vao", "Giờ vào", "Gio vao"), ngay)
        gio_ra = _parse_attendance_datetime(_attendance_value(row, "gio_ra", "Giờ ra", "Gio ra"), ngay)
        if gio_vao:
            log.gio_vao = gio_vao
        if gio_ra:
            log.gio_ra = gio_ra

        numeric_fields = {
            "so_cong": ("so_cong", "Công", "Cong"),
            "so_gio_ot": ("so_gio_ot", "Tăng ca", "Tang ca"),
            "tong_gio_thuc": ("tong_gio_thuc", "Tổng giờ", "Tong gio"),
        }
        for field, aliases in numeric_fields.items():
            value = _attendance_value(row, *aliases)
            if value not in (None, ""):
                setattr(log, field, Decimal(str(value).replace(",", ".")))

        log.trang_thai = _attendance_value(row, "trang_thai", "Trạng thái") or log.trang_thai or "hop_le"
        log.ghi_chu = _attendance_value(row, "ghi_chu", "Ghi chú") or log.ghi_chu
        saved += 1

    if errors:
        db.rollback()
        raise HTTPException(400, {"message": "Import cham cong co loi", "errors": errors})
    db.commit()
    return {"ok": True, "count": saved}

# --- Leave Requests ---
# NOTE: Sprint C — chuyển toàn bộ logic leave-requests sang routers/hr_workflow.py
# (workflow đơn từ thống nhất với 5 loại đơn + security hardening).
# Các endpoint cũ /leave-requests đã được xoá khỏi đây để tránh route conflict.

# --- Payroll Config ---
@router.get("/payroll-configs", response_model=List[schemas.PayrollConfig])
def list_payroll_configs(
    loai: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """List config lương. Có thể filter theo loai:
    - 'san_pham' — Bảng đơn giá theo mã hàng (Điều 6 quy chế)
    - 'gio_quy_doi' — Bảng quy đổi giờ → công (Điều 9 quy chế, Table 5)
    - 'min_wage' — Lương tối thiểu vùng I-IV (NĐ 74/2024)
    - 'config' — Config chung (giờ/ngày chuẩn, vùng áp dụng, hệ số thử việc)
    - 'so_lop_giay' — Hệ số máy sóng theo số lớp giấy (đã có sẵn)
    """
    q = db.query(PayrollConfig)
    if loai:
        q = q.filter(PayrollConfig.loai == loai)
    return q.all()

@router.post("/payroll-configs", response_model=schemas.PayrollConfig)
def create_payroll_config(body: schemas.PayrollConfigCreate, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    db_cfg = db.query(PayrollConfig).filter(PayrollConfig.ma_hang == body.ma_hang).first()
    if db_cfg:
        for k, v in body.model_dump().items():
            setattr(db_cfg, k, v)
    else:
        db_cfg = PayrollConfig(**body.model_dump())
        db.add(db_cfg)
    db.commit()
    db.refresh(db_cfg)
    return db_cfg

@router.put("/payroll-configs/{id}", response_model=schemas.PayrollConfig)
def update_payroll_config(id: int, body: schemas.PayrollConfigCreate, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    db_cfg = db.get(PayrollConfig, id)
    if not db_cfg:
        raise HTTPException(404, "Khong tim thay cau hinh luong")
    for k, v in body.model_dump().items():
        setattr(db_cfg, k, v)
    db.commit()
    db.refresh(db_cfg)
    return db_cfg

@router.post("/payroll-configs/bulk")
def bulk_create_payroll_configs(body: schemas.PayrollConfigBulkCreate, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    created = 0
    updated = 0
    for item in body.items:
        cfg = db.query(PayrollConfig).filter(PayrollConfig.ma_hang == item.ma_hang).first()
        if cfg:
            for k, v in item.model_dump().items():
                setattr(cfg, k, v)
            updated += 1
        else:
            db.add(PayrollConfig(**item.model_dump()))
            created += 1
    db.commit()
    return {"ok": True, "created": created, "updated": updated}

@router.get("/payroll-holidays", response_model=List[schemas.PayrollHoliday])
def list_payroll_holidays(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(PayrollHoliday)
    if from_date:
        q = q.filter(PayrollHoliday.ngay >= from_date)
    if to_date:
        q = q.filter(PayrollHoliday.ngay <= to_date)
    return q.order_by(PayrollHoliday.ngay).all()

@router.post("/payroll-holidays", response_model=schemas.PayrollHoliday)
def create_payroll_holiday(body: schemas.PayrollHolidayCreate, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    holiday = db.query(PayrollHoliday).filter(PayrollHoliday.ngay == body.ngay).first()
    if holiday:
        for k, v in body.model_dump().items():
            setattr(holiday, k, v)
    else:
        holiday = PayrollHoliday(**body.model_dump())
        db.add(holiday)
    db.commit()
    db.refresh(holiday)
    return holiday

@router.delete("/payroll-holidays/{id}")
def delete_payroll_holiday(id: int, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    holiday = db.get(PayrollHoliday, id)
    if not holiday:
        raise HTTPException(404, "Khong tim thay ngay le")
    db.delete(holiday)
    db.commit()
    return {"ok": True}

@router.post("/contracts/import-allowances")
def import_contract_allowances(rows: List[dict], db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    updated = 0
    created = 0
    errors = []
    for idx, row in enumerate(rows, start=1):
        ma_nv = _attendance_value(row, "ma_nv", "Mã NV", "Mã Nhân Viên", "Ma Nhan Vien")
        ma_nv = str(ma_nv).strip() if ma_nv is not None else None
        emp = db.query(Employee).filter(Employee.ma_nv == ma_nv).first() if ma_nv else None
        if not emp:
            errors.append({"row": idx, "ma_nv": ma_nv, "error": "Khong tim thay ma NV"})
            continue

        contract = sorted(emp.contracts, key=lambda x: x.ngay_hieu_luc or x.ngay_ky, reverse=True)
        current = contract[0] if contract else None
        if not current:
            current = LaborContract(
                employee_id=emp.id,
                so_hop_dong=f"AUTO-{emp.ma_nv}",
                loai_hop_dong="khong_thoi_han",
                ngay_ky=date.today(),
                ngay_hieu_luc=emp.ngay_vao_lam or date.today(),
                trang_thai="hieu_luc",
            )
            db.add(current)
            created += 1
        else:
            updated += 1

        base_salary = _attendance_value(row, "luong_co_ban", "Lương cơ bản", "Luong co ban")
        if base_salary not in (None, ""):
            current.luong_co_ban = Decimal(str(base_salary).replace(",", "."))
        current.phu_cap_chuyen_can = _decimal_value(row, "phu_cap_chuyen_can", "Chuyên cần", "Chuyen can")
        current.phu_cap_trach_nhiem = _decimal_value(row, "phu_cap_trach_nhiem", "Trách nhiệm", "Trach nhiem")
        current.phu_cap_nha_o_com = _decimal_value(row, "phu_cap_nha_o_com", "Nhà ở/Cơm", "Nha o/Com")
        current.phu_cap_dien_thoai = _decimal_value(row, "phu_cap_dien_thoai", "Điện thoại", "Dien thoai")
        current.phu_cap_khac = _decimal_value(row, "phu_cap_khac", "Hỗ trợ khác", "Ho tro khac")
        current.phu_cap = _contract_allowance_total(current)

    if errors:
        db.rollback()
        raise HTTPException(400, {"message": "Import phu cap co loi", "errors": errors})
    db.commit()
    return {"ok": True, "created": created, "updated": updated}

# ─── Rate limit issue-account: max 20 NV/giờ/HR ───
# Chống compromised HR bulk-reset tài khoản. In-memory deque cho mỗi user_id.
#
# ⚠ WARNING: Implementation in-memory CHỈ HOẠT ĐỘNG ĐÚNG với 1 worker uvicorn.
# Khi triển khai production với `uvicorn --workers N`, mỗi worker có dict riêng
# → effective limit = 20 × N. Để fix triệt để cần chuyển sang Redis:
#   `await redis.incr(f"rl:issue:{user_id}"); await redis.expire(..., 3600)`
# Hoặc query bảng audit_log đếm số lần POST issue-account trong 1h qua.
# Hiện tại Nam Phương deploy 1-worker nên acceptable.
from collections import deque
from datetime import datetime as _dt, timedelta as _td
_issue_account_history: dict[int, deque] = {}
_ISSUE_ACCOUNT_LIMIT_PER_HOUR = 20


def _check_issue_account_rate_limit(user_id: int) -> None:
    """Throw HTTPException 429 nếu user đã cấp ≥20 tài khoản trong 1 giờ qua."""
    now = _dt.now()
    cutoff = now - _td(hours=1)
    dq = _issue_account_history.setdefault(user_id, deque(maxlen=100))
    # Bỏ các record cũ
    while dq and dq[0] < cutoff:
        dq.popleft()
    if len(dq) >= _ISSUE_ACCOUNT_LIMIT_PER_HOUR:
        raise HTTPException(
            status_code=429,
            detail=f"Quá tải: chỉ được cấp tối đa {_ISSUE_ACCOUNT_LIMIT_PER_HOUR} tài khoản/giờ. Hãy đợi lại sau.",
        )
    dq.append(now)


@router.post("/employees/{id}/issue-account")
def issue_employee_account(id: int, db: Session = Depends(get_db), current_user: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    from app.models.auth import Role, User
    import bcrypt, secrets, string

    # Rate limit trước khi làm gì cả — chống compromised HR bulk-reset
    _check_issue_account_rate_limit(current_user.id)

    emp = db.get(Employee, id)
    if not emp: raise HTTPException(404, "Không tìm thấy nhân viên")
    if emp.user_id: raise HTTPException(400, "Nhân viên đã có tài khoản")

    # Random password 12 ký tự (chữ + số) — KHÔNG còn hardcode "123456"
    # Đảm bảo có ít nhất 1 chữ + 1 số (phòng trường hợp gen ngẫu nhiên ra all-letter/all-digit)
    # Trả về duy nhất 1 lần qua response cho HR đọc cho NV; lưu DB chỉ là hash
    while True:
        alphabet = string.ascii_letters + string.digits
        temp_password = ''.join(secrets.choice(alphabet) for _ in range(12))
        if any(c.isalpha() for c in temp_password) and any(c.isdigit() for c in temp_password):
            break
    hashed = bcrypt.hashpw(temp_password.encode(), bcrypt.gensalt()).decode()

    # Tìm vai trò theo ma trận: Phòng ban + Chức vụ
    # Ma trận dùng mã role có nghĩa (vd SALE_ADMIN), KHÔNG dùng ma_bo_phan/ma_chuc_vu nội bộ
    role = None
    if emp.bo_phan and emp.chuc_vu:
        role_code = f"{emp.bo_phan.ma_bo_phan}_{emp.chuc_vu.ma_chuc_vu}"
        role = db.query(Role).filter(Role.ma_vai_tro == role_code).first()

    # Fallback an toàn: tìm NHAN_VIEN, rồi dừng — KHÔNG fallback .first() vì sẽ gán role ngẫu nhiên
    if not role:
        role = db.query(Role).filter(Role.ma_vai_tro == "NHAN_VIEN").first()
    if not role:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Không tìm thấy role phù hợp cho nhân viên {emp.ma_nv}. "
                "Kiểm tra ma trận phòng ban/chức vụ hoặc đảm bảo role NHAN_VIEN tồn tại."
            ),
        )

    new_user = User(
        username=emp.ma_nv,
        ho_ten=emp.ho_ten,
        password_hash=hashed,
        role_id=role.id if role else 1,
        trang_thai=True,
        must_change_password=True,  # Force đổi pass lần đầu
    )
    db.add(new_user)
    db.flush()

    emp.user_id = new_user.id
    db.commit()
    logger.info("HR issue account emp=%s username=%s by user=%s", emp.id, emp.ma_nv, current_user.id)
    # Mật khẩu random trả VÀO RESPONSE duy nhất 1 lần — HR phải đọc/copy cho NV
    return {
        "status": "success",
        "username": emp.ma_nv,
        "temp_password": temp_password,
        "must_change_password": True,
        "message": "⚠ Mật khẩu chỉ hiển thị 1 lần. Vui lòng đưa cho NV và họ phải đổi khi đăng nhập lần đầu.",
    }

@router.post("/employees/{id}/toggle-account-status")
def toggle_account_status(id: int, db: Session = Depends(get_db), _: User = Depends(require_roles("ADMIN", "NHAN_SU"))):
    from app.models.auth import User
    emp = db.get(Employee, id)
    if not emp or not emp.user_id:
        raise HTTPException(400, "Nhân viên chưa có tài khoản")

    user = db.get(User, emp.user_id)
    user.trang_thai = not user.trang_thai
    db.commit()
    return {"status": "success", "new_status": user.trang_thai}


class LinkUserRequest(BaseModel):
    user_id: int | None


@router.patch("/employees/{id}/link-user")
def link_user_to_employee(
    id: int,
    body: LinkUserRequest,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    """Gán tài khoản người dùng có sẵn vào hồ sơ nhân viên (hoặc gỡ liên kết khi user_id=null)."""
    from app.models.auth import User as AuthUser
    emp = db.get(Employee, id)
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")
    if body.user_id is not None:
        user = db.get(AuthUser, body.user_id)
        if not user:
            raise HTTPException(404, "Không tìm thấy tài khoản")
    emp.user_id = body.user_id
    db.commit()
    return {"ok": True, "user_id": emp.user_id}


class SyncSaleAccountsRequest(BaseModel):
    employee_ids: list[int]


@router.post("/employees/sync-sale-accounts")
def sync_sale_accounts(
    data: SyncSaleAccountsRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    """Tạo hàng loạt tài khoản Sale Admin cho danh sách NV chọn. Password cố định 123456."""
    from app.models.auth import Role, User as AuthUser
    import bcrypt

    role = db.query(Role).filter(Role.ma_vai_tro == "SALE_ADMIN").first()
    if not role:
        raise HTTPException(400, "Không tìm thấy role SALE_ADMIN trong hệ thống")

    hashed_pw = bcrypt.hashpw(b"123456", bcrypt.gensalt()).decode()

    created, skipped, errors = [], [], []

    for emp_id in data.employee_ids:
        emp = db.get(Employee, emp_id)
        if not emp:
            errors.append({"id": emp_id, "ly_do": "Không tìm thấy nhân viên"})
            continue

        if emp.user_id:
            skipped.append({"id": emp_id, "ma_nv": emp.ma_nv, "ho_ten": emp.ho_ten, "ly_do": "Đã có tài khoản"})
            continue

        if db.query(AuthUser).filter(AuthUser.username == emp.ma_nv).first():
            skipped.append({"id": emp_id, "ma_nv": emp.ma_nv, "ho_ten": emp.ho_ten, "ly_do": "Username đã tồn tại"})
            continue

        try:
            new_user = AuthUser(
                username=emp.ma_nv,
                ho_ten=emp.ho_ten,
                password_hash=hashed_pw,
                role_id=role.id,
                trang_thai=True,
                must_change_password=False,
            )
            db.add(new_user)
            db.flush()
            emp.user_id = new_user.id
            db.commit()
            created.append({"id": emp_id, "ma_nv": emp.ma_nv, "ho_ten": emp.ho_ten})
        except Exception as e:
            db.rollback()
            errors.append({"id": emp_id, "ma_nv": getattr(emp, "ma_nv", "?"), "ly_do": str(e)})

    logger.info("sync-sale-accounts created=%d skipped=%d errors=%d by=%s",
                len(created), len(skipped), len(errors), current_user.id)
    return {"created": created, "skipped": skipped, "errors": errors}


# --- Payroll History Import ---

HR_PAYROLL_IMPORT_FIELDS = [
    ImportField("ma_nv", "Ma NV", required=True),
    ImportField("thang", "Thang", required=True, help_text="1-12"),
    ImportField("nam", "Nam", required=True, help_text="VD: 2025"),
    ImportField("luong_co_ban", "Luong co ban"),
    ImportField("luong_san_pham", "Luong san pham"),
    ImportField("phu_cap", "Phu cap"),
    ImportField("phu_cap_chuyen_can", "Chuyen can"),
    ImportField("phu_cap_trach_nhiem", "Trach nhiem"),
    ImportField("tong_thu_nhap", "Tong thu nhap"),
    ImportField("thuong", "Thuong"),
    ImportField("bao_hiem", "Bao hiem"),
    ImportField("thue_tncn", "Thue TNCN"),
    ImportField("tam_ung", "Tam ung"),
    ImportField("luong_thuc_nhan", "Luong thuc nhan"),
    ImportField("ghi_chu", "Ghi chu"),
]


@router.get("/payroll/import-history-template")
def payroll_import_template(_: User = Depends(get_current_user)):
    return build_template_response("mau_import_lich_su_luong.xlsx", HR_PAYROLL_IMPORT_FIELDS)


def _parse_money(value: object) -> Decimal:
    """Parse a money value from Excel cell: strip commas, convert to Decimal. Returns 0 if empty."""
    if value is None or str(value).strip() == "":
        return Decimal("0")
    text = str(value).replace(",", "").strip()
    try:
        return Decimal(text)
    except Exception:
        return Decimal("0")


@router.post("/payroll/import-history")
async def import_payroll_history(
    file: UploadFile = File(...),
    commit: bool = True,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    """
    Import lich su luong (PayrollRun) tu file Excel.
    Upsert theo khoa (employee_id, thang, nam).
    Tra ve: {"created": N, "updated": N, "errors": [...]}
    """
    from io import BytesIO
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Chi chap nhan file Excel .xlsx/.xls")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File rong")

    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Khong doc duoc file Excel: {exc}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # Build header map from first row
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="File khong co du lieu")

    # Normalize header labels to field names using HR_PAYROLL_IMPORT_FIELDS
    label_to_field: dict[str, str] = {
        field.label.strip().lower(): field.name
        for field in HR_PAYROLL_IMPORT_FIELDS
    }
    col_index: dict[str, int] = {}  # field_name -> column index
    for col_i, header_val in enumerate(header_row):
        if header_val is None:
            continue
        normalized = str(header_val).strip().lower()
        field_name = label_to_field.get(normalized)
        if field_name:
            col_index[field_name] = col_i

    # Validate required columns present
    required_fields = [f.name for f in HR_PAYROLL_IMPORT_FIELDS if f.required]
    missing_cols = [f for f in required_fields if f not in col_index]
    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail=f"Thieu cot bat buoc: {', '.join(missing_cols)}"
        )

    created_count = 0
    updated_count = 0
    errors: list[dict] = []

    # Collect valid rows to process (upsert) — only if no row-level errors when commit=True
    pending: list[tuple[object | None, dict]] = []  # (existing_obj_or_None, values_dict)

    for row_i, row in enumerate(rows_iter, start=2):
        # Skip completely blank rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def _cell(field_name: str):
            idx = col_index.get(field_name)
            if idx is None:
                return None
            return row[idx] if idx < len(row) else None

        ma_nv_raw = _cell("ma_nv")
        ma_nv = str(ma_nv_raw).strip() if ma_nv_raw is not None else None

        # Parse thang / nam
        try:
            thang_val = _cell("thang")
            thang = int(float(str(thang_val).replace(",", "").strip()))
            if not 1 <= thang <= 12:
                raise ValueError("Thang phai tu 1 den 12")
        except Exception as exc:
            errors.append({"row": row_i, "ma_nv": ma_nv, "error": f"Thang khong hop le: {exc}"})
            continue

        try:
            nam_val = _cell("nam")
            nam = int(float(str(nam_val).replace(",", "").strip()))
        except Exception as exc:
            errors.append({"row": row_i, "ma_nv": ma_nv, "error": f"Nam khong hop le: {exc}"})
            continue

        # Lookup employee
        if not ma_nv:
            errors.append({"row": row_i, "ma_nv": None, "error": "Ma NV trong"})
            continue

        emp = db.query(Employee).filter(Employee.ma_nv == ma_nv).first()
        if not emp:
            errors.append({"row": row_i, "ma_nv": ma_nv, "error": f"Khong tim thay NV ma_nv={ma_nv!r}"})
            continue

        # Build values dict for money fields
        money_map = {
            "luong_co_ban": "luong_co_ban",
            "luong_san_pham": "luong_san_pham",
            "phu_cap": "phu_cap",
            "phu_cap_chuyen_can": "phu_cap_chuyen_can",
            "phu_cap_trach_nhiem": "phu_cap_trach_nhiem",
            "tong_thu_nhap": "tong_thu_nhap",
            "thuong": "thuong",
            "bao_hiem": "bao_hiem",
            "thue_tncn": "thue_tncn",
            "tam_ung": "tam_ung",
            # "luong_thuc_nhan" in Excel maps to "thuc_linh" in model
            "luong_thuc_nhan": "thuc_linh",
        }

        values: dict = {
            "employee_id": emp.id,
            "thang": thang,
            "nam": nam,
        }
        for excel_field, model_field in money_map.items():
            values[model_field] = _parse_money(_cell(excel_field))

        ghi_chu_val = _cell("ghi_chu")
        if ghi_chu_val is not None and str(ghi_chu_val).strip():
            values["ghi_chu_import"] = str(ghi_chu_val).strip()

        # Upsert lookup
        existing = db.query(PayrollRun).filter(
            PayrollRun.employee_id == emp.id,
            PayrollRun.thang == thang,
            PayrollRun.nam == nam,
        ).first()

        pending.append((existing, values))
        if existing:
            updated_count += 1
        else:
            created_count += 1

    wb.close()

    if commit:
        for existing_obj, vals in pending:
            if existing_obj:
                for field, value in vals.items():
                    if hasattr(existing_obj, field):
                        setattr(existing_obj, field, value)
            else:
                # Only set fields that exist on the model
                safe_vals = {k: v for k, v in vals.items() if hasattr(PayrollRun, k)}
                db.add(PayrollRun(**safe_vals))
        db.commit()

    return {
        "created": created_count,
        "updated": updated_count,
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# Import nhân viên bulk từ Excel
# ---------------------------------------------------------------------------

@router.post("/employees/import")
async def import_employees(
    commit: bool = Query(default=True),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles("ADMIN", "NHAN_SU")),
):
    """
    Import nhân viên bulk từ file Excel (.xlsx/.xls).
    Upsert theo ma_nv: tồn tại → update, chưa có → create.
    Lookup Position theo ten_chuc_vu (ilike), Department theo ten_bo_phan (ilike).
    Trả về: {"created": N, "updated": N, "errors": [{"row": i, "error": "..."}]}
    """
    from io import BytesIO
    from datetime import datetime as _dt
    from decimal import Decimal as _Decimal, InvalidOperation
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Chi chap nhan file Excel .xlsx/.xls")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File rong")

    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Khong doc duoc file Excel: {exc}") from exc

    ws = wb.active
    rows_iter = iter(ws.rows)

    # Đọc header row — map label → field name
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="File khong co du lieu")

    label_to_field: dict[str, str] = {
        f.label.strip().lower(): f.name for f in HR_EMPLOYEE_IMPORT_FIELDS
    }

    col_index: dict[str, int] = {}  # field_name → column index (0-based)
    for idx, cell in enumerate(header_cells):
        raw_label = str(cell.value or "").strip().lower()
        if raw_label in label_to_field:
            col_index[label_to_field[raw_label]] = idx

    missing = [f.label for f in HR_EMPLOYEE_IMPORT_FIELDS if f.required and f.name not in col_index]
    if missing:
        raise HTTPException(status_code=400, detail=f"Thieu cot bat buoc: {', '.join(missing)}")

    def _cell_val(row_cells, field_name: str):
        idx = col_index.get(field_name)
        if idx is None or idx >= len(row_cells):
            return None
        return row_cells[idx].value

    def _parse_date_val(value) -> date | None:
        if value is None:
            return None
        if isinstance(value, _dt):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return _dt.strptime(text, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Khong nhan dang duoc ngay: {text!r}")

    def _parse_decimal_val(value) -> _Decimal | None:
        if value is None:
            return None
        text = str(value).replace(",", ".").strip()
        if not text:
            return None
        try:
            return _Decimal(text)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError("Phai la so thap phan") from exc

    def _parse_text_val(value) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    # Caches để tránh query lặp trong cùng file
    position_cache: dict[str, int | None] = {}
    department_cache: dict[str, int | None] = {}

    n_created = 0
    n_updated = 0
    errors: list[dict] = []
    objects_to_save: list[tuple] = []  # (existing | None, values_dict)

    for row_idx, row_cells in enumerate(rows_iter, start=2):
        # Bỏ qua dòng trắng
        if all(c.value is None or str(c.value).strip() == "" for c in row_cells):
            continue

        row_errors: list[str] = []
        values: dict = {}

        # ma_nv (bắt buộc)
        ma_nv = _parse_text_val(_cell_val(row_cells, "ma_nv"))
        if not ma_nv:
            row_errors.append("Ma NV: bat buoc")
        else:
            values["ma_nv"] = ma_nv

        # ho_ten (bắt buộc)
        ho_ten = _parse_text_val(_cell_val(row_cells, "ho_ten"))
        if not ho_ten:
            row_errors.append("Ho ten: bat buoc")
        else:
            values["ho_ten"] = ho_ten

        # ngay_sinh
        try:
            v = _parse_date_val(_cell_val(row_cells, "ngay_sinh"))
            if v is not None:
                values["ngay_sinh"] = v
        except ValueError as exc:
            row_errors.append(f"Ngay sinh: {exc}")

        # gioi_tinh
        v = _parse_text_val(_cell_val(row_cells, "gioi_tinh"))
        if v is not None:
            values["gioi_tinh"] = v

        # so_dien_thoai
        v = _parse_text_val(_cell_val(row_cells, "so_dien_thoai"))
        if v is not None:
            values["so_dien_thoai"] = v

        # email
        v = _parse_text_val(_cell_val(row_cells, "email"))
        if v is not None:
            values["email"] = v

        # ngay_vao_lam
        try:
            v = _parse_date_val(_cell_val(row_cells, "ngay_vao_lam"))
            if v is not None:
                values["ngay_vao_lam"] = v
        except ValueError as exc:
            row_errors.append(f"Ngay vao lam: {exc}")

        # ngay_nghi_viec
        try:
            v = _parse_date_val(_cell_val(row_cells, "ngay_nghi_viec"))
            if v is not None:
                values["ngay_nghi_viec"] = v
        except ValueError as exc:
            row_errors.append(f"Ngay nghi viec: {exc}")

        # trang_thai
        v = _parse_text_val(_cell_val(row_cells, "trang_thai"))
        if v is not None:
            valid_statuses = {"dang_lam", "tam_nghi", "da_nghi"}
            if v not in valid_statuses:
                row_errors.append(
                    f"Trang thai: phai la {'/'.join(sorted(valid_statuses))}, nhan duoc '{v}'"
                )
            else:
                values["trang_thai"] = v

        # ten_chuc_vu → chuc_vu_id (lookup ilike)
        ten_chuc_vu = _parse_text_val(_cell_val(row_cells, "ten_chuc_vu"))
        if ten_chuc_vu:
            cache_key = ten_chuc_vu.lower()
            if cache_key not in position_cache:
                pos = db.query(Position).filter(
                    Position.ten_chuc_vu.ilike(ten_chuc_vu)
                ).first()
                position_cache[cache_key] = pos.id if pos else None
            chuc_vu_id = position_cache[cache_key]
            if chuc_vu_id is None:
                row_errors.append(f"Chuc vu '{ten_chuc_vu}': khong tim thay trong he thong")
            else:
                values["chuc_vu_id"] = chuc_vu_id

        # ten_bo_phan → bo_phan_id (lookup ilike)
        ten_bo_phan = _parse_text_val(_cell_val(row_cells, "ten_bo_phan"))
        if ten_bo_phan:
            cache_key = ten_bo_phan.lower()
            if cache_key not in department_cache:
                dept = db.query(Department).filter(
                    Department.ten_bo_phan.ilike(ten_bo_phan)
                ).first()
                department_cache[cache_key] = dept.id if dept else None
            bo_phan_id = department_cache[cache_key]
            if bo_phan_id is None:
                row_errors.append(f"Bo phan '{ten_bo_phan}': khong tim thay trong he thong")
            else:
                values["bo_phan_id"] = bo_phan_id

        # he_so_ca_nhan
        try:
            v = _parse_decimal_val(_cell_val(row_cells, "he_so_ca_nhan"))
            if v is not None:
                values["he_so_ca_nhan"] = v
        except ValueError as exc:
            row_errors.append(f"He so ca nhan: {exc}")

        if row_errors:
            errors.append({"row": row_idx, "error": "; ".join(row_errors)})
            continue

        # Upsert theo ma_nv
        existing = db.query(Employee).filter(
            Employee.ma_nv == values.get("ma_nv")
        ).first()

        if existing:
            objects_to_save.append((existing, values))
            n_updated += 1
        else:
            objects_to_save.append((None, values))
            n_created += 1

    wb.close()

    if commit:
        for existing_emp, vals in objects_to_save:
            if existing_emp:
                for field_name, field_val in vals.items():
                    setattr(existing_emp, field_name, field_val)
            else:
                db.add(Employee(**vals))
        db.commit()

    return {"created": n_created, "updated": n_updated, "errors": errors}
