import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user, require_any_permission
from app.models.auth import User
from app.models.accounting import ChartOfAccounts

router = APIRouter(
    prefix="/api/chart-of-accounts",
    dependencies=[Depends(require_any_permission("accounting.general_ledger", "accounting.manage"))],
    tags=["chart-of-accounts"],
)

# Nhãn tiếng Việt cho loại tài khoản — dùng khi xuất Excel.
LOAI_TK_LABELS = {
    "tai_san": "Tài sản",
    "no_phai_tra": "Nợ phải trả",
    "von_chu_so_huu": "Vốn chủ sở hữu",
    "doanh_thu": "Doanh thu",
    "chi_phi": "Chi phí",
}


# ─── Schemas ─────────────────────────────────────────────────────────────────

class ChartOfAccountsBase(BaseModel):
    so_tk: str
    ten_tk: str
    loai_tk: str
    cap: int = 1
    so_tk_cha: str | None = None
    trang_thai: bool = True
    theo_doi_doi_tuong: bool = False
    loai_doi_tuong: str | None = None


class ChartOfAccountsResponse(ChartOfAccountsBase):
    id: int

    class Config:
        from_attributes = True


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.get("", response_model=list[ChartOfAccountsResponse])
def list_chart_of_accounts(
    loai_tk: str | None = Query(default=None),
    q: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    query = db.query(ChartOfAccounts)
    if loai_tk:
        query = query.filter(ChartOfAccounts.loai_tk == loai_tk)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                ChartOfAccounts.so_tk.ilike(like),
                ChartOfAccounts.ten_tk.ilike(like),
            )
        )
    return query.order_by(ChartOfAccounts.so_tk).all()


# Đặt TRƯỚC route "/{id}" — nếu không, "export-excel" sẽ bị bắt làm path param id.
@router.get("/export-excel")
def export_chart_of_accounts_excel(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = db.query(ChartOfAccounts).order_by(ChartOfAccounts.so_tk).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "He thong tai khoan"

    headers = ["Số TK", "Tên tài khoản", "Loại", "Cấp", "TK cha", "Trạng thái"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, _title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for r in rows:
        ws.append([
            r.so_tk,
            r.ten_tk,
            LOAI_TK_LABELS.get(r.loai_tk, r.loai_tk),
            r.cap,
            r.so_tk_cha or "",
            "Đang dùng" if r.trang_thai else "Ngừng",
        ])

    widths = [14, 40, 16, 8, 12, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="he_thong_tai_khoan.xlsx"'},
    )


@router.get("/{id}", response_model=ChartOfAccountsResponse)
def get_chart_of_account(
    id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    obj = db.query(ChartOfAccounts).filter(ChartOfAccounts.id == id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản")
    return obj


@router.post("", response_model=ChartOfAccountsResponse, status_code=201)
def create_chart_of_account(
    data: ChartOfAccountsBase,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    existing = db.query(ChartOfAccounts).filter(ChartOfAccounts.so_tk == data.so_tk).first()
    if existing:
        raise HTTPException(status_code=400, detail="Mã tài khoản đã tồn tại")
    obj = ChartOfAccounts(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.put("/{id}", response_model=ChartOfAccountsResponse)
def update_chart_of_account(
    id: int,
    data: ChartOfAccountsBase,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    obj = db.query(ChartOfAccounts).filter(ChartOfAccounts.id == id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản")
    # Chỉ kiểm tra trùng khi so_tk thực sự đổi — tránh chính bản ghi này tự báo trùng.
    if data.so_tk != obj.so_tk:
        clash = (
            db.query(ChartOfAccounts)
            .filter(ChartOfAccounts.so_tk == data.so_tk, ChartOfAccounts.id != id)
            .first()
        )
        if clash:
            raise HTTPException(status_code=400, detail="Mã tài khoản đã tồn tại")
    for k, v in data.model_dump().items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/{id}")
def delete_chart_of_account(
    id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    obj = db.query(ChartOfAccounts).filter(ChartOfAccounts.id == id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản")
    child = (
        db.query(ChartOfAccounts)
        .filter(ChartOfAccounts.so_tk_cha == obj.so_tk, ChartOfAccounts.id != id)
        .first()
    )
    if child:
        raise HTTPException(
            status_code=400,
            detail="Không thể xóa tài khoản cha đang có tài khoản con",
        )
    db.delete(obj)
    db.commit()
    return {"ok": True}
