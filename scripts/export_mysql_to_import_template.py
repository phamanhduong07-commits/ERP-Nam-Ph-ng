"""
Script: Xuất dữ liệu MySQL → file Excel import tồn kho đầu kỳ cho ERP Nam Phương
Chạy trên máy có kết nối tới MySQL: python export_mysql_to_import_template.py
"""

import sys

# ── Cấu hình kết nối MySQL ────────────────────────────────────────────────────
MYSQL_HOST = "203.162.54.176"
MYSQL_PORT = 1441
MYSQL_USER = "duong"
MYSQL_PASS = "Namphuong123@"
MYSQL_DB   = ""          # ← điền tên database MySQL vào đây

# ── Cấu hình kho ──────────────────────────────────────────────────────────────
MA_KHO_ERP = "hoang_gia-GC"   # mã kho trong ERP (xem tại /api/warehouses)
LOAI_HANG  = "giay"           # giay | nvl | tp

# ── Query lấy tồn kho ─────────────────────────────────────────────────────────
# Sửa query này cho đúng với cấu trúc bảng MySQL của bạn
QUERY = """
SELECT
    ma_vt      AS ma_hang,
    ten_vt     AS ten_hang,
    dvt        AS don_vi,
    ton_kho    AS so_luong,
    gia_nhap   AS don_gia
FROM vat_tu
WHERE loai = 'giay_cuon'
  AND ton_kho > 0
ORDER BY ma_vt
"""

# Nếu không biết tên bảng, dùng query này để xem tất cả bảng:
QUERY_SHOW_TABLES = "SHOW TABLES"

# ─────────────────────────────────────────────────────────────────────────────

def main():
    try:
        import pymysql
    except ImportError:
        print("Cài pymysql: pip install pymysql openpyxl")
        sys.exit(1)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Cài openpyxl: pip install openpyxl")
        sys.exit(1)

    if not MYSQL_DB:
        print("❌ Chưa điền MYSQL_DB. Sửa biến MYSQL_DB ở đầu file.")
        print("\nCác database hiện có:")
        try:
            conn = pymysql.connect(host=MYSQL_HOST, port=MYSQL_PORT,
                                   user=MYSQL_USER, password=MYSQL_PASS,
                                   connect_timeout=10)
            cur = conn.cursor()
            cur.execute("SHOW DATABASES")
            for row in cur.fetchall():
                print(" -", row[0])
            conn.close()
        except Exception as e:
            print("Lỗi kết nối:", e)
        sys.exit(1)

    print(f"Kết nối MySQL {MYSQL_HOST}:{MYSQL_PORT} DB={MYSQL_DB} ...")
    try:
        conn = pymysql.connect(host=MYSQL_HOST, port=MYSQL_PORT,
                               user=MYSQL_USER, password=MYSQL_PASS,
                               database=MYSQL_DB, connect_timeout=10,
                               charset="utf8mb4")
    except Exception as e:
        print("❌ Kết nối thất bại:", e)
        sys.exit(1)

    cur = conn.cursor(pymysql.cursors.DictCursor)

    # Xem bảng nếu cần debug
    cur.execute(QUERY_SHOW_TABLES)
    tables = [list(r.values())[0] for r in cur.fetchall()]
    print("Bảng trong DB:", tables)

    print("\nThực hiện query tồn kho...")
    try:
        cur.execute(QUERY)
        rows = cur.fetchall()
    except Exception as e:
        print("❌ Query lỗi:", e)
        print("Sửa QUERY trong script cho đúng với cấu trúc bảng MySQL.")
        conn.close()
        sys.exit(1)

    conn.close()
    print(f"→ Lấy được {len(rows)} dòng")

    # ── Tạo Excel ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Du lieu import"

    HEADERS = ["Ma hang", "So luong", "Don gia", "Don vi", "Ghi chu (ten hang)"]
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=str(row.get("ma_hang") or ""))
        ws.cell(row=i, column=2, value=float(row.get("so_luong") or 0))
        ws.cell(row=i, column=3, value=float(row.get("don_gia") or 0))
        ws.cell(row=i, column=4, value=str(row.get("don_vi") or "Kg"))
        ws.cell(row=i, column=5, value=str(row.get("ten_hang") or ""))

    # Số format cho cột số
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.##"

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 3
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

    # Sheet hướng dẫn
    guide = wb.create_sheet("Huong dan")
    guide.append(["Thông tin", "Giá trị"])
    guide.append(["Kho ERP (Ma kho)", MA_KHO_ERP])
    guide.append(["Loai hang", LOAI_HANG])
    guide.append(["", ""])
    guide.append(["Cột", "Bắt buộc", "Ghi chú"])
    guide.append(["Ma hang",  "Có",    "Ma chinh của vật tư trong ERP (trùng với ma_vt MySQL)"])
    guide.append(["So luong", "Có",    "Tồn lượng đầu kỳ (kg)"])
    guide.append(["Don gia",  "Không", "Đơn giá bình quân (để trống nếu chưa có)"])
    guide.append(["Don vi",   "Không", "Đơn vị tính — để trống sẽ lấy từ danh mục"])
    guide.append(["", ""])
    guide.append(["Lưu ý", "Cột 'Ghi chu' KHÔNG được import — chỉ để tham khảo tên hàng"])

    out = f"ton_kho_dau_ky_{MA_KHO_ERP}.xlsx"
    wb.save(out)
    print(f"\n✅ Đã tạo file: {out}")
    print(f"   {len(rows)} dòng vật tư giấy cuộn")
    print(f"\nBước tiếp theo:")
    print(f"  1. Kiểm tra cột 'Ma hang' — phải khớp với ma_chinh trong ERP")
    print(f"  2. Điều chỉnh 'So luong' và 'Don gia' nếu cần")
    print(f"  3. Xóa cột 'Ghi chu' trước khi import")
    print(f"  4. Upload file vào ERP: Kho → Tồn kho → Import → Chọn kho '{MA_KHO_ERP}'")


if __name__ == "__main__":
    main()
